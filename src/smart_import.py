import json
import logging
import fnmatch
import re
import traceback
from dataclasses import dataclass
from pathlib import Path

import pandas as pd

from src.imports import clean_identifier, clean_store_number


APP_DIR = Path(__file__).resolve().parents[1]
MAPPING_PATH = APP_DIR / "data" / "import_mappings.json"
LOGGER = logging.getLogger("field_planner.smart_import")

US_STATE_NAMES = {
    "alabama": "AL",
    "alaska": "AK",
    "arizona": "AZ",
    "arkansas": "AR",
    "california": "CA",
    "colorado": "CO",
    "connecticut": "CT",
    "delaware": "DE",
    "florida": "FL",
    "georgia": "GA",
    "hawaii": "HI",
    "idaho": "ID",
    "illinois": "IL",
    "indiana": "IN",
    "iowa": "IA",
    "kansas": "KS",
    "kentucky": "KY",
    "louisiana": "LA",
    "maine": "ME",
    "maryland": "MD",
    "massachusetts": "MA",
    "michigan": "MI",
    "minnesota": "MN",
    "mississippi": "MS",
    "missouri": "MO",
    "montana": "MT",
    "nebraska": "NE",
    "nevada": "NV",
    "new_hampshire": "NH",
    "new_jersey": "NJ",
    "new_mexico": "NM",
    "new_york": "NY",
    "north_carolina": "NC",
    "north_dakota": "ND",
    "ohio": "OH",
    "oklahoma": "OK",
    "oregon": "OR",
    "pennsylvania": "PA",
    "rhode_island": "RI",
    "south_carolina": "SC",
    "south_dakota": "SD",
    "tennessee": "TN",
    "texas": "TX",
    "utah": "UT",
    "vermont": "VT",
    "virginia": "VA",
    "washington": "WA",
    "west_virginia": "WV",
    "wisconsin": "WI",
    "wyoming": "WY",
}
US_STATES = set(US_STATE_NAMES.values())

FIELD_ALIASES = {
    "store_number": [
        "store number", "store #", "store no", "store id", "site number", "site #", "site no",
        "location number", "location id", "unit number", "unit #", "number", "store", "site", "str", "str #",
    ],
    "address": [
        "address", "street address", "store address", "location address", "physical address", "site address",
        "address 1", "address1", "addr", "street", "service address", "property address",
    ],
    "city": ["city", "store city", "location city", "municipality", "town", "site city"],
    "state": ["state", "st", "store state", "location state", "province", "site state"],
    "zip": ["zip", "zip code", "postal code", "store zip", "location zip", "zipcode", "postal"],
    "latitude": ["latitude", "lat", "y", "y coordinate", "gps lat", "store latitude", "location latitude", "site latitude"],
    "longitude": ["longitude", "lon", "lng", "long", "x", "x coordinate", "gps lon", "gps lng", "store longitude", "location longitude", "site longitude"],
    "market": ["market", "region"],
    "zone": ["zone"],
    "type": ["type", "store type", "location type"],
    "active": ["active", "active status", "status", "store status"],
    "first_name": ["first name", "firstname", "given name"],
    "last_name": ["last name", "lastname", "surname", "family name"],
    "full_name": ["employee", "employee name", "full name", "technician", "tech", "tech name", "pmt", "calibration tech", "assigned tech", "assigned technician", "name"],
    "employee_number": ["employee number", "employee no", "employee id", "employee #", "emp id", "emp #", "s number"],
    "role": ["role", "job role", "position", "employee type", "technician type", "type"],
    "team": ["team", "team name", "crew", "group"],
    "phone": ["phone", "mobile", "cell", "phone number", "work phone"],
    "email": ["email", "e-mail", "work email", "employee email"],
    "home_address": ["home address", "employee address", "technician address", "starting address", "start address", "home base", "base address", "address", "street address"],
    "home_city": ["home city", "employee city", "technician city", "base city", "city"],
    "home_state": ["home state", "employee state", "technician state", "base state", "state", "st"],
    "home_zip": ["home zip", "home zip code", "employee zip", "technician zip", "base zip", "zip", "zip code", "zipcode", "postal code"],
    "home_latitude": ["home latitude", "home lat", "employee latitude", "technician latitude", "base latitude"],
    "home_longitude": ["home longitude", "home lon", "home lng", "employee longitude", "technician longitude", "base longitude"],
    "assigned_pmt": ["pmt", "pmt technician", "assigned pmt", "assigned tech", "assigned technician"],
    "assigned_brand": ["brand team", "brand enhancement team", "brand enhancement", "assigned brand", "brand tech"],
    "assigned_calibration": ["calibration technician", "calibration tech", "assigned calibration"],
    "area": ["area", "territory", "managed area"],
}

TARGET_FIELDS = {
    "stores": ["store_number", "address", "city", "state", "zip", "latitude", "longitude", "market", "zone", "type", "active", "assigned_pmt", "assigned_brand", "assigned_calibration", "area"],
    "employees": ["full_name", "first_name", "last_name", "employee_number", "role", "team", "phone", "email", "home_address", "home_city", "home_state", "home_zip", "home_latitude", "home_longitude", "active"],
    "assignments": ["store_number", "full_name", "employee_number", "email", "phone", "assigned_pmt", "assigned_brand", "assigned_calibration", "team", "area", "address", "city", "state", "zip", "latitude", "longitude", "home_address", "home_city", "home_state", "home_zip", "home_latitude", "home_longitude"],
}

REQUIRED_FIELDS = {
    "stores": ["store_number"],
    "employees": ["full_name"],
    "assignments": ["store_number"],
}

SUPPORTED_UPLOAD_EXTENSIONS = {".csv", ".xlsx", ".xls", ".xlsm"}


@dataclass
class ColumnMatch:
    field: str
    column: str
    confidence: int
    reason: str


def key(value):
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


def display_field(field):
    return field.replace("_", " ").title()


def clean_text(value):
    text = "" if pd.isna(value) else str(value)
    return re.sub(r"\s+", " ", text).strip()


def to_number(value):
    text = clean_text(value).replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def normalize_state(value):
    text = clean_text(value)
    if not text:
        return ""
    upper = text.upper()
    if upper in US_STATES:
        return upper
    return US_STATE_NAMES.get(key(text), upper[:2] if len(upper) == 2 else text)


def clean_zip(value):
    text = clean_identifier(value)
    if not text:
        return ""
    match = re.search(r"(\d{5})(?:-\d{4})?", text)
    return match.group(0) if match else text


def valid_lat(value):
    number = to_number(value)
    return number is not None and -90 <= number <= 90


def valid_lon(value):
    number = to_number(value)
    return number is not None and -180 <= number <= 180


def series_nonblank(series):
    sample = series.dropna().astype(str).map(str.strip)
    sample = sample[sample.ne("")]
    return sample.head(100)


def ratio(series, predicate):
    sample = series_nonblank(series)
    if sample.empty:
        return 0.0
    return float(sample.map(predicate).sum()) / float(len(sample))


def header_score(column_name, field):
    column_key = key(column_name)
    aliases = {key(alias) for alias in FIELD_ALIASES.get(field, [])}
    if column_key in aliases:
        return 100
    if any(alias and alias in column_key for alias in aliases):
        return 82
    return 0


def pattern_score(series, field):
    if field in ("latitude", "home_latitude"):
        return int(ratio(series, valid_lat) * 90)
    if field in ("longitude", "home_longitude"):
        return int(ratio(series, valid_lon) * 90)
    if field == "zip" or field == "home_zip":
        return int(ratio(series, lambda value: bool(re.search(r"\b\d{5}(?:-\d{4})?\b", clean_identifier(value)))) * 85)
    if field == "state" or field == "home_state":
        return int(ratio(series, lambda value: normalize_state(value) in US_STATES) * 85)
    if field == "store_number":
        return int(ratio(series, lambda value: bool(clean_store_number(value)) or bool(re.fullmatch(r"\d{4,6}", clean_identifier(value)))) * 88)
    if field == "email":
        return int(ratio(series, lambda value: "@" in clean_text(value)) * 95)
    if field == "phone":
        return int(ratio(series, lambda value: len(re.sub(r"\D", "", clean_text(value))) >= 10) * 90)
    if field in ("address", "home_address"):
        street_words = r"\b(rd|road|st|street|ave|avenue|blvd|dr|drive|ct|court|ln|lane|pkwy|highway|hwy|cir|circle|pl|place)\b"
        return int(ratio(series, lambda value: bool(re.search(r"\d+.*" + street_words, clean_text(value), flags=re.I))) * 90)
    if field in ("full_name", "assigned_pmt", "assigned_brand", "assigned_calibration", "team"):
        return int(ratio(series, lambda value: bool(re.search(r"[A-Za-z]{2,}", clean_text(value)))) * 55)
    if field == "employee_number":
        return int(ratio(series, lambda value: bool(re.search(r"[A-Za-z0-9]{3,}", clean_identifier(value)))) * 65)
    return 0


def detect_columns(df, import_type):
    candidates = []
    fields = TARGET_FIELDS.get(import_type, TARGET_FIELDS["stores"])
    for field in fields:
        for column in df.columns:
            h_score = header_score(column, field)
            p_score = pattern_score(df[column], field)
            score = min(100, max(h_score, p_score, int(h_score * 0.7 + p_score * 0.4)))
            if score >= 45:
                reason = "header and data pattern" if h_score and p_score else "header alias" if h_score else "data pattern"
                candidates.append(ColumnMatch(field, column, score, reason))
    mapping = {}
    used_columns = set()
    for match in sorted(candidates, key=lambda item: item.confidence, reverse=True):
        if match.field not in mapping and match.column not in used_columns:
            mapping[match.field] = match
            used_columns.add(match.column)
    ambiguous = []
    for field in fields:
        field_matches = [match for match in candidates if match.field == field]
        if len(field_matches) > 1:
            top = sorted(field_matches, key=lambda item: item.confidence, reverse=True)[:2]
            if top[0].confidence - top[1].confidence <= 10:
                ambiguous.append(field)
    return mapping, candidates, ambiguous


def data_score(df, import_type):
    mapping, _, _ = detect_columns(df, import_type)
    score = 0
    for field in REQUIRED_FIELDS.get(import_type, []):
        if field in mapping:
            score += mapping[field].confidence + 100
    for field in TARGET_FIELDS.get(import_type, []):
        if field in mapping:
            score += mapping[field].confidence
    return score


def make_unique_columns(columns):
    counts = {}
    unique = []
    for column in columns:
        name = clean_text(column) or "Column"
        counts[name] = counts.get(name, 0) + 1
        unique.append(name if counts[name] == 1 else f"{name} {counts[name]}")
    return unique


def empty_scan_result(sheet, warning="", error="", technical_detail=""):
    return {
        "sheet": sheet or "Upload",
        "header_row": 0,
        "header_confidence": 0,
        "df": pd.DataFrame(),
        "mapping": {},
        "candidates": [],
        "ambiguous": [],
        "score": 0,
        "rows": 0,
        "columns": 0,
        "warning": warning,
        "error": error,
        "technical_detail": technical_detail,
    }


def safe_error_detail(exc):
    return "".join(traceback.format_exception(type(exc), exc, exc.__traceback__))


def valid_header_row(raw, header_row):
    if raw is None or raw.empty:
        return False
    try:
        row_number = int(header_row)
    except (TypeError, ValueError):
        return False
    return 0 <= row_number < len(raw.index)


def nonempty_shape(raw):
    if raw is None or raw.empty:
        return 0, 0
    cleaned = raw.replace(r"^\s*$", pd.NA, regex=True)
    nonempty_rows = int(cleaned.dropna(how="all").shape[0])
    nonempty_columns = int(cleaned.dropna(axis=1, how="all").shape[1])
    return nonempty_rows, nonempty_columns


def dataframe_from_raw(raw, header_row):
    if not valid_header_row(raw, header_row):
        return pd.DataFrame()
    row_number = int(header_row)
    headers = make_unique_columns(raw.iloc[row_number].tolist())
    df = raw.iloc[row_number + 1:].copy()
    df.columns = headers
    return df.dropna(how="all").fillna("").reset_index(drop=True)


def best_header_row(raw, import_type):
    if raw is None or raw.empty:
        return 0, 0
    best_row = 0
    best_score = -1
    limit = min(30, len(raw.index))
    for row_number in range(limit):
        df = dataframe_from_raw(raw, row_number)
        if df.empty:
            continue
        header_text = " ".join(clean_text(value) for value in raw.iloc[row_number].tolist())
        alias_hits = sum(1 for aliases in FIELD_ALIASES.values() for alias in aliases if key(alias) and key(alias) in key(header_text))
        score = data_score(df.head(100), import_type) + alias_hits * 8 + min(len(df), 100)
        if score > best_score:
            best_score = score
            best_row = row_number
    return best_row, max(best_score, 0)


def excel_match_value(cell_value, filter_value):
    cell_text = clean_text(cell_value)
    filter_text = clean_text(filter_value)
    if cell_text == filter_text:
        return True
    if clean_identifier(cell_text) == clean_identifier(filter_text):
        return True
    cell_number = to_number(cell_text)
    filter_number = to_number(filter_text)
    if cell_number is not None and filter_number is not None:
        return cell_number == filter_number
    return False


def excel_custom_filter_match(cell_value, custom_filter):
    operator = getattr(custom_filter, "operator", None) or "equal"
    filter_value = getattr(custom_filter, "val", "")
    cell_text = clean_text(cell_value)
    filter_text = clean_text(filter_value)
    cell_number = to_number(cell_text)
    filter_number = to_number(filter_text)
    if operator in ("equal", "notEqual") and any(char in filter_text for char in ["*", "?"]):
        matched = fnmatch.fnmatch(cell_text.lower(), filter_text.lower())
    elif cell_number is not None and filter_number is not None:
        comparisons = {
            "equal": cell_number == filter_number,
            "notEqual": cell_number != filter_number,
            "greaterThan": cell_number > filter_number,
            "greaterThanOrEqual": cell_number >= filter_number,
            "lessThan": cell_number < filter_number,
            "lessThanOrEqual": cell_number <= filter_number,
        }
        matched = comparisons.get(operator, True)
    else:
        comparisons = {
            "equal": cell_text.lower() == filter_text.lower(),
            "notEqual": cell_text.lower() != filter_text.lower(),
            "greaterThan": cell_text.lower() > filter_text.lower(),
            "greaterThanOrEqual": cell_text.lower() >= filter_text.lower(),
            "lessThan": cell_text.lower() < filter_text.lower(),
            "lessThanOrEqual": cell_text.lower() <= filter_text.lower(),
        }
        matched = comparisons.get(operator, True)
    return matched


def excel_filter_column_match(cell_value, filter_column):
    filters = getattr(filter_column, "filters", None)
    if filters is not None:
        allowed_values = list(getattr(filters, "filter", None) or [])
        include_blank = bool(getattr(filters, "blank", False))
        if include_blank and not clean_text(cell_value):
            return True
        if allowed_values:
            return any(excel_match_value(cell_value, allowed) for allowed in allowed_values)
        date_groups = list(getattr(filters, "dateGroupItem", None) or [])
        if date_groups:
            return True
    custom_filters = getattr(filter_column, "customFilters", None)
    if custom_filters is not None:
        custom_items = list(getattr(custom_filters, "customFilter", None) or [])
        if custom_items:
            matches = [excel_custom_filter_match(cell_value, item) for item in custom_items]
            join_with_and = bool(getattr(custom_filters, "and_", getattr(custom_filters, "_and", False)))
            return all(matches) if join_with_and else any(matches)
    return True


def excel_row_matches_filters(row_values, filter_columns, min_column):
    for filter_column in filter_columns:
        column_id = int(getattr(filter_column, "colId", 0) or 0)
        row_index = (min_column - 1) + column_id
        cell_value = row_values[row_index] if row_index < len(row_values) else ""
        if not excel_filter_column_match(cell_value, filter_column):
            return False
    return True


def worksheet_filter_definition(worksheet):
    sheet_filter_ref = getattr(worksheet.auto_filter, "ref", None)
    sheet_filter_columns = list(getattr(worksheet.auto_filter, "filterColumn", None) or [])
    if sheet_filter_ref and sheet_filter_columns:
        return sheet_filter_ref, sheet_filter_columns, "sheet"
    for table in getattr(worksheet, "tables", {}).values():
        table_filter = getattr(table, "autoFilter", None)
        table_ref = getattr(table_filter, "ref", None) or getattr(table, "ref", None)
        table_filter_columns = list(getattr(table_filter, "filterColumn", None) or [])
        if table_ref and table_filter_columns:
            return table_ref, table_filter_columns, f"table {getattr(table, 'displayName', '')}".strip()
    if sheet_filter_ref:
        return sheet_filter_ref, sheet_filter_columns, "sheet"
    for table in getattr(worksheet, "tables", {}).values():
        table_filter = getattr(table, "autoFilter", None)
        table_ref = getattr(table_filter, "ref", None) or getattr(table, "ref", None)
        if table_ref:
            table_filter_columns = list(getattr(table_filter, "filterColumn", None) or [])
            return table_ref, table_filter_columns, f"table {getattr(table, 'displayName', '')}".strip()
    return None, [], ""


def read_visible_excel_sheet(uploaded_file, sheet_name=None):
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.cell import range_boundaries
    except ImportError:
        return None
    uploaded_file.seek(0)
    workbook = load_workbook(uploaded_file, read_only=False, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active
    filter_ref, filter_columns, filter_source = worksheet_filter_definition(worksheet)
    has_filter = bool(filter_ref)
    hidden_rows = {
        row_number
        for row_number, dimension in worksheet.row_dimensions.items()
        if getattr(dimension, "hidden", False)
    }
    if not has_filter:
        return None
    min_column, min_row, max_column, max_row = range_boundaries(filter_ref)
    if not hidden_rows and not filter_columns:
        rows = []
        sheet_max_column = worksheet.max_column or 0
        for row in worksheet.iter_rows(values_only=True):
            rows.append(["" if value is None else value for value in list(row[:sheet_max_column])])
        df = pd.DataFrame(rows).fillna("")
        df.attrs["filter_detected_unapplied"] = True
        df.attrs["filter_source"] = filter_source
        return df
    rows = []
    skipped_by_filter = 0
    for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if row_number in hidden_rows:
            continue
        values = list(row[:max_column])
        if filter_columns and min_row < row_number <= max_row and not excel_row_matches_filters(values, filter_columns, min_column):
            skipped_by_filter += 1
            continue
        if filter_columns and row_number > max_row:
            skipped_by_filter += 1
            continue
        rows.append(["" if value is None else value for value in values])
    df = pd.DataFrame(rows).fillna("")
    df.attrs["filtered_visible_only"] = True
    df.attrs["hidden_rows_skipped"] = len(hidden_rows)
    df.attrs["filter_rows_skipped"] = skipped_by_filter
    df.attrs["filter_source"] = filter_source
    return df


def read_raw_sheet(uploaded_file, sheet_name=None):
    uploaded_file.seek(0)
    if uploaded_file.name.lower().endswith(".csv"):
        try:
            return pd.read_csv(uploaded_file, header=None, dtype=str, sep=None, engine="python").fillna("")
        except Exception:
            uploaded_file.seek(0)
            return pd.read_csv(uploaded_file, header=None, dtype=str).fillna("")
    suffix = Path(uploaded_file.name or "").suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        visible = read_visible_excel_sheet(uploaded_file, sheet_name)
        if visible is not None:
            return visible
        uploaded_file.seek(0)
    return pd.read_excel(uploaded_file, sheet_name=sheet_name, header=None, dtype=str).fillna("")


def sheet_names(uploaded_file):
    suffix = Path(uploaded_file.name or "").suffix.lower()
    if suffix not in SUPPORTED_UPLOAD_EXTENSIONS:
        raise ValueError("Unsupported file type. Upload a CSV or Excel workbook.")
    if suffix == ".csv":
        return ["CSV file"]
    uploaded_file.seek(0)
    return pd.ExcelFile(uploaded_file).sheet_names


def scan_workbook(uploaded_file, import_type):
    results = []
    if uploaded_file is None:
        return [empty_scan_result("Upload", error="No file was uploaded.")]
    try:
        names = sheet_names(uploaded_file)
    except Exception as exc:
        LOGGER.exception("Import scan could not open uploaded file %s", getattr(uploaded_file, "name", "unknown"))
        return [
            empty_scan_result(
                "Upload",
                error="The file could not be opened. Check that it is a normal CSV or Excel file and is not password protected or corrupt.",
                technical_detail=safe_error_detail(exc),
            )
        ]
    if not names:
        return [empty_scan_result("Upload", error="No sheets were found in this workbook.")]
    for name in names:
        warnings = []
        try:
            raw = read_raw_sheet(uploaded_file, name)
        except Exception as exc:
            LOGGER.exception("Import scan skipped unreadable sheet %s in %s", name, getattr(uploaded_file, "name", "unknown"))
            results.append(
                empty_scan_result(
                    name,
                    error="This sheet could not be read and was skipped.",
                    technical_detail=safe_error_detail(exc),
                )
            )
            continue
        if raw.attrs.get("filtered_visible_only"):
            skipped = int(raw.attrs.get("hidden_rows_skipped", 0) or 0) + int(raw.attrs.get("filter_rows_skipped", 0) or 0)
            warnings.append(
                f"Excel filter detected. {skipped} filtered/hidden row(s) were skipped; only visible rows are included."
            )
        elif raw.attrs.get("filter_detected_unapplied"):
            warnings.append(
                "Excel filter detected, but the saved filter criteria could not be applied automatically. Save the file with filtered rows hidden, or copy the visible filtered rows to a new sheet before uploading."
            )
        nonempty_rows, nonempty_columns = nonempty_shape(raw)
        if raw is None or raw.empty or nonempty_rows == 0 or nonempty_columns == 0:
            df = pd.DataFrame()
            header_row = 0
            header_confidence = 0
            mapping, candidates, ambiguous = {}, [], []
            warnings.append("Sheet is empty and was skipped.")
        elif nonempty_rows < 2 or nonempty_columns < 2:
            df = pd.DataFrame()
            header_row = 0
            header_confidence = 0
            mapping, candidates, ambiguous = {}, [], []
            warnings.append("Sheet does not look like a usable table and was skipped.")
        else:
            header_row, header_confidence = best_header_row(raw, import_type)
            if not valid_header_row(raw, header_row):
                df = pd.DataFrame()
                mapping, candidates, ambiguous = {}, [], []
                warnings.append("The app could not find a valid header row on this sheet.")
            else:
                df = dataframe_from_raw(raw, header_row)
                if df.empty:
                    mapping, candidates, ambiguous = {}, [], []
                    warnings.append("The detected header row had no usable data rows under it.")
                else:
                    mapping, candidates, ambiguous = detect_columns(df, import_type)
                    missing = [field for field in REQUIRED_FIELDS.get(import_type, []) if field not in mapping]
                    if missing:
                        warnings.append(
                            "Missing required field mapping: "
                            + ", ".join(display_field(field) for field in missing)
                            + ". Use Advanced Mapping if this is the correct sheet."
                        )
                    if ambiguous:
                        warnings.append(
                            "Multiple columns look similar for: "
                            + ", ".join(display_field(field) for field in ambiguous)
                            + ". Review Advanced Mapping before importing."
                        )
        score = data_score(df, import_type) + header_confidence if not df.empty else 0
        results.append(
            {
                "sheet": name,
                "header_row": header_row,
                "header_confidence": header_confidence,
                "df": df,
                "mapping": mapping,
                "candidates": candidates,
                "ambiguous": ambiguous,
                "score": score,
                "rows": len(df),
                "columns": len(df.columns),
                "warning": " ".join(warnings),
                "error": "",
                "technical_detail": "",
            }
        )
    results.sort(key=lambda item: item["score"], reverse=True)
    return results


def scan_issue_rows(scans):
    rows = []
    for item in scans or []:
        if item.get("error"):
            rows.append(
                {
                    "Sheet": item.get("sheet", "Upload"),
                    "Severity": "Could Not Read",
                    "Message": item.get("error", ""),
                    "Rows": item.get("rows", 0),
                    "Columns": item.get("columns", 0),
                }
            )
        if item.get("warning"):
            rows.append(
                {
                    "Sheet": item.get("sheet", "Upload"),
                    "Severity": "Warning",
                    "Message": item.get("warning", ""),
                    "Rows": item.get("rows", 0),
                    "Columns": item.get("columns", 0),
                }
            )
    return pd.DataFrame(rows)


def mapped_dataframe(df, mapping):
    mapped = pd.DataFrame(index=df.index)
    for field, column in mapping.items():
        if column and column in df.columns:
            mapped[field] = df[column]
    return clean_mapped_dataframe(mapped)


def clean_mapped_dataframe(df):
    df = df.copy()
    for column in df.columns:
        df[column] = df[column].map(clean_text)
    if "store_number" in df.columns:
        df["store_number"] = df["store_number"].map(lambda value: clean_store_number(value) or clean_identifier(value))
    for column in ["zip", "home_zip"]:
        if column in df.columns:
            df[column] = df[column].map(clean_zip)
    for column in ["state", "home_state"]:
        if column in df.columns:
            df[column] = df[column].map(normalize_state)
    for column in ["email"]:
        if column in df.columns:
            df[column] = df[column].str.lower().str.strip()
    for column in ["latitude", "longitude", "home_latitude", "home_longitude"]:
        if column in df.columns:
            df[column] = df[column].map(lambda value: "" if to_number(value) is None else str(to_number(value)))
    if "full_name" not in df.columns and {"first_name", "last_name"}.issubset(df.columns):
        df["full_name"] = (df["first_name"].fillna("") + " " + df["last_name"].fillna("")).str.strip()
    return df.fillna("")


def review_table(mapped, import_type):
    rows = []
    duplicate_key = "store_number" if import_type in ("stores", "assignments") else None
    if import_type == "employees":
        if "employee_number" in mapped.columns and mapped["employee_number"].astype(str).str.strip().ne("").any():
            duplicate_key = "employee_number"
        elif "email" in mapped.columns and mapped["email"].astype(str).str.strip().ne("").any():
            duplicate_key = "email"
        else:
            duplicate_key = "full_name"
    duplicates = set()
    if duplicate_key and duplicate_key in mapped.columns:
        values = mapped[duplicate_key].astype(str).str.strip()
        duplicates = set(values[values.ne("") & values.duplicated(keep=False)])
    for idx, row in mapped.iterrows():
        row_number = int(idx) + 2
        if import_type in ("stores", "assignments"):
            store_number = clean_text(row.get("store_number", ""))
            if not store_number or not clean_store_number(store_number):
                rows.append([row_number, "Must Fix", "Store number missing or invalid.", store_number, clean_text(row.get("full_name", "")), "store_number", "Select the column with the 4-6 digit store/site number."])
            has_assignee_identity = any(clean_text(row.get(column, "")) for column in ["full_name", "employee_number", "email", "phone"])
            if import_type == "assignments" and not has_assignee_identity:
                rows.append([row_number, "Must Fix", "Assignment identity missing.", store_number, "", "assignment", "Map the technician name, employee number, email, phone, PMT, Brand, Calibration, or team column."])
            lat = to_number(row.get("latitude", ""))
            lon = to_number(row.get("longitude", ""))
            has_coords = lat is not None and lon is not None and -90 <= lat <= 90 and -180 <= lon <= 180
            has_address = any(clean_text(row.get(column, "")) for column in ["address", "city", "state", "zip"])
            if lat is not None and not (-90 <= lat <= 90):
                rows.append([row_number, "Must Fix", "Latitude is outside -90 to 90.", store_number, "", "latitude", "Correct the latitude column or mapping."])
            if lon is not None and not (-180 <= lon <= 180):
                rows.append([row_number, "Must Fix", "Longitude is outside -180 to 180.", store_number, "", "longitude", "Correct the longitude column or mapping."])
            if has_coords and lon and lon > 0 and normalize_state(row.get("state", "")) in US_STATES:
                rows.append([row_number, "Warning", "Longitude is positive for a U.S. location.", store_number, "", "longitude", "Confirm the longitude should be negative."])
            if has_coords and normalize_state(row.get("state", "")) == "OH" and not (38 <= lat <= 42 and -85 <= lon <= -80):
                rows.append([row_number, "Warning", "Coordinates do not appear to be in Ohio.", store_number, "", "coordinates", "Review latitude and longitude."])
            if import_type == "stores" and not has_coords and not has_address:
                rows.append([row_number, "Must Fix", "No usable coordinates or address.", store_number, "", "coordinates/address", "Add latitude/longitude or street/city/state/ZIP."])
            if store_number in duplicates:
                rows.append([row_number, "Info", "Duplicate store number in upload; import will merge/update one store.", store_number, "", "store_number", "Review duplicate rows if values conflict."])
        if import_type == "employees":
            employee = clean_text(row.get("full_name", ""))
            if not employee:
                rows.append([row_number, "Must Fix", "Employee name missing.", "", employee, "full_name", "Map the employee/name column or add a name."])
            home_lat = to_number(row.get("home_latitude", ""))
            home_lon = to_number(row.get("home_longitude", ""))
            if home_lat is not None and not (-90 <= home_lat <= 90):
                rows.append([row_number, "Must Fix", "Home latitude is outside -90 to 90.", "", employee, "home_latitude", "Correct the latitude column or mapping."])
            if home_lon is not None and not (-180 <= home_lon <= 180):
                rows.append([row_number, "Must Fix", "Home longitude is outside -180 to 180.", "", employee, "home_longitude", "Correct the longitude column or mapping."])
            if duplicate_key and clean_text(row.get(duplicate_key, "")) in duplicates:
                rows.append([row_number, "Info", "Duplicate employee in upload; import will update/fill one employee.", "", employee, duplicate_key, "Review duplicate rows if values conflict."])
    return pd.DataFrame(rows, columns=["Row Number", "Severity", "Problem", "Store Number", "Employee Name", "Field Missing", "Suggested Fix"])


def mapping_summary(mapping, required_fields):
    rows = []
    for field, match in mapping.items():
        rows.append({"Field": display_field(field), "Detected Column": match.column, "Confidence": f"{match.confidence}%", "Reason": match.reason})
    for field in required_fields:
        if field not in mapping:
            rows.append({"Field": display_field(field), "Detected Column": "Not found", "Confidence": "0%", "Reason": "Required"})
    return pd.DataFrame(rows)


def preview_summary(mapped, review):
    must_fix = int((review["Severity"] == "Must Fix").sum()) if not review.empty else 0
    warnings = int((review["Severity"] == "Warning").sum()) if not review.empty else 0
    return {
        "rows": len(mapped),
        "ready": max(0, len(mapped) - must_fix),
        "needs_review": must_fix + warnings,
        "must_fix": must_fix,
        "warnings": warnings,
    }


def load_saved_mappings():
    if not MAPPING_PATH.exists():
        return {}
    try:
        return json.loads(MAPPING_PATH.read_text())
    except Exception:
        return {}


def save_mapping_pattern(import_type, name, mapping):
    saved = load_saved_mappings()
    saved.setdefault(import_type, {})[name] = mapping
    MAPPING_PATH.parent.mkdir(exist_ok=True)
    MAPPING_PATH.write_text(json.dumps(saved, indent=2, sort_keys=True))
