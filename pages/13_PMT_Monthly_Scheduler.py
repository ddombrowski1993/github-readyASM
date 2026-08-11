from datetime import date, datetime, timedelta
import io
import json
import re
import time

import pandas as pd
import streamlit as st
import folium
import streamlit.components.v1 as components
from streamlit_folium import st_folium

st.set_page_config(page_title="PMT Monthly Scheduler", layout="wide")


from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Image as ReportLabImage
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import func, insert, select

from src.database import active_employees, log_action, safe_query, session_scope
from src.manager_rollup import manager_rollup_dataframe, manager_rollup_query, manager_rollup_totals
from src.exports import download_table, excel_bytes
from src.geocoding import build_address, geocode_address, local_coordinate_estimate
from src.imports import normalize_columns
from src.maps import map_html, render_plain_table, render_route_preview, render_store_map, stable_color
from src.models import Employee, MapArea, PMTScheduleBacklog, PMTScheduleRun, Schedule, ScheduleItem, Store, Team
from src.pdf_reports import REPORT_DIR, pdf_bytes
from src.scheduler import haversine_miles, is_company_holiday
from src.smart_import import scan_issue_rows, scan_workbook
from src.utils import apply_theme, effective_rollup_user_id, ensure_database_or_stop, is_all_managed_view, metric_help_card, page_header, section_header, sidebar_nav, step_flow


apply_theme()
sidebar_nav()
ensure_database_or_stop()
page_header(
    "PMT Monthly Scheduler",
    "Upload PMT assignments and automatically build monthly PMT schedules from technician home address to assigned stores.",
)

if is_all_managed_view():
    st.caption("Read-only roll-up view. Select a specific workspace from the sidebar to build or edit PMT schedules.")
    _ru_df = manager_rollup_dataframe(effective_rollup_user_id())
    if not _ru_df.empty:
        _ru_t = manager_rollup_totals(_ru_df)
        _m1, _m2, _m3, _m4, _m5 = st.columns(5)
        _m1.metric("Scheduled This Month", _ru_t["PMT Scheduled This Month"])
        _m2.metric("Completed", _ru_t["PMT Completed This Month"])
        _m3.metric("Month Progress", f"{_ru_t['PMT Month Progress']}%")
        _m4.metric("Remaining", _ru_t["PMT Remaining This Month"])
        _m5.metric("Behind Pace", _ru_t["PMT Technicians Behind Pace"])
        _pmt_progress_cols = [c for c in [
            "Managed Area", "PMT Scheduled This Month", "PMT Completed This Month",
            "PMT Month Progress", "PMT Remaining This Month",
            "PMT Carryover Stores", "PMT Stores Not Scheduled",
            "PMT Technicians Behind Pace",
        ] if c in _ru_df.columns]
        st.subheader("PMT Progress by Managed Area")
        st.dataframe(_ru_df[_pmt_progress_cols], use_container_width=True, hide_index=True)
    _pmt_runs = manager_rollup_query(
        effective_rollup_user_id(),
        """
        select r.run_name, r.cycle_start, r.cycle_end, r.months,
               r.technician_count, r.store_count, r.unscheduled_count,
               r.status, r.created_at
        from pmt_schedule_runs r
        order by r.created_at desc, r.id desc
        """,
    )
    st.subheader("Published PMT Runs Across All Managed Areas")
    if _pmt_runs.empty:
        st.info("No published PMT schedule runs found across managed areas.")
    else:
        st.dataframe(_pmt_runs, use_container_width=True, hide_index=True)
    st.stop()

step_flow(
    ["Load assignments", "Validate", "Set targets", "Generate draft", "Review routes", "Publish"],
    hint="Build Schedule tab: create a new PMT schedule. Then use Carryover & Backlog, Manage, and Export tabs as needed.",
)


def clean(value):
    return str(value or "").strip()


def key(value):
    return re.sub(r"[^a-z0-9]", "", clean(value).lower())


def workflow_break(title, body):
    st.markdown(
        f"""
        <div style="
            margin: 2.8rem 0 1.4rem 0;
            padding: 1.15rem 1.25rem;
            border: 2px solid #fecaca;
            border-left: 10px solid #7f1d1d;
            border-radius: 10px;
            background: linear-gradient(90deg, #7f1d1d, #991b1b);
            box-shadow: 0 10px 26px rgba(127, 29, 29, 0.25);
        ">
            <div style="font-size: 0.8rem; font-weight: 900; color: #fecaca; text-transform: uppercase; letter-spacing: .08em;">Management Workflow</div>
            <div style="font-size: 1.45rem; font-weight: 900; color: #ffffff; margin-top: .15rem;">{title}</div>
            <div style="color: #fee2e2; margin-top: .2rem;">{body}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def part_break(part, title, body, color="#1d4ed8"):
    st.markdown(
        f"""
        <div style="
            margin: 2rem 0 1rem 0;
            padding: 1rem 1.15rem;
            border-left: 10px solid {color};
            border-radius: 8px;
            background: #f8fafc;
            border-top: 1px solid #cbd5e1;
            border-right: 1px solid #cbd5e1;
            border-bottom: 1px solid #cbd5e1;
        ">
            <div style="font-size: .78rem; font-weight: 900; color: {color}; text-transform: uppercase; letter-spacing: .08em;">{part}</div>
            <div style="font-size: 1.25rem; font-weight: 900; color: #0f172a; margin-top: .12rem;">{title}</div>
            <div style="color: #475569; margin-top: .2rem;">{body}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def store_number_keys(value):
    raw = clean(value)
    if not raw:
        return []
    compact = re.sub(r"\s+", "", raw).upper()
    no_label = re.sub(r"^(STORE|STORE#|STORENO|STORENUMBER|SITE|SITE#|LOCATION|LOCATION#)[:#-]*", "", compact)
    no_decimal = re.sub(r"\.0+$", "", no_label)
    digits = re.sub(r"\D", "", no_decimal)
    keys = [raw, compact, no_label, no_decimal]
    try:
        numeric_value = float(no_decimal)
        if numeric_value.is_integer():
            keys.append(str(int(numeric_value)))
    except ValueError:
        pass
    if digits:
        keys.extend([digits, digits.lstrip("0") or "0"])
    return list(dict.fromkeys([item for item in keys if item]))


def month_start(value):
    return date(value.year, value.month, 1)


def add_months(value, months):
    month = value.month - 1 + months
    year = value.year + month // 12
    month = month % 12 + 1
    return date(year, month, 1)


def month_label(value):
    return value.strftime("%B %Y")


def first_workday(value, avoid_weekends=True, avoid_holidays=True, employee_id=None, avoid_pto=True):
    current = value
    end = add_months(value, 1)
    pto_dates = set()
    if employee_id and avoid_pto:
        pto = safe_query(
            """
            select event_date, coalesce(end_date, event_date) as end_date
            from calloff_pto
            where employee_id = :employee_id
              and event_date < :end_date
              and coalesce(end_date, event_date) >= :start_date
              and lower(trim(coalesce(status, ''))) not in ('denied','cancelled','canceled')
            """,
            {"employee_id": int(employee_id), "start_date": value, "end_date": end},
        )
        for _, row in pto.iterrows():
            start = pd.to_datetime(row["event_date"]).date()
            stop = pd.to_datetime(row["end_date"]).date()
            for item in pd.date_range(start, stop):
                pto_dates.add(item.date())
    while current < end:
        if avoid_weekends and current.weekday() >= 5:
            current += timedelta(days=1)
            continue
        if avoid_holidays and is_company_holiday(current):
            current += timedelta(days=1)
            continue
        if current in pto_dates:
            current += timedelta(days=1)
            continue
        return current
    return value


def next_workday_after(value):
    return first_workday(value + timedelta(days=1))


def normalize_pmt_assignment_columns(df):
    df = normalize_columns(df)
    aliases = {
        "technician": "technician_name",
        "tech": "technician_name",
        "tech_name": "technician_name",
        "technician_name": "technician_name",
        "employee": "technician_name",
        "employee_name": "technician_name",
        "pmt": "technician_name",
        "pmt_name": "technician_name",
        "primary_tech": "technician_name",
        "assigned_tech": "technician_name",
        "assigned_pmt": "technician_name",
        "home_address": "home_address",
        "technician_address": "home_address",
        "employee_address": "home_address",
        "starting_address": "home_address",
        "start_location": "home_address",
        "home_city": "home_city",
        "home_state": "home_state",
        "home_zip": "home_zip",
        "home_latitude": "home_latitude",
        "home_lat": "home_latitude",
        "home_longitude": "home_longitude",
        "home_lon": "home_longitude",
        "home_lng": "home_longitude",
        "store": "store_number",
        "store_number": "store_number",
        "store_#": "store_number",
        "store_no": "store_number",
        "store_num": "store_number",
        "store_nbr": "store_number",
        "store_id": "store_number",
        "store_code": "store_number",
        "str": "store_number",
        "str_#": "store_number",
        "str_no": "store_number",
        "str_num": "store_number",
        "str_nbr": "store_number",
        "site": "store_number",
        "site_id": "store_number",
        "site_number": "store_number",
        "site_no": "store_number",
        "site_num": "store_number",
        "site_nbr": "store_number",
        "location": "store_number",
        "location_id": "store_number",
        "location_number": "store_number",
        "location_no": "store_number",
        "location_num": "store_number",
        "location_nbr": "store_number",
        "branch": "store_number",
        "branch_number": "store_number",
        "branch_no": "store_number",
        "branch_num": "store_number",
        "branch_nbr": "store_number",
        "assigned_store": "store_number",
        "store_address": "store_address",
        "address": "store_address",
        "store_city": "store_city",
        "city": "store_city",
        "store_state": "store_state",
        "state": "store_state",
        "store_zip": "store_zip",
        "zip": "store_zip",
        "latitude": "latitude",
        "lat": "latitude",
        "longitude": "longitude",
        "lng": "longitude",
        "lon": "longitude",
        "active_status": "active_status",
        "active": "active_status",
    }
    df = df.rename(columns={k: v for k, v in aliases.items() if k in df.columns})
    if df.columns.duplicated().any():
        collapsed = pd.DataFrame(index=df.index)
        for column in dict.fromkeys(df.columns):
            matches = df.loc[:, df.columns == column]
            collapsed[column] = matches.replace("", pd.NA).bfill(axis=1).iloc[:, 0].fillna("")
        df = collapsed
    return df.fillna("")


def upload_sheet_names(uploaded_file):
    return cached_upload_sheet_names(uploaded_file.name, uploaded_file.getvalue())


@st.cache_data(show_spinner=False)
def cached_upload_sheet_names(file_name, file_bytes):
    if file_name.lower().endswith(".csv"):
        return ["CSV file"]
    workbook = pd.ExcelFile(io.BytesIO(file_bytes))
    return workbook.sheet_names


def read_upload_sheet(uploaded_file, sheet_name):
    return cached_read_upload_sheet(uploaded_file.name, uploaded_file.getvalue(), sheet_name)


@st.cache_data(show_spinner=False)
def cached_read_upload_sheet(file_name, file_bytes, sheet_name):
    if file_name.lower().endswith(".csv"):
        return pd.read_csv(io.BytesIO(file_bytes), dtype=str).fillna("")
    return pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, dtype=str).fillna("")


def scan_uploaded_workbook(uploaded_file, import_type):
    return cached_scan_uploaded_workbook(uploaded_file.name, uploaded_file.getvalue(), import_type)


@st.cache_data(show_spinner=False)
def cached_scan_uploaded_workbook(file_name, file_bytes, import_type):
    workbook = io.BytesIO(file_bytes)
    workbook.name = file_name
    return scan_workbook(workbook, import_type)


def column_key(column):
    return re.sub(r"[^a-z0-9]", "_", clean(column).lower()).strip("_")


def default_from_candidates(original_columns, candidates):
    candidate_keys = {column_key(candidate) for candidate in candidates}
    for original in original_columns:
        if column_key(original) in candidate_keys:
            return original
    return ""


def default_column(original_columns, target):
    for original in original_columns:
        probe = normalize_pmt_assignment_columns(pd.DataFrame(columns=[original]))
        if target in probe.columns:
            return original
    return ""


ASSIGNMENT_COLUMN_CANDIDATES = {
    "technician_name": ["PMT", "Technician", "Technician Name", "Tech", "Tech Name", "Employee", "Employee Name", "Assigned Tech", "Assigned PMT"],
    "store_number": ["Site Number", "Store Number", "Store #", "Store", "Location Number", "Location", "Site", "STR", "STR #"],
    "store_address": ["Store Address", "Address", "Street Address", "Location Address", "Site Address"],
    "store_city": ["Store City", "City", "Location City", "Site City"],
    "store_state": ["Store State", "State", "ST"],
    "store_zip": ["Store Zip", "Zip", "Zip Code", "Postal Code"],
    "latitude": ["Lat", "Latitude", "Store Latitude", "Location Latitude"],
    "longitude": ["Lon", "Lng", "Long", "Longitude", "Store Longitude", "Location Longitude"],
}


SCHEDULE_COLUMN_CANDIDATES = {
    "technician_name": ["PMT", "Technician", "Technician Name", "Tech", "Employee", "Employee Name", "Assigned PMT"],
    "store_number": ["Site Number", "Store Number", "Store #", "Store", "Location Number", "Location", "Site", "STR", "STR #"],
    "schedule_date": ["Schedule Date", "Scheduled Date", "Date", "Visit Date", "PMT Date", "Planned Date"],
    "schedule_month": ["Month", "Schedule Month", "PMT Month", "Cycle Month"],
    "sequence_number": ["Stop", "Stop Number", "Sequence", "Sequence Number", "Route Order", "Order"],
    "status": ["Status", "Schedule Status", "State"],
    "notes": ["Notes", "Comments", "Reason", "Schedule Notes"],
}


ADDRESS_COLUMN_CANDIDATES = {
    "technician_name": ["Name", "Full Name", "Employee Name", "Technician", "Technician Name", "Tech", "Tech Name", "PMT", "PMT Name"],
    "home_address": ["Address", "Home Address", "Street Address", "Technician Address", "Employee Address", "Starting Address"],
    "home_city": ["City", "Home City"],
    "home_state": ["State", "Home State", "ST"],
    "home_zip": ["Zip", "Zip Code", "Home Zip", "Postal Code"],
    "home_latitude": ["Home Latitude", "Home Lat", "Latitude", "Lat"],
    "home_longitude": ["Home Longitude", "Home Lon", "Home Lng", "Longitude", "Lon", "Lng"],
}


def best_column(original_columns, target, context):
    if context == "address":
        candidates = ADDRESS_COLUMN_CANDIDATES
    elif context == "schedule":
        candidates = SCHEDULE_COLUMN_CANDIDATES
    else:
        candidates = ASSIGNMENT_COLUMN_CANDIDATES
    return default_from_candidates(original_columns, candidates.get(target, [])) or default_column(original_columns, target)


def selectbox_with_default(container, label, options, default_value, key):
    index = options.index(default_value) if default_value in options else 0
    return container.selectbox(label, options, index=index, key=key)


def score_sheet_columns(columns, context):
    candidate_map = ADDRESS_COLUMN_CANDIDATES if context == "address" else ASSIGNMENT_COLUMN_CANDIDATES
    score = 0
    for target, candidates in candidate_map.items():
        if default_from_candidates(columns, candidates):
            score += 3 if target in ("technician_name", "store_number", "home_address") else 1
    return score


def detected_sheet_index(sheet_names, uploaded_file, context):
    best_index = 0
    best_score = -1
    for index, sheet_name in enumerate(sheet_names):
        try:
            columns = read_upload_sheet(uploaded_file, sheet_name).columns.tolist()
        except Exception:
            columns = []
        score = score_sheet_columns(columns, context)
        if context == "address" and "address" in column_key(sheet_name):
            score += 4
        if context == "assignment" and ("assign" in column_key(sheet_name) or "store" in column_key(sheet_name)):
            score += 4
        if score > best_score:
            best_index = index
            best_score = score
    return best_index


def apply_column_mapping(normalized, incoming, mapping):
    mapped = normalized.copy()
    for target, source in mapping.items():
        if source:
            mapped[target] = incoming[source]
    return mapped


def merge_home_address_sheet(assignments_df, address_df):
    if assignments_df.empty or address_df.empty or "technician_name" not in address_df.columns:
        return assignments_df
    address_fields = ["home_address", "home_city", "home_state", "home_zip", "home_latitude", "home_longitude"]
    available_fields = ["technician_name"] + [field for field in address_fields if field in address_df.columns]
    clean_address_df = address_df[available_fields].copy()
    clean_address_df["tech_key"] = clean_address_df["technician_name"].apply(key)
    clean_address_df = clean_address_df[clean_address_df["tech_key"] != ""].drop_duplicates("tech_key")
    merged = assignments_df.copy()
    merged["tech_key"] = merged["technician_name"].apply(key)
    merged = merged.merge(clean_address_df.drop(columns=["technician_name"]), on="tech_key", how="left", suffixes=("", "_from_address_sheet"))
    for field in address_fields:
        sheet_field = f"{field}_from_address_sheet"
        if sheet_field in merged.columns:
            if field not in merged.columns:
                merged[field] = ""
            merged[field] = merged[field].where(
                merged[field].notna() & (merged[field].astype(str).str.strip() != ""),
                merged[sheet_field],
            )
    drop_cols = [col for col in merged.columns if col.endswith("_from_address_sheet") or col == "tech_key"]
    return merged.drop(columns=drop_cols)


def employee_lookup():
    employees = active_employees()
    lookup = {}
    for row in employees.to_dict("records") if not employees.empty else []:
        full_name = clean(row["full_name"])
        lookup[key(full_name)] = row
        parts = full_name.split()
        if len(parts) >= 2:
            lookup[key(f"{parts[-1]} {' '.join(parts[:-1])}")] = row
            lookup[key(f"{parts[-1]}, {' '.join(parts[:-1])}")] = row
    return employees, lookup


def match_employee_name(name, lookup):
    name_key = key(name)
    if not name_key:
        return None
    if name_key in lookup:
        return lookup[name_key]
    for lookup_key, employee in lookup.items():
        if name_key in lookup_key or lookup_key in name_key:
            return employee
    return None


def employee_name_keys(name):
    clean_name = clean(name)
    values = [clean_name]
    parts = clean_name.split()
    if len(parts) >= 2:
        values.append(f"{parts[-1]} {' '.join(parts[:-1])}")
        values.append(f"{parts[-1]}, {' '.join(parts[:-1])}")
    return [key(value) for value in values if key(value)]


def ensure_uploaded_pmt_employees(mapped_df):
    if mapped_df.empty or "technician_name" not in mapped_df.columns:
        return {"created": 0, "updated": 0}
    created = 0
    updated = set()
    with session_scope() as session:
        lookup = {}
        for employee in session.query(Employee).all():
            if not employee.full_name:
                continue
            for name_key in employee_name_keys(employee.full_name):
                lookup.setdefault(name_key, employee)
        for _, row in mapped_df.iterrows():
            tech_name = clean(row.get("technician_name", ""))
            if not tech_name:
                continue
            employee = None
            for name_key in employee_name_keys(tech_name):
                employee = lookup.get(name_key)
                if employee:
                    break
            if not employee:
                employee = Employee(full_name=tech_name, role="PMT", active=True)
                parts = tech_name.split()
                if parts:
                    employee.first_name = parts[0]
                    employee.last_name = " ".join(parts[1:]) if len(parts) > 1 else ""
                session.add(employee)
                session.flush()
                created += 1
                for name_key in employee_name_keys(tech_name):
                    lookup[name_key] = employee
            employee.active = True
            employee.role = "PMT"
            for source, target in [
                ("home_address", "home_address"),
                ("home_city", "home_city"),
                ("home_state", "home_state"),
                ("home_zip", "home_zip"),
            ]:
                if clean(row.get(source, "")):
                    setattr(employee, target, clean(row.get(source, "")))
            home_lat = to_float(row.get("home_latitude", ""))
            home_lon = to_float(row.get("home_longitude", ""))
            if home_lat is not None and home_lon is not None:
                employee.home_latitude = home_lat
                employee.home_longitude = home_lon
            updated.add(int(employee.id))
    return {"created": created, "updated": len(updated)}


def to_float(value):
    try:
        if clean(value) == "":
            return None
        return float(value)
    except ValueError:
        return None


def current_assignments_from_database():
    return safe_query(
        """
        select e.id as employee_id, e.full_name as technician_name, e.home_address, e.home_city, e.home_state, e.home_zip,
               e.home_latitude, e.home_longitude, coalesce(e.monthly_pmt_store_target, 10) as monthly_target,
               s.id as store_id, s.store_number, s.address as store_address, s.city as store_city, s.state as store_state,
               s.zip as store_zip, s.latitude, s.longitude
        from employees e
        join stores s on s.assigned_pmt_employee_id = e.id
        where e.active = true
          and s.active = true
        order by e.full_name, s.store_number
        """
    )


def active_pmt_employee_summary():
    return safe_query(
        """
        select
            e.id as employee_id,
            e.full_name as technician_name,
            e.home_address,
            e.home_city,
            e.home_state,
            e.home_zip,
            e.home_latitude,
            e.home_longitude,
            count(s.id) as assigned_stores
        from employees e
        left join stores s on s.assigned_pmt_employee_id = e.id and s.active = true
        where e.active = true
          and lower(trim(coalesce(e.role, ''))) in (
              'pmt',
              'pmt technician',
              'pm technician',
              'preventive maintenance technician',
              'preventative maintenance technician'
          )
        group by e.id, e.full_name, e.home_address, e.home_city, e.home_state, e.home_zip, e.home_latitude, e.home_longitude
        order by e.full_name
        """
    )


def prepare_uploaded_assignments(mapped_df):
    employees, lookup = employee_lookup()
    stores = safe_query(
        """
        select id as store_id, store_number, address as db_address, city as db_city, state as db_state,
               zip as db_zip, latitude as db_latitude, longitude as db_longitude, active
        from stores
        """
    )
    store_lookup = {}
    if not stores.empty:
        for store_row in stores.to_dict("records"):
            for store_key in store_number_keys(store_row["store_number"]):
                store_lookup.setdefault(store_key, store_row)
    rows = []
    problems = []
    for index, row in mapped_df.iterrows():
        tech_name = clean(row.get("technician_name", ""))
        store_number = clean(row.get("store_number", ""))
        if not tech_name and not store_number:
            continue
        employee = match_employee_name(tech_name, lookup)
        store = None
        for store_key in store_number_keys(store_number):
            store = store_lookup.get(store_key)
            if store is not None:
                break
        if employee is None:
            problems.append({"Problem": "Technician not matched to active employee", "Detail": tech_name or f"Row {index + 2}"})
        if store is None:
            problems.append({"Problem": "Store not found in store database", "Detail": store_number or f"Row {index + 2}"})
        elif not bool(store.get("active", True)):
            problems.append({"Problem": "Store exists but is inactive", "Detail": store["store_number"]})
        if employee is None or store is None:
            continue
        rows.append(
            {
                "employee_id": int(employee["id"]),
                "technician_name": employee["full_name"],
                "home_address": clean(row.get("home_address", "")),
                "home_city": clean(row.get("home_city", "")),
                "home_state": clean(row.get("home_state", "")),
                "home_zip": clean(row.get("home_zip", "")),
                "home_latitude": to_float(row.get("home_latitude", "")),
                "home_longitude": to_float(row.get("home_longitude", "")),
                "monthly_target": 10,
                "store_id": int(store["store_id"]),
                "store_number": clean(store["store_number"]),
                "store_address": clean(row.get("store_address", "")) or store["db_address"],
                "store_city": clean(row.get("store_city", "")) or store["db_city"],
                "store_state": clean(row.get("store_state", "")) or store["db_state"],
                "store_zip": clean(row.get("store_zip", "")) or store["db_zip"],
                "latitude": to_float(row.get("latitude", "")) if clean(row.get("latitude", "")) else store["db_latitude"],
                "longitude": to_float(row.get("longitude", "")) if clean(row.get("longitude", "")) else store["db_longitude"],
            }
        )
    return pd.DataFrame(rows), pd.DataFrame(problems)


def enrich_assignments(df):
    if df.empty:
        return df
    employee_details = safe_query(
        """
        select id as employee_id, full_name, home_address as saved_home_address, home_city as saved_home_city,
               home_state as saved_home_state, home_zip as saved_home_zip, home_latitude as saved_home_latitude,
               home_longitude as saved_home_longitude, coalesce(monthly_pmt_store_target, 10) as saved_monthly_target
        from employees
        where active = true
        """
    )
    if employee_details.empty:
        return df
    merged = df.merge(employee_details, on="employee_id", how="left")
    for target, saved in [
        ("home_address", "saved_home_address"),
        ("home_city", "saved_home_city"),
        ("home_state", "saved_home_state"),
        ("home_zip", "saved_home_zip"),
        ("home_latitude", "saved_home_latitude"),
        ("home_longitude", "saved_home_longitude"),
        ("monthly_target", "saved_monthly_target"),
    ]:
        if target not in merged.columns:
            merged[target] = ""
        merged[target] = merged[target].where(merged[target].notna() & (merged[target].astype(str).str.strip() != ""), merged[saved])
    drop_cols = [col for col in merged.columns if col.startswith("saved_") or col == "full_name"]
    return merged.drop(columns=drop_cols)


def validation_summary(assignments):
    if assignments.empty:
        return {}, pd.DataFrame()
    dupes = assignments.groupby("store_number").filter(lambda group: group["employee_id"].nunique() > 1)
    problems = []
    for _, row in assignments.drop_duplicates("employee_id").iterrows():
        home_has_coordinates = pd.notna(row.get("home_latitude")) and pd.notna(row.get("home_longitude"))
        home_has_address = bool(clean(row.get("home_address", "")) and clean(row.get("home_city", "")) and clean(row.get("home_state", "")))
        home_has_estimate_source = bool(clean(row.get("home_city", "")) and (clean(row.get("home_state", "")) or clean(row.get("home_zip", ""))))
        if not home_has_coordinates and home_has_address:
            problems.append({"Severity": "Must Fix", "Problem": f"{row['technician_name']} needs home coordinates. Use Find Coordinates From Address.", "Technician": row["technician_name"], "Store": ""})
        elif not home_has_coordinates and home_has_estimate_source:
            problems.append({"Severity": "Must Fix", "Problem": f"{row['technician_name']} needs home coordinates. Use City/ZIP Estimate or enter coordinates manually.", "Technician": row["technician_name"], "Store": ""})
        elif not home_has_coordinates:
            problems.append({"Severity": "Must Fix", "Problem": f"{row['technician_name']} has no usable home location in Employees.", "Technician": row["technician_name"], "Store": ""})
    for _, row in assignments.iterrows():
        store_has_coordinates = pd.notna(row.get("latitude")) and pd.notna(row.get("longitude"))
        store_has_address = bool(clean(row.get("store_address", "")) and clean(row.get("store_city", "")) and clean(row.get("store_state", "")))
        if not store_has_coordinates and not store_has_address:
            problems.append({"Severity": "Must Fix", "Problem": f"Store {row['store_number']} has no usable location. Add coordinates or a full address.", "Technician": row["technician_name"], "Store": row["store_number"]})
        elif not store_has_coordinates:
            problems.append({"Severity": "Must Fix", "Problem": f"Store {row['store_number']} needs coordinates before routing.", "Technician": row["technician_name"], "Store": row["store_number"]})
    for _, row in dupes.drop_duplicates("store_number").iterrows():
        owners = ", ".join(sorted(dupes[dupes["store_number"] == row["store_number"]]["technician_name"].unique()))
        problems.append({"Severity": "Warning", "Problem": f"Store {row['store_number']} is assigned to multiple PMTs: {owners}.", "Technician": owners, "Store": row["store_number"]})
    summary = {
        "Rows": len(assignments),
        "Technicians": assignments["employee_id"].nunique(),
        "Stores": assignments["store_id"].nunique(),
        "Missing Home Coordinates": int(assignments.drop_duplicates("employee_id")[["home_latitude", "home_longitude"]].isna().any(axis=1).sum()),
        "Stores Missing Coordinates": int(assignments[["latitude", "longitude"]].isna().any(axis=1).sum()),
        "Stores With Coordinates": int((assignments[["latitude", "longitude"]].notna().all(axis=1)).sum()),
        "Duplicate Store Assignments": int(dupes["store_number"].nunique()) if not dupes.empty else 0,
    }
    return summary, pd.DataFrame(problems)


def nearest_neighbor_order(stores_df, start_lat, start_lon):
    remaining = stores_df.copy()
    ordered_rows = []
    current_lat = float(start_lat)
    current_lon = float(start_lon)
    while not remaining.empty:
        remaining = remaining.copy()
        remaining["_route_distance"] = remaining.apply(
            lambda row: haversine_miles(current_lat, current_lon, float(row["latitude"]), float(row["longitude"])),
            axis=1,
        )
        next_index = remaining["_route_distance"].idxmin()
        next_row = remaining.loc[next_index].drop(labels=["_route_distance"], errors="ignore")
        ordered_rows.append(next_row)
        current_lat = float(next_row["latitude"])
        current_lon = float(next_row["longitude"])
        remaining = remaining.drop(index=next_index)
    return ordered_rows


def nearest_neighbor_route(stores_df, start_lat, start_lon, limit=None):
    remaining = stores_df.copy()
    ordered_rows = []
    current_lat = float(start_lat)
    current_lon = float(start_lon)
    stop_limit = len(remaining) if limit is None else min(int(limit), len(remaining))
    while not remaining.empty and len(ordered_rows) < stop_limit:
        remaining = remaining.copy()
        remaining["_route_distance"] = remaining.apply(
            lambda row: haversine_miles(current_lat, current_lon, float(row["latitude"]), float(row["longitude"])),
            axis=1,
        )
        next_index = remaining["_route_distance"].idxmin()
        next_row = remaining.loc[next_index].drop(labels=["_route_distance"], errors="ignore").copy()
        next_row["miles_from_previous_stop"] = round(float(remaining.loc[next_index, "_route_distance"]), 1)
        ordered_rows.append(next_row)
        current_lat = float(next_row["latitude"])
        current_lon = float(next_row["longitude"])
        remaining = remaining.drop(index=next_index)
    return ordered_rows


def home_distance_route(stores_df, start_lat, start_lon, limit=None):
    ordered = stores_df.sort_values(["distance_from_home", "store_number"], ascending=[True, True]).copy()
    if limit is not None:
        ordered = ordered.head(int(limit)).copy()
    current_lat = float(start_lat)
    current_lon = float(start_lon)
    routed_rows = []
    for _, row in ordered.iterrows():
        next_row = row.copy()
        next_row["miles_from_previous_stop"] = round(
            haversine_miles(current_lat, current_lon, float(next_row["latitude"]), float(next_row["longitude"])),
            1,
        )
        routed_rows.append(next_row)
        current_lat = float(next_row["latitude"])
        current_lon = float(next_row["longitude"])
    return routed_rows


HOME_ROUTE = "Home-Based Route"
NEXT_ROUTE = "Next-Closest Store Route"
ROUTE_EXPORT_OPTIONS = ["Home-Based Route", "Next-Closest Store Route", "Both Route Options"]


def route_notes(route_type):
    if route_type == HOME_ROUTE:
        return "Best if the PMT starts from home each day and works one store per day."
    return "Best if the PMT finishes one store and drives directly to the next closest store."


def route_source_columns(draft):
    df = draft.copy()
    defaults = {
        "zip": "",
        "distance_from_home": None,
        "miles_from_previous_stop": None,
        "estimated_drive_time": "",
        "latitude": None,
        "longitude": None,
        "home_latitude": None,
        "home_longitude": None,
        "notes": "",
    }
    for column, default in defaults.items():
        if column not in df.columns:
            df[column] = default
    return df


def home_coordinates_for_group(group):
    home_lat = to_float(group.iloc[0].get("home_latitude"))
    home_lon = to_float(group.iloc[0].get("home_longitude"))
    return home_lat, home_lon


def build_route_rows_for_group(group, route_type):
    if group.empty:
        return []
    group = route_source_columns(group)
    home_lat, home_lon = home_coordinates_for_group(group)
    has_store_coordinates = group[["latitude", "longitude"]].notna().all().all() if {"latitude", "longitude"}.issubset(group.columns) else False
    if route_type == NEXT_ROUTE and home_lat is not None and home_lon is not None and has_store_coordinates:
        ordered_rows = nearest_neighbor_route(group, home_lat, home_lon)
    elif route_type == HOME_ROUTE and home_lat is not None and home_lon is not None and has_store_coordinates:
        ordered_rows = home_distance_route(group, home_lat, home_lon) if home_lat is not None and home_lon is not None else [
            row.copy() for _, row in group.sort_values(["distance_from_home", "store_number"], ascending=[True, True]).iterrows()
        ]
    else:
        sort_columns = [column for column in ["distance_from_home", "sequence_number", "store_number"] if column in group.columns]
        ordered_rows = [row.copy() for _, row in group.sort_values(sort_columns, ascending=True).iterrows()]
    rows = []
    previous_lat = home_lat
    previous_lon = home_lon
    for route_order, row in enumerate(ordered_rows, start=1):
        distance_from_home = to_float(row.get("distance_from_home"))
        distance_previous = to_float(row.get("miles_from_previous_stop"))
        lat = to_float(row.get("latitude"))
        lon = to_float(row.get("longitude"))
        if route_type == HOME_ROUTE and previous_lat is not None and previous_lon is not None and lat is not None and lon is not None:
            distance_previous = round(haversine_miles(previous_lat, previous_lon, lat, lon), 1)
        elif route_type == NEXT_ROUTE and distance_previous is None and previous_lat is not None and previous_lon is not None and lat is not None and lon is not None:
            distance_previous = round(haversine_miles(previous_lat, previous_lon, lat, lon), 1)
        rows.append(
            {
                "route_order": route_order,
                "technician": row.get("technician", ""),
                "employee_id": row.get("employee_id"),
                "month": row.get("month", ""),
                "month_start": row.get("month_start"),
                "schedule_date": row.get("schedule_date"),
                "store_id": row.get("store_id"),
                "store_number": row.get("store_number", ""),
                "address": row.get("address", ""),
                "city": row.get("city", ""),
                "state": row.get("state", ""),
                "zip": row.get("zip", ""),
                "latitude": row.get("latitude"),
                "longitude": row.get("longitude"),
                "distance_from_home": round(distance_from_home, 1) if distance_from_home is not None else "",
                "miles_from_previous_stop": round(distance_previous, 1) if distance_previous is not None else "",
                "estimated_drive_time": row.get("estimated_drive_time", ""),
                "route_type": route_type,
                "notes": route_notes(route_type),
                "status": row.get("status", "Scheduled"),
            }
        )
        if lat is not None and lon is not None:
            previous_lat = lat
            previous_lon = lon
    return rows


def route_options_for_draft(draft, route_filter="Both Route Options"):
    if draft.empty:
        return pd.DataFrame()
    df = route_source_columns(draft)
    df["_month_sort"] = pd.to_datetime(df["month_start"], errors="coerce")
    route_types = [HOME_ROUTE, NEXT_ROUTE] if route_filter == "Both Route Options" else [route_filter]
    rows = []
    for _, group in df.sort_values(["_month_sort", "technician", "sequence_number", "store_number"]).groupby(["employee_id", "month"], sort=False):
        for route_type in route_types:
            rows.extend(build_route_rows_for_group(group, route_type))
    return pd.DataFrame(rows)


def route_table_view(routes):
    if routes.empty:
        return pd.DataFrame()
    return routes[
        [
            "route_order",
            "technician",
            "month",
            "store_number",
            "address",
            "city",
            "state",
            "distance_from_home",
            "miles_from_previous_stop",
            "estimated_drive_time",
            "route_type",
            "notes",
        ]
    ].rename(
        columns={
            "route_order": "Route Order",
            "technician": "Technician",
            "month": "Month",
            "store_number": "Store Number",
            "address": "Store Address",
            "city": "City",
            "state": "State",
            "distance_from_home": "Distance From Home",
            "miles_from_previous_stop": "Distance From Previous Stop",
            "estimated_drive_time": "Estimated Drive Time",
            "route_type": "Route Type",
            "notes": "Notes",
        }
    )


def route_map_source_for_export(draft, route_filter):
    if draft.empty:
        return draft
    if route_filter in (HOME_ROUTE, NEXT_ROUTE):
        return draft_with_route_order(draft.copy(), route_filter)
    return draft_with_route_order(draft.copy(), HOME_ROUTE)


def pmt_route_map_png(group, title, width=900, height=520):
    try:
        from PIL import Image, ImageDraw, ImageFont
    except ImportError:
        return None
    image = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(image)
    title_font = ImageFont.load_default()
    label_font = ImageFont.load_default()
    draw.rectangle([(0, 0), (width - 1, height - 1)], outline="#cbd5e1")
    draw.rectangle([(0, 0), (width, 44)], fill="#1f2937")
    draw.text((18, 15), title, fill="white", font=title_font)
    if group.empty:
        draw.text((30, 80), "No route rows available.", fill="#334155", font=label_font)
        buffer = io.BytesIO()
        image.save(buffer, format="PNG")
        buffer.seek(0)
        return buffer

    route = group.copy()
    if "sequence_number" not in route.columns:
        route["sequence_number"] = range(1, len(route) + 1)
    if "store_number" not in route.columns:
        route["store_number"] = ""
    route["sequence_number"] = pd.to_numeric(route["sequence_number"], errors="coerce").fillna(0)
    route = route.sort_values(["sequence_number", "store_number"])
    points = []
    for _, row in route.iterrows():
        lat = to_float(row.get("latitude"))
        lon = to_float(row.get("longitude"))
        if lat is not None and lon is not None:
            points.append((lat, lon, int(row.get("sequence_number") or 0), clean(row.get("store_number"))))
    home_lat, home_lon = home_coordinates_for_group(route)
    all_coords = [(lat, lon) for lat, lon, _, _ in points]
    if home_lat is not None and home_lon is not None:
        all_coords.append((home_lat, home_lon))
    if not all_coords:
        draw.text((30, 80), "No latitude/longitude available for this route.", fill="#334155", font=label_font)
        buffer = io.BytesIO()
        image.save(buffer, format="PNG")
        buffer.seek(0)
        return buffer

    try:
        from staticmap import CircleMarker, Line, StaticMap
        osm_map = StaticMap(width, height, url_template="https://a.tile.openstreetmap.org/{z}/{x}/{y}.png")
        route_coords = [(lon, lat) for lat, lon, _, _ in points]
        if home_lat is not None and home_lon is not None and route_coords:
            osm_map.add_line(Line([(home_lon, home_lat), route_coords[0]], "#64748b", 2))
        if len(route_coords) > 1:
            osm_map.add_line(Line(route_coords, "#2563eb", 4))
        if home_lat is not None and home_lon is not None:
            osm_map.add_marker(CircleMarker((home_lon, home_lat), "#111827", 11))
        for lat, lon, _, _ in points:
            osm_map.add_marker(CircleMarker((lon, lat), "#ef4444", 9))
        image = osm_map.render()
        draw = ImageDraw.Draw(image)
        draw.rectangle([(0, 0), (width, 44)], fill="#1f2937")
        draw.text((18, 15), title, fill="white", font=title_font)

        min_lat = min(lat for lat, _ in all_coords)
        max_lat = max(lat for lat, _ in all_coords)
        min_lon = min(lon for _, lon in all_coords)
        max_lon = max(lon for _, lon in all_coords)
        lat_pad = max((max_lat - min_lat) * 0.15, 0.03)
        lon_pad = max((max_lon - min_lon) * 0.15, 0.03)
        min_lat -= lat_pad
        max_lat += lat_pad
        min_lon -= lon_pad
        max_lon += lon_pad
        left, top, right, bottom = 0, 0, width, height

        def project_tile(lat, lon):
            x = left + (lon - min_lon) / (max_lon - min_lon) * (right - left)
            y = bottom - (lat - min_lat) / (max_lat - min_lat) * (bottom - top)
            return int(x), int(y)

        if home_lat is not None and home_lon is not None:
            hx, hy = project_tile(home_lat, home_lon)
            draw.text((hx + 12, hy - 6), "Home", fill="#111827", font=label_font)
        for lat, lon, stop, store_number in points:
            x, y = project_tile(lat, lon)
            label = str(stop or "")
            draw.ellipse([(x - 10, y - 10), (x + 10, y + 10)], fill="#ef4444", outline="white", width=2)
            draw.text((x - 4, y - 5), label, fill="white", font=label_font)
            draw.text((x + 12, y - 6), store_number, fill="#0f172a", font=label_font)
        buffer = io.BytesIO()
        image.save(buffer, format="PNG")
        buffer.seek(0)
        return buffer
    except Exception:
        pass

    min_lat = min(lat for lat, _ in all_coords)
    max_lat = max(lat for lat, _ in all_coords)
    min_lon = min(lon for _, lon in all_coords)
    max_lon = max(lon for _, lon in all_coords)
    lat_pad = max((max_lat - min_lat) * 0.15, 0.03)
    lon_pad = max((max_lon - min_lon) * 0.15, 0.03)
    min_lat -= lat_pad
    max_lat += lat_pad
    min_lon -= lon_pad
    max_lon += lon_pad
    left, top, right, bottom = 44, 68, width - 44, height - 42
    draw.rectangle([(left, top), (right, bottom)], fill="#f8fafc", outline="#94a3b8")
    for fraction in [0.25, 0.5, 0.75]:
        x = left + int((right - left) * fraction)
        y = top + int((bottom - top) * fraction)
        draw.line([(x, top), (x, bottom)], fill="#e2e8f0")
        draw.line([(left, y), (right, y)], fill="#e2e8f0")

    def project(lat, lon):
        x = left + (lon - min_lon) / (max_lon - min_lon) * (right - left)
        y = bottom - (lat - min_lat) / (max_lat - min_lat) * (bottom - top)
        return int(x), int(y)

    route_xy = [project(lat, lon) for lat, lon, _, _ in points]
    if home_lat is not None and home_lon is not None and route_xy:
        home_xy = project(home_lat, home_lon)
        draw.line([home_xy, route_xy[0]], fill="#64748b", width=2)
    if len(route_xy) > 1:
        draw.line(route_xy, fill="#2563eb", width=4)
    if home_lat is not None and home_lon is not None:
        hx, hy = project(home_lat, home_lon)
        draw.rectangle([(hx - 8, hy - 8), (hx + 8, hy + 8)], fill="#111827", outline="white", width=2)
        draw.text((hx + 12, hy - 6), "Home", fill="#111827", font=label_font)
    for lat, lon, stop, store_number in points:
        x, y = project(lat, lon)
        draw.ellipse([(x - 10, y - 10), (x + 10, y + 10)], fill="#ef4444", outline="white", width=2)
        label = str(stop or "")
        draw.text((x - 4, y - 5), label, fill="white", font=label_font)
        draw.text((x + 12, y - 6), store_number, fill="#0f172a", font=label_font)
    draw.text((18, height - 24), "Generated route image: numbered stops are zoomed to this technician/month.", fill="#475569", font=label_font)
    buffer = io.BytesIO()
    image.save(buffer, format="PNG")
    buffer.seek(0)
    return buffer


def draft_with_route_order(draft, route_type):
    if draft.empty:
        return draft
    routes = route_options_for_draft(draft, route_type)
    if routes.empty:
        return draft
    route_lookup = routes.set_index(["employee_id", "month", "store_id"])["route_order"].to_dict()
    updated = draft.copy()
    updated["sequence_number"] = updated.apply(
        lambda row: int(route_lookup.get((row.get("employee_id"), row.get("month"), row.get("store_id")), row.get("sequence_number", 0))),
        axis=1,
    )
    updated["miles_from_previous_stop"] = updated.apply(
        lambda row: routes.set_index(["employee_id", "month", "store_id"])["miles_from_previous_stop"].to_dict().get(
            (row.get("employee_id"), row.get("month"), row.get("store_id")),
            row.get("miles_from_previous_stop", ""),
        ),
        axis=1,
    )
    updated["notes"] = f"Published route order: {route_type}"
    return updated.sort_values(["month_start", "technician", "sequence_number", "store_number"]).reset_index(drop=True)


def pmt_export_views(draft, route_filter="Both Route Options"):
    if draft.empty:
        return pd.DataFrame(), pd.DataFrame()
    export_df = draft.copy()
    for column in ["zip", "distance_from_home", "miles_from_previous_stop", "estimated_drive_time"]:
        if column not in export_df.columns:
            export_df[column] = ""
    if route_filter in (HOME_ROUTE, NEXT_ROUTE):
        export_df = draft_with_route_order(export_df, route_filter)
    export_df["_month_sort"] = pd.to_datetime(export_df["month_start"], errors="coerce")
    export_df = export_df.sort_values(["_month_sort", "technician", "sequence_number", "store_number"])
    schedule_view = export_df[
        ["technician", "month", "sequence_number", "store_number", "address", "city", "state", "zip"]
    ].rename(
        columns={
            "technician": "Technician",
            "month": "Month",
            "sequence_number": "Stop Number",
            "store_number": "Store/Site Number",
            "address": "Address",
            "city": "City",
            "state": "State",
            "zip": "ZIP",
        }
    )
    route_view = route_table_view(route_options_for_draft(export_df, route_filter))
    return schedule_view, route_view


def safe_sheet_name(value, used):
    base = re.sub(r"[\[\]:*?/\\]", "", str(value))[:31] or "Sheet"
    sheet = base
    suffix = 1
    while sheet in used:
        marker = f" {suffix}"
        sheet = f"{base[:31 - len(marker)]}{marker}"
        suffix += 1
    used.add(sheet)
    return sheet


def pmt_schedule_workbook_bytes(draft, route_filter="Both Route Options"):
    schedule_view, route_view = pmt_export_views(draft, route_filter)
    map_draft = route_map_source_for_export(draft, route_filter)
    buffer = io.BytesIO()
    used_sheets = set()
    image_streams = []
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        try:
            from openpyxl.drawing.image import Image as OpenpyxlImage
        except ImportError:
            OpenpyxlImage = None
        if schedule_view.empty:
            schedule_view.to_excel(writer, index=False, sheet_name="Schedule")
        else:
            month_sort = draft[["month", "month_start"]].drop_duplicates().copy()
            month_sort["_month_sort"] = pd.to_datetime(month_sort["month_start"], errors="coerce")
            for month in month_sort.sort_values("_month_sort")["month"].tolist():
                month_df = schedule_view[schedule_view["Month"] == month]
                sheet_name = safe_sheet_name(str(month).split()[0], used_sheets)
                month_df.to_excel(writer, index=False, sheet_name=sheet_name)
                sheet = writer.sheets[sheet_name]
                for column_letter, width in {
                    "A": 22, "B": 16, "C": 12, "D": 18, "E": 32, "F": 18, "G": 10, "H": 12,
                    "J": 16, "K": 16, "L": 16, "M": 16, "N": 16, "O": 16, "P": 16, "Q": 16,
                    "R": 16, "S": 16, "T": 16, "U": 16, "V": 16, "W": 16, "X": 16, "Y": 16, "Z": 16,
                }.items():
                    sheet.column_dimensions[column_letter].width = width
                sheet["J1"] = "Route Map Screenshots"
                if OpenpyxlImage is None:
                    sheet["J2"] = "Install Pillow to include route map images in Excel exports."
                    continue
                map_month_df = map_draft[map_draft["month"] == month].copy() if not map_draft.empty else pd.DataFrame()
                image_row = 2
                for tech in sorted(map_month_df["technician"].dropna().unique().tolist()) if not map_month_df.empty else []:
                    tech_route = map_month_df[map_month_df["technician"] == tech].copy()
                    image_stream = pmt_route_map_png(tech_route, f"{tech} - {month}")
                    if image_stream is None:
                        sheet[f"J{image_row}"] = "Route map image unavailable."
                        image_row += 3
                        continue
                    image_streams.append(image_stream)
                    image = OpenpyxlImage(image_stream)
                    image.width = 620
                    image.height = 360
                    sheet.add_image(image, f"J{image_row}")
                    image_row += 22
        if route_filter == "Both Route Options" and not route_view.empty:
            for route_type in [HOME_ROUTE, NEXT_ROUTE]:
                route_sheet = route_view[route_view["Route Type"] == route_type]
                route_sheet.to_excel(writer, index=False, sheet_name=safe_sheet_name(route_type[:31], used_sheets))
        else:
            route_view.to_excel(writer, index=False, sheet_name=safe_sheet_name("Recommended Routes", used_sheets))
    return buffer.getvalue()


def build_pmt_schedule_pdf(draft, filename, title, technician=None, route_filter="Both Route Options"):
    schedule_view, route_view = pmt_export_views(draft, route_filter)
    map_draft = route_map_source_for_export(draft, route_filter)
    path = REPORT_DIR / filename
    doc = SimpleDocTemplate(str(path), pagesize=landscape(letter), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    small = ParagraphStyle("Small", parent=styles["Normal"], fontSize=7, leading=8)
    section_style = ParagraphStyle("PMTSection", parent=styles["Heading2"], fontSize=11, leading=13, spaceAfter=6)
    story = [
        Paragraph(title, styles["Title"]),
        Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"]),
    ]
    if technician:
        story.append(Paragraph(f"Technician: {technician}", styles["Normal"]))
    story.append(Spacer(1, 10))
    if schedule_view.empty:
        story.append(Paragraph("No PMT schedule records are available.", styles["Normal"]))
    else:
        for tech_index, tech in enumerate(sorted(schedule_view["Technician"].dropna().unique())):
            if tech_index and not technician:
                story.append(PageBreak())
            tech_schedule = schedule_view[schedule_view["Technician"] == tech]
            for month in tech_schedule["Month"].drop_duplicates().tolist():
                group = tech_schedule[tech_schedule["Month"] == month]
                story.append(Paragraph(f"{tech} - {month}", section_style))
                rows = [["Stop", "Store/Site", "Address", "City", "State"]]
                for _, row in group.iterrows():
                    rows.append([row["Stop Number"], row["Store/Site Number"], Paragraph(str(row["Address"] or ""), small), row["City"], row["State"]])
                table = Table(rows, repeatRows=1, colWidths=[36, 62, 310, 110, 42])
                table.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ]))
                story.append(table)
                map_group = map_draft[(map_draft["technician"] == tech) & (map_draft["month"] == month)].copy() if not map_draft.empty else pd.DataFrame()
                route_image = pmt_route_map_png(map_group, f"{tech} - {month}", width=900, height=430)
                if route_image is not None:
                    story.append(Spacer(1, 6))
                    story.append(ReportLabImage(route_image, width=520, height=248))
                story.append(Spacer(1, 8))
        story.append(PageBreak())
        story.append(Paragraph("Recommended Route Options", styles["Heading1"]))
        for route_type in route_view["Route Type"].drop_duplicates().tolist():
            story.append(Paragraph(route_type, styles["Heading2"]))
            story.append(Paragraph(route_notes(route_type), styles["Normal"]))
            story.append(Spacer(1, 6))
            route_type_view = route_view[route_view["Route Type"] == route_type]
            for tech in sorted(route_type_view["Technician"].dropna().unique()):
                tech_routes = route_type_view[route_type_view["Technician"] == tech]
                for month in tech_routes["Month"].drop_duplicates().tolist():
                    group = tech_routes[tech_routes["Month"] == month]
                    story.append(Paragraph(f"{tech} - {month}", section_style))
                    rows = [["Order", "Store", "Address", "From Home", "From Previous", "Drive Time"]]
                    for _, row in group.iterrows():
                        rows.append([row["Route Order"], row["Store Number"], Paragraph(str(row["Store Address"] or ""), small), row["Distance From Home"], row["Distance From Previous Stop"], row["Estimated Drive Time"] or "Unavailable"])
                    table = Table(rows, repeatRows=1, colWidths=[36, 62, 250, 82, 92, 110])
                    table.setStyle(TableStyle([
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#334155")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 7),
                        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ]))
                    story.append(table)
                    story.append(Spacer(1, 8))
    doc.build(story)
    log_action("PMT schedule PDF exported", "reports", description=title)
    return path


def render_pmt_export_controls(export_draft, key_prefix):
    if export_draft.empty:
        st.info("Generate a PMT draft or select a published PMT schedule run, then the export buttons will appear here.")
        return
    st.subheader("PMT Schedule Exports")
    export_tech_count = int(export_draft["technician"].dropna().nunique()) if "technician" in export_draft.columns else 0
    st.caption(f"Export source contains {len(export_draft)} schedule row(s) for {export_tech_count} technician(s).")
    if export_tech_count == 1:
        only_tech = clean(export_draft["technician"].dropna().iloc[0]) if "technician" in export_draft.columns and not export_draft["technician"].dropna().empty else "the selected technician"
        st.warning(f"This selected export source only contains {only_tech}. Choose a full-team schedule run if you need every PMT.")
    route_filter = st.radio(
        "Route export option",
        ROUTE_EXPORT_OPTIONS,
        horizontal=True,
        index=2,
        key=f"{key_prefix}_route_export_option",
    )
    full_excel, full_pdf = st.columns(2)
    full_excel.download_button(
        "Full Team Excel",
        data=pmt_schedule_workbook_bytes(export_draft, route_filter),
        file_name="pmt_full_team_schedule.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=f"{key_prefix}_full_team_excel",
    )
    if full_pdf.button("Build Full Team PDF", key=f"{key_prefix}_full_team_pdf_button"):
        path = build_pmt_schedule_pdf(export_draft, "pmt_full_team_schedule.pdf", "PMT Full Team Schedule", route_filter=route_filter)
        st.download_button("Download Full Team PDF", data=pdf_bytes(path), file_name="pmt_full_team_schedule.pdf", key=f"{key_prefix}_full_team_pdf_download")
    tech_options = sorted(export_draft["technician"].dropna().unique().tolist())
    if not tech_options:
        st.info("No technician names are available for individual exports.")
        return
    selected_export_tech = st.selectbox("Individual Technician", tech_options, key=f"{key_prefix}_individual_export_tech")
    individual = export_draft[export_draft["technician"] == selected_export_tech].copy()
    ind_excel, ind_pdf = st.columns(2)
    ind_excel.download_button(
        "Individual Excel",
        data=pmt_schedule_workbook_bytes(individual, route_filter),
        file_name=f"pmt_schedule_{key(selected_export_tech) or 'technician'}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=f"{key_prefix}_individual_excel",
    )
    if ind_pdf.button("Build Individual PDF", key=f"{key_prefix}_individual_pdf_button"):
        path = build_pmt_schedule_pdf(individual, f"pmt_schedule_{key(selected_export_tech) or 'technician'}.pdf", "PMT Individual Schedule", selected_export_tech, route_filter)
        st.download_button("Download Individual PDF", data=pdf_bytes(path), file_name=f"pmt_schedule_{key(selected_export_tech) or 'technician'}.pdf", key=f"{key_prefix}_individual_pdf_download")


def published_pmt_run_export_draft(run_id):
    df = safe_query(
        """
        select si.schedule_date, si.sequence_number, si.status, si.completion_notes as notes,
               e.id as employee_id, e.full_name as technician, e.home_latitude, e.home_longitude,
               s.id as store_id, s.store_number, s.address, s.city, s.state, s.zip,
               s.latitude, s.longitude
        from schedule_items si
        left join pmt_schedule_runs r on r.id = si.pmt_schedule_run_id
        left join employees e on e.id = si.employee_id
        left join stores s on s.id = si.store_id
        where si.pmt_schedule_run_id = :run_id
          and si.work_type = 'PMT'
          and coalesce(nullif(lower(trim(si.status)), ''), 'scheduled') not in ('cancelled','canceled','skipped','deleted','transferred','superseded','archived')
          and (r.cycle_start is null or date(si.schedule_date) >= date(r.cycle_start))
          and (r.cycle_end is null or date(si.schedule_date) <= date(r.cycle_end))
        order by si.schedule_date, e.full_name, si.sequence_number, s.store_number
        """,
        {"run_id": int(run_id)},
    )
    if df.empty:
        return pd.DataFrame()
    df = df.copy()
    df["schedule_date"] = pd.to_datetime(df["schedule_date"], errors="coerce")
    df = df.dropna(subset=["schedule_date"])
    if df.empty:
        return pd.DataFrame()
    df["month_start"] = df["schedule_date"].dt.to_period("M").dt.to_timestamp()
    df["month"] = df["month_start"].apply(lambda value: month_label(value.date()))
    df["schedule_date"] = df["schedule_date"].dt.date
    df["month_start"] = df["month_start"].dt.date
    df["work_type"] = "PMT"
    df["estimated_drive_time"] = ""
    df["distance_from_home"] = df.apply(
        lambda row: round(
            haversine_miles(float(row["home_latitude"]), float(row["home_longitude"]), float(row["latitude"]), float(row["longitude"])),
            1,
        )
        if pd.notna(row.get("home_latitude"))
        and pd.notna(row.get("home_longitude"))
        and pd.notna(row.get("latitude"))
        and pd.notna(row.get("longitude"))
        else None,
        axis=1,
    )
    previous_distances = {}
    for group_key, group in df.sort_values(["employee_id", "month_start", "sequence_number", "store_number"]).groupby(["employee_id", "month"], dropna=False):
        prev_lat = to_float(group.iloc[0].get("home_latitude"))
        prev_lon = to_float(group.iloc[0].get("home_longitude"))
        for idx, row in group.iterrows():
            lat = to_float(row.get("latitude"))
            lon = to_float(row.get("longitude"))
            if prev_lat is not None and prev_lon is not None and lat is not None and lon is not None:
                previous_distances[idx] = round(haversine_miles(prev_lat, prev_lon, lat, lon), 1)
                prev_lat = lat
                prev_lon = lon
            else:
                previous_distances[idx] = None
    df["miles_from_previous_stop"] = df.index.map(previous_distances)
    return df


def pmt_carryover_report():
    backlog = safe_query(
        """
        select
            b.id as backlog_id,
            null as schedule_item_id,
            'Backlog' as source,
            e.full_name as technician,
            s.store_number,
            s.city,
            b.status,
            b.reason,
            b.cycles_missed,
            b.priority_score,
            b.last_scheduled_month,
            b.last_completed_date,
            b.notes
        from pmt_schedule_backlog b
        left join employees e on e.id = b.employee_id
        left join stores s on s.id = b.store_id
        where b.status in ('Not Scheduled','Not Completed','Carryover','Overdue','Skipped')
        """
    )
    exceptions = safe_query(
        """
        select
            null as backlog_id,
            si.id as schedule_item_id,
            'Scheduled Item' as source,
            e.full_name as technician,
            s.store_number,
            s.city,
            si.status,
            coalesce(si.completion_notes, '') as reason,
            1 as cycles_missed,
            0 as priority_score,
            si.schedule_date as last_scheduled_month,
            null as last_completed_date,
            si.completion_notes as notes
        from schedule_items si
        left join employees e on e.id = si.employee_id
        left join stores s on s.id = si.store_id
        where si.work_type = 'PMT'
          and si.status in ('Needs Rescheduled','Rescheduled','Rain Delay','Not Completed','Carryover','Overdue','Skipped')
        """
    )
    combined = pd.concat([backlog, exceptions], ignore_index=True)
    if combined.empty:
        return combined
    return combined.sort_values(["technician", "priority_score", "cycles_missed", "store_number"], ascending=[True, False, False, True])


def pmt_stores_not_in_run(run_id):
    return safe_query(
        """
        select
            r.id as run_id,
            r.cycle_start,
            r.cycle_end,
            e.id as employee_id,
            e.full_name as technician,
            s.id as store_id,
            s.store_number,
            s.city,
            s.state,
            'Not Scheduled' as status,
            case
                when s.latitude is null or s.longitude is null then 'Store missing latitude/longitude'
                when e.home_latitude is null or e.home_longitude is null then 'Technician missing home latitude/longitude'
                else 'Assigned PMT store did not fit into this published run'
            end as reason
        from pmt_schedule_runs r
        join stores s on s.active = true
        join employees e on e.id = s.assigned_pmt_employee_id and e.active = true
        where r.id = :run_id
          and not exists (
              select 1
              from schedule_items si
              where si.work_type = 'PMT'
                and si.employee_id = e.id
                and si.store_id = s.id
                and date(si.schedule_date) >= date(r.cycle_start)
                and date(si.schedule_date) <= date(r.cycle_end)
          )
        order by e.full_name, s.store_number
        """,
        {"run_id": int(run_id)},
    )


def pmt_rotation_gaps_for_period(cycle_start, months):
    cycle_start = month_start(cycle_start)
    cycle_end = add_months(cycle_start, int(months)) - timedelta(days=1)
    not_scheduled = safe_query(
        """
        select
            null as run_id,
            :cycle_start as cycle_start,
            :cycle_end as cycle_end,
            null as schedule_item_id,
            'Missing From Selected Period' as source,
            e.id as employee_id,
            e.full_name as technician,
            s.id as store_id,
            s.store_number,
            s.city,
            s.state,
            'Not Scheduled' as status,
            case
                when s.latitude is null or s.longitude is null then 'Store missing latitude/longitude'
                when e.home_latitude is null or e.home_longitude is null then 'Technician missing home latitude/longitude'
                else 'Assigned PMT store did not fit into the selected schedule period'
            end as reason
        from stores s
        join employees e on e.id = s.assigned_pmt_employee_id and e.active = true
        where s.active = true
          and not exists (
              select 1
              from schedule_items si
              where si.work_type = 'PMT'
                and si.employee_id = e.id
                and si.store_id = s.id
                and date(si.schedule_date) >= date(:cycle_start)
                and date(si.schedule_date) <= date(:cycle_end)
          )
        """,
        {"cycle_start": cycle_start, "cycle_end": cycle_end},
    )
    not_completed = safe_query(
        """
        select
            si.pmt_schedule_run_id as run_id,
            :cycle_start as cycle_start,
            :cycle_end as cycle_end,
            si.id as schedule_item_id,
            'Scheduled But Not Completed' as source,
            e.id as employee_id,
            e.full_name as technician,
            s.id as store_id,
            s.store_number,
            s.city,
            s.state,
            si.status,
            coalesce(si.completion_notes, si.status) as reason
        from schedule_items si
        left join employees e on e.id = si.employee_id
        left join stores s on s.id = si.store_id
        where si.work_type = 'PMT'
          and date(si.schedule_date) >= date(:cycle_start)
          and date(si.schedule_date) <= date(:cycle_end)
          and si.status in ('Needs Rescheduled','Rescheduled','Rain Delay','Not Completed','Carryover','Overdue','Skipped','Cancelled')
        """,
        {"cycle_start": cycle_start, "cycle_end": cycle_end},
    )
    combined = pd.concat([not_scheduled, not_completed], ignore_index=True)
    if combined.empty:
        return combined
    return combined.sort_values(["technician", "source", "store_number"])


def pmt_rotation_gap_summary(cycle_start, months):
    cycle_start = month_start(cycle_start)
    cycle_end_exclusive = add_months(cycle_start, int(months))
    return safe_query(
        """
        with technician_rotation as (
            select
                e.id as employee_id,
                e.full_name as technician,
                count(distinct s.id) as assigned_stores,
                count(distinct si.store_id) as unique_stores_scheduled,
                max(coalesce(e.monthly_pmt_store_target, 10)) as monthly_target,
                max(coalesce(e.monthly_pmt_store_target, 10)) * cast(:months as bigint) as period_capacity,
                greatest(0::bigint, count(distinct s.id) - count(distinct si.store_id)) as assigned_stores_not_scheduled,
                count(distinct case when si.status in ('Needs Rescheduled','Rescheduled','Rain Delay','Not Completed','Carryover','Overdue','Skipped','Cancelled','Canceled') then si.id end) as scheduled_not_completed
            from employees e
            join stores s on s.assigned_pmt_employee_id = e.id and s.active = true
            left join schedule_items si
              on si.work_type = 'PMT'
             and si.employee_id = e.id
             and si.store_id = s.id
             and si.schedule_date >= cast(:cycle_start as date)
             and si.schedule_date < cast(:cycle_end_exclusive as date)
            where e.active = true
            group by e.id, e.full_name
        )
        select *
        from technician_rotation
        where assigned_stores_not_scheduled > 0
           or scheduled_not_completed > 0
        order by assigned_stores_not_scheduled desc, scheduled_not_completed desc, technician
        """,
        {"cycle_start": cycle_start, "cycle_end_exclusive": cycle_end_exclusive, "months": int(months)},
    )


def save_pmt_gap_rows(gap_rows, source_description):
    if gap_rows.empty:
        return {"created": 0, "updated": 0}
    created = 0
    updated = 0
    with session_scope() as session:
        for _, row in gap_rows.iterrows():
            employee_id = scalar_int(row.get("employee_id"), 0)
            store_id = scalar_int(row.get("store_id"), 0)
            if not employee_id or not store_id:
                continue
            run_id = scalar_int(row.get("run_id"), 0) or None
            status = clean(row.get("status", "")) or "Not Scheduled"
            if status in ("Needs Rescheduled", "Rescheduled", "Rain Delay"):
                status = "Not Completed"
            existing = session.query(PMTScheduleBacklog).filter(
                PMTScheduleBacklog.employee_id == employee_id,
                PMTScheduleBacklog.store_id == store_id,
                PMTScheduleBacklog.status.in_(PMT_BACKLOG_OPEN_STATUSES),
            ).first()
            record = existing or PMTScheduleBacklog(
                pmt_schedule_run_id=run_id,
                employee_id=employee_id,
                store_id=store_id,
                cycle_start=scalar_date(row.get("cycle_start")) or month_start(date.today()),
                cycle_end=scalar_date(row.get("cycle_end")),
            )
            if not existing:
                session.add(record)
                created += 1
            else:
                updated += 1
            record.pmt_schedule_run_id = record.pmt_schedule_run_id or run_id
            record.status = status
            record.reason = clean(row.get("reason", "")) or source_description
            record.cycles_missed = max(int(record.cycles_missed or 0), 1)
            record.priority_score = max(int(record.priority_score or 0), 1000 if status == "Not Scheduled" else 900)
            record.notes = f"{source_description}: {record.reason}"
    return {"created": created, "updated": updated}


def pmt_manage_run_items(run_id):
    df = safe_query(
        """
        select si.id as schedule_item_id, si.schedule_id, si.schedule_date, si.sequence_number,
               si.employee_id, e.full_name as technician, e.home_latitude, e.home_longitude,
               si.store_id, s.store_number, s.address, s.city, s.state, s.zip,
               s.assigned_pmt_employee_id, owner.full_name as assigned_technician,
               s.latitude, s.longitude, si.work_type, si.status, si.cycle_label,
               si.schedule_source, si.original_schedule_date,
               si.completion_notes as notes
        from schedule_items si
        left join employees e on e.id = si.employee_id
        left join stores s on s.id = si.store_id
        left join employees owner on owner.id = s.assigned_pmt_employee_id
        where si.pmt_schedule_run_id = :run_id
          and si.work_type = 'PMT'
        order by si.schedule_date, e.full_name, si.sequence_number, s.store_number
        """,
        {"run_id": int(run_id)},
    )
    if df.empty:
        return df
    df = df.copy()
    df["schedule_date"] = pd.to_datetime(df["schedule_date"], errors="coerce")
    df = df.dropna(subset=["schedule_date"])
    if df.empty:
        return df
    df["month_start"] = df["schedule_date"].dt.to_period("M").dt.to_timestamp().dt.date
    df["month"] = df["month_start"].apply(month_label)
    df["schedule_date"] = df["schedule_date"].dt.date
    df["original_schedule_date"] = pd.to_datetime(df.get("original_schedule_date"), errors="coerce").dt.date
    for column in ["employee_id", "store_id", "schedule_item_id", "schedule_id", "sequence_number", "assigned_pmt_employee_id"]:
        df[column] = pd.to_numeric(df[column], errors="coerce")
    return df


PMT_ACTIVE_STATUSES = {"scheduled", "needs rescheduled", "rescheduled", "rain delay", "not completed"}
PMT_COMPLETED_STATUSES = {"completed", "complete"}
PMT_CANCELED_STATUSES = {"cancelled", "canceled", "skipped", "deleted", "transferred", "superseded", "archived"}
PMT_ACTIVE_STATUS_VALUES = ["Scheduled", "Needs Rescheduled", "Rescheduled", "Rain Delay", "Not Completed"]


def normalize_schedule_status(value):
    return clean(value).lower()


def pmt_active_item_mask(df):
    statuses = df.get("status", pd.Series([], dtype=str)).apply(normalize_schedule_status)
    return statuses.isin(PMT_ACTIVE_STATUSES)


def pmt_completed_item_mask(df):
    statuses = df.get("status", pd.Series([], dtype=str)).apply(normalize_schedule_status)
    return statuses.isin(PMT_COMPLETED_STATUSES)


def pmt_canceled_item_mask(df):
    statuses = df.get("status", pd.Series([], dtype=str)).apply(normalize_schedule_status)
    return statuses.isin(PMT_CANCELED_STATUSES)


def distinct_store_count(df):
    if df.empty or "store_id" not in df.columns:
        return 0
    return int(df["store_id"].dropna().astype(int).nunique())


def manage_month_options(run_items):
    if run_items.empty:
        return ["All months"]
    months = sorted(run_items["month_start"].dropna().unique().tolist())
    return ["All months"] + months


def filter_manage_scope(df, employee_id=None, selected_month="All months", status_filter="Active"):
    if df.empty:
        return df.copy()
    scope = df.copy()
    if employee_id is not None:
        scope = scope[pd.to_numeric(scope["employee_id"], errors="coerce").fillna(-1).astype(int) == int(employee_id)].copy()
    if selected_month != "All months":
        scope = scope[scope["month_start"] == selected_month].copy()
    if status_filter == "Active":
        scope = scope[pmt_active_item_mask(scope)].copy()
    elif status_filter == "Completed":
        scope = scope[pmt_completed_item_mask(scope)].copy()
    elif status_filter == "Canceled / Skipped":
        scope = scope[pmt_canceled_item_mask(scope)].copy()
    return scope


def assigned_store_ids_for_employee(employee_id):
    assigned = assigned_pmt_store_candidates(employee_id, include_scheduled=True)
    if assigned.empty:
        return set(), assigned
    ids = set(assigned["store_id"].dropna().astype(int).tolist())
    return ids, assigned


def technician_schedule_reconciliation(run_items, employee_id, selected_month="All months"):
    assigned_ids, assigned_df = assigned_store_ids_for_employee(employee_id)
    tech_all = filter_manage_scope(run_items, employee_id=employee_id, selected_month=selected_month, status_filter="All")
    tech_active = filter_manage_scope(run_items, employee_id=employee_id, selected_month=selected_month, status_filter="Active")
    tech_completed = filter_manage_scope(run_items, employee_id=employee_id, selected_month=selected_month, status_filter="Completed")
    tech_canceled = filter_manage_scope(run_items, employee_id=employee_id, selected_month=selected_month, status_filter="Canceled / Skipped")
    active_any = filter_manage_scope(run_items, selected_month=selected_month, status_filter="Active")
    completed_any = filter_manage_scope(run_items, selected_month=selected_month, status_filter="Completed")
    active_store_ids_any = set(active_any["store_id"].dropna().astype(int).tolist()) if not active_any.empty else set()
    completed_store_ids_any = set(completed_any["store_id"].dropna().astype(int).tolist()) if not completed_any.empty else set()
    accounted_store_ids_any = active_store_ids_any | completed_store_ids_any
    assigned_not_scheduled_ids = assigned_ids - accounted_store_ids_any
    if assigned_df.empty:
        assigned_not_scheduled = pd.DataFrame()
    else:
        assigned_not_scheduled = assigned_df[assigned_df["store_id"].astype(int).isin(assigned_not_scheduled_ids)].copy()
    scheduled_no_longer_assigned = tech_active[
        ~pd.to_numeric(tech_active["store_id"], errors="coerce").fillna(-1).astype(int).isin(assigned_ids)
    ].copy() if not tech_active.empty else pd.DataFrame()
    assigned_scheduled_elsewhere = active_any[
        pd.to_numeric(active_any["store_id"], errors="coerce").fillna(-1).astype(int).isin(assigned_ids)
        & (pd.to_numeric(active_any["employee_id"], errors="coerce").fillna(-1).astype(int) != int(employee_id))
    ].copy() if not active_any.empty else pd.DataFrame()
    return {
        "assigned_count": len(assigned_ids),
        "active_count": distinct_store_count(tech_active),
        "completed_count": distinct_store_count(tech_completed),
        "canceled_count": distinct_store_count(tech_canceled),
        "assigned_not_scheduled_count": len(assigned_not_scheduled_ids),
        "scheduled_no_longer_assigned_count": distinct_store_count(scheduled_no_longer_assigned),
        "assigned_scheduled_elsewhere_count": distinct_store_count(assigned_scheduled_elsewhere),
        "assigned_df": assigned_df,
        "tech_all": tech_all,
        "tech_active": tech_active,
        "tech_completed": tech_completed,
        "tech_canceled": tech_canceled,
        "assigned_not_scheduled": assigned_not_scheduled,
        "scheduled_no_longer_assigned": scheduled_no_longer_assigned,
        "assigned_scheduled_elsewhere": assigned_scheduled_elsewhere,
    }


def run_status_counts(run_items):
    if run_items.empty:
        return {"unique_stores": 0, "active_rows": 0, "completed_rows": 0, "canceled_rows": 0}
    return {
        "unique_stores": distinct_store_count(run_items),
        "active_rows": int(pmt_active_item_mask(run_items).sum()),
        "completed_rows": int(pmt_completed_item_mask(run_items).sum()),
        "canceled_rows": int(pmt_canceled_item_mask(run_items).sum()),
    }


def pmt_schedule_conflicts(run_items):
    if run_items.empty:
        return pd.DataFrame()
    active = run_items[pmt_active_item_mask(run_items)].copy()
    if active.empty:
        return pd.DataFrame()
    for column in ["schedule_item_id", "store_id", "employee_id", "assigned_pmt_employee_id", "sequence_number"]:
        if column in active.columns:
            active[column] = pd.to_numeric(active[column], errors="coerce")
    duplicate_store_ids = set(
        active.groupby("store_id")["employee_id"]
        .nunique()
        .loc[lambda series: series > 1]
        .index.dropna()
        .astype(int)
        .tolist()
    )
    row_count_duplicate_ids = set(
        active.groupby("store_id")["schedule_item_id"]
        .count()
        .loc[lambda series: series > 1]
        .index.dropna()
        .astype(int)
        .tolist()
    )
    owner_mismatch = (
        active["assigned_pmt_employee_id"].notna()
        & active["employee_id"].notna()
        & (active["assigned_pmt_employee_id"].astype("Int64") != active["employee_id"].astype("Int64"))
    )
    conflict_store_ids = duplicate_store_ids | row_count_duplicate_ids | set(active.loc[owner_mismatch, "store_id"].dropna().astype(int).tolist())
    if not conflict_store_ids:
        return pd.DataFrame()
    conflicts = active[active["store_id"].astype("Int64").isin(conflict_store_ids)].copy()

    def conflict_type(row):
        reasons = []
        store_id = int(row["store_id"]) if pd.notna(row.get("store_id")) else None
        if store_id in duplicate_store_ids:
            reasons.append("Duplicate active schedule")
        elif store_id in row_count_duplicate_ids:
            reasons.append("Multiple active rows in same run")
        assigned_id = row.get("assigned_pmt_employee_id")
        employee_id = row.get("employee_id")
        if pd.notna(assigned_id) and pd.notna(employee_id) and int(assigned_id) != int(employee_id):
            reasons.append("Scheduled under previous technician")
        return "; ".join(reasons) or "Schedule conflict"

    conflicts["conflict_type"] = conflicts.apply(conflict_type, axis=1)
    conflicts["other_active_technicians"] = conflicts.groupby("store_id")["technician"].transform(
        lambda values: ", ".join(sorted({clean(value) for value in values if clean(value)}))
    )
    conflicts["recommended_technician"] = conflicts["assigned_technician"].fillna("")
    return conflicts.sort_values(["store_number", "schedule_date", "technician", "sequence_number"])


def reconciliation_store_number(value):
    raw = clean(value)
    if raw.endswith(".0"):
        raw = raw[:-2]
    return raw.strip()


def pmt_reconciliation_scan(effective_date, run_id=None, ignore_effective_date=False):
    scan_start_date = date(1900, 1, 1) if ignore_effective_date else month_start(effective_date)
    terminal_statuses = ["completed", "complete", "cancelled", "canceled", "skipped", "deleted", "transferred", "superseded", "archived"]
    params = {
        "effective_date": effective_date,
        "scan_start_date": scan_start_date,
        "run_id": int(run_id) if run_id else None,
    }
    future_items = safe_query(
        """
        select si.id as schedule_item_id, si.schedule_id, si.pmt_schedule_run_id,
               coalesce(r.run_name, sch.schedule_name, '') as schedule_name,
               si.schedule_date, si.sequence_number, si.status, si.cycle_label,
               si.store_id, s.store_number, s.city, s.state, s.latitude, s.longitude,
               si.employee_id as scheduled_employee_id,
               coalesce(se.full_name, 'Unassigned') as scheduled_technician,
               coalesce(se.active, false) as scheduled_employee_active,
               se.home_latitude as scheduled_home_latitude,
               se.home_longitude as scheduled_home_longitude,
               s.assigned_pmt_employee_id as assigned_employee_id,
               coalesce(ae.full_name, 'Unassigned') as assigned_technician,
               ae.home_latitude as assigned_home_latitude,
               ae.home_longitude as assigned_home_longitude,
               s.assigned_pmt_team_id as assigned_team_id,
               si.schedule_source,
               si.completion_notes as notes
        from schedule_items si
        left join schedules sch on sch.id = si.schedule_id
        left join pmt_schedule_runs r on r.id = si.pmt_schedule_run_id
        left join stores s on s.id = si.store_id
        left join employees se on se.id = si.employee_id
        left join employees ae on ae.id = s.assigned_pmt_employee_id
        where upper(coalesce(si.work_type, '')) like '%PMT%'
          and si.schedule_date >= :scan_start_date
          and coalesce(nullif(lower(trim(si.status)), ''), 'scheduled') not in ('completed','complete','cancelled','canceled','skipped','deleted','transferred','superseded','archived')
          and (r.id is null or coalesce(lower(trim(r.status)), '') <> 'deleted')
          and (sch.id is null or coalesce(lower(trim(sch.status)), '') <> 'deleted')
          and (:run_id is null or si.pmt_schedule_run_id = :run_id)
        order by si.schedule_date, scheduled_technician, si.sequence_number, s.store_number
        """,
        params,
        use_cache=False,
    )
    if future_items.empty:
        future_items = pd.DataFrame(
            columns=[
                "schedule_item_id", "schedule_id", "pmt_schedule_run_id", "schedule_name", "schedule_date",
                "sequence_number", "status", "cycle_label", "store_id", "store_number", "city", "state",
                "scheduled_employee_id", "scheduled_technician", "scheduled_employee_active",
                "scheduled_home_latitude", "scheduled_home_longitude", "assigned_employee_id", "assigned_technician",
                "assigned_home_latitude", "assigned_home_longitude", "assigned_team_id", "schedule_source", "notes",
            ]
        )
    else:
        future_items = future_items.copy()
        for column in ["schedule_item_id", "schedule_id", "pmt_schedule_run_id", "store_id", "scheduled_employee_id", "assigned_employee_id", "assigned_team_id", "sequence_number"]:
            future_items[column] = pd.to_numeric(future_items[column], errors="coerce")
        for column in ["latitude", "longitude", "scheduled_home_latitude", "scheduled_home_longitude", "assigned_home_latitude", "assigned_home_longitude"]:
            if column in future_items.columns:
                future_items[column] = pd.to_numeric(future_items[column], errors="coerce")
        future_items["schedule_date"] = pd.to_datetime(future_items["schedule_date"], errors="coerce").dt.date
        future_items["month"] = future_items["schedule_date"].apply(month_label)
        future_items["normalized_store_number"] = future_items["store_number"].apply(reconciliation_store_number)
        future_items["before_effective_date"] = future_items["schedule_date"].apply(lambda value: bool(value and value < effective_date))

    duplicate_store_ids = set()
    conflicts = future_items.copy()
    if conflicts.empty:
        conflicts = pd.DataFrame(columns=list(future_items.columns) + ["conflict_type", "recommended_action", "assignment_effective_date", "resolution"])
    else:
        duplicate_store_ids = set(
            conflicts.groupby("store_id")["schedule_item_id"]
            .count()
            .loc[lambda series: series > 1]
            .index.dropna()
            .astype(int)
            .tolist()
        )

        def conflict_type(row):
            reasons = []
            scheduled_id = row.get("scheduled_employee_id")
            assigned_id = row.get("assigned_employee_id")
            if not bool(row.get("scheduled_employee_active", True)):
                reasons.append("Scheduled to Inactive Technician")
            if pd.isna(assigned_id):
                reasons.append("Unassigned Store Still Scheduled")
            elif pd.notna(scheduled_id) and int(scheduled_id) != int(assigned_id):
                reasons.append("Scheduled Technician Does Not Match Current Assignment")
                if bool(row.get("before_effective_date")) and not ignore_effective_date:
                    reasons.append("Current Month Unfinished Before Effective Date")
                else:
                    reasons.append("Store Scheduled After Reassignment")
            if pd.notna(row.get("store_id")) and int(row["store_id"]) in duplicate_store_ids:
                reasons.append("Duplicate Future Schedule")
            return "; ".join(dict.fromkeys(reasons))

        conflicts["conflict_type"] = conflicts.apply(conflict_type, axis=1)
        conflicts = conflicts[conflicts["conflict_type"].astype(str).str.strip().ne("")].copy()
        conflicts["assignment_effective_date"] = effective_date
        conflicts["work_timing"] = conflicts["schedule_date"].apply(
            lambda value: "Earlier unfinished work" if value and value < effective_date else "Effective-date/future work"
        )
        conflicts["recommended_action"] = conflicts["assigned_employee_id"].apply(
            lambda value: "Transfer unfinished schedule item to current assigned PMT and resequence route" if pd.notna(value) else "Manual review - store is unassigned"
        )
        conflicts["resolution"] = "Unresolved"
        conflicts = conflicts.sort_values(["scheduled_technician", "schedule_date", "store_number"])

    active_assigned = safe_query(
        """
        select s.id as store_id, s.store_number, s.city, s.state,
               s.assigned_pmt_employee_id as assigned_employee_id,
               coalesce(e.full_name, 'Unknown PMT') as assigned_technician,
               coalesce(e.active, false) as assigned_employee_active,
               e.role as assigned_employee_role
        from stores s
        left join employees e on e.id = s.assigned_pmt_employee_id
        where s.active = true
          and s.assigned_pmt_employee_id is not null
        order by assigned_technician, s.store_number
        """,
        use_cache=False,
    )
    assigned_not_scheduled = pd.DataFrame()
    if not active_assigned.empty:
        active_assigned = active_assigned.copy()
        active_assigned["store_id"] = pd.to_numeric(active_assigned["store_id"], errors="coerce")
        active_assigned["assigned_employee_id"] = pd.to_numeric(active_assigned["assigned_employee_id"], errors="coerce")
        active_assigned["normalized_store_number"] = active_assigned["store_number"].apply(reconciliation_store_number)
        scheduled_pairs = set()
        scheduled_store_ids = set()
        scheduled_employee_ids = set()
        if not future_items.empty:
            scheduled_pairs = set(
                tuple(int(value) for value in pair)
                for pair in future_items[["store_id", "scheduled_employee_id"]].dropna().astype(int).drop_duplicates().values.tolist()
            )
            scheduled_store_ids = set(future_items["store_id"].dropna().astype(int).tolist())
            scheduled_employee_ids = set(future_items["scheduled_employee_id"].dropna().astype(int).tolist())

        def missing_schedule_reason(row):
            if pd.isna(row.get("store_id")) or pd.isna(row.get("assigned_employee_id")):
                return ""
            store_id = int(row["store_id"])
            assigned_employee_id = int(row["assigned_employee_id"])
            if store_id in scheduled_store_ids and (store_id, assigned_employee_id) not in scheduled_pairs:
                return "Assigned Store Missing From Receiving PMT Schedule"
            if assigned_employee_id not in scheduled_employee_ids:
                return "New Technician Has Assigned Stores but No Schedule"
            return ""

        active_assigned["conflict_type"] = active_assigned.apply(missing_schedule_reason, axis=1)
        assigned_not_scheduled = active_assigned[active_assigned["conflict_type"].astype(str).str.strip().ne("")].copy()
        assigned_not_scheduled["recommended_action"] = "Build or add schedule only for this PMT"
        assigned_not_scheduled["assignment_effective_date"] = effective_date
        assigned_not_scheduled["work_timing"] = "Assigned store gap"

    active_pmts = safe_query(
        """
        select e.id as employee_id, e.full_name as technician,
               coalesce(assigned.assigned_stores, 0) as assigned_stores,
               coalesce(scheduled.future_schedule_items, 0) as future_schedule_items
        from employees e
        left join (
            select assigned_pmt_employee_id as employee_id, count(*) as assigned_stores
            from stores
            where active = true and assigned_pmt_employee_id is not null
            group by assigned_pmt_employee_id
        ) assigned on assigned.employee_id = e.id
        left join (
            select employee_id, count(*) as future_schedule_items
            from schedule_items
            where upper(coalesce(work_type, '')) like '%PMT%'
              and schedule_date >= :scan_start_date
              and coalesce(nullif(lower(trim(status)), ''), 'scheduled') not in ('completed','complete','cancelled','canceled','skipped','deleted','transferred','superseded','archived')
              and employee_id is not null
            group by employee_id
        ) scheduled on scheduled.employee_id = e.id
        where e.active = true and e.role = 'PMT'
        order by e.full_name
        """,
        {"scan_start_date": scan_start_date},
        use_cache=False,
    )
    affected_ids = set()
    for column in ["scheduled_employee_id", "assigned_employee_id"]:
        if column in conflicts.columns and not conflicts.empty:
            affected_ids |= set(conflicts[column].dropna().astype(int).tolist())
    if not assigned_not_scheduled.empty:
        affected_ids |= set(assigned_not_scheduled["assigned_employee_id"].dropna().astype(int).tolist())
    if active_pmts.empty:
        affected = pd.DataFrame()
        protected = pd.DataFrame()
    else:
        active_pmts = active_pmts.copy()
        active_pmts["employee_id"] = pd.to_numeric(active_pmts["employee_id"], errors="coerce")
        active_pmts["reconciliation_status"] = active_pmts["employee_id"].apply(lambda value: "Affected" if pd.notna(value) and int(value) in affected_ids else "Protected - No assignment conflict detected")
        affected = active_pmts[active_pmts["reconciliation_status"] == "Affected"].copy()
        protected = active_pmts[active_pmts["reconciliation_status"].str.startswith("Protected")].copy()
    schedule_runs_scanned = safe_query(
        """
        select count(*) as count
        from pmt_schedule_runs
        where (:run_id is null or id = :run_id)
          and coalesce(lower(trim(status)), '') <> 'deleted'
        """,
        {"run_id": int(run_id) if run_id else None},
        use_cache=False,
    )
    total_assignments = int(len(active_assigned)) if isinstance(active_assigned, pd.DataFrame) else 0
    total_runs = int(schedule_runs_scanned.iloc[0]["count"] or 0) if not schedule_runs_scanned.empty else 0
    rows_with_store_join = int(future_items["store_id"].notna().sum()) if "store_id" in future_items.columns else 0
    rows_not_joined = int(future_items["store_id"].isna().sum()) if "store_id" in future_items.columns else 0
    null_scheduled_employee_ids = int(future_items["scheduled_employee_id"].isna().sum()) if "scheduled_employee_id" in future_items.columns else 0
    null_assigned_employee_ids = int(future_items["assigned_employee_id"].isna().sum()) if "assigned_employee_id" in future_items.columns else 0
    effective_date_exclusions = int(future_items["before_effective_date"].sum()) if "before_effective_date" in future_items.columns else 0
    diagnostics = {
        "assignment_source": "stores.assigned_pmt_employee_id left joined to employees.id",
        "schedule_source": "schedule_items where work_type contains PMT",
        "technician_identifier": "employees.id",
        "workspace": st.session_state.get("active_account_label") or st.session_state.get("active_account_slug") or "Current workspace",
        "pm_assignment_column": "stores.assigned_pmt_employee_id",
        "assignments_loaded": total_assignments,
        "schedule_runs_scanned": total_runs,
        "schedule_items_loaded": len(future_items),
        "rows_successfully_joined": rows_with_store_join,
        "rows_not_joined": rows_not_joined,
        "conflicts_detected": len(conflicts),
        "assigned_not_scheduled_detected": len(assigned_not_scheduled),
        "duplicate_future_store_numbers": len(duplicate_store_ids),
        "null_scheduled_employee_ids": null_scheduled_employee_ids,
        "null_assigned_employee_ids": null_assigned_employee_ids,
        "effective_date_exclusions": effective_date_exclusions,
        "effective_date": str(effective_date),
        "scan_start_date": str(scan_start_date),
        "ignore_effective_date": bool(ignore_effective_date),
        "terminal_statuses_excluded": terminal_statuses,
        "cache_status": "Bypassed safe_query cache for reconciliation scan",
        "scan_timestamp": datetime.utcnow().isoformat(timespec="seconds") + "Z",
    }
    return {
        "future_items": future_items,
        "conflicts": conflicts,
        "assigned_not_scheduled": assigned_not_scheduled,
        "affected": affected,
        "protected": protected,
        "diagnostics": diagnostics,
    }


def pmt_reconciliation_package_bytes(scan, effective_date, reason=""):
    from openpyxl.styles import Font, PatternFill

    buffer = io.BytesIO()
    conflicts = scan.get("conflicts", pd.DataFrame())
    assigned_not_scheduled = scan.get("assigned_not_scheduled", pd.DataFrame())
    affected = scan.get("affected", pd.DataFrame())
    protected = scan.get("protected", pd.DataFrame())
    color_legend = pd.DataFrame(
        [
            {"Color Code": color_name, "Meaning": meaning, "Excel Fill": color}
            for color, color_name, meaning in RECONCILIATION_COLOR_RULES
        ]
    )

    def add_color_code(df):
        if df.empty:
            return df
        coded = df.copy()
        coded.insert(0, "Color Code", coded.apply(reconciliation_row_color_name, axis=1))
        return coded

    conflict_export = add_color_code(conflicts)
    assigned_export = add_color_code(assigned_not_scheduled)
    summary = pd.DataFrame(
        [
            {"Metric": "Effective Date", "Value": str(effective_date)},
            {"Metric": "Reason", "Value": reason},
            {"Metric": "Schedule Conflicts", "Value": len(conflicts)},
            {"Metric": "Assigned But Not Scheduled", "Value": len(assigned_not_scheduled)},
            {"Metric": "Affected Technicians", "Value": len(affected)},
            {"Metric": "Protected Unaffected Technicians", "Value": len(protected)},
            {"Metric": "Color Coding", "Value": "See Color Legend tab. Colored rows use the same rules as the Reconciliation page."},
        ]
    )
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        summary.to_excel(writer, index=False, sheet_name="Summary")
        color_legend.to_excel(writer, index=False, sheet_name="Color Legend")
        conflict_export.to_excel(writer, index=False, sheet_name="Schedule Conflicts")
        assigned_export.to_excel(writer, index=False, sheet_name="Assigned Not Scheduled")
        affected.to_excel(writer, index=False, sheet_name="Affected Techs")
        protected.to_excel(writer, index=False, sheet_name="Unaffected Techs")
        manual_review = conflicts[
            conflicts.get("assigned_employee_id", pd.Series(dtype=float)).isna()
        ].copy() if not conflicts.empty else pd.DataFrame()
        manual_review_export = add_color_code(manual_review)
        manual_review_export.to_excel(writer, index=False, sheet_name="Manual Review")

        workbook = writer.book
        for sheet_name in ["Summary", "Color Legend", "Schedule Conflicts", "Assigned Not Scheduled", "Affected Techs", "Unaffected Techs", "Manual Review"]:
            sheet = workbook[sheet_name]
            sheet.freeze_panes = "A2"
            if sheet.max_row and sheet.max_column:
                sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.font = Font(bold=True)

        for row_index, (color, _, _) in enumerate(RECONCILIATION_COLOR_RULES, start=2):
            fill = PatternFill("solid", fgColor=color.replace("#", "").upper())
            for cell in workbook["Color Legend"][row_index]:
                cell.fill = fill

        def apply_reconciliation_fills(sheet_name, source_df):
            if source_df.empty:
                return
            sheet = workbook[sheet_name]
            for excel_row, (_, source_row) in enumerate(source_df.iterrows(), start=2):
                color = reconciliation_row_color(source_row)
                if not color:
                    continue
                fill = PatternFill("solid", fgColor=color.replace("#", "").upper())
                for cell in sheet[excel_row]:
                    cell.fill = fill

        apply_reconciliation_fills("Schedule Conflicts", conflicts)
        apply_reconciliation_fills("Assigned Not Scheduled", assigned_not_scheduled)
        apply_reconciliation_fills("Manual Review", manual_review)
    return buffer.getvalue()


RECONCILIATION_COLOR_RULES = [
    ("#fee2e2", "Red", "Scheduled to inactive technician or inactive owner issue"),
    ("#ffedd5", "Orange", "Earlier unfinished work that still needs reassignment"),
    ("#dbeafe", "Blue", "Scheduled PMT does not match current assigned PMT"),
    ("#f3e8ff", "Purple", "Duplicate future/unfinished schedule row"),
    ("#fef9c3", "Yellow", "Assigned store is missing from receiving PMT schedule"),
    ("#e5e7eb", "Gray", "Store is unassigned but still scheduled"),
]


def reconciliation_row_color(row):
    conflict_text = clean(row.get("conflict_type", "")).lower()
    timing_text = clean(row.get("work_timing", "")).lower()
    if "inactive" in conflict_text:
        return "#fee2e2"
    if "earlier unfinished" in timing_text:
        return "#ffedd5"
    if "does not match" in conflict_text or "after reassignment" in conflict_text:
        return "#dbeafe"
    if "duplicate" in conflict_text:
        return "#f3e8ff"
    if "missing" in conflict_text or "no schedule" in conflict_text:
        return "#fef9c3"
    if "unassigned" in conflict_text:
        return "#e5e7eb"
    return ""


def reconciliation_row_color_name(row):
    color = reconciliation_row_color(row)
    for rule_color, color_name, _ in RECONCILIATION_COLOR_RULES:
        if rule_color == color:
            return color_name
    return ""


def reconciliation_styler(df):
    if df.empty:
        return df

    def style_row(row):
        color = reconciliation_row_color(row)
        return [f"background-color: {color}" if color else "" for _ in row]

    return df.style.apply(style_row, axis=1)


def render_reconciliation_color_legend():
    items = "".join(
        f"<span style='display:inline-flex;align-items:center;gap:.35rem;margin-right:1rem;margin-bottom:.35rem;'>"
        f"<span style='display:inline-block;width:1rem;height:1rem;border:1px solid #94a3b8;background:{color};'></span>"
        f"<span><strong>{name}</strong>: {meaning}</span></span>"
        for color, name, meaning in RECONCILIATION_COLOR_RULES
    )
    st.markdown(f"<div style='line-height:1.6'>{items}</div>", unsafe_allow_html=True)


def build_pmt_reconciliation_compare_preview(scan, selected_item_ids, monthly_target_override=None):
    selected_ids = {int(value) for value in selected_item_ids if pd.notna(value)}
    base = scan.get("future_items", pd.DataFrame()).copy()
    conflicts = scan.get("conflicts", pd.DataFrame()).copy()
    protected = scan.get("protected", pd.DataFrame()).copy()
    if base.empty:
        return {"compare": pd.DataFrame(), "old_schedule": pd.DataFrame(), "new_schedule": pd.DataFrame(), "protected": protected}

    for column in ["schedule_item_id", "scheduled_employee_id", "assigned_employee_id", "sequence_number", "store_id"]:
        if column in base.columns:
            base[column] = pd.to_numeric(base[column], errors="coerce")
    base["month_start"] = base["schedule_date"].apply(lambda value: month_start(value) if value else None)
    base["old_employee_id"] = base["scheduled_employee_id"]
    base["old_technician"] = base["scheduled_technician"]
    base["old_sequence_number"] = base["sequence_number"]
    base["old_schedule_date"] = base["schedule_date"]
    base["old_month"] = base["month"]
    base["new_employee_id"] = base["scheduled_employee_id"]
    base["new_technician"] = base["scheduled_technician"]
    base["new_sequence_number"] = base["sequence_number"]
    base["new_schedule_date"] = base["schedule_date"]
    base["new_month_start"] = base["month_start"]
    base["new_month"] = base["month"]
    base["new_home_latitude"] = base.get("scheduled_home_latitude")
    base["new_home_longitude"] = base.get("scheduled_home_longitude")
    base["proposed_change"] = "No change"

    transferable = base["schedule_item_id"].fillna(-1).astype(int).isin(selected_ids) & base["assigned_employee_id"].notna()
    receiving_rebalances = {}
    base.loc[transferable, "new_employee_id"] = base.loc[transferable, "assigned_employee_id"]
    base.loc[transferable, "new_technician"] = base.loc[transferable, "assigned_technician"]
    base.loc[transferable, "new_home_latitude"] = base.loc[transferable, "assigned_home_latitude"]
    base.loc[transferable, "new_home_longitude"] = base.loc[transferable, "assigned_home_longitude"]
    base.loc[transferable, "proposed_change"] = "Transfer to current assigned PMT"
    effective_date = scalar_date(scan.get("diagnostics", {}).get("effective_date")) or date.today()
    for index, row in base.loc[transferable].iterrows():
        old_date = row.get("old_schedule_date")
        assigned_id = row.get("assigned_employee_id")
        if old_date and old_date < month_start(date.today()) and pd.notna(assigned_id):
            new_date = first_workday(month_start(date.today()), employee_id=int(assigned_id))
            base.at[index, "new_schedule_date"] = new_date
            base.at[index, "new_month_start"] = month_start(new_date)
            base.at[index, "new_month"] = month_label(month_start(new_date))
            base.at[index, "proposed_change"] = "Transfer to current assigned PMT and carry forward to current month"
        if pd.notna(assigned_id):
            rebalance_key = int(assigned_id)
            rebalance_start = max(month_start(base.at[index, "new_schedule_date"]), month_start(effective_date))
            receiving_rebalances[rebalance_key] = min(receiving_rebalances.get(rebalance_key, rebalance_start), rebalance_start)

    monthly_targets = {}
    if receiving_rebalances:
        target_ids = sorted(receiving_rebalances.keys())
        placeholders = ", ".join([f":employee_id_{idx}" for idx, _ in enumerate(target_ids)])
        target_params = {f"employee_id_{idx}": employee_id for idx, employee_id in enumerate(target_ids)}
        target_rows = safe_query(
            f"""
            select id as employee_id, coalesce(monthly_pmt_store_target, 10) as monthly_target
            from employees
            where id in ({placeholders})
            """,
            target_params,
            use_cache=False,
        )
        if not target_rows.empty:
            monthly_targets = target_rows.set_index("employee_id")["monthly_target"].fillna(10).astype(int).to_dict()

    for employee_id, start_month_value in receiving_rebalances.items():
        monthly_target = max(1, int(monthly_target_override or monthly_targets.get(employee_id, 10) or 10))
        mask = (
            base["new_employee_id"].notna()
            & (base["new_employee_id"].astype("Int64") == int(employee_id))
            & base["new_month_start"].notna()
            & (base["new_month_start"] >= start_month_value)
        )
        tech_rows = base.loc[mask].copy()
        if tech_rows.empty:
            continue
        home_lat = to_float(tech_rows["new_home_latitude"].dropna().iloc[0]) if tech_rows["new_home_latitude"].notna().any() else None
        home_lon = to_float(tech_rows["new_home_longitude"].dropna().iloc[0]) if tech_rows["new_home_longitude"].notna().any() else None

        def rebalance_key(row):
            lat = to_float(row.get("latitude"))
            lon = to_float(row.get("longitude"))
            distance = haversine_miles(home_lat, home_lon, lat, lon) if home_lat is not None and home_lon is not None and lat is not None and lon is not None else 999999
            old_date = row.get("old_schedule_date") or date.max
            carry_forward = old_date < start_month_value
            return (0 if carry_forward else 1, distance, clean(row.get("store_number")), old_date, int(row.get("old_sequence_number") or 0))

        cursor_month = start_month_value
        sequence_number = 1
        for index in sorted(tech_rows.index.tolist(), key=lambda idx: rebalance_key(tech_rows.loc[idx])):
            if sequence_number > monthly_target:
                cursor_month = add_months(cursor_month, 1)
                sequence_number = 1
            original_month = base.at[index, "new_month_start"]
            original_sequence = int(base.at[index, "new_sequence_number"] or 0)
            base.at[index, "new_schedule_date"] = first_workday(cursor_month, employee_id=int(employee_id))
            base.at[index, "new_month_start"] = cursor_month
            base.at[index, "new_month"] = month_label(cursor_month)
            base.at[index, "new_sequence_number"] = sequence_number
            if original_month != cursor_month or original_sequence != sequence_number:
                old_change = clean(base.at[index, "proposed_change"])
                base.at[index, "proposed_change"] = "Receiving PMT schedule rebalanced" if old_change == "No change" else f"{old_change}; receiving PMT schedule rebalanced"
            sequence_number += 1

    touched_pairs = set()
    for _, row in base.loc[transferable].iterrows():
        month_value = row.get("month_start")
        new_month_value = row.get("new_month_start")
        if pd.notna(row.get("old_employee_id")) and month_value:
            touched_pairs.add((int(row["old_employee_id"]), month_value))
        if pd.notna(row.get("new_employee_id")) and new_month_value:
            touched_pairs.add((int(row["new_employee_id"]), new_month_value))

    for employee_id, month_value in touched_pairs:
        mask = (
            base["new_employee_id"].notna()
            & (base["new_employee_id"].astype("Int64") == int(employee_id))
            & (base["new_month_start"] == month_value)
        )
        month_rows = base.loc[mask].copy()
        if month_rows.empty:
            continue
        home_lat = to_float(month_rows["new_home_latitude"].dropna().iloc[0]) if month_rows["new_home_latitude"].notna().any() else None
        home_lon = to_float(month_rows["new_home_longitude"].dropna().iloc[0]) if month_rows["new_home_longitude"].notna().any() else None

        def route_key(row):
            lat = to_float(row.get("latitude"))
            lon = to_float(row.get("longitude"))
            distance = haversine_miles(home_lat, home_lon, lat, lon) if home_lat is not None and home_lon is not None and lat is not None and lon is not None else 999999
            return (distance, clean(row.get("store_number")), row.get("old_schedule_date") or date.max, int(row.get("old_sequence_number") or 0))

        ordered_indexes = sorted(month_rows.index.tolist(), key=lambda idx: route_key(month_rows.loc[idx]))
        for new_sequence, index in enumerate(ordered_indexes, start=1):
            old_change = clean(base.at[index, "proposed_change"])
            base.at[index, "new_sequence_number"] = new_sequence
            if old_change == "No change" and int(base.at[index, "old_sequence_number"] or 0) != int(new_sequence):
                base.at[index, "proposed_change"] = "Route order adjusted in affected month"

    compare = base.copy()
    compare["change_status"] = compare["proposed_change"]
    current_month = month_start(date.today())
    compare_current_forward = compare[compare["new_month_start"].notna() & (compare["new_month_start"] >= current_month)].copy()
    compare_columns = [
        "change_status", "store_number", "city", "state", "old_technician", "new_technician",
        "old_month", "new_month", "old_schedule_date", "new_schedule_date", "old_sequence_number",
        "new_sequence_number", "status", "schedule_name", "conflict_type",
    ]
    if "conflict_type" not in compare_current_forward.columns and not conflicts.empty:
        conflict_lookup = conflicts.set_index("schedule_item_id")["conflict_type"].to_dict()
        compare_current_forward["conflict_type"] = compare_current_forward["schedule_item_id"].map(conflict_lookup).fillna("")

    old_schedule = compare_current_forward[
        ["old_month", "old_schedule_date", "old_sequence_number", "old_technician", "store_number", "city", "state", "status", "schedule_name"]
    ].rename(columns={
        "old_month": "month",
        "old_schedule_date": "schedule_date",
        "old_sequence_number": "sequence_number",
        "old_technician": "technician",
    }).sort_values(["month", "technician", "sequence_number", "store_number"])
    new_schedule = compare_current_forward[
        ["new_month", "new_schedule_date", "new_sequence_number", "new_technician", "store_number", "city", "state", "status", "schedule_name", "proposed_change"]
    ].rename(columns={
        "new_month": "month",
        "new_schedule_date": "schedule_date",
        "new_sequence_number": "sequence_number",
        "new_technician": "technician",
    }).sort_values(["month", "technician", "sequence_number", "store_number"])
    compare_view = compare_current_forward[
        [col for col in compare_columns if col in compare_current_forward.columns]
    ].sort_values(["old_technician", "old_month", "old_sequence_number", "store_number"])
    return {"compare": compare_view, "old_schedule": old_schedule, "new_schedule": new_schedule, "protected": protected}


def pmt_reconciliation_compare_workbook_bytes(preview):
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        preview.get("compare", pd.DataFrame()).to_excel(writer, index=False, sheet_name="Compare Old vs New")
        preview.get("old_schedule", pd.DataFrame()).to_excel(writer, index=False, sheet_name="Old Schedule")
        preview.get("new_schedule", pd.DataFrame()).to_excel(writer, index=False, sheet_name="New Schedule Preview")
        preview.get("protected", pd.DataFrame()).to_excel(writer, index=False, sheet_name="Protected Techs")
    return buffer.getvalue()


SCHEDULE_EXPORT_COLOR_RULES = [
    ("#dbeafe", "Blue", "Transferred to the current assigned PMT by reconciliation"),
    ("#dcfce7", "Green", "Route order or schedule date changed after reconciliation"),
    ("#f3e8ff", "Purple", "Superseded or transferred off old PMT schedule"),
    ("#e5e7eb", "Gray", "Old schedule snapshot row"),
]


def schedule_export_change_type(row, is_snapshot=False):
    if is_snapshot:
        return "Old schedule snapshot"
    source = clean(row.get("schedule_source", "")).lower()
    status = normalize_schedule_status(row.get("status", ""))
    original_date = row.get("original_schedule_date")
    schedule_date = row.get("schedule_date")
    if status in {"transferred", "superseded"} or "superseded" in source:
        return "Superseded/transferred old row"
    if pd.notna(original_date) and pd.notna(schedule_date) and original_date != schedule_date:
        return "Date changed after reconciliation"
    if "reconciliation" in source or "territory transfer" in source:
        previous = clean(row.get("Previous Technician", ""))
        current = clean(row.get("technician", ""))
        if previous and current and previous != current:
            return "Transferred to current assigned PMT"
        return "Unchanged"
    return "Unchanged"


def schedule_export_row_color(change_type):
    if change_type == "Transferred to current assigned PMT":
        return "#dbeafe"
    if change_type == "Date changed after reconciliation":
        return "#dcfce7"
    if change_type == "Superseded/transferred old row":
        return "#f3e8ff"
    if change_type == "Old schedule snapshot":
        return "#e5e7eb"
    return ""


def prepare_schedule_export_view(df, is_snapshot=False, old_lookup=None):
    if df.empty:
        return df
    view = df.copy()
    view["_month_sort"] = pd.to_datetime(view.get("schedule_date"), errors="coerce")
    view["Previous Technician"] = ""
    view["New Technician"] = view.get("technician", "")
    if is_snapshot:
        view["Previous Technician"] = view.get("technician", "")
        view["New Technician"] = ""
    elif old_lookup:
        def previous_technician(row):
            store_key = reconciliation_store_number(row.get("store_number"))
            old_value = old_lookup.get(store_key, "")
            current_value = clean(row.get("technician"))
            return old_value if old_value and old_value != current_value else ""

        view["Previous Technician"] = view.apply(previous_technician, axis=1)
    view["Change Type"] = view.apply(lambda row: schedule_export_change_type(row, is_snapshot=is_snapshot), axis=1)
    export_columns = [
        "store_number", "Previous Technician", "New Technician", "month", "schedule_date",
        "original_schedule_date", "sequence_number", "status", "cycle_label", "schedule_source",
        "notes", "Change Type",
    ]
    sort_columns = [col for col in ["_month_sort", "New Technician", "Previous Technician", "sequence_number", "store_number"] if col in view.columns]
    view = view.sort_values(sort_columns)
    return view[[col for col in export_columns if col in view.columns]]


def current_month_forward_schedule_rows(df, cycle_start=None, cycle_end=None):
    if df.empty or "schedule_date" not in df.columns:
        return df.copy()
    scoped = df.copy()
    schedule_dates = pd.to_datetime(scoped["schedule_date"], errors="coerce").dt.date
    mask = schedule_dates >= month_start(date.today())
    if cycle_start is not None:
        mask &= schedule_dates >= cycle_start
    if cycle_end is not None:
        mask &= schedule_dates <= cycle_end
    return scoped[mask].copy()


def ordered_export_months(df):
    if df.empty or "month" not in df.columns or "schedule_date" not in df.columns:
        return []
    month_order = (
        df.assign(_month_sort=pd.to_datetime(df["schedule_date"], errors="coerce"))
        .dropna(subset=["_month_sort"])
        .assign(_month_sort=lambda value: value["_month_sort"].dt.to_period("M").dt.to_timestamp())
        .sort_values("_month_sort")
        .drop_duplicates("month")["month"]
        .tolist()
    )
    return month_order


def old_schedule_technician_lookup(old_run_items):
    if old_run_items.empty or not {"store_number", "technician"}.issubset(old_run_items.columns):
        return {}
    old_lookup_df = old_run_items.copy()
    old_lookup_df["_store_key"] = old_lookup_df["store_number"].apply(reconciliation_store_number)
    return (
        old_lookup_df.sort_values(["schedule_date", "sequence_number"])
        .dropna(subset=["_store_key"])
        .drop_duplicates("_store_key", keep="first")
        .set_index("_store_key")["technician"]
        .fillna("")
        .astype(str)
        .to_dict()
    )


def reconciliation_schedule_export_workbook_bytes(new_run_items, old_run_items=None, run_name="", snapshot_name=""):
    from openpyxl.styles import Font, PatternFill

    old_run_items = old_run_items if old_run_items is not None else pd.DataFrame()
    old_lookup = old_schedule_technician_lookup(old_run_items)
    current_export = prepare_schedule_export_view(new_run_items, is_snapshot=False, old_lookup=old_lookup)
    old_export = prepare_schedule_export_view(old_run_items, is_snapshot=True)
    changed_export = current_export[
        current_export.get("Change Type", pd.Series(dtype=str)).astype(str).ne("Unchanged")
    ].copy() if not current_export.empty else pd.DataFrame()
    legend = pd.DataFrame(
        [{"Color Code": name, "Meaning": meaning, "Excel Fill": color} for color, name, meaning in SCHEDULE_EXPORT_COLOR_RULES]
    )
    summary = pd.DataFrame(
        [
            {"Metric": "Updated Schedule Run", "Value": run_name},
            {"Metric": "Old Snapshot Schedule", "Value": snapshot_name or "Not selected"},
            {"Metric": "Current Schedule Rows", "Value": len(current_export)},
            {"Metric": "Changed Current Rows", "Value": len(changed_export)},
            {"Metric": "Old Snapshot Rows", "Value": len(old_export)},
        ]
    )
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        used_sheets = set()
        summary.to_excel(writer, index=False, sheet_name="Summary")
        used_sheets.add("Summary")
        legend.to_excel(writer, index=False, sheet_name="Color Legend")
        used_sheets.add("Color Legend")
        current_export.to_excel(writer, index=False, sheet_name="New Normal Schedule")
        used_sheets.add("New Normal Schedule")
        changed_export.to_excel(writer, index=False, sheet_name="New Schedule Changes")
        used_sheets.add("New Schedule Changes")
        old_export.to_excel(writer, index=False, sheet_name="Old Schedule Snapshot")
        used_sheets.add("Old Schedule Snapshot")

        new_month_sheet_names = []
        old_month_sheet_names = []
        if not current_export.empty and "month" in current_export.columns:
            for month_name in ordered_export_months(current_export):
                month_df = current_export[current_export["month"] == month_name].copy()
                sheet_name = safe_sheet_name(f"New {month_name}", used_sheets)
                used_sheets.add(sheet_name)
                month_df.to_excel(writer, index=False, sheet_name=sheet_name)
                new_month_sheet_names.append(sheet_name)
        if not old_export.empty and "month" in old_export.columns:
            for month_name in ordered_export_months(old_export):
                month_df = old_export[old_export["month"] == month_name].copy()
                sheet_name = safe_sheet_name(f"Old {month_name}", used_sheets)
                used_sheets.add(sheet_name)
                month_df.to_excel(writer, index=False, sheet_name=sheet_name)
                old_month_sheet_names.append(sheet_name)

        workbook = writer.book
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet.freeze_panes = "A2"
            if sheet.max_row and sheet.max_column:
                sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.font = Font(bold=True)

        for row_index, (color, _, _) in enumerate(SCHEDULE_EXPORT_COLOR_RULES, start=2):
            fill = PatternFill("solid", fgColor=color.replace("#", "").upper())
            for cell in workbook["Color Legend"][row_index]:
                cell.fill = fill

        def color_sheet(sheet_name):
            sheet = workbook[sheet_name]
            if sheet.max_row < 2:
                return
            headers = [cell.value for cell in sheet[1]]
            try:
                change_col_index = headers.index("Change Type") + 1
            except ValueError:
                return
            for row_index in range(2, sheet.max_row + 1):
                change_type = sheet.cell(row=row_index, column=change_col_index).value
                color = schedule_export_row_color(change_type)
                if not color:
                    continue
                fill = PatternFill("solid", fgColor=color.replace("#", "").upper())
                for cell in sheet[row_index]:
                    cell.fill = fill

        for sheet_name in ["New Normal Schedule", "New Schedule Changes", "Old Schedule Snapshot"] + new_month_sheet_names + old_month_sheet_names:
            color_sheet(sheet_name)
    return buffer.getvalue()


def pmt_team_for_employee(session, employee):
    if not employee:
        return None
    team = session.scalars(
        select(Team).where(Team.team_name == employee.full_name, Team.team_type == "PMT")
    ).first()
    if not team:
        team = Team(
            team_name=employee.full_name,
            team_type="PMT",
            city=employee.home_city or employee.base_city or "",
            state=employee.home_state or employee.base_state or "",
            active=True,
        )
        session.add(team)
        session.flush()
    else:
        team.active = True
    return team


def is_active_pmt_schedule_item(item):
    if not item:
        return False
    if "PMT" not in clean(getattr(item, "work_type", "")).upper():
        return False
    terminal_statuses = PMT_COMPLETED_STATUSES | PMT_CANCELED_STATUSES
    return normalize_schedule_status(getattr(item, "status", "")) not in terminal_statuses


def create_pmt_reconciliation_snapshots(session, selected_ids, reason=""):
    if not selected_ids:
        return 0
    source_run_ids = {
        int(run_id)
        for run_id in session.scalars(
            select(ScheduleItem.pmt_schedule_run_id).where(
                ScheduleItem.id.in_([int(value) for value in selected_ids]),
                ScheduleItem.pmt_schedule_run_id.is_not(None),
            )
        ).all()
        if run_id
    }
    snapshots_created = 0
    timestamp = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    for run_id in sorted(source_run_ids):
        source_run = session.get(PMTScheduleRun, int(run_id))
        if not source_run:
            continue
        source_items = session.scalars(
            select(ScheduleItem)
            .where(ScheduleItem.pmt_schedule_run_id == int(run_id))
            .order_by(ScheduleItem.schedule_date, ScheduleItem.employee_id, ScheduleItem.sequence_number, ScheduleItem.id)
        ).all()
        if not source_items:
            continue
        snapshot_run = PMTScheduleRun(
            run_name=f"Snapshot before reconciliation - {source_run.run_name} - {timestamp}",
            cycle_start=source_run.cycle_start,
            cycle_end=source_run.cycle_end,
            months=source_run.months,
            default_monthly_target=source_run.default_monthly_target,
            direction=source_run.direction,
            schedule_mode=source_run.schedule_mode,
            distance_method=source_run.distance_method,
            status="Snapshot",
            technician_count=source_run.technician_count,
            store_count=source_run.store_count,
            unscheduled_count=source_run.unscheduled_count,
            created_by=st.session_state.get("username", ""),
            notes=f"Read-only old schedule snapshot before PMT reconciliation. Source run #{source_run.id}. Reason: {clean(reason)}",
        )
        session.add(snapshot_run)
        session.flush()
        snapshot_schedule = Schedule(
            schedule_name=snapshot_run.run_name,
            schedule_type="PMT Reconciliation Snapshot",
            start_date=source_run.cycle_start,
            end_date=source_run.cycle_end,
            status="Snapshot",
            created_by=st.session_state.get("username", ""),
            notes=f"Old schedule snapshot for source PMT run #{source_run.id}.",
        )
        session.add(snapshot_schedule)
        session.flush()
        for item in source_items:
            session.add(
                ScheduleItem(
                    schedule_id=snapshot_schedule.id,
                    schedule_date=item.schedule_date,
                    sequence_number=item.sequence_number,
                    store_id=item.store_id,
                    employee_id=item.employee_id,
                    team_id=item.team_id,
                    work_type=item.work_type,
                    schedule_source="PMT Reconciliation Old Schedule Snapshot",
                    pmt_schedule_run_id=snapshot_run.id,
                    cycle_label=item.cycle_label,
                    deferred_work_order_id=item.deferred_work_order_id,
                    planned_start_time=item.planned_start_time,
                    planned_end_time=item.planned_end_time,
                    status=item.status,
                    original_schedule_date=item.original_schedule_date,
                    rescheduled_from_item_id=item.rescheduled_from_item_id,
                    rain_delay=item.rain_delay,
                    weather_notes=item.weather_notes,
                    completion_notes=item.completion_notes,
                )
            )
        snapshots_created += 1
    return snapshots_created


def apply_pmt_reconciliation_transfers(schedule_item_ids, effective_date, reason="", monthly_target=10):
    selected_ids = [int(value) for value in schedule_item_ids if pd.notna(value)]
    if not selected_ids:
        return {"transferred": 0, "superseded": 0, "resequenced_rows": 0, "snapshots_created": 0}
    transferred = 0
    superseded = 0
    resequenced_rows = 0
    rebuilt_rows = 0
    rebuild_overflow = 0
    touched_months = set()
    affected_rebuilds = {}
    with session_scope("PMT schedule reconciliation transfer") as session:
        snapshots_created = create_pmt_reconciliation_snapshots(session, selected_ids, reason)
        team_cache = {}
        for item_id in selected_ids:
            item = session.get(ScheduleItem, int(item_id))
            if not is_active_pmt_schedule_item(item):
                continue
            store = session.get(Store, int(item.store_id)) if item.store_id else None
            assigned_employee_id = int(store.assigned_pmt_employee_id) if store and store.assigned_pmt_employee_id else None
            if not assigned_employee_id or item.employee_id == assigned_employee_id:
                continue
            duplicate = session.scalars(
                select(ScheduleItem)
                .where(
                    ScheduleItem.id != int(item.id),
                    func.upper(func.coalesce(ScheduleItem.work_type, "")).like("%PMT%"),
                    ScheduleItem.store_id == int(item.store_id),
                    ScheduleItem.employee_id == assigned_employee_id,
                    func.lower(func.trim(func.coalesce(ScheduleItem.status, "scheduled"))).notin_(list(PMT_COMPLETED_STATUSES | PMT_CANCELED_STATUSES)),
                )
                .order_by(ScheduleItem.schedule_date, ScheduleItem.sequence_number, ScheduleItem.id)
            ).first()
            old_employee_id = item.employee_id
            old_month_value = month_start(item.schedule_date)
            target_schedule_date = item.schedule_date
            if item.schedule_date < month_start(date.today()):
                target_schedule_date = first_workday(month_start(date.today()), employee_id=assigned_employee_id)
            if duplicate:
                item.status = "Transferred"
                item.schedule_source = "PMT Reconciliation Superseded"
                note_parts = [
                    clean(item.completion_notes),
                    f"Reconciliation superseded by item #{duplicate.id} for current assigned PMT.",
                    clean(reason),
                ]
                item.completion_notes = " | ".join([part for part in note_parts if part])
                superseded += 1
                touched_months.add((old_employee_id, old_month_value, item.pmt_schedule_run_id))
                continue
            employee = session.get(Employee, assigned_employee_id)
            if assigned_employee_id not in team_cache:
                team_cache[assigned_employee_id] = pmt_team_for_employee(session, employee)
            team = team_cache[assigned_employee_id]
            if item.original_schedule_date is None:
                item.original_schedule_date = item.schedule_date
            item.employee_id = assigned_employee_id
            item.team_id = int(team.id) if team else None
            item.schedule_date = target_schedule_date
            item.cycle_label = month_label(month_start(target_schedule_date))
            item.schedule_source = "PMT Assignment Reconciliation"
            note_parts = [
                clean(item.completion_notes),
                f"Transferred from employee_id={old_employee_id} to employee_id={assigned_employee_id} after PMT assignment reconciliation effective {effective_date}.",
                f"Original schedule date {item.original_schedule_date}; new schedule date {target_schedule_date}.",
                clean(reason),
            ]
            item.completion_notes = " | ".join([part for part in note_parts if part])
            transferred += 1
            touched_months.add((old_employee_id, old_month_value, item.pmt_schedule_run_id))
            touched_months.add((assigned_employee_id, month_start(target_schedule_date), item.pmt_schedule_run_id))
            rebalance_start = max(month_start(target_schedule_date), month_start(effective_date))
            for affected_employee_id in [old_employee_id, assigned_employee_id]:
                rebuild_key = (int(affected_employee_id), item.pmt_schedule_run_id)
                affected_rebuilds[rebuild_key] = min(affected_rebuilds.get(rebuild_key, rebalance_start), rebalance_start)
        for employee_id, month_value, run_id in touched_months:
            if (int(employee_id), run_id) not in affected_rebuilds:
                resequenced_rows += resequence_pmt_month(session, run_id, employee_id, month_value)
        for (employee_id, run_id), start_month_value in affected_rebuilds.items():
            result = rebuild_pmt_employee_from_current_assignments(session, run_id, employee_id, start_month_value, monthly_target, reason)
            rebuilt_rows += int(result.get("scheduled", 0))
            rebuild_overflow += int(result.get("overflow", 0))
            superseded += int(result.get("superseded", 0))
    log_action("pmt schedule reconciliation applied", "schedule_items", description=f"{transferred} transferred; {superseded} superseded; {resequenced_rows} resequenced; {rebuilt_rows} assignment-source rows rebuilt; {rebuild_overflow} overflow; {snapshots_created} snapshot run(s) created. Effective {effective_date}. {reason}")
    return {"transferred": transferred, "superseded": superseded, "resequenced_rows": resequenced_rows, "rebuilt_rows": rebuilt_rows, "rebuild_overflow": rebuild_overflow, "snapshots_created": snapshots_created}


def resequence_pmt_month(session, run_id, employee_id, month_start_value):
    if run_id is None or employee_id is None or month_start_value is None:
        return 0
    start_value = month_start(month_start_value)
    items = session.scalars(
        select(ScheduleItem)
        .where(
            ScheduleItem.pmt_schedule_run_id == int(run_id),
            ScheduleItem.employee_id == int(employee_id),
            func.upper(func.coalesce(ScheduleItem.work_type, "")).like("%PMT%"),
            func.lower(func.trim(func.coalesce(ScheduleItem.status, "scheduled"))).notin_(list(PMT_COMPLETED_STATUSES | PMT_CANCELED_STATUSES)),
            ScheduleItem.schedule_date >= start_value,
            ScheduleItem.schedule_date < add_months(start_value, 1),
        )
        .order_by(ScheduleItem.schedule_date, ScheduleItem.sequence_number, ScheduleItem.store_id, ScheduleItem.id)
    ).all()
    employee = session.get(Employee, int(employee_id))
    home_lat = to_float(getattr(employee, "home_latitude", None)) if employee else None
    home_lon = to_float(getattr(employee, "home_longitude", None)) if employee else None

    def route_sort_key(item):
        store = session.get(Store, int(item.store_id)) if item.store_id else None
        store_lat = to_float(getattr(store, "latitude", None)) if store else None
        store_lon = to_float(getattr(store, "longitude", None)) if store else None
        if home_lat is not None and home_lon is not None and store_lat is not None and store_lon is not None:
            distance = haversine_miles(home_lat, home_lon, store_lat, store_lon)
        else:
            distance = 999999
        return (distance, clean(getattr(store, "store_number", "")) if store else "", item.schedule_date, item.sequence_number or 0, item.id)

    for index, item in enumerate(sorted(items, key=route_sort_key), start=1):
        item.sequence_number = index
    return len(items)


def rebalance_pmt_employee_future_schedule(session, run_id, employee_id, start_month_value, reason=""):
    if run_id is None or employee_id is None or start_month_value is None:
        return {"items": 0, "months": 0}
    run = session.get(PMTScheduleRun, int(run_id))
    employee = session.get(Employee, int(employee_id))
    start_value = month_start(start_month_value)
    if run and run.cycle_start:
        start_value = max(start_value, month_start(run.cycle_start))
    monthly_target = int(getattr(employee, "monthly_pmt_store_target", None) or getattr(run, "default_monthly_target", None) or 10)
    monthly_target = max(1, monthly_target)
    cycle_end = getattr(run, "cycle_end", None)
    items_query = (
        select(ScheduleItem)
        .where(
            ScheduleItem.pmt_schedule_run_id == int(run_id),
            ScheduleItem.employee_id == int(employee_id),
            func.upper(func.coalesce(ScheduleItem.work_type, "")).like("%PMT%"),
            func.lower(func.trim(func.coalesce(ScheduleItem.status, "scheduled"))).notin_(list(PMT_COMPLETED_STATUSES | PMT_CANCELED_STATUSES)),
            ScheduleItem.schedule_date >= start_value,
        )
        .order_by(ScheduleItem.schedule_date, ScheduleItem.sequence_number, ScheduleItem.store_id, ScheduleItem.id)
    )
    if cycle_end:
        items_query = items_query.where(ScheduleItem.schedule_date <= cycle_end)
    items = session.scalars(items_query).all()
    if not items:
        return {"items": 0, "months": 0}
    if cycle_end:
        months_available = 1
        cursor = start_value
        while add_months(cursor, 1) <= month_start(cycle_end):
            months_available += 1
            cursor = add_months(cursor, 1)
        monthly_target = max(monthly_target, (len(items) + months_available - 1) // months_available)
    home_lat = to_float(getattr(employee, "home_latitude", None)) if employee else None
    home_lon = to_float(getattr(employee, "home_longitude", None)) if employee else None
    store_cache = {}

    def item_store(item):
        if item.store_id not in store_cache:
            store_cache[item.store_id] = session.get(Store, int(item.store_id)) if item.store_id else None
        return store_cache[item.store_id]

    def rebalance_sort_key(item):
        store = item_store(item)
        store_lat = to_float(getattr(store, "latitude", None)) if store else None
        store_lon = to_float(getattr(store, "longitude", None)) if store else None
        distance = haversine_miles(home_lat, home_lon, store_lat, store_lon) if home_lat is not None and home_lon is not None and store_lat is not None and store_lon is not None else 999999
        original_date = item.original_schedule_date or item.schedule_date
        carry_forward = original_date < start_value
        return (0 if carry_forward else 1, distance, clean(getattr(store, "store_number", "")) if store else "", original_date, item.sequence_number or 0, item.id)

    cursor_month = start_value
    sequence_number = 1
    touched_months = set()
    schedule_mode = clean(getattr(run, "schedule_mode", "")) if run else ""
    for item in sorted(items, key=rebalance_sort_key):
        if sequence_number > monthly_target:
            cursor_month = add_months(cursor_month, 1)
            sequence_number = 1
        if cycle_end and cursor_month > month_start(cycle_end):
            cursor_month = month_start(cycle_end)
        new_date = first_workday(cursor_month, employee_id=int(employee_id))
        if schedule_mode == "Monthly schedule with daily stops":
            new_date = first_workday(cursor_month + timedelta(days=(sequence_number - 1)), employee_id=int(employee_id))
        old_month = month_start(item.schedule_date)
        if item.original_schedule_date is None:
            item.original_schedule_date = item.schedule_date
        item.schedule_date = new_date
        item.sequence_number = sequence_number
        item.cycle_label = month_label(cursor_month)
        item.schedule_source = "PMT Assignment Reconciliation Rebalanced"
        note_parts = [
            clean(item.completion_notes),
            f"Rebalanced affected PMT future schedule from {start_value} using monthly target {monthly_target}.",
            clean(reason),
        ]
        item.completion_notes = " | ".join([part for part in note_parts if part])
        touched_months.add(old_month)
        touched_months.add(cursor_month)
        sequence_number += 1
    return {"items": len(items), "months": len(touched_months)}


def rebuild_pmt_employee_from_current_assignments(session, run_id, employee_id, start_month_value, monthly_target=10, reason=""):
    run = session.get(PMTScheduleRun, int(run_id)) if run_id is not None else None
    employee = session.get(Employee, int(employee_id)) if employee_id is not None else None
    if run is None or employee is None:
        return {"assigned": 0, "scheduled": 0, "overflow": 0, "created": 0, "updated": 0, "superseded": 0}
    start_value = month_start(start_month_value)
    if run.cycle_start:
        start_value = max(start_value, month_start(run.cycle_start))
    cycle_end = run.cycle_end or add_months(start_value, 6) - timedelta(days=1)
    if start_value > cycle_end:
        return {"assigned": 0, "scheduled": 0, "overflow": 0, "created": 0, "updated": 0, "superseded": 0}
    schedule_id = session.scalar(
        select(ScheduleItem.schedule_id)
        .where(ScheduleItem.pmt_schedule_run_id == int(run_id))
        .order_by(ScheduleItem.schedule_id)
    )
    if schedule_id is None:
        return {"assigned": 0, "scheduled": 0, "overflow": 0, "created": 0, "updated": 0, "superseded": 0}
    stores = session.execute(
        select(Store)
        .where(
            Store.active == True,  # noqa: E712
            Store.assigned_pmt_employee_id == int(employee_id),
        )
        .order_by(Store.store_number)
    ).scalars().all()
    assigned_count = len(stores)
    monthly_target = max(1, int(monthly_target or getattr(employee, "monthly_pmt_store_target", None) or run.default_monthly_target or 10))
    months_available = 1
    cursor = start_value
    while add_months(cursor, 1) <= month_start(cycle_end):
        months_available += 1
        cursor = add_months(cursor, 1)
    capacity = monthly_target * months_available
    schedule_history, backlog_history = pmt_store_history()
    assignment_rows = []
    home_lat = to_float(getattr(employee, "home_latitude", None))
    home_lon = to_float(getattr(employee, "home_longitude", None))
    for store in stores:
        distance = None
        if home_lat is not None and home_lon is not None and store.latitude is not None and store.longitude is not None:
            distance = haversine_miles(home_lat, home_lon, float(store.latitude), float(store.longitude))
        assignment_rows.append(
            {
                "employee_id": int(employee_id),
                "technician_name": employee.full_name,
                "store_id": int(store.id),
                "store_number": clean(store.store_number),
                "store_address": clean(store.address),
                "store_city": clean(store.city),
                "store_state": clean(store.state),
                "store_zip": clean(store.zip),
                "latitude": store.latitude,
                "longitude": store.longitude,
                "home_latitude": home_lat,
                "home_longitude": home_lon,
                "distance_from_home": distance,
            }
        )
    assignments = pd.DataFrame(assignment_rows)
    if assignments.empty:
        active_employee_items = session.scalars(
            select(ScheduleItem).where(
                ScheduleItem.pmt_schedule_run_id == int(run_id),
                ScheduleItem.employee_id == int(employee_id),
                func.upper(func.coalesce(ScheduleItem.work_type, "")).like("%PMT%"),
                func.lower(func.trim(func.coalesce(ScheduleItem.status, "scheduled"))).notin_(list(PMT_COMPLETED_STATUSES | PMT_CANCELED_STATUSES)),
                ScheduleItem.schedule_date >= start_value,
                ScheduleItem.schedule_date <= cycle_end,
            )
        ).all()
        for item in active_employee_items:
            if item.original_schedule_date is None:
                item.original_schedule_date = item.schedule_date
            item.status = "Transferred"
            item.schedule_source = "PMT Reconciliation Superseded - No Current Assignment"
        return {"assigned": 0, "scheduled": 0, "overflow": 0, "created": 0, "updated": 0, "superseded": len(active_employee_items)}
    schedulable = assignments.dropna(subset=["distance_from_home"]).copy()
    unschedulable = assignments[assignments["distance_from_home"].isna()].copy()
    prioritized = apply_pmt_rotation_priority(schedulable, schedule_history, backlog_history, start_value) if not schedulable.empty else pd.DataFrame()
    sort_columns = ["rotation_priority_group", "distance_from_home", "store_number"]
    scheduled = prioritized.sort_values(sort_columns, ascending=[True, True, True]).head(capacity).copy() if not prioritized.empty else pd.DataFrame()
    overflow = prioritized.drop(index=scheduled.index, errors="ignore").copy() if not prioritized.empty else pd.DataFrame()
    if not unschedulable.empty:
        overflow = pd.concat([overflow, unschedulable], ignore_index=True)
    scheduled_store_ids = set(scheduled["store_id"].dropna().astype(int).tolist()) if not scheduled.empty else set()
    assigned_store_ids = set(assignments["store_id"].dropna().astype(int).tolist())
    impacted_items = session.scalars(
        select(ScheduleItem).where(
            ScheduleItem.pmt_schedule_run_id == int(run_id),
            func.upper(func.coalesce(ScheduleItem.work_type, "")).like("%PMT%"),
            func.lower(func.trim(func.coalesce(ScheduleItem.status, "scheduled"))).notin_(list(PMT_COMPLETED_STATUSES | PMT_CANCELED_STATUSES)),
            ScheduleItem.schedule_date >= start_value,
            ScheduleItem.schedule_date <= cycle_end,
        )
    ).all()
    active_by_store = {}
    active_employee_items = []
    for item in impacted_items:
        if item.employee_id == int(employee_id):
            active_employee_items.append(item)
        if item.store_id in assigned_store_ids:
            active_by_store.setdefault(int(item.store_id), []).append(item)
    created = 0
    updated = 0
    superseded = 0
    superseded_item_ids = set()
    now = datetime.utcnow()
    team = pmt_team_for_employee(session, employee)
    current_month = start_value
    sequence_number = 1
    if not scheduled.empty:
        remaining = scheduled.copy()
        for month_index in range(months_available):
            if remaining.empty:
                break
            current_month = add_months(start_value, month_index)
            month_pool = remaining.sort_values(sort_columns, ascending=[True, True, True]).head(monthly_target).copy()
            routed_rows = home_distance_route(month_pool, home_lat, home_lon, limit=monthly_target) if home_lat is not None and home_lon is not None else [
                row.copy() for _, row in month_pool.sort_values(["distance_from_home", "store_number"], na_position="last").iterrows()
            ]
            remaining = remaining.drop(index=[row.name for row in routed_rows], errors="ignore")
            for sequence_number, row in enumerate(routed_rows, start=1):
                store_id = int(row["store_id"])
                existing_rows = active_by_store.get(store_id, [])
                target_rows = [item for item in existing_rows if item.employee_id == int(employee_id)]
                keep_item = target_rows[0] if target_rows else existing_rows[0] if existing_rows else None
                schedule_date = first_workday(current_month, employee_id=int(employee_id))
                if clean(run.schedule_mode) == "Monthly schedule with daily stops":
                    schedule_date = first_workday(current_month + timedelta(days=(sequence_number - 1)), employee_id=int(employee_id))
                if keep_item is None:
                    keep_item = ScheduleItem(
                        schedule_id=int(schedule_id),
                        schedule_date=schedule_date,
                        sequence_number=int(sequence_number),
                        store_id=store_id,
                        employee_id=int(employee_id),
                        team_id=int(team.id) if team else None,
                        work_type="PMT",
                        status="Scheduled",
                        schedule_source="PMT Assignment Reconciliation Rebuild",
                        pmt_schedule_run_id=int(run_id),
                        cycle_label=month_label(current_month),
                        completion_notes=clean(reason),
                        created_at=now,
                    )
                    session.add(keep_item)
                    session.flush()
                    created += 1
                else:
                    if keep_item.original_schedule_date is None and keep_item.schedule_date != schedule_date:
                        keep_item.original_schedule_date = keep_item.schedule_date
                    keep_item.employee_id = int(employee_id)
                    keep_item.team_id = int(team.id) if team else None
                    keep_item.schedule_date = schedule_date
                    keep_item.sequence_number = int(sequence_number)
                    keep_item.status = "Scheduled"
                    keep_item.schedule_source = "PMT Assignment Reconciliation Rebuild"
                    keep_item.cycle_label = month_label(current_month)
                    keep_item.completion_notes = " | ".join([part for part in [clean(keep_item.completion_notes), clean(reason)] if part])
                    updated += 1
                for item in existing_rows:
                    if item.id == keep_item.id:
                        continue
                    if item.original_schedule_date is None:
                        item.original_schedule_date = item.schedule_date
                    item.status = "Transferred"
                    item.schedule_source = "PMT Assignment Reconciliation Rebuild Superseded"
                    item.completion_notes = " | ".join([part for part in [clean(item.completion_notes), f"Superseded by rebuilt schedule item #{keep_item.id}.", clean(reason)] if part])
                    superseded_item_ids.add(int(item.id))
                    superseded += 1
    for store_id, existing_rows in active_by_store.items():
        if store_id in scheduled_store_ids:
            continue
        for item in existing_rows:
            if int(item.id) in superseded_item_ids:
                continue
            if item.original_schedule_date is None:
                item.original_schedule_date = item.schedule_date
            item.status = "Transferred"
            item.schedule_source = "PMT Assignment Reconciliation Rebuild Overflow Removed"
            item.completion_notes = " | ".join([part for part in [clean(item.completion_notes), "Removed from active future schedule because this current assignment did not fit the rebuild capacity.", clean(reason)] if part])
            superseded_item_ids.add(int(item.id))
            superseded += 1
    for item in active_employee_items:
        if item.store_id in scheduled_store_ids:
            continue
        if int(item.id) in superseded_item_ids:
            continue
        if item.original_schedule_date is None:
            item.original_schedule_date = item.schedule_date
        item.status = "Transferred"
        item.schedule_source = "PMT Assignment Reconciliation Rebuild Removed"
        item.completion_notes = " | ".join([part for part in [clean(item.completion_notes), "Removed from active future schedule because the PMT rebuild uses current assignments and monthly capacity.", clean(reason)] if part])
        superseded_item_ids.add(int(item.id))
        superseded += 1
    overflow_count = 0
    if not overflow.empty:
        for _, row in overflow.iterrows():
            store_id = scalar_int(row.get("store_id"), 0)
            if not store_id:
                continue
            existing = session.query(PMTScheduleBacklog).filter(
                PMTScheduleBacklog.pmt_schedule_run_id == int(run_id),
                PMTScheduleBacklog.employee_id == int(employee_id),
                PMTScheduleBacklog.store_id == store_id,
                PMTScheduleBacklog.status == "Not Scheduled",
            ).first()
            backlog = existing or PMTScheduleBacklog(
                pmt_schedule_run_id=int(run_id),
                employee_id=int(employee_id),
                store_id=store_id,
                cycle_start=start_value,
                cycle_end=cycle_end,
            )
            if not existing:
                session.add(backlog)
            backlog.status = "Not Scheduled"
            backlog.reason = "Did not fit PMT reconciliation rebuild capacity or missing coordinates"
            backlog.cycles_missed = max(int(getattr(backlog, "cycles_missed", 0) or 0), scalar_int(row.get("cycles_missed"), 1))
            backlog.priority_score = max(int(getattr(backlog, "priority_score", 0) or 0), scalar_int(row.get("rotation_priority_score"), 1000))
            backlog.last_scheduled_month = scalar_date(row.get("last_scheduled_month"))
            backlog.last_completed_date = scalar_date(row.get("last_completed_date"))
            backlog.last_completed_month = month_start(backlog.last_completed_date) if backlog.last_completed_date else None
            backlog.notes = f"PMT reconciliation rebuild run {run_id}: {backlog.reason}"
            overflow_count += 1
    run.store_count = int(session.scalar(
        select(func.count(func.distinct(ScheduleItem.store_id))).where(
            ScheduleItem.pmt_schedule_run_id == int(run_id),
            ScheduleItem.work_type == "PMT",
            ScheduleItem.status.in_(PMT_ACTIVE_STATUS_VALUES),
        )
    ) or 0)
    run.unscheduled_count = int(session.scalar(
        select(func.count(PMTScheduleBacklog.id)).where(
            PMTScheduleBacklog.pmt_schedule_run_id == int(run_id),
            PMTScheduleBacklog.status.in_(PMT_BACKLOG_OPEN_STATUSES),
        )
    ) or 0)
    return {
        "assigned": assigned_count,
        "scheduled": len(scheduled_store_ids),
        "overflow": overflow_count,
        "created": created,
        "updated": updated,
        "superseded": superseded,
        "monthly_target": monthly_target,
        "months_available": months_available,
    }


def resolve_pmt_conflicts_keep_assigned(run_id, store_ids, notes=""):
    store_ids = [int(value) for value in store_ids if pd.notna(value)]
    if not store_ids:
        return {"stores": 0, "superseded": 0, "moved": 0, "resequenced_rows": 0}
    superseded = 0
    moved = 0
    touched_months = set()
    touched_store_ids = set()
    with session_scope("Resolve PMT schedule conflicts") as session:
        for store_id in sorted(set(store_ids)):
            store = session.get(Store, int(store_id))
            assigned_employee_id = int(store.assigned_pmt_employee_id) if store and store.assigned_pmt_employee_id else None
            active_items = session.scalars(
                select(ScheduleItem)
                .where(
                    ScheduleItem.pmt_schedule_run_id == int(run_id),
                    ScheduleItem.store_id == int(store_id),
                    ScheduleItem.work_type == "PMT",
                    ScheduleItem.status.in_(PMT_ACTIVE_STATUS_VALUES),
                )
                .order_by(ScheduleItem.schedule_date, ScheduleItem.sequence_number, ScheduleItem.id)
            ).all()
            if not active_items or assigned_employee_id is None:
                continue
            owner_items = [item for item in active_items if item.employee_id == assigned_employee_id]
            if owner_items:
                keep_item = owner_items[0]
            else:
                keep_item = active_items[0]
                old_employee_id = keep_item.employee_id
                if keep_item.original_schedule_date is None:
                    keep_item.original_schedule_date = keep_item.schedule_date
                keep_item.employee_id = assigned_employee_id
                keep_item.schedule_source = "PMT Territory Transfer"
                moved += 1
                touched_months.add((old_employee_id, month_start(keep_item.schedule_date)))
            touched_months.add((keep_item.employee_id, month_start(keep_item.schedule_date)))
            for item in active_items:
                if item.id == keep_item.id:
                    continue
                if item.original_schedule_date is None:
                    item.original_schedule_date = item.schedule_date
                item.status = "Transferred"
                item.schedule_source = "PMT Territory Transfer Superseded"
                note_parts = [
                    clean(item.completion_notes),
                    f"Superseded by active PMT schedule item #{keep_item.id} for current assigned technician.",
                ]
                if notes:
                    note_parts.append(clean(notes))
                item.completion_notes = " | ".join([part for part in note_parts if part])
                superseded += 1
                touched_months.add((item.employee_id, month_start(item.schedule_date)))
            touched_store_ids.add(store_id)
        resequenced_rows = 0
        for employee_id, month_value in touched_months:
            resequenced_rows += resequence_pmt_month(session, run_id, employee_id, month_value)
        run = session.get(PMTScheduleRun, int(run_id))
        if run:
            active_store_count = session.scalar(
                select(func.count(func.distinct(ScheduleItem.store_id))).where(
                    ScheduleItem.pmt_schedule_run_id == int(run_id),
                    ScheduleItem.work_type == "PMT",
                    ScheduleItem.status.in_(PMT_ACTIVE_STATUS_VALUES),
                )
            )
            run.store_count = int(active_store_count or 0)
    log_action(
        "pmt schedule conflicts resolved",
        "pmt_schedule_runs",
        int(run_id),
        f"Resolved {len(touched_store_ids)} store conflict(s); superseded {superseded}; moved {moved}; resequenced {resequenced_rows}.",
    )
    return {"stores": len(touched_store_ids), "superseded": superseded, "moved": moved, "resequenced_rows": resequenced_rows}


def pmt_order_remaining_by_home_distance(candidate_stores, selected_store_ids):
    if candidate_stores.empty:
        return []
    remaining = candidate_stores[~candidate_stores["already_scheduled"]].copy()
    selected_set = {int(store_id) for store_id in selected_store_ids}
    remaining = remaining.loc[~remaining["store_id"].astype(int).isin(selected_set)].copy()
    if remaining.empty:
        return []
    remaining["_home_distance_sort"] = pd.to_numeric(remaining.get("distance_from_home"), errors="coerce")
    remaining = remaining.sort_values(["_home_distance_sort", "store_number"], ascending=[True, True], na_position="last")
    return remaining["store_id"].dropna().astype(int).tolist()


def add_preview_leg_distances(preview_df):
    if preview_df.empty:
        return preview_df
    preview_df = preview_df.copy()
    preview_df["Distance From Previous Stop"] = None
    previous_lat = previous_lon = None
    for index, row in preview_df.iterrows():
        lat = to_float(row.get("latitude"))
        lon = to_float(row.get("longitude"))
        if previous_lat is not None and previous_lon is not None and lat is not None and lon is not None:
            preview_df.at[index, "Distance From Previous Stop"] = round(haversine_miles(previous_lat, previous_lon, lat, lon), 1)
        previous_lat, previous_lon = lat, lon
    return preview_df


def split_run_items_by_period(run_items, cycle_start, cycle_end):
    if run_items.empty or (cycle_start is None and cycle_end is None):
        return run_items.copy(), pd.DataFrame()
    dated = pd.to_datetime(run_items["schedule_date"], errors="coerce").dt.date
    in_period = pd.Series(True, index=run_items.index)
    if cycle_start is not None:
        in_period &= dated >= cycle_start
    if cycle_end is not None:
        in_period &= dated <= cycle_end
    return run_items[in_period].copy(), run_items[~in_period].copy()


def dataframe_from_session_records(state_key):
    value = st.session_state.get(state_key, [])
    if isinstance(value, pd.DataFrame):
        return value.copy()
    if isinstance(value, list):
        return pd.DataFrame(value)
    if isinstance(value, dict):
        return pd.DataFrame([value])
    st.session_state.pop(state_key, None)
    return pd.DataFrame()


def pmt_month_capacity(run_items, employee_id, default_value=10):
    tech_items = run_items[run_items["employee_id"] == int(employee_id)].copy()
    if tech_items.empty:
        return int(default_value)
    counts = tech_items.groupby("month_start")["schedule_item_id"].count()
    if counts.empty:
        return int(default_value)
    return max(1, int(counts.max()))


def pmt_store_lookup(store_number):
    keys = store_number_keys(store_number)
    if not keys:
        return pd.DataFrame()
    placeholders = ", ".join([f":key_{idx}" for idx, _ in enumerate(keys)])
    params = {f"key_{idx}": value for idx, value in enumerate(keys)}
    return safe_query(
        f"""
        select id as store_id, store_number, address, city, state, zip, latitude, longitude
        from stores
        where store_number in ({placeholders})
        order by store_number
        """,
        params,
    )


def build_pmt_reflow_preview(run_items, employee_id, selected_item_ids, target_month, monthly_capacity, urgent_store=None):
    if run_items.empty:
        return pd.DataFrame()
    target_month = month_start(target_month)
    selected_ids = {int(value) for value in selected_item_ids if pd.notna(value)}
    active_items = run_items[pmt_active_item_mask(run_items)].copy()
    tech_items = active_items[active_items["employee_id"] == int(employee_id)].copy()
    if tech_items.empty and urgent_store is None:
        return pd.DataFrame()
    downstream = tech_items[(tech_items["month_start"] >= target_month) | (tech_items["schedule_item_id"].isin(selected_ids))].copy()
    if not downstream.empty:
        downstream["is_selected_push"] = downstream["schedule_item_id"].isin(selected_ids)
        downstream["sort_month"] = downstream.apply(lambda row: target_month if row["is_selected_push"] else row["month_start"], axis=1)
        downstream["sort_bucket"] = downstream["is_selected_push"].apply(lambda value: 0 if value else 1)
        downstream["preview_action"] = downstream["is_selected_push"].apply(lambda value: "Push incomplete store" if value else "Shift to preserve route order")
    rows = [downstream]
    if urgent_store is not None and not urgent_store.empty:
        store = urgent_store.iloc[0].to_dict()
        existing_match = run_items[
            (run_items["store_id"] == int(store["store_id"]))
            & (run_items["employee_id"] == int(employee_id))
            & pmt_active_item_mask(run_items)
        ]
        existing_any = run_items[(run_items["store_id"] == int(store["store_id"])) & pmt_active_item_mask(run_items)]
        if not existing_match.empty:
            urgent_row = existing_match.iloc[[0]].copy()
            urgent_row["preview_action"] = "Move urgent store earlier"
        elif not existing_any.empty:
            urgent_row = existing_any.iloc[[0]].copy()
            urgent_row["employee_id"] = int(employee_id)
            tech_name = run_items.loc[run_items["employee_id"] == int(employee_id), "technician"].dropna()
            urgent_row["technician"] = tech_name.iloc[0] if not tech_name.empty else ""
            urgent_row["preview_action"] = "Switch urgent store to selected PMT"
        else:
            schedule_id = tech_items["schedule_id"].dropna().iloc[0] if not tech_items.empty else run_items["schedule_id"].dropna().iloc[0]
            tech_name = run_items.loc[run_items["employee_id"] == int(employee_id), "technician"].dropna()
            home_lat = run_items.loc[run_items["employee_id"] == int(employee_id), "home_latitude"].dropna()
            home_lon = run_items.loc[run_items["employee_id"] == int(employee_id), "home_longitude"].dropna()
            urgent_row = pd.DataFrame(
                [
                    {
                        "schedule_item_id": -int(store["store_id"]),
                        "schedule_id": int(schedule_id),
                        "schedule_date": target_month,
                        "sequence_number": 0,
                        "employee_id": int(employee_id),
                        "technician": tech_name.iloc[0] if not tech_name.empty else "",
                        "home_latitude": home_lat.iloc[0] if not home_lat.empty else None,
                        "home_longitude": home_lon.iloc[0] if not home_lon.empty else None,
                        "store_id": int(store["store_id"]),
                        "store_number": store.get("store_number", ""),
                        "address": store.get("address", ""),
                        "city": store.get("city", ""),
                        "state": store.get("state", ""),
                        "zip": store.get("zip", ""),
                        "latitude": store.get("latitude"),
                        "longitude": store.get("longitude"),
                        "work_type": "PMT",
                        "status": "Scheduled",
                        "cycle_label": "",
                        "notes": "",
                        "month_start": target_month,
                        "month": month_label(target_month),
                        "preview_action": "Add urgent store",
                    }
                ]
            )
        urgent_row["is_selected_push"] = True
        urgent_row["sort_month"] = target_month
        urgent_row["sort_bucket"] = -1
        rows.append(urgent_row)
    preview = pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()
    if preview.empty:
        return preview
    preview = preview.drop_duplicates(subset=["schedule_item_id"], keep="last")
    preview = preview.sort_values(["sort_month", "sort_bucket", "sequence_number", "store_number"]).reset_index(drop=True)
    assigned_rows = []
    cursor_month = target_month
    sequence_number = 1
    monthly_capacity = max(1, int(monthly_capacity))
    for _, row in preview.iterrows():
        next_row = row.copy()
        if sequence_number > monthly_capacity:
            cursor_month = add_months(cursor_month, 1)
            sequence_number = 1
        next_row["new_month_start"] = cursor_month
        next_row["new_month"] = month_label(cursor_month)
        next_row["new_schedule_date"] = first_workday(cursor_month, employee_id=int(employee_id))
        next_row["new_sequence_number"] = sequence_number
        next_row["current_month"] = month_label(row["month_start"]) if pd.notna(row.get("month_start")) else ""
        next_row["current_sequence_number"] = int(row["sequence_number"]) if pd.notna(row.get("sequence_number")) else ""
        next_row["change"] = (
            "No date change"
            if row.get("month_start") == cursor_month and int(row.get("sequence_number") or 0) == sequence_number
            else "Date/route changed"
        )
        assigned_rows.append(next_row)
        sequence_number += 1
    result = pd.DataFrame(assigned_rows)
    result["schedule_date"] = result["new_schedule_date"]
    result["month_start"] = result["new_month_start"]
    result["month"] = result["new_month"]
    result["sequence_number"] = result["new_sequence_number"]
    return result


def apply_pmt_reflow_preview(run_id, preview_df, reason):
    if preview_df.empty:
        return 0
    updated = 0
    reason = clean(reason) or "PMT schedule management adjustment"
    with session_scope() as session:
        for _, row in preview_df.iterrows():
            item_id = int(row["schedule_item_id"])
            schedule_date = pd.to_datetime(row["new_schedule_date"]).date()
            note_parts = [clean(row.get("notes")), f"PMT manager adjustment: {reason}"]
            if item_id < 0:
                item = ScheduleItem(
                    schedule_id=int(row["schedule_id"]),
                    schedule_date=schedule_date,
                    sequence_number=int(row["new_sequence_number"]),
                    store_id=int(row["store_id"]),
                    employee_id=int(row["employee_id"]),
                    work_type="PMT",
                    status="Scheduled",
                    schedule_source="PMT Manager Adjustment",
                    pmt_schedule_run_id=int(run_id),
                    cycle_label=month_label(month_start(schedule_date)),
                    completion_notes=" | ".join([part for part in note_parts if part]),
                )
                session.add(item)
            else:
                item = session.get(ScheduleItem, item_id)
                if not item:
                    continue
                if item.original_schedule_date is None:
                    item.original_schedule_date = item.schedule_date
                item.employee_id = int(row["employee_id"])
                item.schedule_date = schedule_date
                item.sequence_number = int(row["new_sequence_number"])
                item.status = "Scheduled"
                item.schedule_source = "PMT Manager Adjustment"
                item.cycle_label = month_label(month_start(schedule_date))
                item.completion_notes = " | ".join([part for part in note_parts if part])
            updated += 1
    log_action("pmt schedule manager adjustment", "pmt_schedule_runs", int(run_id), f"{updated} PMT items updated. Reason: {reason}")
    return updated


def normalize_existing_pmt_schedule_upload(raw_df, mapping):
    if raw_df.empty:
        return pd.DataFrame(), pd.DataFrame()
    employees, employee_match = employee_lookup()
    stores = safe_query(
        """
        select id as store_id, store_number, address, city, state, zip, latitude, longitude, active
        from stores
        """
    )
    store_lookup = {}
    if not stores.empty:
        for store_row in stores.to_dict("records"):
            for store_key in store_number_keys(store_row["store_number"]):
                store_lookup.setdefault(store_key, store_row)
    rows = []
    problems = []
    first_workday_cache = {}
    for index, source_row in raw_df.fillna("").iterrows():
        tech_name = clean(source_row.get(mapping.get("technician_name", ""), ""))
        store_number = clean(source_row.get(mapping.get("store_number", ""), ""))
        if not tech_name and not store_number:
            continue
        employee = match_employee_name(tech_name, employee_match)
        store = None
        for store_key in store_number_keys(store_number):
            store = store_lookup.get(store_key)
            if store is not None:
                break
        schedule_date = scalar_date(source_row.get(mapping.get("schedule_date", ""), ""))
        schedule_month = scalar_date(source_row.get(mapping.get("schedule_month", ""), ""))
        if schedule_date is None and schedule_month is not None:
            schedule_month_start = month_start(schedule_month)
            if employee:
                cache_key = (int(employee["id"]), schedule_month_start)
                if cache_key not in first_workday_cache:
                    first_workday_cache[cache_key] = first_workday(schedule_month_start, employee_id=int(employee["id"]))
                schedule_date = first_workday_cache[cache_key]
            else:
                schedule_date = schedule_month_start
        sequence_number = scalar_int(source_row.get(mapping.get("sequence_number", ""), ""), 0)
        if employee is None:
            problems.append({"Row": index + 2, "Problem": "Technician not matched to an active employee", "Value": tech_name})
        if store is None:
            problems.append({"Row": index + 2, "Problem": "Store not found in the master store list", "Value": store_number})
        elif not bool(store.get("active", True)):
            problems.append({"Row": index + 2, "Problem": "Store exists but is inactive", "Value": store.get("store_number", store_number)})
        if schedule_date is None:
            problems.append({"Row": index + 2, "Problem": "Schedule date or month could not be read", "Value": clean(source_row.get(mapping.get("schedule_date", ""), "")) or clean(source_row.get(mapping.get("schedule_month", ""), ""))})
        if employee is None or store is None or schedule_date is None:
            continue
        rows.append(
            {
                "employee_id": int(employee["id"]),
                "technician": employee["full_name"],
                "store_id": int(store["store_id"]),
                "store_number": clean(store["store_number"]),
                "address": clean(store.get("address", "")),
                "city": clean(store.get("city", "")),
                "state": clean(store.get("state", "")),
                "zip": clean(store.get("zip", "")),
                "latitude": store.get("latitude"),
                "longitude": store.get("longitude"),
                "schedule_date": schedule_date,
                "month_start": month_start(schedule_date),
                "month": month_label(month_start(schedule_date)),
                "sequence_number": sequence_number,
                "status": clean(source_row.get(mapping.get("status", ""), "")) or "Scheduled",
                "notes": clean(source_row.get(mapping.get("notes", ""), "")),
            }
        )
    normalized = pd.DataFrame(rows)
    if not normalized.empty:
        normalized = normalized.sort_values(["employee_id", "schedule_date", "sequence_number", "store_number"]).reset_index(drop=True)
        normalized["sequence_number"] = normalized.groupby(["employee_id", "month_start"]).cumcount() + 1
        duplicate_mask = normalized.duplicated(["employee_id", "store_id", "month_start"], keep=False)
        if duplicate_mask.any():
            for _, row in normalized.loc[duplicate_mask].drop_duplicates(["employee_id", "store_id", "month_start"]).iterrows():
                problems.append(
                    {
                        "Row": "",
                        "Problem": "Duplicate technician/store/month in uploaded schedule; first row will be used",
                        "Value": f"{row['technician']} - Store {row['store_number']} - {row['month']}",
                    }
                )
            normalized = normalized.drop_duplicates(["employee_id", "store_id", "month_start"], keep="first")
    return normalized, pd.DataFrame(problems)


def import_existing_pmt_schedule(normalized, run_name):
    if normalized.empty:
        return None
    start_date = normalized["schedule_date"].min()
    end_date = normalized["schedule_date"].max()
    months = (end_date.year - start_date.year) * 12 + end_date.month - start_date.month + 1
    cycle_label = f"{month_label(month_start(start_date))} - {month_label(month_start(end_date))}"
    with session_scope() as session:
        run = PMTScheduleRun(
            run_name=run_name,
            cycle_start=month_start(start_date),
            cycle_end=end_date,
            months=months,
            default_monthly_target=0,
            direction="Imported",
            schedule_mode="Imported existing PMT schedule",
            distance_method="Imported order",
            technician_count=int(normalized["employee_id"].nunique()),
            store_count=len(normalized),
            unscheduled_count=0,
            created_by=st.session_state.get("username", ""),
            notes=f"Imported existing PMT schedule | {cycle_label}",
        )
        session.add(run)
        session.flush()
        schedule = Schedule(
            schedule_name=run_name,
            schedule_type="PMT Imported Schedule",
            start_date=month_start(start_date),
            end_date=end_date,
            status="Published",
            created_by=st.session_state.get("username", ""),
            notes=f"Schedule Source: Imported Existing PMT Schedule | Run ID: {run.id} | Cycle: {cycle_label}",
        )
        session.add(schedule)
        session.flush()
        item_rows = []
        now = datetime.utcnow()
        for _, row in normalized.iterrows():
            item_rows.append(
                {
                    "schedule_id": schedule.id,
                    "schedule_date": row["schedule_date"],
                    "sequence_number": int(row["sequence_number"]),
                    "store_id": int(row["store_id"]),
                    "employee_id": int(row["employee_id"]),
                    "work_type": "PMT",
                    "status": clean(row.get("status", "")) or "Scheduled",
                    "schedule_source": "Imported Existing PMT Schedule",
                    "pmt_schedule_run_id": run.id,
                    "cycle_label": cycle_label,
                    "completion_notes": clean(row.get("notes", "")),
                    "created_at": now,
                    "updated_at": now,
                }
            )
        if item_rows:
            session.execute(insert(ScheduleItem), item_rows)
            session.flush()
            saved_count = session.scalar(
                select(func.count(ScheduleItem.id)).where(ScheduleItem.pmt_schedule_run_id == int(run.id))
            )
            if int(saved_count or 0) != len(item_rows):
                raise RuntimeError(f"Expected to save {len(item_rows)} PMT schedule items, but saved {int(saved_count or 0)}.")
        run_id = run.id
    log_action("pmt existing schedule imported", "pmt_schedule_runs", int(run_id), f"{len(normalized)} PMT schedule items imported")
    return {"run_id": run_id, "created": len(normalized)}


def assigned_pmt_store_candidates(employee_id, run_id=None, include_scheduled=False):
    params = {"employee_id": int(employee_id)}
    run_filter = ""
    schedule_join = ""
    schedule_columns = "null as scheduled_item_id, null as scheduled_employee_id, null as scheduled_technician, null as scheduled_date, 0 as scheduled_count"
    if run_id is not None:
        params["run_id"] = int(run_id)
        schedule_columns = """
               min(si.id) as scheduled_item_id,
               coalesce(
                   max(case when si.employee_id <> :employee_id then si.employee_id end),
                   min(si.employee_id)
               ) as scheduled_employee_id,
               coalesce(
                   max(case when si.employee_id <> :employee_id then se.full_name end),
                   max(se.full_name)
               ) as scheduled_technician,
               min(si.schedule_date) as scheduled_date,
               count(si.id) as scheduled_count
        """
        schedule_join = """
        left join schedule_items si
          on si.pmt_schedule_run_id = :run_id
         and si.store_id = s.id
         and si.work_type = 'PMT'
         and si.status in ('Scheduled','Needs Rescheduled','Rescheduled','Rain Delay','Not Completed')
        left join employees se on se.id = si.employee_id
        """
        if not include_scheduled:
            run_filter = """
          and not exists (
              select 1 from schedule_items si
              where si.pmt_schedule_run_id = :run_id
                and si.store_id = s.id
                and si.work_type = 'PMT'
                and si.status in ('Scheduled','Needs Rescheduled','Rescheduled','Rain Delay','Not Completed')
          )
            """
    df = safe_query(
        f"""
        select s.id as store_id, s.store_number, s.address, s.city, s.state, s.zip,
               s.latitude, s.longitude, e.full_name as technician,
               e.home_latitude, e.home_longitude,
               {schedule_columns}
        from stores s
        join employees e on e.id = s.assigned_pmt_employee_id
        {schedule_join}
        where s.active = true
          and s.assigned_pmt_employee_id = :employee_id
          {run_filter}
        group by s.id, s.store_number, s.address, s.city, s.state, s.zip,
                 s.latitude, s.longitude, e.full_name, e.home_latitude, e.home_longitude
        order by s.store_number
        """,
        params,
    )
    if df.empty:
        return df
    df = df.copy()
    df["distance_from_home"] = df.apply(
        lambda row: round(haversine_miles(float(row["home_latitude"]), float(row["home_longitude"]), float(row["latitude"]), float(row["longitude"])), 1)
        if pd.notna(row.get("home_latitude"))
        and pd.notna(row.get("home_longitude"))
        and pd.notna(row.get("latitude"))
        and pd.notna(row.get("longitude"))
        else None,
        axis=1,
    )
    return df


def pmt_route_builder_store_pool(run_items, employee_ids, selected_month="All months", run_id=None):
    employee_ids = [int(employee_id) for employee_id in employee_ids if scalar_int(employee_id, 0)]
    if not employee_ids:
        return pd.DataFrame()
    pool_parts = []
    for employee_id in employee_ids:
        assigned = assigned_pmt_store_candidates(employee_id, run_id=run_id, include_scheduled=True)
        if assigned.empty:
            continue
        assigned = assigned.copy()
        assigned["assigned_employee_id"] = int(employee_id)
        assigned["assigned_technician"] = assigned.get("technician", "")
        assigned["route_layer"] = "Assigned Store"
        pool_parts.append(assigned)
    if not run_items.empty:
        scheduled = run_items[pmt_active_item_mask(run_items)].copy()
        scheduled = scheduled[pd.to_numeric(scheduled.get("employee_id"), errors="coerce").fillna(-1).astype(int).isin(employee_ids)].copy()
        if selected_month != "All months" and "month_start" in scheduled.columns:
            scheduled = scheduled[scheduled["month_start"] == selected_month].copy()
        if not scheduled.empty:
            scheduled = scheduled.rename(columns={"employee_id": "scheduled_employee_id", "technician": "scheduled_technician"})
            scheduled["assigned_employee_id"] = pd.to_numeric(scheduled.get("assigned_pmt_employee_id"), errors="coerce")
            scheduled["route_layer"] = "Existing Schedule"
            if "assigned_technician" not in scheduled.columns:
                scheduled["assigned_technician"] = ""
            pool_parts.append(scheduled)
    if not pool_parts:
        return pd.DataFrame()
    pool = pd.concat(pool_parts, ignore_index=True, sort=False)
    pool["store_id"] = pd.to_numeric(pool.get("store_id"), errors="coerce")
    pool["latitude"] = pd.to_numeric(pool.get("latitude"), errors="coerce")
    pool["longitude"] = pd.to_numeric(pool.get("longitude"), errors="coerce")
    pool = pool.dropna(subset=["store_id", "latitude", "longitude"]).copy()
    if pool.empty:
        return pool
    pool["store_id"] = pool["store_id"].astype(int)
    pool["store_number"] = pool.get("store_number", "").fillna("").astype(str)
    return pool.sort_values(["route_layer", "store_number"]).drop_duplicates("store_id", keep="first")


def nearest_route_builder_store(store_pool, clicked):
    if store_pool.empty or not clicked:
        return None
    clicked_lat = clicked.get("lat")
    clicked_lon = clicked.get("lng")
    if clicked_lat is None or clicked_lon is None:
        return None
    nearest = None
    nearest_distance = None
    for _, row in store_pool.iterrows():
        distance = haversine_miles(float(clicked_lat), float(clicked_lon), float(row["latitude"]), float(row["longitude"]))
        if nearest_distance is None or distance < nearest_distance:
            nearest = row
            nearest_distance = distance
    if nearest is None or nearest_distance is None or nearest_distance > 1.0:
        return None
    return nearest


def render_pmt_route_builder_map(store_pool, employees_df, route_df, show_assigned_layer=True, show_existing_layer=True, key=None):
    mapped_pool = store_pool.copy()
    mapped_pool["latitude"] = pd.to_numeric(mapped_pool.get("latitude"), errors="coerce")
    mapped_pool["longitude"] = pd.to_numeric(mapped_pool.get("longitude"), errors="coerce")
    mapped_pool = mapped_pool.dropna(subset=["latitude", "longitude"])
    if mapped_pool.empty:
        st.info("No mapped stores found for the selected PMT layer.")
        return {}
    fmap = folium.Map(location=[float(mapped_pool["latitude"].mean()), float(mapped_pool["longitude"].mean())], zoom_start=9, tiles="OpenStreetMap")
    if employees_df is not None and not employees_df.empty:
        employees_df = employees_df.copy()
        employees_df["home_latitude"] = pd.to_numeric(employees_df.get("home_latitude"), errors="coerce")
        employees_df["home_longitude"] = pd.to_numeric(employees_df.get("home_longitude"), errors="coerce")
        for _, row in employees_df.dropna(subset=["home_latitude", "home_longitude"]).iterrows():
            folium.Marker(
                [float(row["home_latitude"]), float(row["home_longitude"])],
                icon=folium.Icon(color="black", icon="home", prefix="fa"),
                popup=folium.Popup(f"<b>{row.get('technician_name', '')}</b><br>Home", max_width=260),
                tooltip=f"{row.get('technician_name', '')} home",
            ).add_to(fmap)
    assigned_group = folium.FeatureGroup(name="Assigned Store Ownership", show=show_assigned_layer)
    schedule_group = folium.FeatureGroup(name="Existing Schedule Route", show=show_existing_layer)
    for _, row in mapped_pool.iterrows():
        assigned_name = clean(row.get("assigned_technician", "")) or clean(row.get("technician", "")) or "Unassigned"
        scheduled_name = clean(row.get("scheduled_technician", "")) or ""
        popup = f"""
        <b>Store {row.get('store_number', '')}</b><br>
        {row.get('address', '')}<br>
        {row.get('city', '')}, {row.get('state', '')}<br>
        Assigned PMT: {assigned_name}<br>
        Scheduled PMT: {scheduled_name or 'Not active in selected schedule'}<br>
        Click the dot to add it to the proposed route.
        """
        folium.CircleMarker(
            [float(row["latitude"]), float(row["longitude"])],
            radius=7,
            color="#ffffff",
            weight=1,
            fill=True,
            fill_color=stable_color(assigned_name),
            fill_opacity=0.9,
            popup=folium.Popup(popup, max_width=340),
            tooltip=f"Store {row.get('store_number', '')} | Assigned: {assigned_name}",
        ).add_to(assigned_group)
        if scheduled_name:
            folium.CircleMarker(
                [float(row["latitude"]), float(row["longitude"])],
                radius=10,
                color="#111827",
                weight=3,
                fill=False,
                popup=folium.Popup(popup, max_width=340),
                tooltip=f"Scheduled: {scheduled_name} | Store {row.get('store_number', '')}",
            ).add_to(schedule_group)
    assigned_group.add_to(fmap)
    schedule_group.add_to(fmap)
    proposed = route_df.copy() if route_df is not None else pd.DataFrame()
    if not proposed.empty:
        proposed["latitude"] = pd.to_numeric(proposed.get("latitude"), errors="coerce")
        proposed["longitude"] = pd.to_numeric(proposed.get("longitude"), errors="coerce")
        proposed = proposed.dropna(subset=["latitude", "longitude"]).sort_values(["Proposed Stop", "store_number"])
        proposed_points = proposed[["latitude", "longitude"]].astype(float).values.tolist()
        if len(proposed_points) >= 2:
            folium.PolyLine(proposed_points, color="#dc2626", weight=5, opacity=0.82, tooltip="Proposed route").add_to(fmap)
        for _, row in proposed.iterrows():
            stop_number = scalar_int(row.get("Proposed Stop"), 0)
            folium.Marker(
                [float(row["latitude"]), float(row["longitude"])],
                icon=folium.DivIcon(
                    html=f"""
                    <div style="background:#dc2626;color:white;border:2px solid white;border-radius:999px;
                    width:28px;height:28px;line-height:24px;text-align:center;font-size:13px;font-weight:900;
                    box-shadow:0 1px 5px rgba(0,0,0,.4);">{stop_number}</div>
                    """
                ),
                tooltip=f"Proposed Stop {stop_number}: Store {row.get('store_number', '')}",
            ).add_to(fmap)
    folium.LayerControl(collapsed=False).add_to(fmap)
    return st_folium(fmap, width=None, height=620, key=key, returned_objects=["last_object_clicked"])


def render_fast_pmt_route_picker_map(store_pool, employees_df, show_assigned_layer=True, show_existing_layer=True, component_key="pmt_route_picker"):
    mapped_pool = store_pool.copy()
    mapped_pool["latitude"] = pd.to_numeric(mapped_pool.get("latitude"), errors="coerce")
    mapped_pool["longitude"] = pd.to_numeric(mapped_pool.get("longitude"), errors="coerce")
    mapped_pool = mapped_pool.dropna(subset=["latitude", "longitude"])
    if mapped_pool.empty:
        st.info("No mapped stores found for the selected PMT layer.")
        return
    fmap = folium.Map(location=[float(mapped_pool["latitude"].mean()), float(mapped_pool["longitude"].mean())], zoom_start=9, tiles="OpenStreetMap")
    click_scripts = []
    if employees_df is not None and not employees_df.empty:
        employees_df = employees_df.copy()
        employees_df["home_latitude"] = pd.to_numeric(employees_df.get("home_latitude"), errors="coerce")
        employees_df["home_longitude"] = pd.to_numeric(employees_df.get("home_longitude"), errors="coerce")
        for _, row in employees_df.dropna(subset=["home_latitude", "home_longitude"]).iterrows():
            folium.Marker(
                [float(row["home_latitude"]), float(row["home_longitude"])],
                icon=folium.Icon(color="black", icon="home", prefix="fa"),
                popup=folium.Popup(f"<b>{row.get('technician_name', '')}</b><br>Home", max_width=260),
                tooltip=f"{row.get('technician_name', '')} home",
            ).add_to(fmap)
    assigned_group = folium.FeatureGroup(name="Assigned Store Ownership", show=show_assigned_layer)
    schedule_group = folium.FeatureGroup(name="Existing Schedule Route", show=show_existing_layer)
    for _, row in mapped_pool.iterrows():
        assigned_name = clean(row.get("assigned_technician", "")) or clean(row.get("technician", "")) or "Unassigned"
        scheduled_name = clean(row.get("scheduled_technician", "")) or ""
        store_payload = {
            "store_id": scalar_int(row.get("store_id"), 0),
            "store_number": clean(row.get("store_number", "")),
            "city": clean(row.get("city", "")),
            "state": clean(row.get("state", "")),
            "assigned_technician": assigned_name,
            "scheduled_technician": scheduled_name,
        }
        popup = f"""
        <b>Store {row.get('store_number', '')}</b><br>
        {row.get('address', '')}<br>
        {row.get('city', '')}, {row.get('state', '')}<br>
        Assigned PMT: {assigned_name}<br>
        Scheduled PMT: {scheduled_name or 'Not active in selected schedule'}<br>
        Fast mode: click to add this store to the route list inside the map.
        """
        marker = folium.CircleMarker(
            [float(row["latitude"]), float(row["longitude"])],
            radius=8,
            color="#ffffff",
            weight=1,
            fill=True,
            fill_color=stable_color(assigned_name),
            fill_opacity=0.9,
            popup=folium.Popup(popup, max_width=340),
            tooltip=f"Store {row.get('store_number', '')} | Assigned: {assigned_name}",
        ).add_to(assigned_group)
        click_scripts.append(f"{marker.get_name()}.on('click', function() {{ addRouteStore({json.dumps(store_payload)}); }});")
        if scheduled_name:
            folium.CircleMarker(
                [float(row["latitude"]), float(row["longitude"])],
                radius=11,
                color="#111827",
                weight=3,
                fill=False,
                popup=folium.Popup(popup, max_width=340),
                tooltip=f"Scheduled: {scheduled_name} | Store {row.get('store_number', '')}",
            ).add_to(schedule_group)
    assigned_group.add_to(fmap)
    schedule_group.add_to(fmap)
    folium.LayerControl(collapsed=False).add_to(fmap)
    safe_key = key(component_key)
    panel_html = f"""
    <div id="{safe_key}_route_panel" style="
        position:absolute; top:12px; right:12px; z-index:9999; width:320px; max-height:560px;
        background:white; border:1px solid #cbd5e1; border-radius:8px; box-shadow:0 8px 24px rgba(15,23,42,.18);
        padding:10px; font-family:Arial, sans-serif; font-size:13px;">
        <div style="font-weight:800; font-size:14px; margin-bottom:6px;">Fast Route Click List</div>
        <div style="color:#475569; margin-bottom:8px;">Click stores on the map. No page reload happens until you paste the final route below the map.</div>
        <ol id="{safe_key}_route_list" style="margin:0 0 8px 22px; padding:0; max-height:220px; overflow:auto;"></ol>
        <textarea id="{safe_key}_route_text" style="width:100%; height:88px; box-sizing:border-box; font-size:12px;" placeholder="Store numbers appear here when you click stores"></textarea>
        <div style="display:flex; gap:6px; margin-top:8px;">
            <button type="button" onclick="copyRouteText()" style="flex:1; padding:6px; font-weight:700;">Done / Copy Route</button>
            <button type="button" onclick="clearRouteStores()" style="padding:6px;">Clear</button>
        </div>
    </div>
    """
    script = f"""
    <script>
    var routeStores = [];
    function refreshRoutePanel() {{
        var list = document.getElementById("{safe_key}_route_list");
        var text = document.getElementById("{safe_key}_route_text");
        if (!list || !text) return;
        list.innerHTML = "";
        routeStores.forEach(function(store, index) {{
            var item = document.createElement("li");
            item.textContent = store.store_number + " - " + store.city + ", " + store.state;
            list.appendChild(item);
        }});
        text.value = routeStores.map(function(store) {{ return store.store_number; }}).join("\\n");
    }}
    function addRouteStore(store) {{
        if (routeStores.some(function(existing) {{ return existing.store_id === store.store_id; }})) return;
        routeStores.push(store);
        refreshRoutePanel();
    }}
    function clearRouteStores() {{
        routeStores = [];
        refreshRoutePanel();
    }}
    function copyRouteText() {{
        refreshRoutePanel();
        var text = document.getElementById("{safe_key}_route_text");
        text.focus();
        text.select();
        try {{ document.execCommand("copy"); }} catch (err) {{}}
    }}
    {chr(10).join(click_scripts)}
    </script>
    """
    fmap.get_root().html.add_child(folium.Element(panel_html))
    fmap.get_root().script.add_child(folium.Element(script))
    components.html(fmap.get_root().render(), height=660, scrolling=False)


def move_scheduled_stores_to_pmt(run_id, employee_id, store_ids, target_month, notes=""):
    if not store_ids:
        return 0
    target_month = month_start(target_month)
    moved = 0
    with session_scope() as session:
        max_sequence = session.scalar(
            select(ScheduleItem.sequence_number)
            .where(
                ScheduleItem.pmt_schedule_run_id == int(run_id),
                ScheduleItem.employee_id == int(employee_id),
                ScheduleItem.work_type == "PMT",
                ScheduleItem.schedule_date >= target_month,
                ScheduleItem.schedule_date < add_months(target_month, 1),
            )
            .order_by(ScheduleItem.sequence_number.desc())
        ) or 0
        for store_id in [int(value) for value in store_ids]:
            active_items = session.scalars(
                select(ScheduleItem).where(
                    ScheduleItem.pmt_schedule_run_id == int(run_id),
                    ScheduleItem.store_id == int(store_id),
                    ScheduleItem.work_type == "PMT",
                    ScheduleItem.status.in_(PMT_ACTIVE_STATUS_VALUES),
                )
            ).all()
            if not active_items:
                continue
            target_items = [item for item in active_items if item.employee_id == int(employee_id)]
            other_items = [item for item in active_items if item.employee_id != int(employee_id)]
            if target_items:
                # The selected PMT already has this store; retire duplicate future rows without erasing history.
                keep_item = target_items[0]
                for duplicate_item in other_items:
                    if duplicate_item.original_schedule_date is None:
                        duplicate_item.original_schedule_date = duplicate_item.schedule_date
                    duplicate_item.status = "Transferred"
                    duplicate_item.schedule_source = "PMT Schedule Conflict Superseded"
                    note_parts = [
                        clean(duplicate_item.completion_notes),
                        f"Superseded by active PMT schedule item #{keep_item.id} for selected PMT.",
                    ]
                    if notes:
                        note_parts.append(clean(notes))
                    duplicate_item.completion_notes = " | ".join([part for part in note_parts if part])
                    moved += 1
                continue
            item = other_items[0] if other_items else None
            if not item:
                continue
            for duplicate_item in other_items[1:]:
                if duplicate_item.original_schedule_date is None:
                    duplicate_item.original_schedule_date = duplicate_item.schedule_date
                duplicate_item.status = "Transferred"
                duplicate_item.schedule_source = "PMT Schedule Conflict Superseded"
                note_parts = [
                    clean(duplicate_item.completion_notes),
                    f"Superseded by transferred PMT schedule item #{item.id}.",
                ]
                if notes:
                    note_parts.append(clean(notes))
                duplicate_item.completion_notes = " | ".join([part for part in note_parts if part])
                moved += 1
            if item.employee_id == int(employee_id):
                continue
            max_sequence += 1
            if item.original_schedule_date is None:
                item.original_schedule_date = item.schedule_date
            item.employee_id = int(employee_id)
            item.schedule_date = first_workday(target_month, employee_id=int(employee_id))
            item.sequence_number = int(max_sequence)
            item.status = "Scheduled"
            item.schedule_source = "PMT Schedule Conflict Move"
            item.cycle_label = month_label(target_month)
            note_parts = [clean(item.completion_notes), "Moved because store assignment belongs to selected PMT"]
            if notes:
                note_parts.append(clean(notes))
            item.completion_notes = " | ".join([part for part in note_parts if part])
            moved += 1
    log_action("pmt scheduled stores moved to assigned pmt", "pmt_schedule_runs", int(run_id), f"{moved} scheduled item(s) moved to employee_id={int(employee_id)}")
    return moved


def apply_pmt_manage_build_preview(run_id, employee_id, preview_records, notes=""):
    preview_df = pd.DataFrame(preview_records or [])
    if preview_df.empty:
        return {"saved": 0, "superseded": 0, "created": 0, "updated": 0, "resequenced_rows": 0}
    preview_df = preview_df.copy()
    preview_df["store_id"] = pd.to_numeric(preview_df.get("store_id"), errors="coerce")
    preview_df["Proposed Stop"] = pd.to_numeric(preview_df.get("Proposed Stop"), errors="coerce")
    preview_df = preview_df.dropna(subset=["store_id"]).sort_values(["Proposed Stop", "store_number"])
    if preview_df.empty:
        return {"saved": 0, "superseded": 0, "created": 0, "updated": 0, "resequenced_rows": 0}
    saved = 0
    created = 0
    updated = 0
    superseded = 0
    touched_months = set()
    now = datetime.utcnow()
    with session_scope("Apply PMT manage schedule preview") as session:
        schedule_id = session.scalar(
            select(ScheduleItem.schedule_id)
            .where(ScheduleItem.pmt_schedule_run_id == int(run_id))
            .order_by(ScheduleItem.schedule_id)
        )
        if schedule_id is None:
            return {"saved": 0, "superseded": 0, "created": 0, "updated": 0, "resequenced_rows": 0}
        for _, row in preview_df.iterrows():
            store_id = int(row["store_id"])
            proposed_date = pd.to_datetime(row.get("Proposed Date") or row.get("schedule_date"), errors="coerce")
            schedule_date = proposed_date.date() if pd.notna(proposed_date) else first_workday(date.today(), employee_id=int(employee_id))
            proposed_stop = scalar_int(row.get("Proposed Stop"), 0) or 1
            active_items = session.scalars(
                select(ScheduleItem)
                .where(
                    ScheduleItem.pmt_schedule_run_id == int(run_id),
                    ScheduleItem.store_id == int(store_id),
                    ScheduleItem.work_type == "PMT",
                    ScheduleItem.status.in_(PMT_ACTIVE_STATUS_VALUES),
                )
                .order_by(ScheduleItem.schedule_date, ScheduleItem.sequence_number, ScheduleItem.id)
            ).all()
            target_items = [item for item in active_items if item.employee_id == int(employee_id)]
            if target_items:
                keep_item = target_items[0]
                updated += 1
            else:
                keep_item = ScheduleItem(
                    schedule_id=int(schedule_id),
                    schedule_date=schedule_date,
                    sequence_number=int(proposed_stop),
                    pmt_schedule_run_id=int(run_id),
                    employee_id=int(employee_id),
                    store_id=store_id,
                    work_type="PMT",
                    status="Scheduled",
                    schedule_source="PMT Manage Schedule Preview",
                    cycle_label=month_label(month_start(schedule_date)),
                    completion_notes=clean(notes),
                    created_at=now,
                )
                session.add(keep_item)
                created += 1
            if keep_item.original_schedule_date is None and keep_item.schedule_date and keep_item.schedule_date != schedule_date:
                keep_item.original_schedule_date = keep_item.schedule_date
            keep_item.schedule_date = schedule_date
            keep_item.sequence_number = int(proposed_stop)
            keep_item.status = "Scheduled"
            keep_item.schedule_source = "PMT Manage Schedule Preview"
            keep_item.cycle_label = month_label(month_start(schedule_date))
            keep_item.completion_notes = clean(notes) or keep_item.completion_notes
            session.flush()
            touched_months.add((int(employee_id), month_start(schedule_date)))
            for item in active_items:
                if item.id == keep_item.id:
                    continue
                if item.original_schedule_date is None:
                    item.original_schedule_date = item.schedule_date
                item.status = "Transferred"
                item.schedule_source = "PMT Manage Schedule Preview Superseded"
                note_parts = [
                    clean(item.completion_notes),
                    f"Superseded by preview schedule item #{keep_item.id} for the selected PMT.",
                ]
                if notes:
                    note_parts.append(clean(notes))
                item.completion_notes = " | ".join([part for part in note_parts if part])
                touched_months.add((item.employee_id, month_start(item.schedule_date)))
                superseded += 1
            saved += 1
        resequenced_rows = 0
        for touched_employee_id, touched_month in touched_months:
            resequenced_rows += resequence_pmt_month(session, run_id, touched_employee_id, touched_month)
        run = session.get(PMTScheduleRun, int(run_id))
        if run:
            active_store_count = session.scalar(
                select(func.count(func.distinct(ScheduleItem.store_id))).where(
                    ScheduleItem.pmt_schedule_run_id == int(run_id),
                    ScheduleItem.work_type == "PMT",
                    ScheduleItem.status.in_(PMT_ACTIVE_STATUS_VALUES),
                )
            )
            run.store_count = int(active_store_count or 0)
    log_action(
        "pmt manage schedule preview applied",
        "pmt_schedule_runs",
        int(run_id),
        f"Saved {saved} preview store(s); created {created}; updated {updated}; superseded {superseded}.",
    )
    return {"saved": saved, "superseded": superseded, "created": created, "updated": updated, "resequenced_rows": resequenced_rows}


def add_assigned_stores_auto_fill_to_pmt_run(run_id, employee_id, store_ids, target_month, fill_end_month, monthly_capacity, notes=""):
    if not store_ids:
        return {"added": 0, "skipped": 0}
    target_month = month_start(target_month)
    fill_end_month = month_start(fill_end_month)
    monthly_capacity = max(1, int(monthly_capacity))
    added = 0
    skipped = 0
    cursor_month = target_month
    now = datetime.utcnow()
    with session_scope() as session:
        schedule_id = session.scalar(
            select(ScheduleItem.schedule_id)
            .where(ScheduleItem.pmt_schedule_run_id == int(run_id))
            .order_by(ScheduleItem.schedule_id)
        )
        run = session.get(PMTScheduleRun, int(run_id))
        if schedule_id is None:
            return {"added": 0, "skipped": len(store_ids)}
        monthly_sequences = {}
        def current_month_sequence(schedule_month):
            if schedule_month not in monthly_sequences:
                existing_max = session.scalar(
                    select(ScheduleItem.sequence_number)
                    .where(
                        ScheduleItem.pmt_schedule_run_id == int(run_id),
                        ScheduleItem.employee_id == int(employee_id),
                        ScheduleItem.work_type == "PMT",
                        ScheduleItem.schedule_date >= schedule_month,
                        ScheduleItem.schedule_date < add_months(schedule_month, 1),
                    )
                    .order_by(ScheduleItem.sequence_number.desc())
                ) or 0
                monthly_sequences[schedule_month] = int(existing_max)
            return monthly_sequences[schedule_month]

        for store_id in store_ids:
            if cursor_month > fill_end_month:
                skipped += 1
                continue
            duplicate = session.scalar(
                select(ScheduleItem.id).where(
                    ScheduleItem.pmt_schedule_run_id == int(run_id),
                    ScheduleItem.store_id == int(store_id),
                    ScheduleItem.work_type == "PMT",
                    ScheduleItem.status.in_(PMT_ACTIVE_STATUS_VALUES),
                )
            )
            if duplicate:
                skipped += 1
                continue
            if current_month_sequence(cursor_month) >= monthly_capacity:
                cursor_month = add_months(cursor_month, 1)
                if cursor_month > fill_end_month:
                    skipped += 1
                    continue
                current_month_sequence(cursor_month)
            monthly_sequences[cursor_month] += 1
            schedule_date = first_workday(cursor_month, employee_id=int(employee_id))
            session.add(
                ScheduleItem(
                    schedule_id=int(schedule_id),
                    schedule_date=schedule_date,
                    sequence_number=int(monthly_sequences[cursor_month]),
                    store_id=int(store_id),
                    employee_id=int(employee_id),
                    work_type="PMT",
                    status="Scheduled",
                    schedule_source="PMT Manual And Auto Fill",
                    pmt_schedule_run_id=int(run_id),
                    cycle_label=month_label(cursor_month),
                    completion_notes=clean(notes),
                    created_at=now,
                    updated_at=now,
                )
            )
            added += 1
        if run and added:
            run.store_count = int(run.store_count or 0) + added
            run.cycle_end = max(run.cycle_end or target_month, add_months(fill_end_month, 1) - timedelta(days=1))
    log_action("pmt assigned stores auto filled", "pmt_schedule_runs", int(run_id), f"{added} stores added and {skipped} skipped")
    return {"added": added, "skipped": skipped}


def add_assigned_stores_to_pmt_run(run_id, employee_id, store_ids, target_month, notes=""):
    if not store_ids:
        return 0
    target_month = month_start(target_month)
    schedule_date = first_workday(target_month, employee_id=int(employee_id))
    added = 0
    with session_scope() as session:
        schedule_id = session.scalar(
            select(ScheduleItem.schedule_id)
            .where(ScheduleItem.pmt_schedule_run_id == int(run_id))
            .order_by(ScheduleItem.schedule_id)
        )
        run = session.get(PMTScheduleRun, int(run_id))
        if schedule_id is None:
            return 0
        max_sequence = session.scalar(
            select(ScheduleItem.sequence_number)
            .where(
                ScheduleItem.pmt_schedule_run_id == int(run_id),
                ScheduleItem.employee_id == int(employee_id),
                ScheduleItem.work_type == "PMT",
                ScheduleItem.schedule_date >= target_month,
                ScheduleItem.schedule_date < add_months(target_month, 1),
            )
            .order_by(ScheduleItem.sequence_number.desc())
        ) or 0
        for store_id in store_ids:
            duplicate = session.scalar(
                select(ScheduleItem.id).where(
                    ScheduleItem.pmt_schedule_run_id == int(run_id),
                    ScheduleItem.store_id == int(store_id),
                    ScheduleItem.work_type == "PMT",
                    ScheduleItem.status.in_(PMT_ACTIVE_STATUS_VALUES),
                )
            )
            if duplicate:
                continue
            max_sequence += 1
            session.add(
                ScheduleItem(
                    schedule_id=int(schedule_id),
                    schedule_date=schedule_date,
                    sequence_number=int(max_sequence),
                    store_id=int(store_id),
                    employee_id=int(employee_id),
                    work_type="PMT",
                    status="Scheduled",
                    schedule_source="PMT Manual Assigned Store Add",
                    pmt_schedule_run_id=int(run_id),
                    cycle_label=month_label(target_month),
                    completion_notes=clean(notes),
                )
            )
            added += 1
        if run:
            run.store_count = int(run.store_count or 0) + added
            run.cycle_end = max(run.cycle_end or schedule_date, add_months(target_month, 1) - timedelta(days=1))
    log_action("pmt assigned stores added to schedule", "pmt_schedule_runs", int(run_id), f"{added} assigned PMT stores added manually")
    return added


def save_manual_pmt_schedule_edits(edited_df):
    if edited_df.empty:
        return 0
    updated = 0
    with session_scope() as session:
        for _, row in edited_df.iterrows():
            item_id = scalar_int(row.get("schedule_item_id"), 0)
            if not item_id:
                continue
            item = session.get(ScheduleItem, int(item_id))
            if not item:
                continue
            new_date = scalar_date(row.get("schedule_date")) or item.schedule_date
            new_sequence = max(1, scalar_int(row.get("sequence_number"), item.sequence_number or 1))
            if item.original_schedule_date is None and item.schedule_date != new_date:
                item.original_schedule_date = item.schedule_date
            item.schedule_date = new_date
            item.sequence_number = new_sequence
            item.status = clean(row.get("status", "")) or item.status
            item.cycle_label = month_label(month_start(new_date))
            item.completion_notes = clean(row.get("notes", "")) or item.completion_notes
            updated += 1
    log_action("pmt manual schedule order updated", "schedule_items", description=f"{updated} PMT schedule items manually updated")
    return updated


def delete_pmt_schedule_items(item_ids, reason=""):
    item_ids = [int(item_id) for item_id in item_ids if scalar_int(item_id, 0)]
    if not item_ids:
        return 0
    deleted = 0
    deleted_by_run = {}
    schedule_ids = set()
    with session_scope() as session:
        items = session.scalars(select(ScheduleItem).where(ScheduleItem.id.in_(item_ids))).all()
        for item in items:
            if item.work_type != "PMT":
                continue
            if item.pmt_schedule_run_id:
                run_id = int(item.pmt_schedule_run_id)
                deleted_by_run[run_id] = deleted_by_run.get(run_id, 0) + 1
            if item.schedule_id:
                schedule_ids.add(int(item.schedule_id))
            session.delete(item)
            deleted += 1
        for run_id, run_deleted in deleted_by_run.items():
            run = session.get(PMTScheduleRun, int(run_id))
            if run:
                run.store_count = max(0, int(run.store_count or 0) - int(run_deleted))
        for schedule_id in schedule_ids:
            remaining = session.scalar(select(ScheduleItem.id).where(ScheduleItem.schedule_id == int(schedule_id)))
            if remaining is None:
                schedule = session.get(Schedule, int(schedule_id))
                if schedule:
                    session.delete(schedule)
    log_action("pmt schedule items deleted", "schedule_items", description=f"{deleted} PMT schedule item(s) deleted. Reason: {clean(reason)}")
    return deleted


def archive_out_of_period_pmt_schedule_items(run_id, item_ids, reason=""):
    item_ids = [int(item_id) for item_id in item_ids if scalar_int(item_id, 0)]
    if not item_ids:
        return 0
    archived = 0
    with session_scope("Archive out-of-period PMT schedule rows") as session:
        items = session.scalars(
            select(ScheduleItem).where(
                ScheduleItem.id.in_(item_ids),
                ScheduleItem.pmt_schedule_run_id == int(run_id),
                ScheduleItem.work_type == "PMT",
            )
        ).all()
        for item in items:
            if normalize_schedule_status(item.status) in PMT_COMPLETED_STATUSES:
                continue
            if item.original_schedule_date is None:
                item.original_schedule_date = item.schedule_date
            item.status = "Archived"
            item.schedule_source = "PMT Out-of-Period Cleanup"
            note_parts = [
                clean(item.completion_notes),
                "Archived because this row falls outside the selected PMT schedule run date range.",
            ]
            if reason:
                note_parts.append(clean(reason))
            item.completion_notes = " | ".join([part for part in note_parts if part])
            archived += 1
        run = session.get(PMTScheduleRun, int(run_id))
        if run:
            active_store_count = session.scalar(
                select(func.count(func.distinct(ScheduleItem.store_id))).where(
                    ScheduleItem.pmt_schedule_run_id == int(run_id),
                    ScheduleItem.work_type == "PMT",
                    ScheduleItem.status.in_(PMT_ACTIVE_STATUS_VALUES),
                )
            )
            run.store_count = int(active_store_count or 0)
    log_action("pmt out-of-period rows archived", "schedule_items", description=f"{archived} PMT schedule item(s) archived. Reason: {clean(reason)}")
    return archived


def delete_pmt_technician_schedule(run_id, employee_id, month_start_value=None, active_only=True, reason=""):
    run_items = pmt_manage_run_items(run_id)
    if run_items.empty:
        return 0
    scope = run_items[run_items["employee_id"].astype("Int64") == int(employee_id)].copy()
    if month_start_value is not None:
        scope = scope[scope["month_start"] == month_start(month_start_value)].copy()
    if active_only:
        scope = scope[pmt_active_item_mask(scope)].copy()
    item_ids = scope["schedule_item_id"].dropna().astype(int).tolist()
    return delete_pmt_schedule_items(item_ids, reason)


def delete_pmt_schedule_run(run_id):
    with session_scope() as session:
        items = session.scalars(select(ScheduleItem).where(ScheduleItem.pmt_schedule_run_id == int(run_id))).all()
        schedule_ids = {item.schedule_id for item in items}
        for item in items:
            session.delete(item)
        run = session.get(PMTScheduleRun, int(run_id))
        if run:
            run.status = "Deleted"
            run.store_count = 0
        for schedule_id in schedule_ids:
            remaining = session.scalar(select(ScheduleItem.id).where(ScheduleItem.schedule_id == schedule_id))
            if remaining is None:
                schedule = session.get(Schedule, int(schedule_id))
                if schedule:
                    session.delete(schedule)
        deleted = len(items)
    log_action("pmt schedule run deleted", "pmt_schedule_runs", int(run_id), f"{deleted} PMT schedule items deleted")
    return deleted


PMT_EXCEPTION_STATUSES = ["Needs Rescheduled", "Rescheduled", "Rain Delay", "Not Completed", "Carryover", "Overdue", "Skipped", "Cancelled"]
PMT_BACKLOG_OPEN_STATUSES = ["Not Scheduled", "Not Completed", "Carryover", "Overdue", "Skipped"]


def pmt_store_history():
    schedule_history = safe_query(
        """
        select
            employee_id,
            store_id,
            max(schedule_date) as last_scheduled_month,
            max(case when status = 'Completed' then schedule_date end) as last_completed_date,
            sum(case when status in ('Needs Rescheduled','Rescheduled','Rain Delay','Not Completed','Carryover','Overdue','Skipped') then 1 else 0 end) as exception_count
        from schedule_items
        where work_type = 'PMT'
          and employee_id is not null
          and store_id is not null
        group by employee_id, store_id
        """
    )
    backlog_history = safe_query(
        """
        select
            employee_id,
            store_id,
            max(cycles_missed) as backlog_cycles_missed,
            sum(case when status = 'Not Scheduled' then 1 else 0 end) as not_scheduled_count,
            sum(case when status in ('Not Completed','Carryover','Overdue','Skipped') then 1 else 0 end) as carryover_count,
            max(last_scheduled_month) as backlog_last_scheduled_month,
            max(last_completed_date) as backlog_last_completed_date
        from pmt_schedule_backlog
        where status in ('Not Scheduled','Not Completed','Carryover','Overdue','Skipped')
        group by employee_id, store_id
        """
    )
    return schedule_history, backlog_history


def apply_pmt_rotation_priority(stores_df, schedule_history, backlog_history, cycle_start):
    priority = stores_df.copy()
    if not schedule_history.empty:
        priority = priority.merge(schedule_history, on=["employee_id", "store_id"], how="left")
    if not backlog_history.empty:
        priority = priority.merge(backlog_history, on=["employee_id", "store_id"], how="left")
    for column in ["exception_count", "backlog_cycles_missed", "not_scheduled_count", "carryover_count"]:
        if column not in priority.columns:
            priority[column] = 0
        priority[column] = pd.to_numeric(priority[column], errors="coerce").fillna(0)
    for column in ["last_scheduled_month", "last_completed_date", "backlog_last_scheduled_month", "backlog_last_completed_date"]:
        if column not in priority.columns:
            priority[column] = pd.NaT
        priority[column] = pd.to_datetime(priority[column], errors="coerce")
    priority["last_scheduled_month"] = priority["last_scheduled_month"].fillna(priority["backlog_last_scheduled_month"])
    priority["last_completed_date"] = priority["last_completed_date"].fillna(priority["backlog_last_completed_date"])
    cycle_ts = pd.to_datetime(cycle_start)
    priority["days_since_completed"] = (cycle_ts - priority["last_completed_date"]).dt.days
    priority["days_since_completed"] = priority["days_since_completed"].fillna(9999).clip(lower=0)
    priority["cycles_missed"] = priority[["backlog_cycles_missed", "exception_count"]].max(axis=1).astype(int)
    priority["rotation_priority_score"] = (
        (priority["not_scheduled_count"] > 0).astype(int) * 1000
        + (priority["carryover_count"] > 0).astype(int) * 900
        + (priority["exception_count"] > 0).astype(int) * 800
        + priority["last_completed_date"].isna().astype(int) * 700
        + priority["cycles_missed"].clip(upper=12) * 50
        + (priority["days_since_completed"].clip(upper=730) / 10).round().astype(int)
    )
    priority["rotation_reason"] = "Normal rotation"
    priority.loc[priority["last_completed_date"].isna(), "rotation_reason"] = "Never completed"
    priority.loc[priority["exception_count"] > 0, "rotation_reason"] = "Prior month not completed or exception"
    priority.loc[priority["carryover_count"] > 0, "rotation_reason"] = "Carryover from prior cycle"
    priority.loc[priority["not_scheduled_count"] > 0, "rotation_reason"] = "Did not fit prior schedule"
    priority["front_of_line_carryover"] = (
        (priority["not_scheduled_count"] > 0)
        | (priority["carryover_count"] > 0)
        | (priority["exception_count"] > 0)
    )
    priority["rotation_priority_group"] = priority["front_of_line_carryover"].apply(lambda value: 0 if value else 1)
    return priority


def scalar_int(value, default=0):
    parsed = pd.to_numeric(value, errors="coerce")
    if pd.isna(parsed):
        return int(default)
    return int(parsed)


def scalar_date(value):
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.date()


def build_pmt_draft(assignments, start_month, months, targets, direction, avoid_weekends, avoid_holidays, avoid_pto, schedule_mode):
    rows = []
    unscheduled = []
    schedule_history, backlog_history = pmt_store_history()
    for employee_id, tech_df in assignments.groupby("employee_id"):
        tech_df = tech_df.copy()
        tech_name = tech_df.iloc[0]["technician_name"]
        home_lat = to_float(tech_df.iloc[0]["home_latitude"])
        home_lon = to_float(tech_df.iloc[0]["home_longitude"])
        if home_lat is None or home_lon is None:
            unscheduled.extend(
                {
                    "employee_id": int(employee_id),
                    "store_id": int(row["store_id"]),
                    "Technician": tech_name,
                    "Store Number": row["store_number"],
                    "City": row.get("store_city", ""),
                    "Reason": "Technician missing home latitude/longitude",
                    "Status": "Not Scheduled",
                    "Priority Score": 0,
                    "Cycles Missed": 1,
                }
                for _, row in tech_df.iterrows()
            )
            continue
        tech_df["distance_from_home"] = tech_df.apply(
            lambda row: haversine_miles(home_lat, home_lon, float(row["latitude"]), float(row["longitude"]))
            if pd.notna(row["latitude"]) and pd.notna(row["longitude"]) else None,
            axis=1,
        )
        missing_store_coords = tech_df[tech_df["distance_from_home"].isna()]
        for _, row in missing_store_coords.iterrows():
            unscheduled.append({
                "employee_id": int(employee_id),
                "store_id": int(row["store_id"]),
                "Technician": tech_name,
                "Store Number": row["store_number"],
                "City": row.get("store_city", ""),
                "Reason": "Store missing latitude/longitude",
                "Status": "Not Scheduled",
                "Priority Score": 0,
                "Cycles Missed": 1,
            })
        schedulable = tech_df.dropna(subset=["distance_from_home"]).copy()
        target = int(targets.get(int(employee_id), 10))
        capacity = target * int(months)
        prioritized = apply_pmt_rotation_priority(schedulable, schedule_history, backlog_history, start_month)
        schedule_sort_columns = ["rotation_priority_group", "distance_from_home", "store_number"]
        scheduled = prioritized.sort_values(schedule_sort_columns, ascending=[True, True, True]).head(capacity).copy()
        scheduled_indexes = scheduled.index.tolist()
        left = prioritized.drop(index=scheduled_indexes, errors="ignore")
        assigned_count = int(schedulable["store_id"].nunique())
        for _, row in left.iterrows():
            unscheduled.append({
                "employee_id": int(employee_id),
                "store_id": int(row["store_id"]),
                "Technician": tech_name,
                "Store Number": row["store_number"],
                "City": row.get("store_city", ""),
                "Reason": f"Too many stores to fit into selected months. Assigned: {assigned_count}; capacity: {capacity}",
                "Status": "Not Scheduled",
                "Priority Score": int(row.get("rotation_priority_score", 0) or 0),
                "Cycles Missed": int(row.get("cycles_missed", 0) or 0) + 1,
                "Last Scheduled Month": row.get("last_scheduled_month"),
                "Last Completed Date": row.get("last_completed_date"),
            })
        remaining = scheduled.copy()
        for month_index in range(int(months)):
            if remaining.empty or not target:
                continue
            cycle_month = add_months(start_month, month_index)
            month_pool = remaining.sort_values(schedule_sort_columns, ascending=[True, True, True]).head(target).copy()
            routed_rows = home_distance_route(month_pool, home_lat, home_lon, limit=target)
            remaining = remaining.drop(index=[row.name for row in routed_rows], errors="ignore")
            for sequence_number, row in enumerate(routed_rows, start=1):
                schedule_date = first_workday(cycle_month, avoid_weekends, avoid_holidays, int(employee_id), avoid_pto)
                if schedule_mode == "Monthly schedule with daily stops":
                    schedule_date = first_workday(cycle_month + timedelta(days=(sequence_number - 1)), avoid_weekends, avoid_holidays, int(employee_id), avoid_pto)
                rows.append(
                    {
                        "technician": tech_name,
                        "employee_id": int(employee_id),
                        "month": month_label(cycle_month),
                        "month_start": cycle_month,
                        "schedule_date": schedule_date,
                        "sequence_number": sequence_number,
                        "store_id": int(row["store_id"]),
                        "store_number": row["store_number"],
                        "address": row.get("store_address", ""),
                        "city": row.get("store_city", ""),
                        "state": row.get("store_state", ""),
                        "zip": row.get("store_zip", ""),
                        "home_latitude": home_lat,
                        "home_longitude": home_lon,
                        "latitude": row.get("latitude"),
                        "longitude": row.get("longitude"),
                        "distance_from_home": round(float(row["distance_from_home"]), 1),
                        "miles_from_previous_stop": round(float(row.get("miles_from_previous_stop", row["distance_from_home"])), 1),
                        "estimated_drive_time": "",
                        "work_type": "PMT",
                        "status": "Scheduled",
                        "rotation_priority_score": int(row.get("rotation_priority_score", 0) or 0),
                        "front_of_line_carryover": bool(row.get("front_of_line_carryover", False)),
                        "rotation_priority_group": int(row.get("rotation_priority_group", 1) or 1),
                        "rotation_reason": row.get("rotation_reason", "Normal rotation"),
                        "cycles_missed": int(row.get("cycles_missed", 0) or 0),
                        "last_completed_date": row.get("last_completed_date"),
                        "last_scheduled_month": row.get("last_scheduled_month"),
                        "notes": f"{row.get('rotation_reason', 'Normal rotation')} | Priority score: {int(row.get('rotation_priority_score', 0) or 0)}",
                    }
                )
    return pd.DataFrame(rows), pd.DataFrame(unscheduled)


def publish_draft(preview, unscheduled_preview, run_name, start_month, months, default_target, direction, schedule_mode, replace_existing):
    if preview.empty:
        return None
    cycle_end = add_months(start_month, months) - timedelta(days=1)
    cycle_label = f"{month_label(start_month)} - {month_label(add_months(start_month, months - 1))}"
    with session_scope() as session:
        if replace_existing:
            existing_items = session.scalars(
                select(ScheduleItem).where(
                    ScheduleItem.work_type == "PMT",
                    ScheduleItem.schedule_date >= start_month,
                    ScheduleItem.schedule_date <= cycle_end,
                )
            ).all()
            for item in existing_items:
                session.delete(item)
        run = PMTScheduleRun(
            run_name=run_name,
            cycle_start=start_month,
            cycle_end=cycle_end,
            months=months,
            default_monthly_target=default_target,
            direction=direction,
            schedule_mode=schedule_mode,
            distance_method="Estimated straight-line distance",
            technician_count=int(preview["employee_id"].nunique()),
            store_count=len(preview),
            unscheduled_count=len(unscheduled_preview) if unscheduled_preview is not None else 0,
            created_by=st.session_state.get("username", ""),
            notes=cycle_label,
        )
        session.add(run)
        session.flush()
        schedule = Schedule(
            schedule_name=run_name,
            schedule_type="PMT Monthly Auto-Scheduler",
            start_date=start_month,
            end_date=cycle_end,
            status="Published",
            created_by=st.session_state.get("username", ""),
            notes=f"Schedule Source: PMT Monthly Auto-Scheduler | Run ID: {run.id} | Cycle: {cycle_label}",
        )
        session.add(schedule)
        session.flush()
        created = 0
        skipped = 0
        for _, row in preview.iterrows():
            month_start_value = pd.to_datetime(row["month_start"]).date() if not isinstance(row["month_start"], date) else row["month_start"]
            schedule_date_value = pd.to_datetime(row["schedule_date"]).date() if not isinstance(row["schedule_date"], date) else row["schedule_date"]
            month_end = add_months(month_start_value, 1) - timedelta(days=1)
            duplicate = session.scalar(
                select(ScheduleItem.id).where(
                    ScheduleItem.employee_id == int(row["employee_id"]),
                    ScheduleItem.store_id == int(row["store_id"]),
                    ScheduleItem.work_type == "PMT",
                    ScheduleItem.schedule_date >= month_start_value,
                    ScheduleItem.schedule_date <= month_end,
                )
            )
            if duplicate and not replace_existing:
                skipped += 1
                continue
            session.add(
                ScheduleItem(
                    schedule_id=schedule.id,
                    schedule_date=schedule_date_value,
                    sequence_number=int(row["sequence_number"]),
                    store_id=int(row["store_id"]),
                    employee_id=int(row["employee_id"]),
                    work_type="PMT",
                    status=row.get("status", "Scheduled"),
                    schedule_source="PMT Monthly Auto-Scheduler",
                    pmt_schedule_run_id=run.id,
                    cycle_label=cycle_label,
                    completion_notes=row.get("notes", ""),
                )
            )
            open_backlog = session.query(PMTScheduleBacklog).filter(
                PMTScheduleBacklog.employee_id == int(row["employee_id"]),
                PMTScheduleBacklog.store_id == int(row["store_id"]),
                PMTScheduleBacklog.status.in_(PMT_BACKLOG_OPEN_STATUSES),
            ).all()
            for backlog in open_backlog:
                backlog.status = "Scheduled"
                backlog.notes = f"{backlog.notes or ''}\nScheduled in PMT run {run.id}.".strip()
            created += 1
        if unscheduled_preview is not None and not unscheduled_preview.empty:
            for _, row in unscheduled_preview.iterrows():
                employee_id = pd.to_numeric(row.get("employee_id"), errors="coerce")
                store_id = pd.to_numeric(row.get("store_id"), errors="coerce")
                if pd.isna(employee_id) or pd.isna(store_id):
                    continue
                existing = session.query(PMTScheduleBacklog).filter(
                    PMTScheduleBacklog.pmt_schedule_run_id == run.id,
                    PMTScheduleBacklog.employee_id == int(employee_id),
                    PMTScheduleBacklog.store_id == int(store_id),
                    PMTScheduleBacklog.status == "Not Scheduled",
                ).first()
                if existing:
                    backlog = existing
                else:
                    backlog = PMTScheduleBacklog(
                        pmt_schedule_run_id=run.id,
                        employee_id=int(employee_id),
                        store_id=int(store_id),
                        cycle_start=start_month,
                        cycle_end=cycle_end,
                    )
                    session.add(backlog)
                backlog.status = clean(row.get("Status", "")) or "Not Scheduled"
                backlog.reason = clean(row.get("Reason", "")) or "Too many stores to fit into selected months"
                backlog.cycles_missed = scalar_int(row.get("Cycles Missed", 1), 1)
                backlog.priority_score = scalar_int(row.get("Priority Score", 0), 0)
                backlog.last_scheduled_month = scalar_date(row.get("Last Scheduled Month"))
                backlog.last_completed_date = scalar_date(row.get("Last Completed Date"))
                backlog.last_completed_month = month_start(backlog.last_completed_date) if backlog.last_completed_date else None
                backlog.notes = f"PMT run {run.id}: {backlog.reason}"
        run.store_count = created
        run.unscheduled_count = len(unscheduled_preview) if unscheduled_preview is not None else 0
        run_id = run.id
    log_action("pmt schedule run published", "pmt_schedule_runs", int(run_id), f"{created} items created, {skipped} skipped")
    return {"run_id": run_id, "created": created, "skipped": skipped}


def pmt_publish_conflicts(preview, start_month, months):
    if preview.empty or not {"employee_id", "store_id", "month_start"}.issubset(preview.columns):
        return pd.DataFrame()
    cycle_end = add_months(start_month, months) - timedelta(days=1)
    checks = preview[["employee_id", "store_id", "month_start"]].dropna().copy()
    checks["employee_id"] = pd.to_numeric(checks["employee_id"], errors="coerce")
    checks["store_id"] = pd.to_numeric(checks["store_id"], errors="coerce")
    checks["month_start"] = pd.to_datetime(checks["month_start"], errors="coerce").dt.date
    checks = checks.dropna(subset=["employee_id", "store_id", "month_start"])
    if checks.empty:
        return pd.DataFrame()
    checks["employee_id"] = checks["employee_id"].astype(int)
    checks["store_id"] = checks["store_id"].astype(int)
    checks = checks.drop_duplicates()
    existing = safe_query(
        """
        select si.id as existing_item_id, si.schedule_id, sch.schedule_name, si.employee_id,
               e.full_name as technician, si.store_id, st.store_number, st.city,
               si.schedule_date, si.status
        from schedule_items si
        left join schedules sch on sch.id = si.schedule_id
        left join employees e on e.id = si.employee_id
        left join stores st on st.id = si.store_id
        where si.work_type = 'PMT'
          and si.status in ('Scheduled','Needs Rescheduled','Rescheduled','Rain Delay','Not Completed')
          and si.schedule_date >= :start_month
          and si.schedule_date <= :cycle_end
        order by e.full_name, st.store_number, si.schedule_date
        """,
        {"start_month": start_month, "cycle_end": cycle_end},
    )
    if existing.empty:
        return existing
    existing = existing.copy()
    existing["month_start"] = pd.to_datetime(existing["schedule_date"], errors="coerce").dt.date.apply(lambda value: date(value.year, value.month, 1) if pd.notna(value) else value)
    return existing.merge(checks, on=["employee_id", "store_id", "month_start"], how="inner")



tab_build, tab_carryover, tab_manage_fix, tab_export = st.tabs([
    "📋  Build Schedule",
    "📊  Carryover & Backlog",
    "🛠️  Manage & Fix Schedule",
    "📥  Export",
])

with tab_manage_fix:
    tab_health, tab_reconcile, tab_rebuild, tab_manage, tab_preview, tab_history = st.tabs([
        "Overview & Health",
        "Territory Reconciliation",
        "Rebuild / Balance",
        "Manual Edit",
        "Preview Changes",
        "History & Revisions",
    ])

with tab_build:
    section_header("Build Step 1: Choose Assignment Source", "Use assignments already saved in the app, or upload a PMT assigned-store file. Employee home addresses come from Employees by default.", "blue", focus_key="pmt_focus_step", focus_value=1)
    source_choice = st.radio("Assignment source", ["Use existing PMT assignments in the app", "Upload PMT assignment Excel/CSV"], horizontal=True)
    uploaded_assignments = pd.DataFrame()
    upload_problems = pd.DataFrame()

    if source_choice == "Upload PMT assignment Excel/CSV":
        section_header(
            "Build Step 2: Upload Assignment File",
            "Upload the PMT assigned-store file. The app will detect PMT, site number, latitude, and longitude automatically.",
            "gray",
            focus_key="pmt_focus_step",
            focus_value=2,
        )
        upload = st.file_uploader("Upload PMT assignment file", type=["xlsx", "xls", "xlsm", "csv"], key="pmt_assignment_upload")
        if upload:
            st.warning("Uploaded files are not saved yet. After the file validates, use Step 3A to save these PMT assignments into the app.")
            assignment_scans = scan_uploaded_workbook(upload, "assignments")
            scan_issues = scan_issue_rows(assignment_scans)
            if not scan_issues.empty:
                with st.expander("Upload scan warnings", expanded=False):
                    st.dataframe(scan_issues, use_container_width=True, hide_index=True)
                    if st.session_state.get("account_role") == "Admin":
                        technical = [item.get("technical_detail") for item in assignment_scans if item.get("technical_detail")]
                        if technical:
                            st.caption("Admin debug details")
                            st.code("\n\n".join(technical))
            if not assignment_scans or all(item["df"].empty for item in assignment_scans):
                st.error("No usable rows were found in this upload. Check that the workbook has a visible sheet with assignment data.")
                st.stop()
            sheet_names = [item["sheet"] for item in assignment_scans]
            if st.button("Auto-Detect Columns", type="secondary"):
                for state_key in list(st.session_state.keys()):
                    if state_key.startswith("pmt_map_") or state_key.startswith("pmt_addr_") or state_key in {"pmt_assignment_sheet", "pmt_address_sheet"}:
                        st.session_state.pop(state_key, None)
                st.rerun()
            st.caption("The app automatically matches common column names like PMT, Site Number, Lat, Lon, State, and Type. Only open Advanced Column Mapping if something looks wrong.")
            assignment_default_index = 0
            address_default_index = detected_sheet_index(sheet_names, upload, "address")
            home_source = st.radio(
                "Home Address Source",
                ["Use employee addresses already saved in the app", "Upload a separate home address sheet"],
                horizontal=True,
                index=0,
            )
            assignment_sheet = st.selectbox("Assigned store sheet", sheet_names, index=assignment_default_index, key="pmt_assignment_sheet")
            assignment_scan = next(item for item in assignment_scans if item["sheet"] == assignment_sheet)
            incoming = assignment_scan["df"]
            normalized = normalize_pmt_assignment_columns(incoming)
            original_columns = incoming.columns.tolist()
            mapping_options = [""] + original_columns
            smart_defaults = {field: match.column for field, match in assignment_scan["mapping"].items()}
            defaults = {
                "technician_name": smart_defaults.get("full_name") or smart_defaults.get("assigned_pmt") or best_column(original_columns, "technician_name", "assignment"),
                "store_number": smart_defaults.get("store_number") or best_column(original_columns, "store_number", "assignment"),
                "store_state": smart_defaults.get("state") or best_column(original_columns, "store_state", "assignment"),
                "latitude": smart_defaults.get("latitude") or best_column(original_columns, "latitude", "assignment"),
                "longitude": smart_defaults.get("longitude") or best_column(original_columns, "longitude", "assignment"),
                "store_address": smart_defaults.get("address") or best_column(original_columns, "store_address", "assignment"),
                "store_city": smart_defaults.get("city") or best_column(original_columns, "store_city", "assignment"),
                "store_zip": smart_defaults.get("zip") or best_column(original_columns, "store_zip", "assignment"),
            }
            missing_required_columns = not defaults["technician_name"] or not defaults["store_number"]
            det1, det2, det3, det4, det5 = st.columns(5)
            det1.metric("Assigned Store Sheet", assignment_sheet)
            st.caption(f"Header row detected: {assignment_scan['header_row'] + 1}. Rows detected: {assignment_scan['rows']:,}.")
            det2.metric("Technician Column", defaults["technician_name"] or "Not found")
            det3.metric("Store Number Column", defaults["store_number"] or "Not found")
            det4.metric("Latitude Column", defaults["latitude"] or "Database fallback")
            det5.metric("Longitude Column", defaults["longitude"] or "Database fallback")
            if missing_required_columns:
                st.error("The app could not find the PMT technician column or store number column. Open Advanced Column Mapping and choose those two columns.")
            else:
                st.success("Ready to validate. Advanced mapping is only needed if the detected columns are wrong.")

            with st.expander("Advanced Column Mapping", expanded=missing_required_columns):
                st.caption("Only Technician Name and Store Number are required. Store coordinates are used from the upload when present, otherwise from the saved Stores database.")
                c1, c2 = st.columns(2)
                tech_col = selectbox_with_default(c1, "Technician / PMT Name", mapping_options, defaults["technician_name"], "pmt_map_tech_col")
                store_col = selectbox_with_default(c2, "Store / Site Number", mapping_options, defaults["store_number"], "pmt_map_store_col")
                c3, c4, c5 = st.columns(3)
                lat_col = selectbox_with_default(c3, "Store Latitude", mapping_options, defaults["latitude"], "pmt_map_lat_col")
                lon_col = selectbox_with_default(c4, "Store Longitude", mapping_options, defaults["longitude"], "pmt_map_lon_col")
                store_state_col = selectbox_with_default(c5, "Store State", mapping_options, defaults["store_state"], "pmt_map_store_state_col")
                show_optional_store_address = st.checkbox("Show optional store address fields", value=False)
                store_address_col = defaults["store_address"]
                store_city_col = defaults["store_city"]
                store_zip_col = defaults["store_zip"]
                if show_optional_store_address:
                    oc1, oc2, oc3 = st.columns(3)
                    store_address_col = selectbox_with_default(oc1, "Store Address", mapping_options, defaults["store_address"], "pmt_map_store_address_col")
                    store_city_col = selectbox_with_default(oc2, "Store City", mapping_options, defaults["store_city"], "pmt_map_store_city_col")
                    store_zip_col = selectbox_with_default(oc3, "Store Zip", mapping_options, defaults["store_zip"], "pmt_map_store_zip_col")
            tech_col = st.session_state.get("pmt_map_tech_col", defaults["technician_name"])
            store_col = st.session_state.get("pmt_map_store_col", defaults["store_number"])
            lat_col = st.session_state.get("pmt_map_lat_col", defaults["latitude"])
            lon_col = st.session_state.get("pmt_map_lon_col", defaults["longitude"])
            store_state_col = st.session_state.get("pmt_map_store_state_col", defaults["store_state"])
            store_address_col = st.session_state.get("pmt_map_store_address_col", defaults["store_address"])
            store_city_col = st.session_state.get("pmt_map_store_city_col", defaults["store_city"])
            store_zip_col = st.session_state.get("pmt_map_store_zip_col", defaults["store_zip"])
            selected = {
                "technician_name": tech_col,
                "store_number": store_col,
                "store_address": store_address_col,
                "store_city": store_city_col,
                "store_state": store_state_col,
                "store_zip": store_zip_col,
                "latitude": lat_col,
                "longitude": lon_col,
            }
            mapped = apply_column_mapping(normalized, incoming, selected)
            if not tech_col or not store_col:
                st.error("The assigned-store sheet needs a Technician Name column and a Store Number column before it can match records.")

            if home_source == "Upload a separate home address sheet":
                st.markdown("#### Optional Home Address Sheet")
                st.caption("Use this only when Employee Admin does not already have home addresses. These values will be matched by technician name.")
                address_sheet_choices = sheet_names
                address_sheet = st.selectbox(
                    "Sheet with PMT home addresses",
                    address_sheet_choices,
                    index=address_default_index,
                    key="pmt_address_sheet",
                )
                address_incoming = read_upload_sheet(upload, address_sheet)
                address_normalized = normalize_pmt_assignment_columns(address_incoming)
                address_columns = address_incoming.columns.tolist()
                address_options = [""] + address_columns
                addr_missing = not best_column(address_columns, "technician_name", "address") or not best_column(address_columns, "home_address", "address")
                with st.expander("Home Address Column Mapping", expanded=addr_missing):
                    a1, a2, a3 = st.columns(3)
                    address_tech_col = selectbox_with_default(a1, "Technician Name", address_options, best_column(address_columns, "technician_name", "address"), "pmt_addr_tech_col")
                    address_home_col = selectbox_with_default(a2, "Home Address", address_options, best_column(address_columns, "home_address", "address"), "pmt_addr_home_col")
                    address_city_col = selectbox_with_default(a3, "Home City", address_options, best_column(address_columns, "home_city", "address"), "pmt_addr_city_col")
                    a4, a5, a6 = st.columns(3)
                    address_state_col = selectbox_with_default(a4, "Home State", address_options, best_column(address_columns, "home_state", "address"), "pmt_addr_state_col")
                    address_zip_col = selectbox_with_default(a5, "Home Zip", address_options, best_column(address_columns, "home_zip", "address"), "pmt_addr_zip_col")
                    address_home_lat_col = selectbox_with_default(a6, "Home Latitude", address_options, best_column(address_columns, "home_latitude", "address"), "pmt_addr_home_lat_col")
                    a7, _ = st.columns(2)
                    address_home_lon_col = selectbox_with_default(a7, "Home Longitude", address_options, best_column(address_columns, "home_longitude", "address"), "pmt_addr_home_lon_col")
                address_selected = {
                    "technician_name": st.session_state.get("pmt_addr_tech_col", best_column(address_columns, "technician_name", "address")),
                    "home_address": st.session_state.get("pmt_addr_home_col", best_column(address_columns, "home_address", "address")),
                    "home_city": st.session_state.get("pmt_addr_city_col", best_column(address_columns, "home_city", "address")),
                    "home_state": st.session_state.get("pmt_addr_state_col", best_column(address_columns, "home_state", "address")),
                    "home_zip": st.session_state.get("pmt_addr_zip_col", best_column(address_columns, "home_zip", "address")),
                    "home_latitude": st.session_state.get("pmt_addr_home_lat_col", best_column(address_columns, "home_latitude", "address")),
                    "home_longitude": st.session_state.get("pmt_addr_home_lon_col", best_column(address_columns, "home_longitude", "address")),
                }
                address_mapped = apply_column_mapping(address_normalized, address_incoming, address_selected)
                mapped = merge_home_address_sheet(mapped, address_mapped)
                preview_cols = [col for col in ["technician_name", "home_address", "home_city", "home_state", "home_zip", "home_latitude", "home_longitude"] if col in address_mapped.columns]
                if preview_cols:
                    st.dataframe(address_mapped[preview_cols].drop_duplicates().head(25), use_container_width=True, hide_index=True)

            employee_import_result = ensure_uploaded_pmt_employees(mapped)
            if employee_import_result["created"] or employee_import_result["updated"]:
                st.info(
                    f"Prepared PMT employee records from upload: "
                    f"{employee_import_result['created']} created, {employee_import_result['updated']} updated/reactivated."
                )
            uploaded_assignments, upload_problems = prepare_uploaded_assignments(mapped)
            upload_count = int(mapped["store_number"].astype(str).str.strip().ne("").sum()) if "store_number" in mapped.columns else len(mapped)
            match_cols = st.columns(4)
            match_cols[0].metric("Rows in Upload", f"{len(mapped):,}")
            match_cols[1].metric("Store Rows Found", f"{upload_count:,}")
            match_cols[2].metric("Matched Stores", f"{uploaded_assignments['store_id'].nunique():,}" if not uploaded_assignments.empty else "0")
            match_cols[3].metric("Matched PMTs", f"{uploaded_assignments['employee_id'].nunique():,}" if not uploaded_assignments.empty else "0")
            if not uploaded_assignments.empty:
                st.info("Next: review Step 3 below, then click 'Save Uploaded PMT Assignments to App'. Generating a schedule is a later step.")
            preview_columns = [
                col
                for col in [
                    "technician_name",
                    "store_number",
                    "store_state",
                    "latitude",
                    "longitude",
                ]
                if col in mapped.columns
            ]
            if home_source == "Upload a separate home address sheet":
                matched_home_addresses = mapped["home_address"].astype(str).str.strip().ne("").sum() if "home_address" in mapped.columns else 0
                st.success(f"Matched home address data onto {matched_home_addresses:,} assigned-store rows.")
            st.dataframe(mapped[preview_columns].head(50) if preview_columns else mapped.head(50), use_container_width=True, hide_index=True)
    else:
        section_header("Build Step 2: Confirm Saved PMT Assignments", "Review existing PMT assignments already saved in Stores and Employees.", "gray", focus_key="pmt_focus_step", focus_value=2)
        uploaded_assignments = current_assignments_from_database()
        if uploaded_assignments.empty:
            st.info("No PMT assignments were found. To continue, either upload a PMT assignment file or assign stores to PMTs under Stores / Areas and Maps.")
            b1, b2, b3 = st.columns(3)
            b1.page_link("pages/13_PMT_Monthly_Scheduler.py", label="Upload PMT Assignment File")
            b2.page_link("pages/4_Map_Center.py", label="Open Areas and Maps")
            b3.page_link("pages/2_Employees.py", label="Open Employees")
        else:
            st.dataframe(uploaded_assignments.head(100), use_container_width=True, hide_index=True)

    active_pmt_summary = active_pmt_employee_summary()
    zero_store_pmts = active_pmt_summary[active_pmt_summary["assigned_stores"].fillna(0).astype(int) == 0] if not active_pmt_summary.empty else pd.DataFrame()
    assignments = enrich_assignments(uploaded_assignments)
    section_header("Build Step 3: Validate PMT Assignments And Locations", "Confirm the upload matched active PMTs, saved stores, store coordinates, and employee home locations.", "yellow", focus_key="pmt_focus_step", focus_value=3)
    summary, problems = validation_summary(assignments)
    if assignments.empty:
        st.warning("No PMT assignments found yet. Upload an assignment file or assign stores to PMTs in Stores / Areas and Maps.")
        if not active_pmt_summary.empty:
            st.metric("Active PMTs In Employees", int(len(active_pmt_summary)))
            st.dataframe(
                active_pmt_summary[["employee_id", "technician_name", "home_city", "home_state", "home_latitude", "home_longitude", "assigned_stores"]],
                use_container_width=True,
                hide_index=True,
            )
        cta1, cta2, cta3 = st.columns(3)
        cta1.page_link("pages/13_PMT_Monthly_Scheduler.py", label="Upload PMT Assignment File")
        cta2.page_link("pages/4_Map_Center.py", label="Open Areas and Maps")
        cta3.page_link("pages/2_Employees.py", label="Open Employees")
    else:
        m1, m2, m3, m4, m5, m6, m7 = st.columns(7)
        m1.metric("Rows", summary["Rows"])
        m2.metric("Active PMTs", int(len(active_pmt_summary)))
        m3.metric("PMTs With Stores", summary["Technicians"])
        m4.metric("Stores Assigned", summary["Stores"])
        m5.metric("Stores With Coordinates", summary["Stores With Coordinates"])
        with m6:
            metric_help_card("Stores Missing Location", summary["Stores Missing Coordinates"], "Assigned PMT stores missing latitude/longitude. These cannot be routed until coordinates are fixed.")
        with m7:
            metric_help_card("PMTs Missing Home Coordinates", summary["Missing Home Coordinates"], "Active PMTs with assigned stores but no usable home/base coordinates. Routing starts from this location.")
        if not zero_store_pmts.empty:
            st.warning(f"{len(zero_store_pmts)} active PMT technician(s) have no assigned stores. Use Areas and Maps -> PMT -> Staffing Change & Territory Rebalance to assign nearby stores.")
            st.dataframe(
                zero_store_pmts[["employee_id", "technician_name", "home_city", "home_state", "home_latitude", "home_longitude", "assigned_stores"]],
                use_container_width=True,
                hide_index=True,
            )
            st.page_link("pages/4_Map_Center.py", label="Open Areas and Maps Rebalance")
        if not upload_problems.empty:
            st.warning("Some upload rows need review. The table below shows whether the issue is a store match, inactive store, or employee match.")
            st.dataframe(upload_problems, use_container_width=True, hide_index=True)
        if source_choice == "Upload PMT assignment Excel/CSV" and not assignments.empty:
            section_header("Build Step 3A: Save Uploaded PMT Assignments", "This writes the uploaded PMT-to-store assignments into the app. It does not publish a schedule.", "green", focus_key="pmt_focus_step", focus_value=3)
            uploaded_schedule_date_col = best_column(original_columns, "schedule_date", "schedule")
            uploaded_schedule_month_col = best_column(original_columns, "schedule_month", "schedule")
            uploaded_schedule_sequence_col = best_column(original_columns, "sequence_number", "schedule")
            uploaded_schedule_status_col = best_column(original_columns, "status", "schedule")
            uploaded_schedule_notes_col = best_column(original_columns, "notes", "schedule")
            schedule_like_upload = bool(uploaded_schedule_date_col or uploaded_schedule_month_col)
            if schedule_like_upload:
                st.error(
                    "This upload appears to contain schedule dates or months. Saving assignments from this file will change store ownership. "
                    "Use the schedule import section below if you only want to create or manage schedule rows."
                )
                confirm_assignment_overwrite = st.checkbox(
                    "I understand this will change store ownership assignments, not just schedules.",
                    key="pmt_confirm_schedule_like_assignment_save",
                )
            else:
                st.warning("Required after upload: click this save button before relying on these assignments in Areas and Maps, reports, or future scheduler runs.")
                confirm_assignment_overwrite = True
            save_cols = st.columns([0.35, 0.65])
            if save_cols[0].button("Save Technician-to-Store Assignments", type="primary", disabled=not confirm_assignment_overwrite, key="pmt_save_uploaded_assignments"):
                with session_scope() as session:
                    saved_stores = 0
                    saved_employees = set()
                    employee_team_ids = {}
                    employee_store_ids = {}
                    for _, row in assignments.iterrows():
                        employee = session.get(Employee, int(row["employee_id"]))
                        store = session.get(Store, int(row["store_id"]))
                        if employee:
                            if clean(row.get("home_address", "")):
                                employee.home_address = clean(row.get("home_address", ""))
                            if clean(row.get("home_city", "")):
                                employee.home_city = clean(row.get("home_city", ""))
                            if clean(row.get("home_state", "")):
                                employee.home_state = clean(row.get("home_state", ""))
                            if clean(row.get("home_zip", "")):
                                employee.home_zip = clean(row.get("home_zip", ""))
                            if pd.notna(row.get("home_latitude")):
                                employee.home_latitude = float(row.get("home_latitude"))
                            if pd.notna(row.get("home_longitude")):
                                employee.home_longitude = float(row.get("home_longitude"))
                            saved_employees.add(int(row["employee_id"]))
                            team = session.query(Team).filter(Team.team_name == employee.full_name, Team.team_type == "PMT").first()
                            if not team:
                                team = Team(team_name=employee.full_name, team_type="PMT", city=employee.home_city or "", state=employee.home_state or "", active=True)
                                session.add(team)
                                session.flush()
                            else:
                                team.active = True
                            employee_team_ids[int(employee.id)] = int(team.id)
                        if store:
                            store.assigned_pmt_employee_id = int(row["employee_id"])
                            if int(row["employee_id"]) in employee_team_ids:
                                store.assigned_pmt_team_id = employee_team_ids[int(row["employee_id"])]
                            employee_store_ids.setdefault(int(row["employee_id"]), []).append(int(store.id))
                            saved_stores += 1
                    for employee_id, team_id in employee_team_ids.items():
                        employee = session.get(Employee, int(employee_id))
                        area = session.query(MapArea).filter(MapArea.team_id == int(team_id), MapArea.area_type == "PMT", MapArea.active == True).first()
                        store_ids = sorted(employee_store_ids.get(int(employee_id), []))
                        if area:
                            area.area_name = employee.full_name if employee else area.area_name
                            area.employee_id = int(employee_id)
                            area.assigned_store_ids = json.dumps(store_ids)
                        else:
                            session.add(
                                MapArea(
                                    area_name=employee.full_name if employee else f"PMT {employee_id}",
                                    area_type="PMT",
                                    team_id=int(team_id),
                                    employee_id=int(employee_id),
                                    assignment_type="PMT area",
                                    team_members=json.dumps([int(employee_id)]),
                                    home_base=", ".join([value for value in [employee.home_city if employee else "", employee.home_state if employee else ""] if value]),
                                    geometry_json=json.dumps({"type": "Polygon", "coordinates": [[]]}),
                                    assigned_store_ids=json.dumps(store_ids),
                                    color=(employee.color if employee else None) or stable_color(employee.full_name if employee else str(employee_id)),
                                    active=True,
                                )
                            )
                st.success(
                    f"Saved PMT assignments successfully. Technicians updated: {len(saved_employees)}. "
                    f"Stores assigned: {saved_stores}. Stores skipped: {len(upload_problems)}. Problems found: {len(upload_problems) + len(problems)}."
                )
                st.rerun()
            save_cols[1].caption("This saves store ownership only. It does not create schedule dates, route stops, or a manageable schedule run.")
            if uploaded_schedule_date_col or uploaded_schedule_month_col:
                st.markdown("**This upload also looks like a schedule**")
                st.caption("Use this section when the file already has schedule dates or schedule months and you want it to appear in Manage & Fix Schedule.")
                with st.expander("Import this upload as an existing PMT schedule", expanded=True):
                    schedule_options = [""] + original_columns
                    sc1, sc2, sc3 = st.columns(3)
                    inline_schedule_tech_col = selectbox_with_default(sc1, "Technician", schedule_options, tech_col, "pmt_inline_schedule_tech_col")
                    inline_schedule_store_col = selectbox_with_default(sc2, "Store Number", schedule_options, store_col, "pmt_inline_schedule_store_col")
                    inline_schedule_date_col = selectbox_with_default(sc3, "Schedule Date", schedule_options, uploaded_schedule_date_col, "pmt_inline_schedule_date_col")
                    sc4, sc5, sc6, sc7 = st.columns(4)
                    inline_schedule_month_col = selectbox_with_default(sc4, "Month if no date", schedule_options, uploaded_schedule_month_col, "pmt_inline_schedule_month_col")
                    inline_schedule_sequence_col = selectbox_with_default(sc5, "Stop / Sequence", schedule_options, uploaded_schedule_sequence_col, "pmt_inline_schedule_sequence_col")
                    inline_schedule_status_col = selectbox_with_default(sc6, "Status", schedule_options, uploaded_schedule_status_col, "pmt_inline_schedule_status_col")
                    inline_schedule_notes_col = selectbox_with_default(sc7, "Notes", schedule_options, uploaded_schedule_notes_col, "pmt_inline_schedule_notes_col")
                    inline_mapping = {
                        "technician_name": inline_schedule_tech_col,
                        "store_number": inline_schedule_store_col,
                        "schedule_date": inline_schedule_date_col,
                        "schedule_month": inline_schedule_month_col,
                        "sequence_number": inline_schedule_sequence_col,
                        "status": inline_schedule_status_col,
                        "notes": inline_schedule_notes_col,
                    }
                    inline_missing = []
                    if not inline_schedule_tech_col:
                        inline_missing.append("Technician")
                    if not inline_schedule_store_col:
                        inline_missing.append("Store Number")
                    if not inline_schedule_date_col and not inline_schedule_month_col:
                        inline_missing.append("Schedule Date or Month")
                    inline_preview, inline_problems = normalize_existing_pmt_schedule_upload(incoming, inline_mapping) if not inline_missing else (pd.DataFrame(), pd.DataFrame())
                    if inline_missing:
                        st.error("Missing required schedule mapping: " + ", ".join(inline_missing))
                    elif inline_preview.empty:
                        st.warning("No valid schedule rows were found after matching technicians, stores, and schedule dates.")
                    else:
                        sm1, sm2, sm3, sm4 = st.columns(4)
                        sm1.metric("Schedule Rows Ready", len(inline_preview))
                        sm2.metric("Technicians", inline_preview["employee_id"].nunique())
                        sm3.metric("Unique Stores", inline_preview["store_id"].nunique())
                        sm4.metric("Warnings / Problems", len(inline_problems))
                        if not inline_problems.empty:
                            with st.expander("Schedule import warnings", expanded=False):
                                st.dataframe(inline_problems, use_container_width=True, hide_index=True)
                        schedule_preview_columns = ["technician", "month", "schedule_date", "sequence_number", "store_number", "city", "state", "status", "notes"]
                        st.dataframe(inline_preview[schedule_preview_columns].head(100), use_container_width=True, hide_index=True)
                        inline_start = inline_preview["schedule_date"].min()
                        inline_end = inline_preview["schedule_date"].max()
                        inline_run_name = st.text_input(
                            "Schedule run name",
                            value=f"Imported PMT Schedule {month_label(month_start(inline_start))} - {month_label(month_start(inline_end))}",
                            key="pmt_inline_schedule_run_name",
                        )
                        inline_confirm = st.checkbox("I reviewed this schedule import and want to create a manageable PMT schedule run.", key="pmt_inline_schedule_confirm")
                        if st.button("Import This Upload As Existing PMT Schedule", type="primary", disabled=not inline_confirm, key="pmt_inline_schedule_import_button"):
                            try:
                                with st.spinner(f"Creating PMT schedule run with {len(inline_preview):,} item(s)..."):
                                    result = import_existing_pmt_schedule(inline_preview, inline_run_name)
                                st.success(f"Imported PMT schedule run #{result['run_id']} with {result['created']} schedule item(s). Open Manage & Fix Schedule -> Manual Edit to review and edit it.")
                            except Exception as exc:
                                st.error(f"PMT schedule import failed: {exc}")
                                st.stop()
            else:
                st.info("If this file is your already-made schedule, it needs a mapped Schedule Date or Month column. The assignment save button will not create a manageable schedule.")
        if problems.empty:
            st.success("Data check passed. You can generate a draft schedule.")
        else:
            st.dataframe(problems, use_container_width=True, hide_index=True)
            st.info("Only Must Fix items block scheduling. Use the tools below for employee home locations, or open Stores to fix store coordinates.")

        section_header("Build Step 4: Fix Blocking Problems", "Fix only the items that stop scheduling, like missing employee home coordinates or stores with no usable location.", "yellow", focus_key="pmt_focus_step", focus_value=4)
        missing_home = assignments.drop_duplicates("employee_id")
        missing_home = missing_home[missing_home[["home_latitude", "home_longitude"]].isna().any(axis=1)]
        if not missing_home.empty:
            with st.expander("Enter Missing Technician Home Coordinates", expanded=True):
                st.caption("Use address lookup for exact home coordinates. If it cannot find the street, use City/ZIP Estimate so the PMT scheduler can keep moving.")
                editor_ids = "_".join(str(int(value)) for value in sorted(missing_home["employee_id"].dropna().unique()))
                edits = st.data_editor(
                    missing_home[["employee_id", "technician_name", "home_address", "home_city", "home_state", "home_zip", "home_latitude", "home_longitude"]],
                    use_container_width=True,
                    hide_index=True,
                    disabled=["employee_id", "technician_name"],
                    key=f"pmt_home_coord_editor_{editor_ids}",
                )
                geo_col, estimate_col, save_col = st.columns(3)
                if geo_col.button("Find Coordinates From Address", type="primary"):
                    found = 0
                    not_found = []
                    found_details = []
                    with session_scope() as session:
                        for _, row in edits.iterrows():
                            employee = session.get(Employee, int(row["employee_id"]))
                            if not employee:
                                continue
                            if employee.home_latitude is not None and employee.home_longitude is not None:
                                continue
                            if to_float(row.get("home_latitude")) is not None and to_float(row.get("home_longitude")) is not None:
                                continue
                            result = geocode_address(
                                row.get("home_address", ""),
                                row.get("home_city", ""),
                                row.get("home_state", ""),
                                row.get("home_zip", ""),
                            )
                            if not result:
                                result = local_coordinate_estimate(
                                    row.get("home_city", ""),
                                    row.get("home_state", ""),
                                    row.get("home_zip", ""),
                                )
                            if not result:
                                not_found.append(
                                    {
                                        "Technician": row["technician_name"],
                                        "Address Tried": build_address(
                                            row.get("home_address", ""),
                                            row.get("home_city", ""),
                                            row.get("home_state", ""),
                                            row.get("home_zip", ""),
                                        ),
                                    }
                                )
                                continue
                            employee.home_address = clean(row.get("home_address", "")) or employee.home_address
                            employee.home_city = clean(row.get("home_city", "")) or employee.home_city
                            employee.home_state = clean(row.get("home_state", "")) or employee.home_state
                            employee.home_zip = clean(row.get("home_zip", "")) or employee.home_zip
                            employee.home_latitude = float(result["latitude"])
                            employee.home_longitude = float(result["longitude"])
                            found += 1
                            found_details.append(
                                {
                                    "Technician": row["technician_name"],
                                    "Match": result.get("match_quality", "Address match"),
                                    "Found Location": result.get("display_name", ""),
                                }
                            )
                            time.sleep(1)
                    if found:
                        st.success(f"Found and saved coordinates for {found} technician(s).")
                        if found_details:
                            st.dataframe(pd.DataFrame(found_details), use_container_width=True, hide_index=True)
                    if not_found:
                        st.warning("These addresses still could not be found. Check spelling, city, state, and ZIP, or enter coordinates manually.")
                        st.dataframe(pd.DataFrame(not_found), use_container_width=True, hide_index=True)
                    st.rerun()
                if estimate_col.button("Use City/ZIP Estimate", type="secondary"):
                    estimated = 0
                    not_estimated = []
                    estimated_details = []
                    with session_scope() as session:
                        for _, row in edits.iterrows():
                            employee = session.get(Employee, int(row["employee_id"]))
                            if not employee:
                                continue
                            if employee.home_latitude is not None and employee.home_longitude is not None:
                                continue
                            if to_float(row.get("home_latitude")) is not None and to_float(row.get("home_longitude")) is not None:
                                continue
                            result = local_coordinate_estimate(
                                row.get("home_city", ""),
                                row.get("home_state", ""),
                                row.get("home_zip", ""),
                            )
                            if not result:
                                not_estimated.append(
                                    {
                                        "Technician": row["technician_name"],
                                        "City": row.get("home_city", ""),
                                        "State": row.get("home_state", ""),
                                        "Zip": row.get("home_zip", ""),
                                    }
                                )
                                continue
                            employee.home_address = clean(row.get("home_address", "")) or employee.home_address
                            employee.home_city = clean(row.get("home_city", "")) or employee.home_city
                            employee.home_state = clean(row.get("home_state", "")) or employee.home_state
                            employee.home_zip = clean(row.get("home_zip", "")) or employee.home_zip
                            employee.home_latitude = float(result["latitude"])
                            employee.home_longitude = float(result["longitude"])
                            estimated += 1
                            estimated_details.append(
                                {
                                    "Technician": row["technician_name"],
                                    "Estimate Used": result.get("display_name", ""),
                                    "Latitude": result["latitude"],
                                    "Longitude": result["longitude"],
                                }
                            )
                    if estimated:
                        st.success(f"Saved city/ZIP coordinate estimates for {estimated} technician(s).")
                        st.dataframe(pd.DataFrame(estimated_details), use_container_width=True, hide_index=True)
                    if not_estimated:
                        st.warning("These technicians did not have enough nearby saved store data for a city/ZIP estimate.")
                        st.dataframe(pd.DataFrame(not_estimated), use_container_width=True, hide_index=True)
                    st.rerun()
                if save_col.button("Save Manual Home Coordinates", type="secondary"):
                    with session_scope() as session:
                        saved = 0
                        for _, row in edits.iterrows():
                            lat = to_float(row.get("home_latitude"))
                            lon = to_float(row.get("home_longitude"))
                            if lat is None or lon is None:
                                continue
                            employee = session.get(Employee, int(row["employee_id"]))
                            if employee:
                                employee.home_address = clean(row.get("home_address", "")) or employee.home_address
                                employee.home_city = clean(row.get("home_city", "")) or employee.home_city
                                employee.home_state = clean(row.get("home_state", "")) or employee.home_state
                                employee.home_zip = clean(row.get("home_zip", "")) or employee.home_zip
                                employee.home_latitude = lat
                                employee.home_longitude = lon
                                saved += 1
                    st.success(f"Saved home coordinates for {saved} technician(s).")
                    st.rerun()

    section_header("Build Step 5: Choose Schedule Settings", "Pick the cycle and monthly targets. PMT stores are selected by carryover/rotation priority first, then ordered by route distance.", "blue", focus_key="pmt_focus_step", focus_value=5)
    settings_disabled = assignments.empty
    s1, s2, s3 = st.columns(3)
    month_options = [add_months(month_start(date.today()), index) for index in range(18)]
    selected_start_month = s1.selectbox(
        "Schedule Start Month",
        month_options,
        format_func=month_label,
        disabled=settings_disabled,
    )
    start_month = selected_start_month
    months = s2.selectbox("Number of Months", [1, 2, 3, 4, 5, 6], index=5, disabled=settings_disabled)
    default_target = s3.number_input("Default Stores / Tech / Month", min_value=1, max_value=60, value=10, disabled=settings_disabled)
    direction = "Closest to home first"
    s5, s6, s7, s8 = st.columns(4)
    avoid_weekends = s5.checkbox("Avoid weekends", value=True, disabled=settings_disabled)
    avoid_holidays = s6.checkbox("Avoid company holidays", value=True, disabled=settings_disabled)
    avoid_pto = s7.checkbox("Avoid PTO / call-off days", value=True, disabled=settings_disabled)
    schedule_mode = s8.selectbox("Schedule Mode", ["Monthly store list only", "Monthly schedule with dates", "Monthly schedule with daily stops"], disabled=settings_disabled)
    publish_mode = st.session_state.get("pmt_publish_mode", "Create Draft Only")

    targets = {}
    if not assignments.empty:
        section_header("Monthly Targets", "Adjust monthly targets by technician before generating the draft.", "gray")
        target_df = (
            assignments.groupby(["employee_id", "technician_name"], as_index=False)
            .agg(assigned_stores=("store_id", "nunique"), current_target=("monthly_target", "max"))
        )
        target_df["monthly_target"] = target_df["current_target"].fillna(default_target).astype(int)
        target_df["total_cycle_capacity"] = target_df["monthly_target"] * int(months)
        target_df["leftover_stores"] = (target_df["assigned_stores"] - target_df["total_cycle_capacity"]).clip(lower=0)
        edited_targets = st.data_editor(
            target_df[["employee_id", "technician_name", "assigned_stores", "monthly_target", "total_cycle_capacity", "leftover_stores"]],
            use_container_width=True,
            hide_index=True,
            disabled=["employee_id", "technician_name", "assigned_stores", "total_cycle_capacity", "leftover_stores"],
            key="pmt_monthly_targets",
        )
        edited_targets["total_cycle_capacity"] = edited_targets["monthly_target"].astype(int) * int(months)
        edited_targets["leftover_stores"] = (edited_targets["assigned_stores"].astype(int) - edited_targets["total_cycle_capacity"]).clip(lower=0)
        over_capacity = edited_targets[edited_targets["leftover_stores"] > 0].copy()
        if not over_capacity.empty:
            total_leftover = int(over_capacity["leftover_stores"].sum())
            st.warning(f"{total_leftover} assigned PMT store(s) do not fit in the selected month range. They will be saved as PMT Stores Not Scheduled after publishing and prioritized in the next cycle.")
        targets = {int(row["employee_id"]): int(row["monthly_target"]) for _, row in edited_targets.iterrows()}
        if st.button("Save Monthly Targets to Employees", type="secondary"):
            with session_scope() as session:
                for employee_id, target in targets.items():
                    employee = session.get(Employee, int(employee_id))
                    if employee:
                        employee.monthly_pmt_store_target = int(target)
            st.success("Monthly PMT targets saved to employee profiles.")

    section_header("Build Step 6: Generate Draft PMT Schedule", "The app builds each technician's monthly store list using carryover, not-scheduled stores, oldest completion history, then route distance.", "green", focus_key="pmt_focus_step", focus_value=6)
    can_generate = not assignments.empty and problems[problems["Severity"] == "Must Fix"].empty if not problems.empty else not assignments.empty
    if st.button("Generate Draft PMT Schedule", disabled=not can_generate, type="primary"):
        draft, unscheduled = build_pmt_draft(assignments, start_month, int(months), targets, direction, avoid_weekends, avoid_holidays, avoid_pto, schedule_mode)
        st.session_state["pmt_schedule_draft"] = draft.to_dict("records")
        st.session_state["pmt_schedule_unscheduled"] = unscheduled.to_dict("records")
        st.session_state["pmt_schedule_draft_settings"] = {
            "start_month": start_month.isoformat(),
            "months": int(months),
            "default_target": int(default_target),
            "direction": direction,
            "schedule_mode": schedule_mode,
        }
        st.success(f"Draft generated with {len(draft)} scheduled stores and {len(unscheduled)} unscheduled stores.")
        st.rerun()
    if not can_generate and not assignments.empty:
        st.error("Fix must-fix data problems before generating the draft.")

    draft_df = pd.DataFrame(st.session_state.get("pmt_schedule_draft", []))
    unscheduled_df = pd.DataFrame(st.session_state.get("pmt_schedule_unscheduled", []))

    section_header("Build Step 7: Review Draft Routes", "Review the draft by technician and month. Route options and maps are here before publishing.", "green", focus_key="pmt_focus_step", focus_value=7)
    if draft_df.empty:
        st.info("Generate a draft schedule first.")
    else:
        d1, d2, d3 = st.columns(3)
        d1.metric("Draft Stores", len(draft_df))
        d2.metric("Technicians", draft_df["employee_id"].nunique())
        with d3:
            metric_help_card("Unscheduled Stores", len(unscheduled_df), "Assigned PMT stores that did not fit into the draft based on the selected months and monthly target/capacity.")
        working_draft = draft_df.copy()
        for column in ["miles_from_previous_stop", "estimated_drive_time"]:
            if column not in working_draft.columns:
                working_draft[column] = ""
        working_draft["_month_sort"] = pd.to_datetime(working_draft["month_start"], errors="coerce")
        ordered_months = (
            working_draft[["month", "_month_sort"]]
            .drop_duplicates()
            .sort_values("_month_sort")["month"]
            .tolist()
        )
        review_mode = st.radio("Draft review section", ["Month Summary", "Route Options", "Full Draft / Edit"], horizontal=True, key="pmt_draft_review_section")
        if review_mode == "Month Summary":
            selected_month = st.selectbox("Month", ordered_months, key="pmt_month_summary_select")
            month_df = working_draft[working_draft["month"] == selected_month].sort_values(["technician", "sequence_number", "store_number"])
            st.subheader(f"{selected_month} PMT Stores")
            m1, m2 = st.columns(2)
            m1.metric("Stores This Month", len(month_df))
            m2.metric("PMTs This Month", month_df["technician"].nunique())
            month_summary = (
                month_df.groupby("technician")
                .agg(Store_Count=("store_number", "count"), Stores=("store_number", lambda values: ", ".join(values.astype(str))))
                .reset_index()
                .rename(columns={"technician": "PMT"})
            )
            render_plain_table(month_summary)
            month_details = month_df[
                ["technician", "sequence_number", "store_number", "address", "city", "state", "distance_from_home", "miles_from_previous_stop"]
            ].rename(
                columns={
                    "technician": "PMT",
                    "sequence_number": "Recommended Stop",
                    "store_number": "Store",
                    "address": "Address",
                    "city": "City",
                    "state": "State",
                    "distance_from_home": "Miles From Home",
                    "miles_from_previous_stop": "Miles From Previous Stop",
                }
            )
            render_plain_table(month_details)

        elif review_mode == "Route Options":
            route_df = working_draft.sort_values(["technician", "_month_sort", "sequence_number", "store_number"]).copy()
            st.subheader("Recommended Route Options")
            route_filters = st.columns(2)
            route_tech = route_filters[0].selectbox("Technician", sorted(route_df["technician"].dropna().unique().tolist()), key="pmt_route_tech_filter")
            route_month = route_filters[1].selectbox("Month", ordered_months, key="pmt_route_month_filter")
            selected_month_stores = route_df[(route_df["technician"] == route_tech) & (route_df["month"] == route_month)].copy()
            st.metric("Stores Scheduled", len(selected_month_stores))
            route_type = st.radio("Route option", [HOME_ROUTE, NEXT_ROUTE], horizontal=True, key="pmt_route_option_select")
            st.caption(route_notes(route_type))
            selected_routes = route_table_view(route_options_for_draft(selected_month_stores, route_type))
            render_plain_table(selected_routes)
            map_preview = route_options_for_draft(selected_month_stores, route_type)
            if not map_preview.empty and {"latitude", "longitude"}.issubset(map_preview.columns):
                map_preview = map_preview.rename(columns={"route_order": "sequence_number"}).copy()
                map_preview["team_name"] = map_preview["technician"]
                map_preview["notes"] = map_preview.apply(
                    lambda row: f"PMT: {row['technician']}<br>Month: {row['month']}<br>Route: {row['route_type']}<br>Stop: {row['sequence_number']}",
                    axis=1,
                )
                try:
                    draft_map, _ = render_store_map(
                        map_preview,
                        color_by="technician",
                        show_homes=False,
                        height=520,
                        key=f"pmt_route_map_{route_type}",
                        cluster=False,
                        show_route_path=True,
                        static_preview=True,
                    )
                    if draft_map:
                        st.download_button(
                            f"Download {route_type} Map HTML",
                            data=map_html(draft_map),
                            file_name=f"pmt_{key(route_type)}_map.html",
                            mime="text/html",
                            key=f"pmt_{key(route_type)}_map_download",
                        )
                except Exception as exc:
                    st.warning("Interactive map could not load. Static backup preview is shown below. Please check the app logs for details.")
                    with st.expander("Map render error. Open debug details.", expanded=False):
                        st.code(str(exc))
                    route_csv = render_route_preview(map_preview, height=520)
                    if route_csv:
                        st.download_button(
                            f"Download {route_type} Route CSV",
                            data=route_csv.encode("utf-8"),
                            file_name=f"pmt_{key(route_type)}_route.csv",
                            mime="text/csv",
                            key=f"pmt_{key(route_type)}_route_download",
                        )
            all_route_options = route_table_view(route_options_for_draft(route_df, "Both Route Options"))
            st.download_button("Export Both Route Options Excel", data=excel_bytes(all_route_options), file_name="pmt_route_options.xlsx")

        else:
            editor_columns = [
                col
                for col in [
                    "technician",
                    "month",
                    "schedule_date",
                    "sequence_number",
                    "store_number",
                    "address",
                    "city",
                    "state",
                    "distance_from_home",
                    "miles_from_previous_stop",
                    "estimated_drive_time",
                    "status",
                    "notes",
                    "employee_id",
                    "store_id",
                    "month_start",
                    "zip",
                    "home_latitude",
                    "home_longitude",
                    "latitude",
                    "longitude",
                    "work_type",
                ]
                if col in draft_df.columns
            ]
            render_plain_table(draft_df[editor_columns])
            enable_full_editor = st.checkbox("Edit full draft table", value=False, key="pmt_enable_full_draft_editor")
            edited_draft = draft_df[editor_columns].copy()
            if enable_full_editor:
                edited_draft = st.data_editor(
                    draft_df[editor_columns],
                    use_container_width=True,
                    hide_index=True,
                    disabled=[
                        col
                        for col in [
                            "technician",
                            "store_number",
                            "address",
                            "city",
                            "state",
                            "distance_from_home",
                            "miles_from_previous_stop",
                            "employee_id",
                            "store_id",
                            "month_start",
                            "zip",
                            "home_latitude",
                            "home_longitude",
                            "latitude",
                            "longitude",
                            "work_type",
                        ]
                        if col in editor_columns
                    ],
                    key="pmt_draft_editor",
                )
                if st.button("Apply Draft Edits", type="secondary", key="pmt_apply_draft_edits"):
                    st.session_state["pmt_schedule_draft"] = edited_draft.to_dict("records")
                    st.success("PMT draft edits saved.")
                    st.rerun()
            if not unscheduled_df.empty:
                st.warning("Some stores could not be scheduled.")
                st.dataframe(unscheduled_df, use_container_width=True, hide_index=True)
            e1, e2, e3 = st.columns(3)
            e1.download_button("Export Full Draft Excel", data=excel_bytes(edited_draft), file_name="pmt_monthly_draft.xlsx")
            if not unscheduled_df.empty:
                e2.download_button("Export Unscheduled Stores", data=excel_bytes(unscheduled_df), file_name="pmt_unscheduled_stores.xlsx")
            if e3.button("Clear Draft", type="secondary"):
                st.session_state.pop("pmt_schedule_draft", None)
                st.session_state.pop("pmt_schedule_unscheduled", None)
                st.rerun()

    section_header("Build Step 8: Publish Schedule", "Publishing adds PMT schedule records to the active schedule system.", "yellow", focus_key="pmt_focus_step", focus_value=8)
    if draft_df.empty:
        st.info("No draft is ready to publish.")
    else:
        draft_settings = st.session_state.get("pmt_schedule_draft_settings", {})
        publish_start_month = pd.to_datetime(draft_settings.get("start_month", start_month.isoformat())).date()
        publish_months = int(draft_settings.get("months", months))
        publish_default_target = int(draft_settings.get("default_target", default_target))
        publish_direction = draft_settings.get("direction", direction)
        publish_schedule_mode = draft_settings.get("schedule_mode", schedule_mode)
        run_name = st.text_input(
            "Schedule Run Name",
            value=f"PMT Monthly Schedule {month_label(publish_start_month)} - {month_label(add_months(publish_start_month, publish_months - 1))}",
        )
        publish_mode = st.radio(
            "Publish option",
            ["Create Draft Only", "Add only new schedule items", "Replace existing PMT schedule for selected months"],
            horizontal=True,
            index=["Create Draft Only", "Add only new schedule items", "Replace existing PMT schedule for selected months"].index(st.session_state.get("pmt_publish_mode", "Create Draft Only")),
            key="pmt_publish_mode",
        )
        publish_route_order = st.radio(
            "Publish Route Order",
            ["Use Home-Based Route", "Use Next-Closest Store Route"],
            horizontal=True,
            index=0,
            key="pmt_publish_route_order",
        )
        if publish_mode == "Create Draft Only":
            st.info("Safe mode is selected. Change this to Add only new schedule items when you are ready to publish.")
        edited = pd.DataFrame(st.session_state.get("pmt_schedule_draft", []))
        selected_route_type = HOME_ROUTE if publish_route_order == "Use Home-Based Route" else NEXT_ROUTE
        edited = draft_with_route_order(edited, selected_route_type)
        publish_conflicts = pmt_publish_conflicts(edited, publish_start_month, publish_months)
        draft_keys = edited[["employee_id", "store_id", "month_start"]].drop_duplicates() if {"employee_id", "store_id", "month_start"}.issubset(edited.columns) else pd.DataFrame()
        all_items_already_exist = publish_mode == "Add only new schedule items" and not publish_conflicts.empty and len(publish_conflicts[["employee_id", "store_id", "month_start"]].drop_duplicates()) >= len(draft_keys)
        if publish_mode == "Add only new schedule items" and not publish_conflicts.empty:
            st.warning("Some PMT draft items already exist for the same technician, store, and month. Existing items will not be duplicated.")
            st.dataframe(
                publish_conflicts[["schedule_id", "schedule_name", "technician", "store_number", "city", "schedule_date", "status"]].head(100),
                use_container_width=True,
                hide_index=True,
            )
        if all_items_already_exist:
            st.error("This PMT draft is already fully scheduled for the selected months. Delete/replace the existing run or change the schedule range before publishing.")
        confirm_publish = st.checkbox("I have reviewed this schedule and confirm I am ready to publish it.", key="pmt_confirm_publish")
        if st.button("Publish PMT Schedule", disabled=not confirm_publish or publish_mode == "Create Draft Only" or all_items_already_exist, type="primary"):
            result = publish_draft(
                edited,
                unscheduled_df,
                run_name,
                publish_start_month,
                publish_months,
                publish_default_target,
                selected_route_type,
                publish_schedule_mode,
                replace_existing=publish_mode == "Replace existing PMT schedule for selected months",
            )
            st.success(f"Published PMT schedule run #{result['run_id']}. Created {result['created']} schedule items. Skipped {result['skipped']} duplicates.")
            st.rerun()


with tab_carryover:
    section_header("Completion Step 1: PMT Carryover / Not Scheduled Stores", "Review PMT stores that did not fit, were not completed, or need priority in the next monthly cycle.", "orange")
    st.caption(
        "PMT status guide: Scheduled = on the current PMT route. Not Scheduled = assigned to a PMT but did not fit in the selected cycle. "
        "Not Completed = scheduled but missed. Carryover = saved for priority in the next PMT cycle."
    )
    gap_runs = safe_query(
        """
        select r.id, r.run_name, r.created_at, r.cycle_start, r.cycle_end, r.months, r.technician_count,
               r.store_count, r.unscheduled_count, r.status
        from pmt_schedule_runs r
        order by r.created_at desc, r.id desc
        """
    )
    if not gap_runs.empty:
        st.markdown("**Mark PMT Stores Completed**")
        st.caption("Use this when a PMT finishes stores during the month. Check the stores that are complete and save.")
        complete_run = st.selectbox(
            "Published PMT run",
            gap_runs["id"].tolist(),
            format_func=lambda x: f"#{x} - {gap_runs.set_index('id').loc[x, 'run_name']}",
            key="pmt_complete_run",
        )
        completion_items = pmt_manage_run_items(complete_run)
        if completion_items.empty:
            st.info("This PMT run does not have stores to complete.")
        else:
            complete_cols = st.columns(3)
            complete_techs = (
                completion_items[["employee_id", "technician"]]
                .dropna(subset=["employee_id"])
                .drop_duplicates()
                .sort_values("technician")
            )
            complete_employee = complete_cols[0].selectbox(
                "PMT",
                complete_techs["employee_id"].astype(int).tolist(),
                format_func=lambda value: complete_techs.set_index("employee_id").loc[value, "technician"],
                key=f"pmt_complete_employee_{complete_run}",
            )
            complete_months = (
                completion_items.loc[completion_items["employee_id"].astype("Int64") == int(complete_employee), ["month_start", "month"]]
                .drop_duplicates()
                .sort_values("month_start")
            )
            if complete_months.empty:
                st.info("This PMT has no scheduled months in the selected run.")
            else:
                complete_month = complete_cols[1].selectbox(
                    "Month",
                    complete_months["month_start"].tolist(),
                    format_func=lambda value: complete_months.set_index("month_start").loc[value, "month"],
                    key=f"pmt_complete_month_{complete_run}_{complete_employee}",
                )
                completed_on = complete_cols[2].date_input("Completed date", value=date.today(), key=f"pmt_completed_on_{complete_run}_{complete_employee}_{complete_month}")
                month_items = completion_items[
                    (completion_items["employee_id"].astype("Int64") == int(complete_employee))
                    & (completion_items["month_start"] == complete_month)
                ].copy()
                month_items["Completed"] = month_items["status"].eq("Completed")
                completion_view = month_items[["Completed", "schedule_item_id", "sequence_number", "store_number", "city", "status", "notes"]].rename(
                    columns={
                        "sequence_number": "Stop",
                        "store_number": "Store",
                        "city": "City",
                        "status": "Current Status",
                        "notes": "Notes",
                    }
                )
                edited_completion = st.data_editor(
                    completion_view,
                    use_container_width=True,
                    hide_index=True,
                    disabled=["schedule_item_id", "Stop", "Store", "City", "Current Status", "Notes"],
                    column_config={
                        "Completed": st.column_config.CheckboxColumn("Completed"),
                        "schedule_item_id": None,
                    },
                    key=f"pmt_completion_editor_{complete_run}_{complete_employee}_{complete_month}",
                )
                selected_completed = edited_completion.loc[edited_completion["Completed"].astype(bool), "schedule_item_id"].dropna().astype(int).tolist()
                completion_note = st.text_input("Completion note optional", key=f"pmt_completion_note_{complete_run}_{complete_employee}_{complete_month}")
                if st.button("Save Completed PMT Stores", type="primary", disabled=not selected_completed, key=f"save_pmt_completed_{complete_run}_{complete_employee}_{complete_month}"):
                    updated = 0
                    with session_scope() as session:
                        for item_id in selected_completed:
                            item = session.get(ScheduleItem, int(item_id))
                            if not item:
                                continue
                            item.status = "Completed"
                            note_parts = [clean(item.completion_notes), f"Completed on {completed_on}"]
                            if completion_note:
                                note_parts.append(completion_note)
                            item.completion_notes = " | ".join([part for part in note_parts if part])
                            updated += 1
                    log_action("pmt stores marked completed", "schedule_items", description=f"{updated} PMT schedule item(s) marked completed")
                    st.success(f"Marked {updated} PMT store(s) completed.")
                    st.rerun()

    if not gap_runs.empty:
        latest_run = gap_runs.iloc[0]
        latest_run_id = int(latest_run["id"])
        latest_run_start = scalar_date(latest_run.get("cycle_start"))
        latest_run_months = scalar_int(latest_run.get("months"), 1)
        latest_run_end = scalar_date(latest_run.get("cycle_end"))
        latest_run_missing = pmt_stores_not_in_run(latest_run_id)
        latest_run_summary = pmt_rotation_gap_summary(latest_run_start, latest_run_months) if latest_run_start else pd.DataFrame()
        st.markdown("**Latest Published PMT Run: Stores Not Scheduled**")
        st.caption(f"Run #{latest_run_id}: {latest_run.get('run_name', '')} | Period: {latest_run_start} to {latest_run_end} | Missing assigned stores: {len(latest_run_missing)}")
        if latest_run_summary.empty and latest_run_missing.empty:
            st.success("The latest published PMT run has no assigned-store scheduling gaps.")
        else:
            lr1, lr2 = st.columns(2)
            with lr1:
                metric_help_card("Assigned Stores Not Scheduled In Latest Run", len(latest_run_missing), "Assigned PMT stores that do not appear anywhere in the latest published run period. These should be prioritized next cycle.")
            lr2.metric("Affected PMTs", latest_run_missing["employee_id"].nunique() if not latest_run_missing.empty else 0)
            if not latest_run_summary.empty:
                render_plain_table(latest_run_summary[["technician", "assigned_stores", "unique_stores_scheduled", "period_capacity", "assigned_stores_not_scheduled", "scheduled_not_completed"]])
            if not latest_run_missing.empty:
                st.caption("The summary above is the main view. Open the detail below only when you need the actual store list.")
                with st.expander("View missing store details for latest run", expanded=False):
                    missing_techs = sorted(latest_run_missing["technician"].dropna().unique().tolist())
                    selected_missing_tech = st.selectbox("PMT", ["All PMTs"] + missing_techs, key="latest_run_missing_tech")
                    latest_missing_detail = latest_run_missing if selected_missing_tech == "All PMTs" else latest_run_missing[latest_run_missing["technician"] == selected_missing_tech]
                    display_missing = latest_missing_detail[["technician", "store_number", "city", "state", "reason"]].rename(
                        columns={
                            "technician": "PMT",
                            "store_number": "Store",
                            "city": "City",
                            "state": "State",
                            "reason": "Why It Is On This List",
                        }
                    )
                    render_plain_table(display_missing)
                st.download_button("Export Latest Run Not Scheduled Stores", data=excel_bytes(latest_run_missing), file_name=f"pmt_latest_run_not_scheduled_{latest_run_id}.xlsx")

    st.markdown("**Completion Step 2: Check A PMT Period For Gaps**")
    gap_period_cols = st.columns(2)
    run_month_options = []
    if not gap_runs.empty and "cycle_start" in gap_runs.columns:
        run_month_options = [scalar_date(value) for value in gap_runs["cycle_start"].dropna().tolist()]
    run_month_options = [value for value in run_month_options if value is not None]
    calendar_month_options = [add_months(month_start(date.today()), index) for index in range(-12, 25)]
    gap_month_options = sorted(set(run_month_options + calendar_month_options))
    latest_gap_start = scalar_date(gap_runs.iloc[0]["cycle_start"]) if not gap_runs.empty else start_month
    latest_gap_months = scalar_int(gap_runs.iloc[0]["months"], int(months)) if not gap_runs.empty else int(months)
    gap_start_index = gap_month_options.index(latest_gap_start) if latest_gap_start in gap_month_options else (gap_month_options.index(start_month) if start_month in gap_month_options else 0)
    gap_start_month = gap_period_cols[0].selectbox(
        "Schedule period starts",
        gap_month_options,
        index=gap_start_index,
        format_func=month_label,
        key="pmt_rotation_gap_start_month",
    )
    gap_month_count = gap_period_cols[1].selectbox(
        "Months to check",
        [1, 2, 3, 4, 5, 6, 9, 12],
        index=[1, 2, 3, 4, 5, 6, 9, 12].index(int(latest_gap_months)) if int(latest_gap_months) in [1, 2, 3, 4, 5, 6, 9, 12] else 5,
        key="pmt_rotation_gap_month_count",
    )
    period_gaps = pmt_rotation_gaps_for_period(gap_start_month, int(gap_month_count))
    period_gap_summary = pmt_rotation_gap_summary(gap_start_month, int(gap_month_count))
    st.markdown("**PMT Rotation Summary By Technician**")
    if period_gap_summary.empty:
        st.info("No PMT rotation gaps found for the selected period.")
    else:
        render_plain_table(period_gap_summary[["technician", "assigned_stores", "unique_stores_scheduled", "period_capacity", "assigned_stores_not_scheduled", "scheduled_not_completed"]])
    if period_gaps.empty:
        st.success("All assigned PMT stores are either scheduled in this period or have no not-completed exceptions in this period.")
    else:
        pg1, pg2, pg3 = st.columns(3)
        with pg1:
            metric_help_card("Total PMT Rotation Gaps", len(period_gaps), "Combined count of PMT stores that either did not fit into this period or were scheduled but not completed.")
        with pg2:
            metric_help_card("Assigned Stores Not Scheduled", int((period_gaps["source"] == "Missing From Selected Period").sum()), "Assigned PMT stores that did not appear anywhere in the selected schedule period.")
        with pg3:
            metric_help_card("Scheduled But Not Completed", int((period_gaps["source"] == "Scheduled But Not Completed").sum()), "PMT stores that were scheduled in the selected period but have an exception/not-completed status.")
        st.warning("These stores should feed PMT carryover. Missing stores did not appear anywhere in the selected period. Not-completed stores were scheduled but have an exception status.")
        with st.expander("View store-level rotation gap details", expanded=False):
            gap_techs = sorted(period_gaps["technician"].dropna().unique().tolist())
            selected_gap_tech = st.selectbox("PMT", ["All PMTs"] + gap_techs, key="period_gap_detail_tech")
            gap_detail = period_gaps if selected_gap_tech == "All PMTs" else period_gaps[period_gaps["technician"] == selected_gap_tech]
            display_gaps = gap_detail[["technician", "store_number", "city", "state", "source", "reason"]].rename(
                columns={
                    "technician": "PMT",
                    "store_number": "Store",
                    "city": "City",
                    "state": "State",
                    "source": "Gap Type",
                    "reason": "Why It Is On This List",
                }
            )
            render_plain_table(display_gaps)
        save_period_gaps, export_period_gaps = st.columns(2)
        if save_period_gaps.button("Save Period Gaps to PMT Carryover", type="secondary", key="save_pmt_period_gaps"):
            summary = save_pmt_gap_rows(period_gaps, f"Detected from {month_label(gap_start_month)} through {month_label(add_months(gap_start_month, int(gap_month_count) - 1))}")
            log_action("pmt period gaps saved to carryover", "pmt_schedule_backlog", description=str(summary))
            st.success(f"Saved {summary['created']} new and updated {summary['updated']} existing PMT carryover record(s).")
            st.rerun()
        export_period_gaps.download_button("Export PMT Rotation Gaps", data=excel_bytes(period_gaps), file_name="pmt_rotation_gaps.xlsx")

    if not gap_runs.empty:
        st.markdown("**Stores Missing From Published PMT Run**")
        selected_gap_run = st.selectbox(
            "Published PMT run to check",
            gap_runs["id"].tolist(),
            format_func=lambda x: f"#{x} - {gap_runs.set_index('id').loc[x, 'run_name']}",
            key="pmt_gap_run_select",
        )
        missing_from_run = pmt_stores_not_in_run(selected_gap_run)
        if missing_from_run.empty:
            st.success("All assigned PMT stores are included in the selected published run.")
        else:
            m1, m2 = st.columns(2)
            m1.metric("Assigned Stores Not In This Run", len(missing_from_run))
            m2.metric("Affected PMTs", missing_from_run["employee_id"].nunique())
            st.warning("These stores are assigned to PMTs but are not in the selected published schedule run. Save them to carryover so the next PMT draft prioritizes them.")
            render_plain_table(missing_from_run[["technician", "store_number", "city", "state", "status", "reason"]])
            save_gap, export_gap = st.columns(2)
            if save_gap.button("Save Missing Run Stores to PMT Carryover", type="secondary", key=f"save_pmt_gap_{selected_gap_run}"):
                created = 0
                updated = 0
                with session_scope() as session:
                    for _, row in missing_from_run.iterrows():
                        existing = session.query(PMTScheduleBacklog).filter(
                            PMTScheduleBacklog.pmt_schedule_run_id == int(row["run_id"]),
                            PMTScheduleBacklog.employee_id == int(row["employee_id"]),
                            PMTScheduleBacklog.store_id == int(row["store_id"]),
                            PMTScheduleBacklog.status == "Not Scheduled",
                        ).first()
                        record = existing or PMTScheduleBacklog(
                            pmt_schedule_run_id=int(row["run_id"]),
                            employee_id=int(row["employee_id"]),
                            store_id=int(row["store_id"]),
                            cycle_start=scalar_date(row.get("cycle_start")) or month_start(date.today()),
                            cycle_end=scalar_date(row.get("cycle_end")),
                        )
                        if not existing:
                            session.add(record)
                            created += 1
                        else:
                            updated += 1
                        record.status = "Not Scheduled"
                        record.reason = clean(row.get("reason", "")) or "Assigned PMT store did not fit into this published run"
                        record.cycles_missed = max(int(record.cycles_missed or 0), 1)
                        record.priority_score = max(int(record.priority_score or 0), 1000)
                        record.notes = f"Detected from published PMT run {int(row['run_id'])}."
                log_action("pmt run gaps saved to carryover", "pmt_schedule_backlog", int(selected_gap_run), f"{created} created, {updated} updated")
                st.success(f"Saved {created} new and updated {updated} existing PMT carryover record(s).")
                st.rerun()
            export_gap.download_button("Export Stores Missing From Run", data=excel_bytes(missing_from_run), file_name=f"pmt_stores_missing_from_run_{selected_gap_run}.xlsx")

    pmt_carryover = pmt_carryover_report()
    if not period_gaps.empty:
        live_gap_rows = period_gaps.copy()
        live_gap_rows["backlog_id"] = None
        if "schedule_item_id" not in live_gap_rows.columns:
            live_gap_rows["schedule_item_id"] = None
        live_gap_rows["source"] = live_gap_rows["source"].replace({"Missing From Selected Period": "Live Period Gap"})
        live_gap_rows["cycles_missed"] = 1
        live_gap_rows["priority_score"] = live_gap_rows["status"].apply(lambda value: 1000 if value == "Not Scheduled" else 900)
        live_gap_rows["last_scheduled_month"] = None
        live_gap_rows["last_completed_date"] = None
        live_gap_rows["notes"] = "Live period gap. Save period gaps to make this permanent carryover."
        live_gap_rows = live_gap_rows[[col for col in pmt_carryover.columns if col in live_gap_rows.columns]] if not pmt_carryover.empty else live_gap_rows
        pmt_carryover = pd.concat([pmt_carryover, live_gap_rows], ignore_index=True)
        if {"technician", "store_number", "status", "source"}.issubset(pmt_carryover.columns):
            pmt_carryover = pmt_carryover.drop_duplicates(["technician", "store_number", "status", "source"], keep="first")
    if pmt_carryover.empty:
        st.success("No PMT carryover or not-scheduled stores are currently open.")
    else:
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            metric_help_card("Open PMT Backlog", len(pmt_carryover), "All open PMT backlog rows: not scheduled, carryover, not completed, skipped, or overdue.")
        with c2:
            metric_help_card("Not Scheduled", int((pmt_carryover["status"] == "Not Scheduled").sum()), "Assigned PMT stores that did not fit into a schedule period and need to be picked up next cycle.")
        with c3:
            metric_help_card("Carryover / Not Completed", int(pmt_carryover["status"].isin(["Carryover", "Not Completed", "Needs Rescheduled", "Rescheduled", "Rain Delay", "Skipped"]).sum()), "PMT stores missed, pushed, skipped, or marked not completed. These are prioritized before normal route distance.")
        with c4:
            metric_help_card("Overdue", int((pmt_carryover["status"] == "Overdue").sum()), "PMT backlog stores marked overdue after being missed too long.")
        st.caption("These stores are prioritized before normal distance routing the next time you generate a PMT draft.")
        editable = pmt_carryover.copy()
        visible_columns = [
            "technician",
            "store_number",
            "city",
            "source",
            "status",
            "reason",
            "cycles_missed",
            "priority_score",
            "last_scheduled_month",
            "last_completed_date",
            "notes",
            "backlog_id",
            "schedule_item_id",
        ]
        carryover_columns = [col for col in visible_columns if col in editable.columns]
        carryover_disabled = [
            col for col in ["technician", "store_number", "city", "source", "cycles_missed", "priority_score", "last_scheduled_month", "last_completed_date", "backlog_id", "schedule_item_id"]
            if col in carryover_columns
        ]
        edited_carryover = st.data_editor(
            editable[carryover_columns],
            use_container_width=True,
            hide_index=True,
            disabled=carryover_disabled,
            column_config={
                "status": st.column_config.SelectboxColumn("Status", options=["Not Scheduled", "Not Completed", "Carryover", "Overdue", "Skipped", "Completed", "Cancelled", "Scheduled"]),
                "backlog_id": None,
                "schedule_item_id": None,
            },
            key="pmt_carryover_editor",
        )
        save_carryover, export_carryover = st.columns(2)
        if save_carryover.button("Save PMT Carryover Status Updates", type="secondary"):
            updated = 0
            with session_scope() as session:
                for _, row in edited_carryover.iterrows():
                    backlog_id = scalar_int(row.get("backlog_id"), 0)
                    schedule_item_id = scalar_int(row.get("schedule_item_id"), 0)
                    if backlog_id:
                        backlog = session.get(PMTScheduleBacklog, backlog_id)
                        if backlog:
                            backlog.status = clean(row.get("status", "")) or backlog.status
                            backlog.reason = clean(row.get("reason", "")) or backlog.reason
                            backlog.notes = clean(row.get("notes", "")) or backlog.notes
                            updated += 1
                    elif schedule_item_id:
                        item = session.get(ScheduleItem, schedule_item_id)
                        if item:
                            item.status = clean(row.get("status", "")) or item.status
                            item.completion_notes = clean(row.get("reason", "")) or item.completion_notes
                            updated += 1
            log_action("pmt carryover statuses updated", "pmt_schedule_backlog", description=f"{updated} PMT carryover/backlog records updated")
            st.success(f"Updated {updated} PMT carryover/backlog record(s).")
            st.rerun()
        export_carryover.download_button("Export PMT Carryover / Not Scheduled", data=excel_bytes(pmt_carryover), file_name="pmt_carryover_not_scheduled.xlsx")


with tab_health:
    section_header(
        "Overview & Schedule Health",
        "Start here when a PMT schedule looks wrong. This view tells you which repair tab to use.",
        "orange",
    )
    st.info(
        "Most schedule problems fall into two buckets: territory changes that have not been reconciled yet, or one PMT's schedule needing rebuilt from current Areas and Maps assignments."
    )
    repair_runs = safe_query(
        """
        select r.id, r.run_name, r.created_at, r.cycle_start, r.cycle_end, r.months, r.technician_count,
               r.store_count, r.unscheduled_count, r.status
        from pmt_schedule_runs r
        where coalesce(lower(trim(r.status)), '') not in ('deleted','snapshot')
        order by r.created_at desc, r.id desc
        """,
        use_cache=False,
    )
    if repair_runs.empty:
        st.warning("No published or imported PMT schedule runs are available to repair.")
    else:
        problem_options = [
            "A PMT has too few stores or uneven months",
            "Territories changed and schedules do not match assignments",
            "A store is on the wrong PMT or duplicated",
            "I just need the corrected Excel/PDF export",
        ]
        problem_choice = st.radio("What looks wrong?", problem_options, key="pmt_fix_problem_choice")
        repair_cols = st.columns([0.5, 0.5])
        repair_run_id = repair_cols[0].selectbox(
            "Schedule run to fix",
            repair_runs["id"].tolist(),
            format_func=lambda value: f"#{value} - {repair_runs.set_index('id').loc[value, 'run_name']}",
            key="pmt_fix_run",
        )
        repair_run_row = repair_runs.set_index("id").loc[repair_run_id]
        repair_run_items = pmt_manage_run_items(repair_run_id)
        repair_cycle_start = scalar_date(repair_run_row.get("cycle_start"))
        repair_cycle_end = scalar_date(repair_run_row.get("cycle_end"))
        repair_run_items, repair_outside_items = split_run_items_by_period(repair_run_items, repair_cycle_start, repair_cycle_end)
        repair_tech_options = (
            repair_run_items[["employee_id", "technician"]].dropna(subset=["employee_id"]).drop_duplicates()
            if not repair_run_items.empty and {"employee_id", "technician"}.issubset(repair_run_items.columns)
            else pd.DataFrame(columns=["employee_id", "technician"])
        )
        all_repair_pmts = active_pmt_employee_summary()[["employee_id", "technician_name"]].rename(columns={"technician_name": "technician"})
        repair_tech_options = pd.concat([repair_tech_options, all_repair_pmts], ignore_index=True).dropna(subset=["employee_id"]).drop_duplicates("employee_id").sort_values("technician")
        if repair_tech_options.empty:
            st.warning("No active PMT technicians are available.")
        else:
            repair_employee_id = repair_cols[1].selectbox(
                "PMT to check/fix",
                repair_tech_options["employee_id"].astype(int).tolist(),
                format_func=lambda value: repair_tech_options.set_index("employee_id").loc[value, "technician"],
                key="pmt_fix_employee",
            )
            repair_tech_name = repair_tech_options.set_index("employee_id").loc[repair_employee_id, "technician"]
            repair_rec = technician_schedule_reconciliation(repair_run_items, repair_employee_id, "All months")
            assigned_count = int(repair_rec.get("assigned_count", 0))
            active_count = int(repair_rec.get("active_count", 0))
            completed_count = int(repair_rec.get("completed_count", 0))
            missing_count = int(repair_rec.get("assigned_not_scheduled_count", 0))
            wrong_count = int(repair_rec.get("scheduled_no_longer_assigned_count", 0))
            elsewhere_count = int(repair_rec.get("assigned_scheduled_elsewhere_count", 0))
            metric_cols = st.columns(6)
            metric_cols[0].metric("Assigned Now", assigned_count)
            metric_cols[1].metric("Active Scheduled", active_count)
            metric_cols[2].metric("Completed", completed_count)
            metric_cols[3].metric("Missing From Schedule", missing_count)
            metric_cols[4].metric("Scheduled Not Assigned", wrong_count)
            metric_cols[5].metric("Assigned Elsewhere", elsewhere_count)

            team_health_rows = []
            for _, tech_row in repair_tech_options.iterrows():
                tech_employee_id = scalar_int(tech_row.get("employee_id"), 0)
                if not tech_employee_id:
                    continue
                tech_rec = technician_schedule_reconciliation(repair_run_items, tech_employee_id, "All months")
                tech_assigned = int(tech_rec.get("assigned_count", 0))
                tech_active = int(tech_rec.get("active_count", 0))
                tech_missing = int(tech_rec.get("assigned_not_scheduled_count", 0))
                tech_wrong = int(tech_rec.get("scheduled_no_longer_assigned_count", 0))
                tech_elsewhere = int(tech_rec.get("assigned_scheduled_elsewhere_count", 0))
                if tech_wrong or tech_elsewhere:
                    recommendation = "Territory Reconciliation Required"
                elif tech_missing or tech_active < tech_assigned:
                    recommendation = "Rebuild / Balance Required"
                else:
                    recommendation = "Protected - No Changes Needed"
                team_health_rows.append(
                    {
                        "Technician": tech_row.get("technician"),
                        "Current Assigned Stores": tech_assigned,
                        "Future Scheduled Stores": tech_active,
                        "Completed": int(tech_rec.get("completed_count", 0)),
                        "Missing From Schedule": tech_missing,
                        "Scheduled Not Assigned": tech_wrong,
                        "Assigned Elsewhere": tech_elsewhere,
                        "Monthly Target": 10,
                        "Recommended Action": recommendation,
                    }
                )
            team_health = pd.DataFrame(team_health_rows)
            if not team_health.empty:
                needs_review = team_health[team_health["Recommended Action"] != "Protected - No Changes Needed"]
                if needs_review.empty:
                    st.success("Schedule Health: GREEN - schedule looks healthy. All reviewed PMTs are protected.")
                elif (team_health["Recommended Action"] == "Territory Reconciliation Required").any():
                    st.error(f"Schedule Health: RED - territory reconciliation required for {len(needs_review)} PMT(s).")
                else:
                    st.warning(f"Schedule Health: YELLOW - {len(needs_review)} PMT(s) need rebuild/balance review.")
                st.markdown("**PMT Schedule Health**")
                st.dataframe(team_health, use_container_width=True, hide_index=True)

            if problem_choice == "I just need the corrected Excel/PDF export":
                st.success("Go to the Export tab after repairs are complete. Use Published PMT Schedule Run, then download Full Team Excel or PDF.")
                st.caption("If the count above still looks wrong, repair the schedule before exporting.")
            elif problem_choice == "Territories changed and schedules do not match assignments":
                st.warning(
                    "Use Territory Reconciliation when assignments just changed and old schedule rows are still under the wrong PMTs. "
                    "Set the effective date, set monthly target, run the scan, select the conflict rows, and apply."
                )
                st.markdown(
                    """
                    **Quick rule**

                    Use `Territory Reconciliation` when stores moved between PMTs and you need old PMT rows superseded and affected PMTs rebuilt.

                    Use `Rebuild / Balance` when reconciliation was already applied but one PMT, like Anthony, still has too few stores.
                    """
                )
            elif problem_choice == "A store is on the wrong PMT or duplicated":
                repair_conflicts = pmt_schedule_conflicts(repair_run_items)
                if repair_conflicts.empty:
                    st.success("No duplicate or wrong-PMT active rows were found in this schedule run.")
                else:
                    st.error(f"{distinct_store_count(repair_conflicts)} store conflict(s) were found. Open Manual Edit -> Review and Resolve Schedule Conflicts for this run.")
                    st.dataframe(
                        repair_conflicts[["store_number", "assigned_technician", "technician", "month", "schedule_date", "status", "conflict_type"]],
                        use_container_width=True,
                        hide_index=True,
                    )

            if problem_choice in [
                "A PMT has too few stores or uneven months",
                "Territories changed and schedules do not match assignments",
            ]:
                if missing_count or active_count < assigned_count:
                    st.warning(f"{repair_tech_name} has {missing_count} assigned store(s) missing from the active future schedule.")
                    st.markdown("**Recommended:** open `Rebuild / Balance`.")
                if wrong_count or elsewhere_count:
                    st.warning(f"{repair_tech_name} has territory/schedule ownership mismatch rows.")
                    st.markdown("**Recommended:** open `Territory Reconciliation` if territories changed, or `Manual Edit` for a one-off correction.")


with tab_rebuild:
    section_header(
        "Rebuild / Balance PMT Schedule",
        "Use this when one technician owns the correct stores but their future schedule is incomplete, uneven, or needs rebuilt from current Areas and Maps assignments.",
        "orange",
    )
    rebuild_runs = safe_query(
        """
        select r.id, r.run_name, r.created_at, r.cycle_start, r.cycle_end, r.months, r.technician_count,
               r.store_count, r.unscheduled_count, r.status
        from pmt_schedule_runs r
        where coalesce(lower(trim(r.status)), '') not in ('deleted','snapshot')
        order by r.created_at desc, r.id desc
        """,
        use_cache=False,
    )
    if rebuild_runs.empty:
        st.warning("No current PMT schedule runs are available to rebuild.")
    else:
        rebuild_select_cols = st.columns([0.45, 0.35, 0.2])
        rebuild_run_id = rebuild_select_cols[0].selectbox(
            "Current active schedule run",
            rebuild_runs["id"].tolist(),
            format_func=lambda value: f"#{value} - {rebuild_runs.set_index('id').loc[value, 'run_name']}",
            key="pmt_rebuild_run",
        )
        rebuild_run_row = rebuild_runs.set_index("id").loc[rebuild_run_id]
        rebuild_run_items = pmt_manage_run_items(rebuild_run_id)
        rebuild_cycle_start = scalar_date(rebuild_run_row.get("cycle_start"))
        rebuild_cycle_end = scalar_date(rebuild_run_row.get("cycle_end"))
        rebuild_run_items, _rebuild_outside_items = split_run_items_by_period(rebuild_run_items, rebuild_cycle_start, rebuild_cycle_end)
        rebuild_tech_options = (
            rebuild_run_items[["employee_id", "technician"]].dropna(subset=["employee_id"]).drop_duplicates()
            if not rebuild_run_items.empty and {"employee_id", "technician"}.issubset(rebuild_run_items.columns)
            else pd.DataFrame(columns=["employee_id", "technician"])
        )
        all_rebuild_pmts = active_pmt_employee_summary()[["employee_id", "technician_name"]].rename(columns={"technician_name": "technician"})
        rebuild_tech_options = pd.concat([rebuild_tech_options, all_rebuild_pmts], ignore_index=True).dropna(subset=["employee_id"]).drop_duplicates("employee_id").sort_values("technician")
        if rebuild_tech_options.empty:
            st.warning("No active PMT technicians are available.")
        else:
            rebuild_employee_id = rebuild_select_cols[1].selectbox(
                "PMT technician",
                rebuild_tech_options["employee_id"].astype(int).tolist(),
                format_func=lambda value: rebuild_tech_options.set_index("employee_id").loc[value, "technician"],
                key="pmt_rebuild_employee",
            )
            rebuild_tech_name = rebuild_tech_options.set_index("employee_id").loc[rebuild_employee_id, "technician"]
            rebuild_start_options = manage_month_options(rebuild_run_items)
            rebuild_default_start = month_start(date.today())
            if rebuild_default_start not in rebuild_start_options:
                rebuild_start_options.append(rebuild_default_start)
                rebuild_start_options = ["All months"] + sorted([value for value in rebuild_start_options if value != "All months"])
            rebuild_start_month = rebuild_select_cols[2].selectbox(
                "Start month",
                rebuild_start_options,
                index=rebuild_start_options.index(rebuild_default_start) if rebuild_default_start in rebuild_start_options else 0,
                format_func=lambda value: value if value == "All months" else month_label(value),
                key="pmt_rebuild_start_month",
            )
            if rebuild_start_month == "All months":
                rebuild_start_month = rebuild_default_start
            rebuild_rec = technician_schedule_reconciliation(rebuild_run_items, rebuild_employee_id, "All months")
            rebuild_metric_cols = st.columns(6)
            rebuild_metric_cols[0].metric("Current Assigned", int(rebuild_rec.get("assigned_count", 0)))
            rebuild_metric_cols[1].metric("Future Scheduled", int(rebuild_rec.get("active_count", 0)))
            rebuild_metric_cols[2].metric("Completed", int(rebuild_rec.get("completed_count", 0)))
            rebuild_metric_cols[3].metric("Missing", int(rebuild_rec.get("assigned_not_scheduled_count", 0)))
            rebuild_metric_cols[4].metric("Scheduled Not Assigned", int(rebuild_rec.get("scheduled_no_longer_assigned_count", 0)))
            rebuild_metric_cols[5].metric("Assigned Elsewhere", int(rebuild_rec.get("assigned_scheduled_elsewhere_count", 0)))
            rebuild_controls = st.columns([0.18, 0.32, 0.34, 0.16])
            rebuild_monthly_target = rebuild_controls[0].number_input(
                "Monthly target",
                min_value=1,
                max_value=50,
                value=10,
                step=1,
                key="pmt_rebuild_monthly_target",
            )
            available_months = max(1, ((rebuild_cycle_end.year - rebuild_start_month.year) * 12 + rebuild_cycle_end.month - rebuild_start_month.month + 1) if rebuild_cycle_end else 6)
            expected_capacity = int(rebuild_monthly_target) * available_months
            assigned_now = int(rebuild_rec.get("assigned_count", 0))
            expected_scheduled = min(assigned_now, expected_capacity)
            expected_overflow = max(0, assigned_now - expected_capacity)
            rebuild_controls[1].caption(
                f"Expected distribution: {expected_scheduled} scheduled, {expected_overflow} overflow/not scheduled across {available_months} month(s)."
            )
            rebuild_reason = rebuild_controls[2].text_input(
                "Reason",
                value="Repair PMT schedule from current Areas and Maps assignments.",
                key="pmt_rebuild_reason",
            )
            rebuild_confirm = rebuild_controls[3].checkbox("Confirm", key="pmt_rebuild_confirm")
            if st.button(
                "Apply Rebuilt Schedule",
                type="primary",
                disabled=not rebuild_confirm,
                key="pmt_rebuild_apply",
            ):
                with session_scope("PMT rebuild/balance from current assignments") as session:
                    result = rebuild_pmt_employee_from_current_assignments(
                        session,
                        rebuild_run_id,
                        rebuild_employee_id,
                        rebuild_start_month,
                        rebuild_monthly_target,
                        rebuild_reason,
                    )
                st.success(
                    f"Rebuilt {result.get('scheduled', 0)} of {result.get('assigned', 0)} current assigned store(s) for {rebuild_tech_name}. "
                    f"Overflow/not scheduled: {result.get('overflow', 0)}. Created {result.get('created', 0)}, updated {result.get('updated', 0)}, superseded {result.get('superseded', 0)}."
                )
                st.rerun()


with tab_reconcile:
    section_header(
        "Territory Reconciliation",
        "Use this when PMT store ownership changed and existing future schedules need to align with current Areas and Maps assignments. Affected PMTs are rebuilt from current assignments; protected PMTs are not changed.",
        "yellow",
    )
    st.caption("Use this after a PMT quits, a new PMT is hired, or territories are realigned. Completed historical rows are preserved, and unaffected technicians are protected.")
    reconcile_runs = safe_query(
        """
        select id, run_name, cycle_start, cycle_end, status, created_at
        from pmt_schedule_runs
        order by created_at desc, id desc
        """
    )
    recon_cols = st.columns([0.24, 0.28, 0.48])
    effective_choice = recon_cols[0].selectbox(
        "Assignment effective date",
        ["Immediately", "Beginning of next workday", "Beginning of next month", "Custom"],
        key="pmt_reconciliation_effective_choice",
    )
    if effective_choice == "Beginning of next month":
        reconciliation_effective_date = add_months(month_start(date.today()), 1)
    elif effective_choice == "Beginning of next workday":
        reconciliation_effective_date = next_workday_after(date.today())
    elif effective_choice == "Custom":
        reconciliation_effective_date = recon_cols[1].date_input("Custom effective date", value=date.today(), key="pmt_reconciliation_effective_custom")
    else:
        reconciliation_effective_date = date.today()
    run_options = [None] + (reconcile_runs["id"].tolist() if not reconcile_runs.empty else [])
    selected_reconcile_run = recon_cols[1].selectbox(
        "Schedule run scope",
        run_options,
        format_func=lambda value: "All future PMT schedules" if value is None else f"#{value} - {reconcile_runs.set_index('id').loc[value, 'run_name']}",
        key="pmt_reconciliation_run_scope",
    )
    reconciliation_reason = recon_cols[2].text_input(
        "Reason / note",
        value="Territory realignment after PMT assignment changes",
        key="pmt_reconciliation_reason",
    )
    reconciliation_monthly_target = st.number_input(
        "Monthly PMT store target for rebuilt affected schedules",
        min_value=1,
        max_value=50,
        value=10,
        step=1,
        help="Used when reconciliation rebuilds affected PMTs from current Areas and Maps assignments. Example: Anthony has 33 stores and target 10 means 10/10/10/3 if four months are available.",
        key="pmt_reconciliation_monthly_target",
    )
    scan_controls = st.columns([0.34, 0.33, 0.33])
    ignore_effective_date = scan_controls[0].checkbox(
        "Ignore effective date and scan all unfinished PMT items",
        value=True,
        key="pmt_reconciliation_full_scan_ignore_effective_date",
    )
    if scan_controls[1].button("Run Full PMT Assignment-to-Schedule Scan", type="primary", key="pmt_reconciliation_full_rescan"):
        st.cache_data.clear()
        st.session_state["pmt_reconciliation_last_refresh"] = datetime.utcnow().isoformat(timespec="seconds") + "Z"
        st.rerun()
    if scan_controls[2].button("Refresh Assignments and Rescan", key="pmt_reconciliation_refresh_rescan"):
        st.cache_data.clear()
        st.session_state["pmt_reconciliation_last_refresh"] = datetime.utcnow().isoformat(timespec="seconds") + "Z"
        st.rerun()

    scan = pmt_reconciliation_scan(reconciliation_effective_date, selected_reconcile_run, ignore_effective_date=ignore_effective_date)
    conflicts = scan["conflicts"].copy()
    assigned_not_scheduled = scan["assigned_not_scheduled"].copy()
    affected = scan["affected"].copy()
    protected = scan["protected"].copy()
    diagnostics = scan.get("diagnostics", {})

    st.caption(
        " | ".join(
            [
                f"Current Workspace: {diagnostics.get('workspace', 'Current workspace')}",
                f"Assignments Found: {diagnostics.get('assignments_loaded', 0)}",
                f"Schedule Runs Found: {diagnostics.get('schedule_runs_scanned', 0)}",
                f"Future/unfinished Schedule Items Found: {diagnostics.get('schedule_items_loaded', 0)}",
                f"Last refreshed: {diagnostics.get('scan_timestamp', 'Not scanned')}",
            ]
        )
    )
    if st.session_state.get("account_role") == "Admin":
        with st.expander("Reconciliation Scan Diagnostics", expanded=False):
            st.json(diagnostics)

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Schedule Conflicts", len(conflicts))
    m2.metric("Assigned Not Scheduled", len(assigned_not_scheduled))
    m3.metric("Affected PMTs", len(affected))
    m4.metric("Protected PMTs", len(protected))
    st.download_button(
        "Export PMT Territory and Schedule Change Package",
        data=pmt_reconciliation_package_bytes(scan, reconciliation_effective_date, reconciliation_reason),
        file_name=f"pmt_territory_schedule_change_package_{reconciliation_effective_date}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        disabled=conflicts.empty and assigned_not_scheduled.empty,
    )

    if conflicts.empty and assigned_not_scheduled.empty:
        st.success("No future PMT assignment/schedule conflicts were found for this scope.")
    else:
        filter_cols = st.columns(4)
        tech_filter_options = ["All"]
        if not conflicts.empty:
            tech_filter_options += sorted(set(conflicts["scheduled_technician"].dropna().astype(str).tolist()) | set(conflicts["assigned_technician"].dropna().astype(str).tolist()))
        selected_conflict_tech = filter_cols[0].selectbox("Technician filter", tech_filter_options, key="pmt_reconciliation_tech_filter")
        conflict_type_options = ["All"] + (sorted(conflicts["conflict_type"].dropna().astype(str).unique().tolist()) if not conflicts.empty else [])
        selected_conflict_type = filter_cols[1].selectbox("Conflict type", conflict_type_options, key="pmt_reconciliation_type_filter")
        month_options = ["All"] + (sorted(conflicts["month"].dropna().astype(str).unique().tolist()) if "month" in conflicts.columns and not conflicts.empty else [])
        selected_conflict_month = filter_cols[2].selectbox("Month", month_options, key="pmt_reconciliation_month_filter")
        only_transferable = filter_cols[3].checkbox("Show transferable only", value=False, key="pmt_reconciliation_transferable_only")

        filtered_conflicts = conflicts.copy()
        if selected_conflict_tech != "All" and not filtered_conflicts.empty:
            filtered_conflicts = filtered_conflicts[(filtered_conflicts["scheduled_technician"].astype(str) == selected_conflict_tech) | (filtered_conflicts["assigned_technician"].astype(str) == selected_conflict_tech)]
        if selected_conflict_type != "All" and not filtered_conflicts.empty:
            filtered_conflicts = filtered_conflicts[filtered_conflicts["conflict_type"].astype(str) == selected_conflict_type]
        if selected_conflict_month != "All" and "month" in filtered_conflicts.columns and not filtered_conflicts.empty:
            filtered_conflicts = filtered_conflicts[filtered_conflicts["month"].astype(str) == selected_conflict_month]
        if only_transferable and not filtered_conflicts.empty:
            filtered_conflicts = filtered_conflicts[filtered_conflicts["assigned_employee_id"].notna()]

        st.markdown("**Schedule Conflicts**")
        render_reconciliation_color_legend()
        if filtered_conflicts.empty:
            st.info("No conflicts match the selected filters.")
        else:
            decision_tab, compare_tab, old_new_tab, earlier_tab, future_tab = st.tabs([
                "Transfer Decisions",
                "Compare Old vs New",
                "New Schedule Preview",
                "Earlier Unfinished Work",
                "Effective-Date / Future Work",
            ])
            with decision_tab:
                conflict_view = filtered_conflicts.copy()
                conflict_view["Can Transfer"] = conflict_view["assigned_employee_id"].notna()
                bulk_select_transfers = st.checkbox(
                    "Select all transferable rows currently shown",
                    value=False,
                    key="pmt_reconciliation_select_all_visible_transfers",
                )
                conflict_view["Apply"] = conflict_view["Can Transfer"].astype(bool) if bulk_select_transfers else False
                view_cols = [
                    "Apply", "Can Transfer", "work_timing", "schedule_item_id", "store_number", "city", "state", "scheduled_technician",
                    "assigned_technician", "schedule_name", "schedule_date", "status", "conflict_type",
                    "assignment_effective_date", "recommended_action", "resolution",
                ]
                editor_key_parts = [
                    key(selected_conflict_tech),
                    key(selected_conflict_type),
                    key(selected_conflict_month),
                    str(int(bool(only_transferable))),
                    str(int(bool(bulk_select_transfers))),
                ]
                edited_conflicts = st.data_editor(
                    conflict_view[[col for col in view_cols if col in conflict_view.columns]],
                    use_container_width=True,
                    hide_index=True,
                    disabled=[col for col in view_cols if col != "Apply"],
                    column_config={
                        "Apply": st.column_config.CheckboxColumn("Apply Transfer", default=False),
                        "Can Transfer": st.column_config.CheckboxColumn("Can Transfer", disabled=True),
                        "work_timing": st.column_config.TextColumn("Timing"),
                        "schedule_item_id": st.column_config.NumberColumn("Item ID", disabled=True),
                        "scheduled_technician": st.column_config.TextColumn("Original Scheduled PMT"),
                        "assigned_technician": st.column_config.TextColumn("Current Assigned PMT"),
                    },
                    key=f"pmt_reconciliation_conflict_editor_{'_'.join(editor_key_parts)}",
                )
                selected_item_ids = edited_conflicts.loc[
                    edited_conflicts["Apply"].astype(bool) & edited_conflicts["Can Transfer"].astype(bool),
                    "schedule_item_id",
                ].dropna().astype(int).tolist()
                st.caption(f"Selected unfinished schedule rows to transfer and resequence: {len(selected_item_ids)}")
                compare_preview = build_pmt_reconciliation_compare_preview(scan, selected_item_ids, reconciliation_monthly_target)
                confirm_reconciliation = st.checkbox(
                    "I reviewed the PMT assignment conflicts and confirm I am ready to transfer the checked unfinished schedule rows.",
                    key="pmt_reconciliation_confirm_apply",
                )
                if st.button("Apply Selected Reconciliation Transfers", type="primary", disabled=not selected_item_ids or not confirm_reconciliation, key="pmt_reconciliation_apply_transfers"):
                    result = apply_pmt_reconciliation_transfers(selected_item_ids, reconciliation_effective_date, reconciliation_reason, reconciliation_monthly_target)
                    st.success(
                        f"Reconciliation applied. Transferred {result['transferred']} row(s); "
                        f"superseded {result['superseded']} duplicate row(s); resequenced {result.get('resequenced_rows', 0)} source route row(s); "
                        f"rebuilt {result.get('rebuilt_rows', 0)} row(s) from current assignments; "
                        f"{result.get('rebuild_overflow', 0)} store(s) went to Not Scheduled/overflow; "
                        f"created {result.get('snapshots_created', 0)} old-schedule snapshot run(s)."
                    )
                    st.rerun()
            display_cols = [
                "work_timing", "schedule_item_id", "store_number", "city", "state", "scheduled_technician",
                "assigned_technician", "schedule_name", "schedule_date", "status", "conflict_type", "recommended_action",
            ]
            with compare_tab:
                if not selected_item_ids:
                    st.info("Check one or more rows in Transfer Decisions to preview old schedule vs proposed new schedule.")
                else:
                    compare_df = compare_preview.get("compare", pd.DataFrame())
                    protected_df = compare_preview.get("protected", pd.DataFrame())
                    c1, c2, c3, c4 = st.columns(4)
                    c1.metric("Checked Transfers", len(selected_item_ids))
                    c2.metric("Changed / Resequenced Rows", int(compare_df["change_status"].astype(str).ne("No change").sum()) if not compare_df.empty and "change_status" in compare_df.columns else 0)
                    c3.metric("Protected PMTs", len(protected_df))
                    c4.metric("Rows Compared", len(compare_df))
                    st.download_button(
                        "Export Old vs New Schedule Compare",
                        data=pmt_reconciliation_compare_workbook_bytes(compare_preview),
                        file_name=f"pmt_reconciliation_old_vs_new_compare_{reconciliation_effective_date}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="pmt_reconciliation_compare_export",
                    )
                    if compare_df.empty:
                        st.info("No current-month-forward rows were found for the selected transfer preview.")
                    else:
                        st.dataframe(compare_df, use_container_width=True, hide_index=True)
                    with st.expander("Protected PMTs - No Proposed Change", expanded=False):
                        if protected_df.empty:
                            st.info("No protected PMTs are listed for this scan.")
                        else:
                            st.dataframe(protected_df, use_container_width=True, hide_index=True)
            with old_new_tab:
                if not selected_item_ids:
                    st.info("Check rows in Transfer Decisions to preview the proposed new schedule.")
                else:
                    old_schedule = compare_preview.get("old_schedule", pd.DataFrame())
                    new_schedule = compare_preview.get("new_schedule", pd.DataFrame())
                    old_tab, new_tab = st.tabs(["Old Schedule", "New Schedule Preview"])
                    with old_tab:
                        if old_schedule.empty:
                            st.info("No old schedule rows found from the current month forward.")
                        else:
                            for month_name in old_schedule["month"].dropna().drop_duplicates().tolist():
                                with st.expander(str(month_name), expanded=False):
                                    st.dataframe(old_schedule[old_schedule["month"] == month_name], use_container_width=True, hide_index=True)
                    with new_tab:
                        if new_schedule.empty:
                            st.info("No new schedule rows found from the current month forward.")
                        else:
                            for month_name in new_schedule["month"].dropna().drop_duplicates().tolist():
                                with st.expander(str(month_name), expanded=False):
                                    st.dataframe(new_schedule[new_schedule["month"] == month_name], use_container_width=True, hide_index=True)
            with earlier_tab:
                earlier_conflicts = filtered_conflicts[filtered_conflicts.get("work_timing", "").astype(str) == "Earlier unfinished work"].copy()
                if earlier_conflicts.empty:
                    st.success("No earlier unfinished reassignment conflicts match the selected filters.")
                else:
                    st.dataframe(reconciliation_styler(earlier_conflicts[[col for col in display_cols if col in earlier_conflicts.columns]]), use_container_width=True, hide_index=True)
            with future_tab:
                future_conflicts = filtered_conflicts[filtered_conflicts.get("work_timing", "").astype(str) != "Earlier unfinished work"].copy()
                if future_conflicts.empty:
                    st.success("No effective-date or future conflicts match the selected filters.")
                else:
                    st.dataframe(reconciliation_styler(future_conflicts[[col for col in display_cols if col in future_conflicts.columns]]), use_container_width=True, hide_index=True)

        st.markdown("**Assigned Stores With No Future Schedule For Their Current PMT**")
        if assigned_not_scheduled.empty:
            st.success("No assigned-but-not-scheduled PMT stores found.")
        else:
            assigned_gap_view = assigned_not_scheduled[
                [col for col in ["work_timing", "assigned_technician", "store_number", "city", "state", "conflict_type", "recommended_action"] if col in assigned_not_scheduled.columns]
            ].copy()
            st.dataframe(
                reconciliation_styler(assigned_gap_view),
                use_container_width=True,
                hide_index=True,
            )

    with st.expander("Affected and Protected PMTs", expanded=not affected.empty):
        cols = st.columns(2)
        with cols[0]:
            st.markdown("**Affected PMTs**")
            if affected.empty:
                st.success("No affected PMTs found.")
            else:
                st.dataframe(affected, use_container_width=True, hide_index=True)
        with cols[1]:
            st.markdown("**Protected Unaffected PMTs**")
            if protected.empty:
                st.info("No protected PMTs found.")
            else:
                st.dataframe(protected, use_container_width=True, hide_index=True)

    st.divider()
    st.markdown(
        """
        <style>
        .pmt-next-steps {
            border: 2px solid #f59e0b;
            background: #fff7ed;
            padding: 1.1rem 1.25rem;
            border-radius: 8px;
            margin: .75rem 0 1rem 0;
            color: #111827;
        }
        .pmt-next-steps h3 {
            margin: 0 0 .65rem 0;
            font-size: 1.45rem;
            line-height: 1.2;
        }
        .pmt-next-steps ol {
            margin: .3rem 0 0 1.25rem;
            padding: 0;
            font-size: 1.08rem;
            line-height: 1.65;
        }
        .pmt-next-steps strong {
            font-weight: 800;
        }
        .pmt-action-label {
            font-size: 1.05rem;
            font-weight: 800;
            margin-bottom: .25rem;
            color: #111827;
        }
        </style>
        <div class="pmt-next-steps">
            <h3>Where To Get The New Schedule</h3>
            <ol>
                <li><strong>Before applying:</strong> use <strong>Compare Old vs New</strong> to review what will change.</li>
                <li><strong>After applying:</strong> the updated active schedule is in <strong>Manage Existing PMT Schedule</strong>.</li>
                <li>Select the same schedule plan, then open <strong>View all schedule rows</strong> and click <strong>Export Full Schedule</strong>.</li>
                <li>If you need the old version later, select the schedule plan marked <strong>Snapshot</strong>.</li>
            </ol>
        </div>
        """,
        unsafe_allow_html=True,
    )
    next_cols = st.columns([0.34, 0.33, 0.33])
    next_cols[0].markdown('<div class="pmt-action-label">Jump to Manage</div>', unsafe_allow_html=True)
    if next_cols[0].button(
        "Set Manage Tab To This Schedule",
        disabled=selected_reconcile_run is None,
        key="pmt_reconciliation_prepare_manage_schedule",
    ):
        st.session_state["pmt_manage_selected_run"] = int(selected_reconcile_run)
        st.session_state["pmt_reconciliation_manage_notice"] = (
            "Open `Manage Existing PMT Schedule` above. The schedule plan you reconciled is preselected there."
        )
        st.success(st.session_state["pmt_reconciliation_manage_notice"])
    if selected_reconcile_run is None:
        next_cols[0].warning("Pick one schedule run scope above first.")
    updated_schedule_export = pd.DataFrame()
    if selected_reconcile_run is not None:
        reconcile_run_row = reconcile_runs.set_index("id").loc[selected_reconcile_run] if not reconcile_runs.empty and selected_reconcile_run in reconcile_runs["id"].tolist() else {}
        updated_schedule_export = current_month_forward_schedule_rows(
            pmt_manage_run_items(selected_reconcile_run),
            scalar_date(reconcile_run_row.get("cycle_start")) if hasattr(reconcile_run_row, "get") else None,
            scalar_date(reconcile_run_row.get("cycle_end")) if hasattr(reconcile_run_row, "get") else None,
        )
        export_cols = [
            col for col in [
                "schedule_date", "sequence_number", "technician", "assigned_technician", "store_number",
                "address", "city", "state", "zip", "status", "cycle_label", "schedule_source", "notes"
            ] if col in updated_schedule_export.columns
        ]
        if export_cols:
            updated_schedule_export = updated_schedule_export[export_cols]
    next_cols[1].markdown('<div class="pmt-action-label">Get New Schedule</div>', unsafe_allow_html=True)
    next_cols[1].download_button(
        "Export Updated Schedule",
        data=excel_bytes(updated_schedule_export),
        file_name=f"pmt_updated_schedule_run_{selected_reconcile_run or 'all'}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        disabled=selected_reconcile_run is None or updated_schedule_export.empty,
        key="pmt_reconciliation_export_updated_schedule",
    )
    next_cols[2].markdown('<div class="pmt-action-label">Get Change Package</div>', unsafe_allow_html=True)
    next_cols[2].download_button(
        "Export Reconciliation Package",
        data=pmt_reconciliation_package_bytes(scan, reconciliation_effective_date, reconciliation_reason),
        file_name=f"pmt_territory_schedule_change_package_{reconciliation_effective_date}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        disabled=conflicts.empty and assigned_not_scheduled.empty,
        key="pmt_reconciliation_bottom_compare_export",
    )


with tab_manage:
    section_header(
        "Manual Schedule Edit",
        "Use this when you know exactly which schedule item needs to change: add, remove, move, reorder, reschedule, or update status/notes. One-off edits do not rebuild another PMT unless explicitly selected.",
        "blue",
    )
    if st.session_state.get("pmt_reconciliation_manage_notice"):
        st.success(st.session_state.pop("pmt_reconciliation_manage_notice"))
    with st.expander("Optional: Import an Existing PMT Schedule", expanded=False):
        st.caption(
            "Use this only when a PMT schedule was created outside this application and needs to be brought into the system. "
            "Skip this section when managing a schedule that is already in the app."
        )
        schedule_upload = st.file_uploader(
            "Upload existing PMT schedule Excel/CSV",
            type=["xlsx", "xls", "xlsm", "csv"],
            key="pmt_existing_schedule_upload",
        )
        if schedule_upload:
            schedule_sheets = upload_sheet_names(schedule_upload)
            schedule_sheet = st.selectbox("Schedule sheet", schedule_sheets, key="pmt_existing_schedule_sheet")
            schedule_raw = read_upload_sheet(schedule_upload, schedule_sheet)
            original_columns = schedule_raw.columns.tolist()
            mapping_options = [""] + original_columns
            schedule_defaults = {
                field: best_column(original_columns, field, "schedule")
                for field in SCHEDULE_COLUMN_CANDIDATES
            }
            st.caption(f"Rows detected: {len(schedule_raw):,}. Map the columns below before importing.")
            mc1, mc2, mc3 = st.columns(3)
            import_tech_col = selectbox_with_default(mc1, "Technician", mapping_options, schedule_defaults["technician_name"], "pmt_existing_schedule_tech_col")
            import_store_col = selectbox_with_default(mc2, "Store Number", mapping_options, schedule_defaults["store_number"], "pmt_existing_schedule_store_col")
            import_date_col = selectbox_with_default(mc3, "Schedule Date", mapping_options, schedule_defaults["schedule_date"], "pmt_existing_schedule_date_col")
            mc4, mc5, mc6, mc7 = st.columns(4)
            import_month_col = selectbox_with_default(mc4, "Month if no date", mapping_options, schedule_defaults["schedule_month"], "pmt_existing_schedule_month_col")
            import_sequence_col = selectbox_with_default(mc5, "Stop / Sequence", mapping_options, schedule_defaults["sequence_number"], "pmt_existing_schedule_sequence_col")
            import_status_col = selectbox_with_default(mc6, "Status", mapping_options, schedule_defaults["status"], "pmt_existing_schedule_status_col")
            import_notes_col = selectbox_with_default(mc7, "Notes", mapping_options, schedule_defaults["notes"], "pmt_existing_schedule_notes_col")
            import_mapping = {
                "technician_name": import_tech_col,
                "store_number": import_store_col,
                "schedule_date": import_date_col,
                "schedule_month": import_month_col,
                "sequence_number": import_sequence_col,
                "status": import_status_col,
                "notes": import_notes_col,
            }
            required_missing = [label for label, field in [("Technician", "technician_name"), ("Store Number", "store_number")] if not import_mapping.get(field)]
            if not import_mapping.get("schedule_date") and not import_mapping.get("schedule_month"):
                required_missing.append("Schedule Date or Month")
            imported_preview, import_problems = normalize_existing_pmt_schedule_upload(schedule_raw, import_mapping) if not required_missing else (pd.DataFrame(), pd.DataFrame())
            if required_missing:
                st.error("Missing required schedule mapping: " + ", ".join(required_missing))
            else:
                pv1, pv2, pv3, pv4 = st.columns(4)
                pv1.metric("Rows Ready", len(imported_preview))
                pv2.metric("Technicians", imported_preview["employee_id"].nunique() if not imported_preview.empty else 0)
                pv3.metric("Stores", imported_preview["store_id"].nunique() if not imported_preview.empty else 0)
                pv4.metric("Warnings / Problems", len(import_problems))
                if not import_problems.empty:
                    with st.expander("Import warnings and skipped rows", expanded=True):
                        st.dataframe(import_problems, use_container_width=True, hide_index=True)
                if not imported_preview.empty:
                    preview_columns = ["technician", "month", "schedule_date", "sequence_number", "store_number", "city", "state", "status", "notes"]
                    st.dataframe(imported_preview[preview_columns].head(100), use_container_width=True, hide_index=True)
                    imported_start = imported_preview["schedule_date"].min()
                    imported_end = imported_preview["schedule_date"].max()
                    import_run_name = st.text_input(
                        "Imported schedule run name",
                        value=f"Imported PMT Schedule {month_label(month_start(imported_start))} - {month_label(month_start(imported_end))}",
                        key="pmt_existing_schedule_run_name",
                    )
                    confirm_import = st.checkbox("I reviewed this imported schedule and want to create a PMT schedule run.", key="pmt_confirm_existing_schedule_import")
                    import_success = st.session_state.get("pmt_existing_schedule_import_success")
                    if import_success:
                        st.success(import_success)
                        st.info("The imported run is saved. Refresh the page or clear the upload to select it below.")
                    if st.button("Import Existing PMT Schedule", type="primary", disabled=not confirm_import, key="pmt_import_existing_schedule_button"):
                        try:
                            with st.spinner(f"Importing {len(imported_preview):,} PMT schedule item(s)..."):
                                result = import_existing_pmt_schedule(imported_preview, import_run_name)
                            success_message = f"Imported PMT schedule run #{result['run_id']} with {result['created']} schedule item(s)."
                            st.session_state["pmt_existing_schedule_import_success"] = success_message
                            st.success(success_message)
                        except Exception as exc:
                            st.error(f"PMT schedule import failed: {exc}")
                            st.stop()

    section_header("Step 1: Select the Schedule You Want to Manage", "Choose one published or imported PMT schedule. The rest of this workspace follows that selection.", "gray")
    runs = safe_query(
        """
        select r.id, r.run_name, r.created_at, r.cycle_start, r.cycle_end, r.months, r.technician_count,
               r.store_count, r.unscheduled_count, r.status
        from pmt_schedule_runs r
        where coalesce(r.status, '') <> 'Deleted'
        order by r.created_at desc, r.id desc
        """
    )
    if runs.empty:
        st.info("No PMT schedule plans have been published or imported yet.")
    else:
        run_live_counts = safe_query(
            """
            select
                r.id,
                count(si.id) filter (
                    where si.status in ('Scheduled','Needs Rescheduled','Rescheduled','Rain Delay','Not Completed')
                      and (r.cycle_start is null or date(si.schedule_date) >= date(r.cycle_start))
                      and (r.cycle_end is null or date(si.schedule_date) <= date(r.cycle_end))
                ) as live_active_rows,
                count(distinct si.store_id) filter (
                    where si.status in ('Scheduled','Needs Rescheduled','Rescheduled','Rain Delay','Not Completed')
                      and (r.cycle_start is null or date(si.schedule_date) >= date(r.cycle_start))
                      and (r.cycle_end is null or date(si.schedule_date) <= date(r.cycle_end))
                ) as live_active_stores,
                count(distinct si.employee_id) filter (
                    where si.status in ('Scheduled','Needs Rescheduled','Rescheduled','Rain Delay','Not Completed')
                      and (r.cycle_start is null or date(si.schedule_date) >= date(r.cycle_start))
                      and (r.cycle_end is null or date(si.schedule_date) <= date(r.cycle_end))
                ) as live_active_pmts
            from pmt_schedule_runs r
            left join schedule_items si
              on si.pmt_schedule_run_id = r.id
             and si.work_type = 'PMT'
            where coalesce(r.status, '') <> 'Deleted'
            group by r.id
            """,
            use_cache=False,
        )
        if not run_live_counts.empty:
            runs = runs.merge(run_live_counts, on="id", how="left")
        for column in ["live_active_rows", "live_active_stores", "live_active_pmts"]:
            if column not in runs.columns:
                runs[column] = 0
            runs[column] = pd.to_numeric(runs[column], errors="coerce").fillna(0).astype(int)
        run_lookup = runs.set_index("id")

        def manage_run_label(value):
            row = run_lookup.loc[value]
            start_text = month_label(scalar_date(row.get("cycle_start"))) if scalar_date(row.get("cycle_start")) else "No start"
            end_text = month_label(scalar_date(row.get("cycle_end"))) if scalar_date(row.get("cycle_end")) else "No end"
            return (
                f"#{value} - {row['run_name']} | {start_text} to {end_text} | "
                f"{int(row.get('live_active_stores') or 0)} active stores, {int(row.get('live_active_pmts') or 0)} PMTs"
            )

        selected_run = st.selectbox(
            "Schedule plan",
            runs["id"].tolist(),
            format_func=manage_run_label,
            key="pmt_manage_selected_run",
        )
        run_row = run_lookup.loc[selected_run]
        raw_run_items = pmt_manage_run_items(selected_run)
        run_cycle_start = scalar_date(run_row.get("cycle_start"))
        run_cycle_end = scalar_date(run_row.get("cycle_end"))
        run_items, out_of_period_items = split_run_items_by_period(raw_run_items, run_cycle_start, run_cycle_end)
        run_counts = run_status_counts(run_items)
        current_pmt_summary = active_pmt_employee_summary()
        current_assigned_store_count = (
            int(pd.to_numeric(current_pmt_summary.get("assigned_stores"), errors="coerce").fillna(0).sum())
            if not current_pmt_summary.empty
            else 0
        )
        run_cols = st.columns(6)
        run_cols[0].metric("Status", clean(run_row.get("status", "")) or "Published")
        run_cols[1].metric("Start", month_label(scalar_date(run_row.get("cycle_start"))) if scalar_date(run_row.get("cycle_start")) else "N/A")
        run_cols[2].metric("End", month_label(scalar_date(run_row.get("cycle_end"))) if scalar_date(run_row.get("cycle_end")) else "N/A")
        run_cols[3].metric("PMTs in Selected Plan", int(run_row.get("live_active_pmts") or 0))
        run_cols[4].metric("Active Stores in Selected Plan", distinct_store_count(run_items[pmt_active_item_mask(run_items)]) if not run_items.empty else 0)
        run_cols[5].metric("Unscheduled", scalar_int(run_row.get("unscheduled_count"), 0))
        row_cols = st.columns(3)
        row_cols[0].metric("Active Rows in Selected Plan", run_counts["active_rows"])
        row_cols[1].metric("Completed Rows", run_counts["completed_rows"])
        row_cols[2].metric("Canceled / Skipped Rows", run_counts["canceled_rows"])
        st.caption(
            f"This section is showing schedule run #{selected_run} only. Current active PMT assignments across all technicians: "
            f"{current_assigned_store_count:,} stores."
        )
        if current_assigned_store_count and run_counts["active_rows"] < current_assigned_store_count * 0.5:
            st.warning(
                f"The selected schedule plan only has {run_counts['active_rows']:,} active row(s), but current PMT assignments contain "
                f"{current_assigned_store_count:,} store(s). If you expected the full 400+ store schedule, select a full-team run or publish/import the full schedule run."
            )
        with st.expander("View all schedule rows", expanded=False):
            run_item_view = run_items[
                ["schedule_date", "sequence_number", "technician", "assigned_technician", "store_number", "address", "city", "state", "zip", "status", "cycle_label", "notes"]
            ] if not run_items.empty else pd.DataFrame()
            st.dataframe(run_item_view, use_container_width=True, hide_index=True)
            st.download_button("Export Full Schedule", data=excel_bytes(run_item_view), file_name=f"pmt_schedule_plan_{selected_run}.xlsx", key=f"export_selected_pmt_run_{selected_run}")
        active_outside_items = out_of_period_items[pmt_active_item_mask(out_of_period_items)].copy() if not out_of_period_items.empty else pd.DataFrame()
        history_outside_items = out_of_period_items[~pmt_active_item_mask(out_of_period_items)].copy() if not out_of_period_items.empty else pd.DataFrame()
        if not active_outside_items.empty:
            st.error(
                f"{len(active_outside_items)} active row(s) are outside this schedule plan's date range. "
                "These should be archived so they do not look like current schedule work."
            )
            with st.expander("Fix Active Rows Outside This Schedule Plan Date Range", expanded=True):
                outside_view = active_outside_items[
                    ["schedule_item_id", "schedule_date", "sequence_number", "technician", "assigned_technician", "store_number", "city", "state", "status", "schedule_source", "notes"]
                ].copy()
                st.dataframe(outside_view, use_container_width=True, hide_index=True)
                st.warning(
                    "Archiving keeps the audit trail but removes these rows from active schedule logic. "
                    "Completed historical rows are not changed."
                )
                cleanup_note = st.text_input(
                    "Cleanup note",
                    value="Out-of-period rows from this PMT schedule run were archived.",
                    key=f"pmt_manage_outside_cleanup_note_{selected_run}",
                )
                confirm_outside_archive = st.checkbox(
                    "Confirm archive active rows outside this schedule plan date range.",
                    key=f"pmt_manage_outside_archive_confirm_{selected_run}",
                )
                if st.button(
                    "Archive Active Out-of-Period Rows",
                    type="secondary",
                    disabled=not confirm_outside_archive,
                    key=f"pmt_manage_outside_archive_button_{selected_run}",
                ):
                    archived = archive_out_of_period_pmt_schedule_items(
                        selected_run,
                        active_outside_items["schedule_item_id"].dropna().astype(int).tolist(),
                        cleanup_note,
                    )
                    st.success(f"Archived {archived} active out-of-period row(s).")
                    st.rerun()
        elif not history_outside_items.empty:
            with st.expander("Archived or Historical Rows Outside This Schedule Plan", expanded=False):
                st.caption(
                    "These rows are not active schedule work and are excluded from Manual Edit counts, maps, and current exports."
                )
                outside_view = out_of_period_items[
                    ["schedule_item_id", "schedule_date", "sequence_number", "technician", "assigned_technician", "store_number", "city", "state", "status", "schedule_source", "notes"]
                ].copy()
                st.dataframe(outside_view, use_container_width=True, hide_index=True)

        conflict_rows = pmt_schedule_conflicts(run_items)
        if not conflict_rows.empty:
            conflict_store_count = distinct_store_count(conflict_rows)
            st.error(
                f"{conflict_store_count} schedule conflict(s) found. These stores are actively scheduled more than once "
                "or are scheduled under a PMT who no longer owns them."
            )
            with st.expander("Review and Resolve Schedule Conflicts", expanded=True):
                st.caption(
                    "Recommended fix: keep the schedule row for the store's current assigned PMT, then mark the other active future rows as Transferred."
                )
                conflict_view = conflict_rows[
                    [
                        "store_id",
                        "store_number",
                        "assigned_technician",
                        "technician",
                        "other_active_technicians",
                        "month",
                        "schedule_date",
                        "sequence_number",
                        "status",
                        "conflict_type",
                    ]
                ].copy()
                conflict_view.insert(0, "Resolve", True)
                edited_conflicts = st.data_editor(
                    conflict_view,
                    use_container_width=True,
                    hide_index=True,
                    disabled=[
                        "store_id",
                        "store_number",
                        "assigned_technician",
                        "technician",
                        "other_active_technicians",
                        "month",
                        "schedule_date",
                        "sequence_number",
                        "status",
                        "conflict_type",
                    ],
                    column_config={
                        "Resolve": st.column_config.CheckboxColumn("Resolve", help="Resolve this store by keeping the current assigned PMT's active schedule."),
                        "store_id": None,
                        "store_number": "Store",
                        "assigned_technician": "Current Assigned PMT",
                        "technician": "Scheduled PMT",
                        "other_active_technicians": "All Active Scheduled PMTs",
                        "month": "Month",
                        "schedule_date": "Date",
                        "sequence_number": "Stop",
                        "status": "Status",
                        "conflict_type": "Conflict",
                    },
                    key=f"pmt_manage_conflict_editor_{selected_run}",
                )
                selected_conflict_store_ids = (
                    edited_conflicts.loc[edited_conflicts["Resolve"], "store_id"].dropna().astype(int).unique().tolist()
                    if not edited_conflicts.empty
                    else []
                )
                st.write(
                    f"Selected stores to resolve: {len(selected_conflict_store_ids)}. "
                    "Completed historical rows are not changed."
                )
                conflict_notes = st.text_input(
                    "Resolution note",
                    value="Territory transfer conflict resolved from Manual Edit.",
                    key=f"pmt_manage_conflict_notes_{selected_run}",
                )
                confirm_conflict_resolution = st.checkbox(
                    "Confirm: keep each store with its current assigned PMT and transfer/supersede the other active future rows.",
                    key=f"pmt_manage_conflict_confirm_{selected_run}",
                )
                if st.button(
                    "Resolve Selected Conflicts",
                    type="primary",
                    disabled=not selected_conflict_store_ids or not confirm_conflict_resolution,
                    key=f"pmt_manage_conflict_resolve_{selected_run}",
                ):
                    result = resolve_pmt_conflicts_keep_assigned(selected_run, selected_conflict_store_ids, conflict_notes)
                    for state_key in [
                        "pmt_manage_build_preview",
                        "pmt_manage_build_preview_ids",
                        "pmt_manage_build_preview_conflict_ids",
                        "pmt_manage_build_preview_method",
                    ]:
                        st.session_state.pop(state_key, None)
                    st.success(
                        f"Resolved {result['stores']} store conflict(s), transferred/superseded {result['superseded']} old row(s), "
                        f"moved {result['moved']} row(s), and resequenced {result['resequenced_rows']} active row(s)."
                    )
                    st.rerun()

        if run_items.empty:
            st.info("This selected schedule plan has no PMT schedule rows inside its selected date range.")
        else:
            section_header("Step 2: Select Technician and Month", "One shared selection controls the summary, table, map, add-store tools, and reorder tools below.", "blue")
            context_cols = st.columns([0.32, 0.24, 0.22, 0.12, 0.1])
            tech_options = (
                run_items[["employee_id", "technician"]]
                .dropna(subset=["employee_id"])
                .drop_duplicates()
                .sort_values("technician")
            )
            all_pmt_people = active_pmt_employee_summary()[["employee_id", "technician_name"]].rename(columns={"technician_name": "technician"})
            tech_options = pd.concat([tech_options, all_pmt_people], ignore_index=True).dropna(subset=["employee_id"]).drop_duplicates("employee_id").sort_values("technician")
            selected_employee = context_cols[0].selectbox(
                "PMT Technician",
                tech_options["employee_id"].astype(int).tolist(),
                format_func=lambda value: tech_options.set_index("employee_id").loc[value, "technician"],
                key="pmt_manage_context_employee",
            )
            selected_month = context_cols[1].selectbox(
                "Month",
                manage_month_options(run_items),
                format_func=lambda value: value if value == "All months" else month_label(value),
                key="pmt_manage_context_month",
            )
            status_filter = context_cols[2].selectbox(
                "Status",
                ["Active", "All", "Completed", "Canceled / Skipped"],
                key="pmt_manage_context_status",
            )
            show_map = context_cols[3].checkbox("Show Map", value=True, key="pmt_manage_context_show_map")
            show_future_months = context_cols[4].checkbox("Future", value=False, key="pmt_manage_context_future")
            selected_tech_name = tech_options.set_index("employee_id").loc[selected_employee, "technician"]
            month_text = selected_month if selected_month == "All months" else month_label(selected_month)
            st.markdown(f"**Managing: {selected_tech_name} - {month_text}**")

            rec = technician_schedule_reconciliation(run_items, selected_employee, selected_month)
            selected_scope = filter_manage_scope(run_items, selected_employee, selected_month, status_filter).sort_values(["schedule_date", "sequence_number", "store_number"])
            overview_tab, map_builder_tab, build_tab, reorder_tab = st.tabs(["Overview", "Map Route Builder", "Build or Add Stores", "Reorder or Remove Stores"])

            with overview_tab:
                card_cols = st.columns(6)
                card_cols[0].metric("Currently Assigned", rec["assigned_count"])
                card_cols[1].metric("Actively Scheduled", rec["active_count"])
                card_cols[2].metric("Completed", rec["completed_count"])
                card_cols[3].metric("Canceled / Skipped", rec["canceled_count"])
                card_cols[4].metric("Assigned Not Scheduled", rec["assigned_not_scheduled_count"])
                card_cols[5].metric("Scheduled Not Assigned", rec["scheduled_no_longer_assigned_count"])
                if rec["assigned_scheduled_elsewhere_count"]:
                    st.warning(
                        f"{rec['assigned_scheduled_elsewhere_count']} store(s) currently assigned to {selected_tech_name} are active in this schedule under another PMT. "
                        "Open the detail list below before rebuilding this route."
                    )
                explanation = (
                    f"{selected_tech_name} currently owns {rec['assigned_count']} store(s). "
                    f"In this selected schedule context, {rec['active_count']} are actively scheduled under {selected_tech_name}, "
                    f"{rec['completed_count']} are completed, {rec['canceled_count']} are canceled/skipped, "
                    f"and {rec['assigned_not_scheduled_count']} assigned store(s) do not have an active schedule row."
                )
                if rec["scheduled_no_longer_assigned_count"]:
                    explanation += f" {rec['scheduled_no_longer_assigned_count']} store(s) remain scheduled under {selected_tech_name} but are no longer assigned to this PMT."
                st.info(explanation)
                st.caption("Use the Rebuild / Balance sub-tab if this PMT needs to be rebuilt from current Areas and Maps assignments.")
                if rec["completed_count"]:
                    completed_rows = rec["tech_completed"].copy()
                    completed_dates = pd.to_datetime(completed_rows.get("schedule_date"), errors="coerce")
                    completed_start = completed_dates.min()
                    completed_end = completed_dates.max()
                    completed_sources = []
                    if "schedule_source" in completed_rows.columns:
                        completed_sources = completed_rows["schedule_source"].fillna("Blank source").astype(str).value_counts().head(3).to_dict()
                    date_text = ""
                    if pd.notna(completed_start) and pd.notna(completed_end):
                        date_text = f" Dates: {completed_start.date()} through {completed_end.date()}."
                    source_text = ""
                    if completed_sources:
                        source_text = " Sources: " + ", ".join(f"{source}: {count}" for source, count in completed_sources.items()) + "."
                    st.warning(
                        f"{rec['completed_count']} completed store(s) are counted because this selected schedule plan contains rows for {selected_tech_name} with status `Completed`."
                        f"{date_text}{source_text} These are history records, not active scheduled work, and they do not block adding new active stores."
                    )
                detail_options = {
                    "Visible schedule rows": selected_scope,
                    "Assigned not scheduled": rec["assigned_not_scheduled"],
                    "Scheduled but no longer assigned": rec["scheduled_no_longer_assigned"],
                    "Assigned here but scheduled under another PMT": rec["assigned_scheduled_elsewhere"],
                    "Completed rows": rec["tech_completed"],
                    "Canceled / skipped rows": rec["tech_canceled"],
                }
                detail_choice = st.selectbox("View store details", list(detail_options.keys()), key="pmt_manage_overview_detail")
                detail_df = detail_options[detail_choice].copy()
                if detail_df.empty:
                    st.success("No stores in this detail view.")
                else:
                    display_cols = [
                        col for col in [
                            "schedule_date", "sequence_number", "technician", "assigned_technician", "store_number", "address", "city", "state", "status", "schedule_source", "scheduled_technician", "scheduled_date", "distance_from_home", "notes"
                        ] if col in detail_df.columns
                    ]
                    st.dataframe(detail_df[display_cols], use_container_width=True, hide_index=True)
                if show_map:
                    map_scope = selected_scope.copy()
                    if show_future_months and selected_month != "All months":
                        map_scope = filter_manage_scope(run_items, selected_employee, "All months", "Active")
                        map_scope = map_scope[map_scope["month_start"] >= selected_month].copy()
                    if map_scope.empty:
                        st.info("No rows match the current map filters.")
                    else:
                        render_store_map(
                            map_scope.sort_values(["month_start", "schedule_date", "sequence_number", "store_number"]),
                            color_by="month" if show_future_months or selected_month == "All months" else "status",
                            show_route_path=not show_future_months,
                            max_route_points=200,
                            static_preview=True,
                            height=560,
                        )

            with map_builder_tab:
                st.markdown("**Click-To-Build Route Map**")
                st.caption(
                    "Use this as a fast manual scheduling workspace. Click store dots inside the map without waiting for the page to reload, then copy the finished route into the box below."
                )
                if selected_month == "All months":
                    month_values = [value for value in manage_month_options(run_items) if value != "All months"]
                    if not month_values:
                        month_values = [month_start(date.today())]
                    route_month = st.selectbox(
                        "Route month",
                        month_values,
                        index=0,
                        format_func=month_label,
                        key=f"pmt_map_route_month_{selected_run}_{selected_employee}",
                    )
                else:
                    route_month = selected_month
                    st.caption(f"Route month: {month_label(route_month)}")
                layer_cols = st.columns([0.42, 0.18, 0.18, 0.22])
                extra_employee_options = [
                    int(value)
                    for value in tech_options["employee_id"].dropna().astype(int).tolist()
                    if int(value) != int(selected_employee)
                ]
                extra_layer_employee_ids = layer_cols[0].multiselect(
                    "Add PMT layers",
                    extra_employee_options,
                    format_func=lambda value: tech_options.set_index("employee_id").loc[value, "technician"],
                    key=f"pmt_map_route_extra_pmts_{selected_run}_{selected_employee}_{route_month}",
                )
                show_assigned_layer = layer_cols[1].checkbox("Assigned layer", value=True, key=f"pmt_map_route_assigned_layer_{selected_run}_{selected_employee}_{route_month}")
                show_existing_layer = layer_cols[2].checkbox("Existing route layer", value=True, key=f"pmt_map_route_existing_layer_{selected_run}_{selected_employee}_{route_month}")
                route_date = layer_cols[3].date_input(
                    "Route date",
                    value=first_workday(route_month, employee_id=int(selected_employee)),
                    key=f"pmt_map_route_date_{selected_run}_{selected_employee}_{route_month}",
                )
                route_employee_ids = [int(selected_employee)] + [int(value) for value in extra_layer_employee_ids]
                route_pool = pmt_route_builder_store_pool(run_items, route_employee_ids, route_month, run_id=selected_run)
                route_state_key = f"pmt_manual_map_route_{selected_run}_{selected_employee}_{route_month}"
                route_click_queue_key = f"{route_state_key}_click_queue"
                route_records = st.session_state.get(route_state_key, [])
                queued_route_records = st.session_state.get(route_click_queue_key, [])
                route_df = pd.DataFrame(route_records)

                command_cols = st.columns(4)
                if command_cols[0].button("Load Existing Route", key=f"pmt_map_load_existing_{selected_run}_{selected_employee}_{route_month}"):
                    existing_route = filter_manage_scope(run_items, selected_employee, route_month, "Active").sort_values(["schedule_date", "sequence_number", "store_number"]).copy()
                    if existing_route.empty:
                        st.info("No existing active route was found for this PMT/month.")
                    else:
                        existing_route["Proposed Stop"] = range(1, len(existing_route) + 1)
                        existing_route["Proposed Date"] = route_date
                        existing_route["Proposed Month"] = month_label(route_month)
                        existing_route["Manual or Auto-Filled"] = "Loaded existing route"
                        st.session_state[route_state_key] = existing_route.to_dict("records")
                        st.session_state[route_click_queue_key] = existing_route.to_dict("records")
                        st.rerun()
                if command_cols[1].button("Clear Proposed Route", key=f"pmt_map_clear_route_{selected_run}_{selected_employee}_{route_month}"):
                    st.session_state.pop(route_state_key, None)
                    st.session_state.pop(route_click_queue_key, None)
                    st.session_state.pop(f"{route_state_key}_last_click", None)
                    st.rerun()
                command_cols[2].metric("Generated Stops", len(route_records))
                command_cols[3].metric("Map Stores", len(route_pool))

                employee_layers = active_pmt_employee_summary()
                if not employee_layers.empty:
                    employee_layers = employee_layers[
                        pd.to_numeric(employee_layers["employee_id"], errors="coerce").fillna(-1).astype(int).isin(route_employee_ids)
                    ].copy()
                render_fast_pmt_route_picker_map(
                    route_pool,
                    employee_layers,
                    show_assigned_layer=show_assigned_layer,
                    show_existing_layer=show_existing_layer,
                    component_key=f"pmt_route_builder_map_{selected_run}_{selected_employee}_{route_month}",
                )
                st.caption("After clicking stores in the map, click Done / Copy Route in the map panel, paste the store numbers below, then generate the editable route list.")
                pasted_route_text = st.text_area(
                    "Paste route store numbers from the map",
                    placeholder="One store per line, or separated by commas/spaces",
                    height=110,
                    key=f"pmt_map_route_paste_{selected_run}_{selected_employee}_{route_month}",
                )
                if st.button("Generate / Refresh Route List From Pasted Route", type="primary", key=f"pmt_map_generate_route_list_{selected_run}_{selected_employee}_{route_month}"):
                    pasted_tokens = [token for token in re.split(r"[\s,;|]+", clean(pasted_route_text)) if token]
                    route_lookup_rows = []
                    route_pool_lookup = route_pool.copy()
                    route_pool_lookup["_store_keys"] = route_pool_lookup["store_number"].astype(str).apply(lambda value: set(store_number_keys(value)))
                    used_store_ids = set()
                    missing_tokens = []
                    for token in pasted_tokens:
                        token_keys = set(store_number_keys(token))
                        matches = route_pool_lookup[
                            route_pool_lookup["_store_keys"].apply(lambda values: bool(values & token_keys))
                            & ~route_pool_lookup["store_id"].astype(int).isin(used_store_ids)
                        ].copy()
                        if matches.empty:
                            missing_tokens.append(token)
                            continue
                        picked = matches.iloc[0].to_dict()
                        picked["Proposed Stop"] = len(route_lookup_rows) + 1
                        picked["Proposed Date"] = route_date
                        picked["Proposed Month"] = month_label(route_month)
                        picked["technician"] = selected_tech_name
                        picked["Manual or Auto-Filled"] = "Fast map selected"
                        route_lookup_rows.append(picked)
                        used_store_ids.add(int(picked["store_id"]))
                    st.session_state[route_state_key] = route_lookup_rows
                    st.session_state[route_click_queue_key] = route_lookup_rows
                    if missing_tokens:
                        st.warning(f"These pasted stores were not found on the selected map layer: {', '.join(missing_tokens[:20])}")
                    st.rerun()

                route_df = pd.DataFrame(st.session_state.get(route_state_key, []))
                if route_df.empty:
                    st.info("Click store dots in the map panel, click Done / Copy Route in the map, paste the route into the box, then generate the editable list.")
                else:
                    route_df = route_df.copy()
                    route_df["Remove"] = False
                    if "Proposed Stop" in route_df.columns:
                        route_df["Proposed Stop"] = pd.to_numeric(route_df["Proposed Stop"], errors="coerce")
                    else:
                        route_df["Proposed Stop"] = pd.NA
                    route_df["Proposed Stop"] = route_df["Proposed Stop"].where(route_df["Proposed Stop"].notna(), pd.Series(range(1, len(route_df) + 1), index=route_df.index)).astype(int)
                    route_df = route_df.sort_values(["Proposed Stop", "store_number"]).reset_index(drop=True)
                    route_editor_cols = [
                        "Remove", "Proposed Stop", "store_id", "store_number", "city", "state",
                        "assigned_technician", "scheduled_technician", "scheduled_date", "distance_from_home", "Manual or Auto-Filled",
                    ]
                    edited_route = st.data_editor(
                        route_df[[col for col in route_editor_cols if col in route_df.columns]],
                        use_container_width=True,
                        hide_index=True,
                        disabled=["store_id", "store_number", "city", "state", "assigned_technician", "scheduled_technician", "scheduled_date", "distance_from_home", "Manual or Auto-Filled"],
                        column_config={
                            "Remove": st.column_config.CheckboxColumn("Remove"),
                            "Proposed Stop": st.column_config.NumberColumn("Stop #", min_value=1, step=1),
                            "store_id": None,
                            "store_number": st.column_config.TextColumn("Store"),
                            "assigned_technician": st.column_config.TextColumn("Assigned PMT"),
                            "scheduled_technician": st.column_config.TextColumn("Currently Scheduled PMT"),
                            "scheduled_date": st.column_config.DateColumn("Current Scheduled Date"),
                            "distance_from_home": st.column_config.NumberColumn("Miles From Home", format="%.1f"),
                        },
                        key=f"pmt_map_route_editor_{selected_run}_{selected_employee}_{route_month}",
                    )
                    update_cols = st.columns([0.22, 0.26, 0.32, 0.2])
                    if update_cols[0].button("Update Route List", key=f"pmt_map_update_route_list_{selected_run}_{selected_employee}_{route_month}"):
                        edited_ids = edited_route.loc[~edited_route["Remove"].astype(bool), "store_id"].dropna().astype(int).tolist()
                        edited_stops = edited_route.loc[~edited_route["Remove"].astype(bool), ["store_id", "Proposed Stop"]].copy()
                        stop_lookup = dict(zip(edited_stops["store_id"].astype(int), pd.to_numeric(edited_stops["Proposed Stop"], errors="coerce").fillna(9999).astype(int)))
                        updated_route = route_df[pd.to_numeric(route_df["store_id"], errors="coerce").fillna(-1).astype(int).isin(edited_ids)].copy()
                        updated_route["Proposed Stop"] = updated_route["store_id"].astype(int).map(stop_lookup)
                        updated_route = updated_route.sort_values(["Proposed Stop", "store_number"]).reset_index(drop=True)
                        updated_route["Proposed Stop"] = range(1, len(updated_route) + 1)
                        updated_route["Proposed Date"] = route_date
                        updated_route["Proposed Month"] = month_label(route_month)
                        st.session_state[route_state_key] = updated_route.drop(columns=["Remove"], errors="ignore").to_dict("records")
                        st.rerun()
                    route_note = update_cols[1].text_input(
                        "Apply note",
                        value="Manual map route builder schedule update.",
                        key=f"pmt_map_route_note_{selected_run}_{selected_employee}_{route_month}",
                    )
                    apply_confirm = update_cols[2].checkbox(
                        "I reviewed this map-built route and want to apply it.",
                        key=f"pmt_map_route_apply_confirm_{selected_run}_{selected_employee}_{route_month}",
                    )
                    if update_cols[3].button(
                        "Apply Map Route",
                        type="primary",
                        disabled=not apply_confirm,
                        key=f"pmt_map_route_apply_{selected_run}_{selected_employee}_{route_month}",
                    ):
                        apply_df = route_df.drop(columns=["Remove"], errors="ignore").copy()
                        apply_df["Proposed Date"] = route_date
                        apply_df["Proposed Month"] = month_label(route_month)
                        apply_df["technician"] = selected_tech_name
                        apply_df = apply_df.sort_values(["Proposed Stop", "store_number"]).reset_index(drop=True)
                        apply_df["Proposed Stop"] = range(1, len(apply_df) + 1)
                        result = apply_pmt_manage_build_preview(selected_run, selected_employee, apply_df.to_dict("records"), route_note)
                        st.success(
                            f"Applied map route: saved {result['saved']} store(s), created {result['created']}, updated {result['updated']}, "
                            f"superseded/transferred {result['superseded']}, resequenced {result['resequenced_rows']} row(s)."
                        )
                        st.session_state.pop(route_state_key, None)
                        st.session_state.pop(f"{route_state_key}_last_click", None)
                        st.rerun()

            with build_tab:
                st.markdown("**Scheduling Method**")
                method_options = ["Manual First + Auto-Fill Remaining", "Manual Only"]
                selected_method = st.radio(
                    "Scheduling Method",
                    method_options,
                    horizontal=True,
                    key="pmt_manage_build_method",
                )
                if selected_method == "Manual First + Auto-Fill Remaining":
                    st.caption("Manually choose the first stores. The app then fills the remaining eligible stores from this PMT's home outward by miles from home.")
                else:
                    st.caption("Manually select every store to add. No remaining stores will be added automatically.")
                if selected_month == "All months":
                    add_month = month_start(date.today())
                    st.warning("Choose a specific month in Step 2 for the clearest add-store workflow. New stores will default to the current month until then.")
                else:
                    add_month = selected_month
                build_cols = st.columns(3)
                sort_choice = build_cols[0].selectbox("Suggested order", ["Closest to home first", "Farthest from home first", "Store number"], key="pmt_manage_build_sort")
                add_limit = build_cols[1].number_input("Show first", min_value=1, max_value=250, value=50, step=1, key="pmt_manage_build_limit")
                include_scheduled_review = build_cols[2].checkbox("Show stores already active in this schedule", value=False, key="pmt_manage_build_include_scheduled")
                all_assigned_stores = assigned_pmt_store_candidates(selected_employee, selected_run, include_scheduled=True)
                candidate_stores = assigned_pmt_store_candidates(selected_employee, selected_run, include_scheduled=include_scheduled_review)
                if all_assigned_stores.empty:
                    st.info("This PMT has no assigned active stores.")
                else:
                    all_assigned_stores = all_assigned_stores.copy()
                    all_assigned_stores["scheduled_count"] = pd.to_numeric(all_assigned_stores.get("scheduled_count", 0), errors="coerce").fillna(0).astype(int)
                    conflicts = all_assigned_stores[
                        (all_assigned_stores["scheduled_count"] > 0)
                        & (pd.to_numeric(all_assigned_stores.get("scheduled_employee_id"), errors="coerce").fillna(0).astype(int) != int(selected_employee))
                    ].copy()
                    available_to_add = int((all_assigned_stores["scheduled_count"] == 0).sum())
                    build_metric_cols = st.columns(4)
                    build_metric_cols[0].metric("Assigned stores", len(all_assigned_stores))
                    build_metric_cols[1].metric("Available to schedule", available_to_add)
                    build_metric_cols[2].metric("Already active", int((all_assigned_stores["scheduled_count"] > 0).sum()))
                    build_metric_cols[3].metric("Conflicts", len(conflicts))
                    if not conflicts.empty:
                        with st.expander("Stores assigned here but active under another PMT", expanded=True):
                            conflict_cols = ["store_number", "city", "state", "scheduled_technician", "scheduled_date", "distance_from_home"]
                            st.dataframe(conflicts[conflict_cols], use_container_width=True, hide_index=True)
                    move_assigned_conflicts = False
                    if not conflicts.empty:
                        move_assigned_conflicts = st.checkbox(
                            "Move all active scheduled stores that are assigned to this PMT off the other PMT",
                            value=selected_method == "Manual First + Auto-Fill Remaining",
                            key="pmt_manage_build_move_all_conflicts",
                        )
                    if move_assigned_conflicts and not conflicts.empty:
                        candidate_stores = pd.concat([candidate_stores, conflicts], ignore_index=True).drop_duplicates("store_id", keep="first")
                    if candidate_stores.empty:
                        st.info("No stores match the current add-store filters.")
                    else:
                        candidate_stores = candidate_stores.copy()
                        candidate_stores["already_scheduled"] = pd.to_numeric(candidate_stores.get("scheduled_count", 0), errors="coerce").fillna(0).astype(int) > 0
                        candidate_stores["scheduled_employee_id"] = pd.to_numeric(candidate_stores.get("scheduled_employee_id"), errors="coerce")
                        candidate_stores["scheduled_technician"] = candidate_stores.get("scheduled_technician", "").fillna("").astype(str)
                        if sort_choice == "Farthest from home first":
                            candidate_stores = candidate_stores.sort_values(["distance_from_home", "store_number"], ascending=[False, True], na_position="last")
                        elif sort_choice == "Store number":
                            candidate_stores = candidate_stores.sort_values("store_number")
                        else:
                            candidate_stores = candidate_stores.sort_values(["distance_from_home", "store_number"], ascending=[True, True], na_position="last")
                        bulk_cols = st.columns([0.45, 0.55])
                        bulk_options = candidate_stores["store_number"].astype(str).tolist()
                        bulk_selected = bulk_cols[0].multiselect("Select manual first stores", bulk_options, key="pmt_manage_build_bulk_select")
                        pasted_store_text = bulk_cols[1].text_area("Paste store numbers", placeholder="Separate stores with spaces, commas, or new lines", height=90, key="pmt_manage_build_paste")
                        pasted_keys = {key(value) for value in re.split(r"[\s,;|]+", clean(pasted_store_text)) if clean(value)}
                        selected_keys = {key(value) for value in bulk_selected} | pasted_keys
                        matched_precheck = candidate_stores[candidate_stores["store_number"].astype(str).apply(lambda value: key(value) in selected_keys)].copy()
                        base_view = candidate_stores.head(int(add_limit)).copy()
                        candidate_view = pd.concat([matched_precheck, base_view], ignore_index=True).drop_duplicates("store_id", keep="first")
                        candidate_view["Add"] = candidate_view["store_number"].astype(str).apply(lambda value: key(value) in selected_keys)
                        editor_columns = ["Add", "already_scheduled", "scheduled_employee_id", "scheduled_technician", "scheduled_date", "store_id", "store_number", "city", "state", "distance_from_home", "address"]
                        edited_candidates = st.data_editor(
                            candidate_view[editor_columns],
                            use_container_width=True,
                            hide_index=True,
                            disabled=["already_scheduled", "scheduled_employee_id", "scheduled_technician", "scheduled_date", "store_id", "store_number", "city", "state", "distance_from_home", "address"],
                            column_config={
                                "Add": st.column_config.CheckboxColumn("Add"),
                                "already_scheduled": st.column_config.CheckboxColumn("Already Active"),
                                "scheduled_employee_id": None,
                                "scheduled_technician": st.column_config.TextColumn("Active Under"),
                                "scheduled_date": st.column_config.DateColumn("Active Date"),
                                "store_id": None,
                                "distance_from_home": st.column_config.NumberColumn("Miles From Home", format="%.1f"),
                            },
                            key=f"pmt_manage_build_editor_{selected_run}_{selected_employee}_{selected_month}_{sort_choice}_{add_limit}",
                        )
                        selected_rows = edited_candidates[edited_candidates["Add"].astype(bool)].copy()
                        manual_store_ids = selected_rows["store_id"].dropna().astype(int).tolist()
                        fill_capacity = st.number_input("Stores per month", min_value=1, max_value=100, value=10, step=1, key="pmt_manage_build_capacity")
                        fill_end_options = [add_months(add_month, offset) for offset in range(0, 13)]
                        fill_end_month = st.selectbox("Fill through", fill_end_options, index=min(5, len(fill_end_options) - 1), format_func=month_label, key="pmt_manage_build_fill_end")
                        selected_set = set(manual_store_ids)
                        remaining_candidates = candidate_stores.loc[~candidate_stores["store_id"].astype(int).isin(selected_set)].copy()
                        if not move_assigned_conflicts:
                            remaining_candidates = remaining_candidates.loc[~remaining_candidates["already_scheduled"].astype(bool)].copy()
                        remaining_candidates["_home_distance_sort"] = pd.to_numeric(remaining_candidates.get("distance_from_home"), errors="coerce")
                        remaining_ids = (
                            remaining_candidates.sort_values(["_home_distance_sort", "store_number"], ascending=[True, True], na_position="last")["store_id"]
                            .dropna()
                            .astype(int)
                            .tolist()
                        )
                        fill_store_ids = manual_store_ids + (remaining_ids if selected_method == "Manual First + Auto-Fill Remaining" else [])
                        route_source = candidate_stores[candidate_stores["store_id"].astype(int).isin(fill_store_ids)].copy()
                        route_scheduled_employee_ids = pd.to_numeric(route_source.get("scheduled_employee_id"), errors="coerce").fillna(0).astype(int)
                        selected_conflict_ids = route_source.loc[
                            route_source["already_scheduled"].astype(bool)
                            & (route_scheduled_employee_ids != int(selected_employee)),
                            "store_id",
                        ].dropna().astype(int).tolist()
                        summary_cols = st.columns(4)
                        summary_cols[0].metric("Manual selected", len(manual_store_ids))
                        summary_cols[1].metric("Auto-fill", len(remaining_ids) if selected_method == "Manual First + Auto-Fill Remaining" else 0)
                        summary_cols[2].metric("Total result", len(fill_store_ids))
                        summary_cols[3].metric("Selected conflicts", len(selected_conflict_ids))
                        add_notes = st.text_input("Add note", value="Manually managed from PMT schedule workspace", key="pmt_manage_build_notes")
                        if selected_conflict_ids:
                            st.warning(
                                f"{len(selected_conflict_ids)} selected store(s) are currently active under another PMT. "
                                "The preview will show those stores moving to this PMT and the other PMT's remaining schedule."
                            )
                        if st.button("Preview Schedule Changes", type="primary", disabled=not fill_store_ids and not selected_conflict_ids, key="pmt_manage_build_preview_button"):
                            preview_source = route_source.copy()
                            preview_source["move_from_employee_id"] = preview_source["scheduled_employee_id"]
                            preview_source["move_from_technician"] = preview_source["scheduled_technician"]
                            order_lookup = {int(store_id): index for index, store_id in enumerate(fill_store_ids)}
                            preview_source["_proposed_order"] = preview_source["store_id"].astype(int).map(order_lookup)
                            preview_source = preview_source.sort_values("_proposed_order").drop(columns=["_proposed_order"], errors="ignore")
                            preview_source["Manual or Auto-Filled"] = preview_source.apply(
                                lambda row: (
                                    f"{'Manual' if int(row['store_id']) in selected_set else 'Auto-filled'} - move from {row.get('scheduled_technician') or 'another PMT'}"
                                    if bool(row.get("already_scheduled")) and scalar_int(row.get("scheduled_employee_id"), 0) != int(selected_employee)
                                    else ("Manual" if int(row["store_id"]) in selected_set else "Auto-filled")
                                ),
                                axis=1,
                            )
                            proposed_dates = []
                            proposed_month_labels = []
                            proposed_stops = []
                            kept_indices = []
                            cursor_month = add_month
                            month_stop = 0
                            for row_index in preview_source.index:
                                if selected_method == "Manual First + Auto-Fill Remaining" and month_stop >= int(fill_capacity):
                                    cursor_month = add_months(cursor_month, 1)
                                    month_stop = 0
                                if selected_method == "Manual First + Auto-Fill Remaining" and cursor_month > fill_end_month:
                                    continue
                                month_stop += 1
                                kept_indices.append(row_index)
                                proposed_month_labels.append(month_label(cursor_month))
                                proposed_dates.append(first_workday(cursor_month, employee_id=int(selected_employee)))
                                proposed_stops.append(month_stop)
                            preview_source = preview_source.loc[kept_indices].copy()
                            preview_source["Proposed Month"] = proposed_month_labels
                            preview_source["Proposed Date"] = proposed_dates
                            preview_source["Proposed Stop"] = proposed_stops
                            preview_source["technician"] = selected_tech_name
                            preview_source["schedule_date"] = preview_source["Proposed Date"]
                            preview_source["sequence_number"] = preview_source["Proposed Stop"]
                            preview_source["status"] = preview_source["Manual or Auto-Filled"]
                            preview_source = add_preview_leg_distances(preview_source)
                            st.session_state["pmt_manage_build_preview"] = preview_source.to_dict("records")
                            st.session_state["pmt_manage_build_preview_ids"] = fill_store_ids
                            st.session_state["pmt_manage_build_preview_conflict_ids"] = selected_conflict_ids
                            st.session_state["pmt_manage_build_preview_method"] = selected_method
                        preview_df = dataframe_from_session_records("pmt_manage_build_preview")
                        if not preview_df.empty:
                            preview_cols = ["technician", "store_number", "city", "state", "Proposed Month", "Proposed Date", "Proposed Stop", "Manual or Auto-Filled", "distance_from_home", "Distance From Previous Stop", "scheduled_technician"]
                            st.dataframe(preview_df[[col for col in preview_cols if col in preview_df.columns]], use_container_width=True, hide_index=True)
                            st.markdown("**Proposed route map**")
                            map_preview = preview_df.copy()
                            if {"latitude", "longitude"}.issubset(map_preview.columns):
                                render_store_map(
                                    map_preview,
                                    color_by="Manual or Auto-Filled",
                                    show_route_path=True,
                                    max_route_points=200,
                                    static_preview=True,
                                    height=560,
                                )
                            move_conflict_ids = st.session_state.get("pmt_manage_build_preview_conflict_ids", [])
                            if move_conflict_ids:
                                st.markdown("**Impact on other PMT schedules**")
                                moved_rows = preview_df[pd.to_numeric(preview_df.get("store_id"), errors="coerce").fillna(-1).astype(int).isin(move_conflict_ids)].copy()
                                if not moved_rows.empty:
                                    impact_cols = ["store_number", "city", "state", "move_from_technician", "scheduled_date", "Proposed Date", "Proposed Stop"]
                                    st.warning("These stores will be moved off the PMT currently shown in `move_from_technician` and into the proposed route above.")
                                    st.dataframe(moved_rows[[col for col in impact_cols if col in moved_rows.columns]], use_container_width=True, hide_index=True)
                                    for moved_employee_id in pd.to_numeric(moved_rows.get("move_from_employee_id"), errors="coerce").dropna().astype(int).unique().tolist():
                                        remaining_schedule = filter_manage_scope(run_items, moved_employee_id, selected_month, "Active")
                                        remaining_schedule = remaining_schedule[
                                            ~pd.to_numeric(remaining_schedule["store_id"], errors="coerce").fillna(-1).astype(int).isin(move_conflict_ids)
                                        ].copy()
                                        moved_name = moved_rows.loc[pd.to_numeric(moved_rows.get("move_from_employee_id"), errors="coerce").fillna(-1).astype(int) == moved_employee_id, "move_from_technician"].dropna()
                                        moved_name = moved_name.iloc[0] if not moved_name.empty else f"PMT #{moved_employee_id}"
                                        with st.expander(f"Updated active schedule for {moved_name} after move", expanded=False):
                                            if remaining_schedule.empty:
                                                st.info("No active stores remain for this PMT in the selected month/status context.")
                                            else:
                                                remaining_cols = ["schedule_date", "sequence_number", "store_number", "city", "state", "status"]
                                                st.dataframe(remaining_schedule[remaining_cols], use_container_width=True, hide_index=True)
                                                render_store_map(
                                                    remaining_schedule.sort_values(["schedule_date", "sequence_number", "store_number"]),
                                                    color_by="status",
                                                    show_route_path=True,
                                                    max_route_points=200,
                                                    static_preview=True,
                                                    height=420,
                                                )
                            confirm_apply = st.checkbox("I reviewed this preview and want to apply these schedule changes.", key="pmt_manage_build_confirm_apply")
                            if st.button("Apply Schedule Changes", type="primary", disabled=not confirm_apply, key="pmt_manage_build_apply"):
                                preview_records = st.session_state.get("pmt_manage_build_preview", [])
                                result = apply_pmt_manage_build_preview(selected_run, selected_employee, preview_records, add_notes)
                                st.success(
                                    f"Saved {result['saved']} preview store(s): created {result['created']}, updated {result['updated']}, "
                                    f"transferred/superseded {result['superseded']} conflicting row(s), and resequenced {result['resequenced_rows']} active row(s)."
                                )
                                st.session_state.pop("pmt_manage_build_preview", None)
                                st.session_state.pop("pmt_manage_build_preview_ids", None)
                                st.session_state.pop("pmt_manage_build_preview_conflict_ids", None)
                                st.session_state.pop("pmt_manage_build_preview_method", None)
                                st.rerun()

            with reorder_tab:
                reorder_scope = filter_manage_scope(run_items, selected_employee, selected_month, "Active").sort_values(["schedule_date", "sequence_number", "store_number"])
                if reorder_scope.empty:
                    st.info("No active schedule rows match the selected PMT and month.")
                else:
                    st.caption("To reorder, change the Stop # values below, or use Move one store. Choose the route position by store; the app will renumber the stops.")
                    move_cols = st.columns(3)
                    move_options = reorder_scope["schedule_item_id"].dropna().astype(int).tolist()
                    reorder_lookup = reorder_scope.set_index("schedule_item_id")
                    move_item_id = move_cols[0].selectbox(
                        "Move store",
                        move_options,
                        format_func=lambda value: f"Stop {int(reorder_lookup.loc[value, 'sequence_number'])} - Store {reorder_lookup.loc[value, 'store_number']}",
                        key=f"pmt_manage_move_store_{selected_run}_{selected_employee}_{selected_month}",
                    )
                    target_position_options = ["First stop"]
                    target_position_lookup = {"First stop": None}
                    target_position_options.extend(
                        [
                            f"After Stop {int(row['sequence_number'])} - Store {row['store_number']}"
                            for _, row in reorder_scope.iterrows()
                            if int(row["schedule_item_id"]) != int(move_item_id)
                        ]
                    )
                    for _, row in reorder_scope.iterrows():
                        if int(row["schedule_item_id"]) == int(move_item_id):
                            continue
                        target_position_lookup[f"After Stop {int(row['sequence_number'])} - Store {row['store_number']}"] = int(row["schedule_item_id"])
                    move_target_position = move_cols[1].selectbox(
                        "Move to",
                        target_position_options,
                        index=len(target_position_options) - 1,
                        key=f"pmt_manage_move_stop_{selected_run}_{selected_employee}_{selected_month}",
                    )
                    if move_cols[2].button("Preview Move", type="secondary", key=f"pmt_manage_preview_move_{selected_run}_{selected_employee}_{selected_month}"):
                        moved_order = reorder_scope.copy().reset_index(drop=True)
                        current_index = moved_order.index[moved_order["schedule_item_id"].astype(int) == int(move_item_id)].tolist()
                        if current_index:
                            moved_row = moved_order.loc[current_index[0]].copy()
                            moved_order = moved_order.drop(index=current_index[0]).reset_index(drop=True)
                            if move_target_position == "First stop":
                                insert_at = 0
                            else:
                                target_item_id = target_position_lookup.get(move_target_position)
                                target_matches = moved_order.index[moved_order["schedule_item_id"].astype(int) == int(target_item_id)].tolist() if target_item_id else []
                                insert_at = target_matches[0] + 1 if target_matches else len(moved_order)
                            top = moved_order.iloc[:insert_at]
                            bottom = moved_order.iloc[insert_at:]
                            moved_order = pd.concat([top, pd.DataFrame([moved_row]), bottom], ignore_index=True)
                            moved_order["sequence_number"] = range(1, len(moved_order) + 1)
                            st.session_state["pmt_manage_reorder_preview"] = moved_order.to_dict("records")
                    preview_order = dataframe_from_session_records("pmt_manage_reorder_preview")
                    if not preview_order.empty:
                        st.markdown("**Move preview**")
                        preview_cols = ["sequence_number", "store_number", "city", "state", "schedule_date", "status"]
                        st.dataframe(preview_order[[col for col in preview_cols if col in preview_order.columns]], use_container_width=True, hide_index=True)
                        if st.button("Apply Move Preview", type="primary", key=f"pmt_manage_apply_move_{selected_run}_{selected_employee}_{selected_month}"):
                            updated = save_manual_pmt_schedule_edits(preview_order)
                            st.session_state.pop("pmt_manage_reorder_preview", None)
                            st.success(f"Updated {updated} PMT schedule item(s).")
                            st.rerun()
                    reorder_view = reorder_scope[
                        ["schedule_item_id", "schedule_date", "sequence_number", "store_number", "city", "state", "status", "notes"]
                    ].rename(columns={"sequence_number": "Stop #", "store_number": "Store", "city": "City", "state": "State"})
                    reorder_view.insert(0, "Remove", False)
                    edited_order = st.data_editor(
                        reorder_view,
                        use_container_width=True,
                        hide_index=True,
                        disabled=["schedule_item_id", "Store", "City", "State"],
                        column_config={
                            "Remove": st.column_config.CheckboxColumn("Remove"),
                            "schedule_item_id": None,
                            "schedule_date": st.column_config.DateColumn("Schedule Date"),
                            "Stop #": st.column_config.NumberColumn("Stop #", min_value=1, step=1),
                            "status": st.column_config.SelectboxColumn("Status", options=["Scheduled", "Needs Rescheduled", "Rescheduled", "Rain Delay", "Not Completed", "Completed", "Skipped", "Cancelled"]),
                            "notes": st.column_config.TextColumn("Notes"),
                        },
                        key=f"pmt_manage_reorder_editor_{selected_run}_{selected_employee}_{selected_month}",
                    )
                    remove_ids = edited_order.loc[edited_order["Remove"].astype(bool), "schedule_item_id"].dropna().astype(int).tolist()
                    edited_without_remove = edited_order.drop(columns=["Remove"], errors="ignore")
                    edited_without_remove = edited_without_remove.rename(columns={"Stop #": "sequence_number"})
                    st.caption("Removing a store here removes it from this schedule only. It does not change PMT ownership assignment.")
                    action_cols = st.columns(3)
                    if action_cols[0].button("Save Date / Stop Changes", type="primary", key="pmt_manage_reorder_save"):
                        updated = save_manual_pmt_schedule_edits(edited_without_remove)
                        st.success(f"Updated {updated} PMT schedule item(s).")
                        st.rerun()
                    confirm_remove = action_cols[1].checkbox("Confirm remove selected", key="pmt_manage_reorder_confirm_remove")
                    if action_cols[1].button("Remove Selected From Schedule", type="secondary", disabled=not remove_ids or not confirm_remove, key="pmt_manage_reorder_remove"):
                        deleted = delete_pmt_schedule_items(remove_ids, "Removed from selected PMT route")
                        st.success(f"Removed {deleted} store(s) from this schedule. Store assignments were not changed.")
                        st.rerun()
                    action_cols[2].caption("Use the Carryover & Backlog tab for unfinished-work carryover.")

            with st.expander("Danger Zone - Delete Entire Schedule Run", expanded=False):
                st.warning("This marks the selected schedule plan as Deleted and removes its PMT schedule rows. This is separate from removing one store from a technician route.")
                st.write(f"Schedule: {run_row.get('run_name', '')}")
                st.write(f"Affected rows: {len(run_items)}")
                confirm_delete_text = st.text_input("Type DELETE to confirm", key=f"pmt_manage_delete_run_text_{selected_run}")
                confirm_delete_check = st.checkbox("I understand this deletes the entire selected schedule plan.", key=f"pmt_manage_delete_run_check_{selected_run}")
                if st.button("Delete Entire Schedule Plan", type="secondary", disabled=confirm_delete_text != "DELETE" or not confirm_delete_check, key=f"pmt_manage_delete_run_button_{selected_run}"):
                    deleted = delete_pmt_schedule_run(selected_run)
                    st.success(f"Deleted {deleted} PMT schedule item(s) from {run_row.get('run_name', '')}.")
                    st.rerun()


with tab_preview:
    section_header(
        "Before & After Preview",
        "Review pending reconciliation, rebuild, or manual-edit previews before applying major schedule changes.",
        "blue",
    )
    st.caption("Major repair tools show their own confirmation before applying. This tab collects currently staged previews and explains where to review them.")
    preview_sources = [
        ("Manual add/rebuild preview", dataframe_from_session_records("pmt_manage_build_preview")),
        ("Manual route move preview", dataframe_from_session_records("pmt_manage_reorder_preview")),
    ]
    visible_preview = False
    for preview_name, preview_df in preview_sources:
        if preview_df.empty:
            continue
        visible_preview = True
        with st.expander(preview_name, expanded=True):
            st.dataframe(preview_df, use_container_width=True, hide_index=True)
    if not visible_preview:
        st.info(
            "No staged preview is currently waiting. Use Territory Reconciliation for assignment changes, Rebuild / Balance for one PMT rebuilds, or Manual Edit for one-off changes."
        )


with tab_history:
    section_header(
        "Schedule History & Revisions",
        "View current active PMT schedules separately from historical snapshots and revisions.",
        "gray",
    )
    history_runs = safe_query(
        """
        select r.id, r.run_name, r.status, r.cycle_start, r.cycle_end, r.months,
               r.technician_count, r.store_count, r.unscheduled_count, r.created_by, r.created_at, r.notes
        from pmt_schedule_runs r
        order by r.created_at desc, r.id desc
        """,
        use_cache=False,
    )
    if history_runs.empty:
        st.info("No PMT schedule history is available yet.")
    else:
        history_runs = history_runs.copy()
        history_runs["Schedule Type"] = history_runs["status"].fillna("").astype(str).str.lower().apply(
            lambda value: "HISTORICAL / SNAPSHOT" if value == "snapshot" else "DELETED" if value == "deleted" else "CURRENT ACTIVE"
        )
        active_history = history_runs[history_runs["Schedule Type"] == "CURRENT ACTIVE"].copy()
        snapshot_history = history_runs[history_runs["Schedule Type"] != "CURRENT ACTIVE"].copy()
        st.markdown("**Current Active Schedules**")
        if active_history.empty:
            st.info("No current active PMT schedules found.")
        else:
            st.dataframe(
                active_history[["id", "run_name", "Schedule Type", "cycle_start", "cycle_end", "technician_count", "store_count", "unscheduled_count", "created_at", "notes"]],
                use_container_width=True,
                hide_index=True,
            )
        st.markdown("**Historical / Snapshot Schedules**")
        if snapshot_history.empty:
            st.info("No historical snapshots found.")
        else:
            st.dataframe(
                snapshot_history[["id", "run_name", "Schedule Type", "cycle_start", "cycle_end", "technician_count", "store_count", "unscheduled_count", "created_at", "notes"]],
                use_container_width=True,
                hide_index=True,
            )
            st.caption("Snapshots are kept for audit/history and do not appear in the normal Export schedule selector.")


with tab_export:
    section_header("Export Step 1: Export PMT Schedule", "Download full-team or individual PMT schedules from the current draft or a published PMT schedule run.", "green")
    latest_export_draft = pd.DataFrame(st.session_state.get("pmt_schedule_draft", []))
    _export_runs = safe_query(
        """
        select r.id, r.run_name, r.created_at, r.cycle_start, r.cycle_end, r.months, r.technician_count,
               r.store_count, r.unscheduled_count, r.status
        from pmt_schedule_runs r
        order by r.created_at desc, r.id desc
        """
    )
    export_source_options = []
    if not latest_export_draft.empty:
        export_source_options.append("Current Draft Schedule")
    normal_export_runs = _export_runs[
        ~_export_runs["status"].fillna("").astype(str).str.lower().str.strip().isin(["deleted", "snapshot"])
    ].copy() if not _export_runs.empty else pd.DataFrame()
    if not normal_export_runs.empty:
        export_source_options.append("Published PMT Schedule Run")

    if not export_source_options:
        st.info("Generate a PMT draft or publish a PMT schedule run, then export buttons will appear here.")
    else:
        default_source = "Published PMT Schedule Run" if "Published PMT Schedule Run" in export_source_options else export_source_options[0]
        export_source = st.radio(
            "Export source",
            export_source_options,
            horizontal=True,
            index=export_source_options.index(default_source),
            key="pmt_export_source",
        )
        if export_source == "Current Draft Schedule":
            render_pmt_export_controls(latest_export_draft, "pmt_bottom_export_draft")
        else:
            run_options = normal_export_runs["id"].tolist()
            selected_export_run = st.selectbox(
                "Published PMT schedule run",
                run_options,
                format_func=lambda value: f"#{value} - {normal_export_runs.set_index('id').loc[value, 'run_name']}",
                key="pmt_bottom_export_run",
            )
            export_run_row = normal_export_runs.set_index("id").loc[selected_export_run]
            export_raw_items = pmt_manage_run_items(selected_export_run)
            export_cycle_start = scalar_date(export_run_row.get("cycle_start"))
            export_cycle_end = scalar_date(export_run_row.get("cycle_end"))
            export_run_items, _export_outside_items = split_run_items_by_period(export_raw_items, export_cycle_start, export_cycle_end)
            export_conflicts = pmt_schedule_conflicts(export_run_items)
            if not export_conflicts.empty:
                st.error(
                    f"Export blocked: duplicate or wrong-technician active PMT schedule assignments remain for "
                    f"{distinct_store_count(export_conflicts)} store(s). Open Manage & Fix Schedule -> Manual Edit and resolve the conflicts before exporting."
                )
                export_conflict_view = export_conflicts[
                    ["store_number", "assigned_technician", "technician", "month", "schedule_date", "sequence_number", "status", "conflict_type"]
                ].copy()
                st.dataframe(export_conflict_view, use_container_width=True, hide_index=True)
            else:
                published_export_draft = published_pmt_run_export_draft(selected_export_run)
                render_pmt_export_controls(published_export_draft, f"pmt_bottom_export_run_{selected_export_run}")

        st.divider()
        st.markdown("**Reconciliation Schedule Exports**")
        st.caption(
            "Use this after PMT reconciliation. This is separate from the normal schedule export and includes the old snapshot, "
            "the new normal schedule from the current month forward, month-by-month tabs, changed rows, color legend, and reconciliation change columns."
        )
        if _export_runs.empty:
            st.info("No PMT schedule runs are available for reconciliation exports.")
        else:
            reconciled_runs = _export_runs[
                ~_export_runs["status"].fillna("").astype(str).str.lower().eq("snapshot")
            ].copy()
            snapshot_runs = _export_runs[
                _export_runs["status"].fillna("").astype(str).str.lower().eq("snapshot")
            ].copy()
            if reconciled_runs.empty:
                st.info("No active PMT schedule run is available for the new schedule export.")
            else:
                recon_export_cols = st.columns([0.5, 0.5])
                selected_recon_export_run = recon_export_cols[0].selectbox(
                    "New / updated schedule run",
                    reconciled_runs["id"].tolist(),
                    format_func=lambda value: f"#{value} - {reconciled_runs.set_index('id').loc[value, 'run_name']}",
                    key="pmt_reconciliation_export_new_run",
                )
                inferred_snapshot_ids = []
                if not snapshot_runs.empty:
                    notes_by_id = snapshot_runs.set_index("id")["run_name"].fillna("").astype(str).to_dict()
                    snapshot_notes = safe_query(
                        """
                        select id, run_name, notes
                        from pmt_schedule_runs
                        where lower(coalesce(status, '')) = 'snapshot'
                        order by created_at desc, id desc
                        """,
                        use_cache=False,
                    )
                    if not snapshot_notes.empty:
                        snapshot_notes["notes_text"] = snapshot_notes["notes"].fillna("").astype(str)
                        inferred_snapshot_ids = snapshot_notes[
                            snapshot_notes["notes_text"].str.contains(f"Source run #{int(selected_recon_export_run)}", regex=False)
                        ]["id"].tolist()
                        notes_by_id.update(snapshot_notes.set_index("id")["run_name"].fillna("").astype(str).to_dict())
                snapshot_options = [None] + (inferred_snapshot_ids if inferred_snapshot_ids else (snapshot_runs["id"].tolist() if not snapshot_runs.empty else []))
                selected_old_snapshot_run = recon_export_cols[1].selectbox(
                    "Old schedule snapshot",
                    snapshot_options,
                    format_func=lambda value: "No old snapshot selected" if value is None else f"#{value} - {notes_by_id.get(value, 'Snapshot')}",
                    key="pmt_reconciliation_export_old_snapshot",
                )
                selected_recon_run_row = reconciled_runs.set_index("id").loc[selected_recon_export_run]
                recon_cycle_start = scalar_date(selected_recon_run_row.get("cycle_start"))
                recon_cycle_end = scalar_date(selected_recon_run_row.get("cycle_end"))
                new_schedule_items = current_month_forward_schedule_rows(
                    pmt_manage_run_items(selected_recon_export_run),
                    recon_cycle_start,
                    recon_cycle_end,
                )
                old_schedule_items = current_month_forward_schedule_rows(
                    pmt_manage_run_items(selected_old_snapshot_run),
                    recon_cycle_start,
                    recon_cycle_end,
                ) if selected_old_snapshot_run else pd.DataFrame()
                recon_export_metrics = st.columns(4)
                changed_preview = prepare_schedule_export_view(
                    new_schedule_items,
                    old_lookup=old_schedule_technician_lookup(old_schedule_items),
                )
                changed_count = int(changed_preview.get("Change Type", pd.Series(dtype=str)).astype(str).ne("Unchanged").sum()) if not changed_preview.empty else 0
                recon_export_metrics[0].metric("New Schedule Rows", len(new_schedule_items))
                recon_export_metrics[1].metric("Changed Rows", changed_count)
                recon_export_metrics[2].metric("Old Snapshot Rows", len(old_schedule_items))
                recon_export_metrics[3].metric("Snapshot Found", "Yes" if selected_old_snapshot_run else "No")
                st.download_button(
                    "Export Reconciled Schedule Package",
                    data=reconciliation_schedule_export_workbook_bytes(
                        new_schedule_items,
                        old_schedule_items,
                        run_name=selected_recon_run_row["run_name"],
                        snapshot_name="" if selected_old_snapshot_run is None else notes_by_id.get(selected_old_snapshot_run, "Snapshot"),
                    ),
                    file_name=f"pmt_reconciled_schedule_package_run_{selected_recon_export_run}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    disabled=new_schedule_items.empty,
                    key="pmt_reconciliation_schedule_package_export",
                )
